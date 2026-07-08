import os
import re
import json
import uuid
import math
import unicodedata
import difflib
import copy
import shutil
import sys
import subprocess
import time
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser, simpledialog
from PIL import Image, ImageTk, ImageOps, ImageDraw, ImageFont
import pillow_heif
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import gpxpy
import tkintermapview
import piexif

pillow_heif.register_heif_opener()


class BisseManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestionnaire de Bisses - Édition Pro")
        self.root.geometry("1580x920")
        self.root.minsize(1280, 760)

        # Données globales du logiciel
        # Ces fichiers ne décrivent pas un bisse précis : ils restent donc
        # à côté de gestion_bisses.py, dans Gestion_Bisses_Data/.
        self.app_folder = self.get_application_folder()
        self.app_data_folder = os.path.join(self.app_folder, "Gestion_Bisses_Data")
        os.makedirs(self.app_data_folder, exist_ok=True)

        # Référentiel global des catégories de segments.
        self.segment_categories_file = os.path.join(
            self.app_data_folder,
            "segment_categories.json"
        )
        self.global_segment_categories = []
        self.settings_window = None
        self.segment_category_manager_tree = None
        self.segment_category_manager_usage = {}

        # Assistant de nettoyage / migration des catégories.
        self.category_cleanup_window = None
        self.category_cleanup_tree = None
        self.category_cleanup_inventory = {}
        self.category_cleanup_plan = {}
        self.category_cleanup_catalogues = []
        self.category_cleanup_target_var = None
        self.category_cleanup_target_lookup = {}
        self.category_cleanup_extra_root = ""

        self.ensure_global_segment_categories_file()

        # Dossier actif
        self.base_folder = ""
        self.photos_folder = ""
        self.manual_photos_folder = ""
        self.export_folder = ""
        self.catalog_path = ""
        self.local_catalog_path = ""
        self.current_project_id = ""
        self.current_project_dir = ""
        self.current_project_catalog_path = ""
        self.catalog_data = []
        self.catalog_container = None
        self.gpx_file = None
        self.gpx_files = []
        self.gpx_sync_max_gap_minutes = 30
        self.write_exif_after_gpx_sync = False
        self.gpx_folder = ""

        # Données cartographiques du bisse
        self.map_trace_paths = []

        # Fonds Swisstopo : mode automatique inspiré de map.geo.admin.
        # Selon le niveau de zoom du widget, on bascule entre :
        # - carte standard mixte pour les vues larges ;
        # - carte nationale 1:25'000 pour le travail local ;
        # - carte nationale 1:10'000 pour le travail très détaillé.
        self.swisstopo_auto_enabled = True
        self.swisstopo_last_layer = {"photo": None, "gpx": None}
        self.swisstopo_auto_after_id = {"photo": None, "gpx": None}
        self.show_trace_ciel_var = tk.BooleanVar(value=True)
        self.show_trace_canalise_var = tk.BooleanVar(value=True)
        self.show_trace_abandonne_var = tk.BooleanVar(value=True)
        self.show_trace_topo_var = tk.BooleanVar(value=False)
        self.show_trace_inconnu_var = tk.BooleanVar(value=False)
        self.show_photos_on_map_var = tk.BooleanVar(value=True)

        # Icônes de repères photo numérotés créées à la volée.
        # Le cache évite de recréer les mêmes pastilles à chaque rafraîchissement.
        self.photo_marker_icon_cache = {}

        # Fuseau horaire des photos
        self.photo_timezone_name = "Europe/Zurich"

        # Carte / édition
        self.map_widget = None
        self.map_markers = []
        self.geolocated_photos = []
        self.current_photo = None

        # Couche photo commune aux cartes de l'atelier Photos et de l'atelier GPX.
        # Les clusters sont uniquement visuels : les coordonnées réelles ne changent jamais.
        self.photo_cluster_icon_cache = {}
        self.photo_discrete_icon_cache = {}
        self.photo_anchor_icon_cache = {}
        self.photo_layer_after_id = {"photo": None, "gpx": None}
        self.photo_layer_last_signature = {"photo": None, "gpx": None}
        self.photo_layer_clusters = {"photo": [], "gpx": []}
        self.photo_spider_state = {"photo": None, "gpx": None}
        self.photo_spider_paths = {"photo": [], "gpx": []}
        self.photo_cluster_threshold_px = 42
        self.photo_zoom_animation_token = 0
        # Les commandes de marqueur et de carte peuvent être déclenchées par
        # le même clic dans tkintermapview. Ce garde empêche le clic de carte
        # de refermer immédiatement un déploiement tout juste ouvert.
        self.photo_map_click_guard_until = {"photo": 0.0, "gpx": 0.0}

        # Visionneuse intégrée
        self.viewer_canvas = None
        self.viewer_original_image = None
        self.viewer_display_image = None
        self.viewer_zoom = 1.0
        self.viewer_image_on_canvas = None
        self.viewer_info_var = tk.StringVar(value="")
        self.viewer_zoom_var = tk.StringVar(value="Zoom : —")

        # Variables UI métadonnées
        self.photo_title_entry = None
        self.photo_desc_text = None
        self.photo_filename_var = tk.StringVar(value="")
        self.photo_meta_var = tk.StringVar(value="")
        self.photo_coords_var = tk.StringVar(value="")
        self.photo_status_var = tk.StringVar(value="")
        self.photo_index_var = tk.StringVar(value="")
        self.platform_selected_var = tk.BooleanVar(value=False)
        self.platform_order_var = tk.IntVar(value=0)

        # Publication / export plateforme
        self.publication_tree = None
        self.publication_status_var = tk.StringVar(value="")
        self.last_platform_export_root = ""

        # Renommage
        self.rename_tree = None
        self.rename_prefix_var = tk.StringVar(value="")
        self.rename_year_var = tk.StringVar(value=str(datetime.now().year))
        self.rename_start_var = tk.IntVar(value=1)
        self.rename_plan = []

        # Atelier GPX restructuré
        self.gpx_editor_map = None
        self.gpx_editor_paths = []
        self.gpx_editor_markers = []
        self.gpx_editor_photo_markers = []
        self.gpx_source_tree = None
        self.gpx_segment_tree = None
        self.gpx_category_tree = None
        self.gpx_category_combo = None
        self.gpx_bicolor_a_combo = None
        self.gpx_bicolor_b_combo = None
        self.gpx_advanced_options_container = None
        self.gpx_advanced_options_visible = False
        self.gpx_display_mode_var = tk.StringVar(value="single")
        self.gpx_workshop_status_var = tk.StringVar(value="")
        self.gpx_workshop_selected_source_var = tk.StringVar(value="Aucune branche sélectionnée")
        self.gpx_workshop_selected_segment_var = tk.StringVar(value="Aucun segment sélectionné")
        self.gpx_workshop_show_photos_var = tk.BooleanVar(value=False)
        self.gpx_photo_display_mode_var = tk.StringVar(value="hidden")
        self.gpx_photo_mode_before_edit = None
        self.gpx_workshop_photos = []
        self.gpx_workshop_active = False
        self.gpx_workshop_show_endpoints_var = tk.BooleanVar(value=True)
        self.gpx_endpoint_toggle_button = None
        self.gpx_endpoint_icon_cache = {}
        self.gpx_workshop_selected_source_id = None
        self.gpx_workshop_click_mode = None
        self.gpx_workshop_pending_segment_id = None
        self.gpx_workshop_undo_stack = []
        self.gpx_workshop_redo_stack = []

        # Visionneuse flottante unique de l'atelier GPX.
        self.gpx_photo_viewer_window = None
        self.gpx_photo_viewer_canvas = None
        self.gpx_photo_viewer_image = None
        self.gpx_photo_viewer_current_photo = None
        self.gpx_photo_viewer_index = -1
        self.gpx_photo_viewer_geometry = ""
        self.gpx_photo_viewer_pil_cache = {}
        self.gpx_photo_viewer_index_var = tk.StringVar(value="")
        self.gpx_photo_viewer_filename_var = tk.StringVar(value="")
        self.gpx_photo_viewer_meta_var = tk.StringVar(value="")
        self.gpx_photo_viewer_integrated_var = tk.BooleanVar(value=False)
        self.gpx_map_viewer_paned = None
        self.gpx_map_holder = None
        self.gpx_photo_integrated_frame = None
        self.gpx_photo_integrated_canvas = None

        style = ttk.Style()
        style.theme_use("clam")

        # Structure globale robuste :
        # - en-tête fixe ;
        # - zone centrale extensible ;
        # - journal toujours réservé en bas, repliable si nécessaire ;
        # - barre de progression fixe.
        root.grid_columnconfigure(0, weight=1)
        root.grid_rowconfigure(2, weight=1)

        header = tk.Frame(root)
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(8, 3))

        tk.Label(
            header,
            text="🏔️ Gestionnaire de Bisses",
            font=("Arial", 20, "bold")
        ).pack(side="left")

        tk.Button(
            header,
            text="📂 Ouvrir un dossier bisse",
            command=self.select_base_folder
        ).pack(side="right", padx=5)

        tk.Button(
            header,
            text="🏠 Mes bisses",
            command=self.show_workspace_home
        ).pack(side="right", padx=5)

        tk.Button(
            header,
            text="⚙️ Paramètres",
            command=self.show_settings_dialog
        ).pack(side="right", padx=5)

        self.status_header = tk.Label(
            root,
            text="En attente d'un dossier...",
            fg="gray",
            font=("Arial", 12)
        )
        self.status_header.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))

        self.main_frame = tk.Frame(root)
        self.main_frame.grid(row=2, column=0, sticky="nsew", padx=16, pady=(4, 6))

        self.log_shell = tk.Frame(root)
        self.log_shell.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 6))
        self.log_shell.grid_columnconfigure(0, weight=1)

        log_header = tk.Frame(self.log_shell)
        log_header.grid(row=0, column=0, sticky="ew")
        log_header.grid_columnconfigure(0, weight=1)

        tk.Label(
            log_header,
            text="Journal d'activité",
            font=("Arial", 10, "bold")
        ).grid(row=0, column=0, sticky="w")

        self.log_visible_var = tk.BooleanVar(value=False)
        self.log_toggle_button = tk.Button(
            log_header,
            text="▼ Afficher le journal",
            command=self.toggle_journal
        )
        self.log_toggle_button.grid(row=0, column=1, sticky="e")

        self.log_content_frame = tk.Frame(self.log_shell)
        self.log_content_frame.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self.log_content_frame.grid_columnconfigure(0, weight=1)

        self.log_text = tk.Text(
            self.log_content_frame,
            height=5,
            width=150,
            state="disabled",
            bg="#f4f4f4",
            font=("Consolas", 9)
        )
        self.log_text.grid(row=0, column=0, sticky="ew")
        self.log_content_frame.grid_remove()

        self.progress = ttk.Progressbar(
            root,
            orient="horizontal",
            length=900,
            mode="determinate"
        )
        self.progress.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 5))

        # Navigation photo globale au clavier. Les champs de texte gardent
        # naturellement l'usage de leurs flèches.
        self.root.bind("<Left>", self.handle_photo_navigation_key, add="+")
        self.root.bind("<Right>", self.handle_photo_navigation_key, add="+")
        self.root.bind("<Escape>", self.handle_global_escape_key, add="+")

        self.create_welcome_screen()

    # ============================================================
    # OUTILS GÉNÉRAUX
    # ============================================================

    def log(self, message):
        self.log_text.config(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.root.update_idletasks()

    def clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.config(state="disabled")

    def toggle_journal(self):
        """
        Réduit / réaffiche le journal sans que la zone centrale ne fasse disparaître
        durablement les logs. Cela stabilise les modes à plusieurs panneaux.
        """
        visible = self.log_visible_var.get()
        if visible:
            self.log_content_frame.grid_remove()
            self.log_toggle_button.config(text="▼ Afficher le journal")
            self.log_visible_var.set(False)
        else:
            self.log_content_frame.grid()
            self.log_toggle_button.config(text="▲ Réduire le journal")
            self.log_visible_var.set(True)

    def clear_main_frame(self):
        self.stop_swisstopo_auto_watch("photo")
        self.stop_swisstopo_auto_watch("gpx")
        self.stop_photo_layer_watch("photo")
        self.stop_photo_layer_watch("gpx")

        try:
            self.cancel_gpx_cut_mode(silent=True)
        except Exception:
            pass

        try:
            self.close_gpx_photo_viewer(destroy=True)
        except Exception:
            pass

        for widget in self.main_frame.winfo_children():
            widget.destroy()

        self.gpx_workshop_active = False
        self.map_widget = None
        self.gpx_editor_map = None
        self.viewer_canvas = None
        self.viewer_original_image = None
        self.viewer_display_image = None
        self.viewer_image_on_canvas = None
        self.map_markers = []
        self.map_trace_paths = []
        self.gpx_editor_photo_markers = []
        self.photo_spider_paths = {"photo": [], "gpx": []}
        self.photo_spider_state = {"photo": None, "gpx": None}
        self.photo_layer_last_signature = {"photo": None, "gpx": None}
        self.current_photo = None
        self.gpx_map_viewer_paned = None
        self.gpx_map_holder = None
        self.gpx_photo_integrated_frame = None
        self.gpx_photo_integrated_canvas = None

    def make_scrollable_page(self, padx=0, pady=0, bind_mousewheel=True):
        """
        Crée une page principale avec défilement vertical global.

        À utiliser pour les pages administratives / tableaux de bord :
        - accueil Mes bisses ;
        - bibliothèque ;
        - publication ;
        - tableau de bord du bisse.

        On évite volontairement de l'utiliser pour les ateliers carte,
        où la molette sert souvent au zoom/déplacement.
        """
        container = tk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)

        content = tk.Frame(canvas, padx=padx, pady=pady)
        inner = canvas.create_window((0, 0), window=content, anchor="nw")

        def on_content_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            # Le contenu prend toujours la largeur visible du canvas.
            canvas.itemconfigure(inner, width=event.width)

        content.bind("<Configure>", on_content_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        if bind_mousewheel:
            def on_mousewheel(event):
                try:
                    if getattr(event, "num", None) == 4:
                        canvas.yview_scroll(-3, "units")
                    elif getattr(event, "num", None) == 5:
                        canvas.yview_scroll(3, "units")
                    else:
                        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except Exception:
                    pass
                return "break"

            def bind_wheel(_event=None):
                canvas.bind_all("<MouseWheel>", on_mousewheel)
                canvas.bind_all("<Button-4>", on_mousewheel)
                canvas.bind_all("<Button-5>", on_mousewheel)

            def unbind_wheel(_event=None):
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")

            canvas.bind("<Enter>", bind_wheel)
            canvas.bind("<Leave>", unbind_wheel)
            content.bind("<Enter>", bind_wheel)
            content.bind("<Leave>", unbind_wheel)

        # Petites références utiles pour debug/ajustements futurs.
        content._scroll_canvas = canvas
        content._scroll_container = container
        content._scrollbar = scrollbar
        return content

    def create_welcome_screen(self):
        """
        Page d'accueil : liste des bisses de travail déjà ouverts.
        """
        self.show_workspace_home()


    def get_application_folder(self):
        """
        Dossier où se trouve le script principal.

        Tous les fichiers globaux du logiciel sont regroupés dans
        Gestion_Bisses_Data/ à côté de gestion_bisses.py, au lieu d'être
        dispersés dans les dossiers source des bisses.
        """
        try:
            if "__file__" in globals():
                return os.path.dirname(os.path.abspath(__file__))
        except Exception:
            pass
        try:
            return os.getcwd()
        except Exception:
            return "."

    def get_app_data_path(self, filename):
        """
        Chemin d'un fichier global du logiciel.
        """
        folder = getattr(self, "app_data_folder", "")
        if not folder:
            folder = os.path.join(self.get_application_folder(), "Gestion_Bisses_Data")
            self.app_data_folder = folder
        os.makedirs(folder, exist_ok=True)
        return os.path.join(folder, filename)


    # ============================================================
    # PROJETS PORTABLES DANS Gestion_Bisses_Data
    # ============================================================

    def get_projects_root(self):
        path = os.path.join(self.app_data_folder, "projects")
        os.makedirs(path, exist_ok=True)
        return path

    def get_project_paths(self, project_id):
        project_id = self.slugify(project_id or "bisse")
        project_dir = os.path.join(self.get_projects_root(), project_id)
        return {
            "dir": project_dir,
            "project": os.path.join(project_dir, "project.json"),
            "catalogue": os.path.join(project_dir, "catalogue.json"),
            "backups": os.path.join(project_dir, "backups")
        }

    def read_json_file_safe(self, path, default=None):
        if default is None:
            default = {}
        if not path or not os.path.exists(path):
            return copy.deepcopy(default)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data
        except Exception:
            return copy.deepcopy(default)

    def write_json_file_safe(self, path, data):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    def normalize_json_for_compare(self, data):
        try:
            return json.dumps(data, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            return str(data)

    def json_data_equal(self, a, b):
        return self.normalize_json_for_compare(a) == self.normalize_json_for_compare(b)

    def parse_catalog_timestamp(self, container, fallback_path=None):
        candidates = []
        if isinstance(container, dict):
            candidates.append(container.get("updated_at"))
            project = container.get("project") if isinstance(container.get("project"), dict) else {}
            candidates.append(project.get("updated_at"))
            platform_export = container.get("platform_export") if isinstance(container.get("platform_export"), dict) else {}
            candidates.append(platform_export.get("last_export_at"))

        for value in candidates:
            if not value:
                continue
            try:
                return datetime.fromisoformat(str(value).replace("Z", "+00:00")).timestamp()
            except Exception:
                pass

        try:
            if fallback_path and os.path.exists(fallback_path):
                return os.path.getmtime(fallback_path)
        except Exception:
            pass
        return 0

    def backup_catalog_file(self, project_id, path, label):
        if not project_id or not path or not os.path.exists(path):
            return ""
        try:
            paths = self.get_project_paths(project_id)
            os.makedirs(paths["backups"], exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_label = self.slugify(label)
            backup_path = os.path.join(paths["backups"], f"{safe_label}_{stamp}.json")
            shutil.copy2(path, backup_path)
            return backup_path
        except Exception as exc:
            self.log(f"⚠️ Sauvegarde de sécurité impossible : {exc}")
            return ""

    def read_catalog_container_from_path(self, catalog_path, folder=None):
        """
        Lit un catalogue depuis un chemin arbitraire en réutilisant la normalisation
        existante de read_catalog_container(), sans changer durablement l'état actif.
        """
        previous_base = self.base_folder
        previous_catalog_path = self.catalog_path
        previous_container = self.catalog_container
        previous_data = self.catalog_data

        try:
            self.base_folder = folder or previous_base
            self.catalog_path = catalog_path
            container = self.read_catalog_container()
            return copy.deepcopy(container)
        finally:
            self.base_folder = previous_base
            self.catalog_path = previous_catalog_path
            self.catalog_container = previous_container
            self.catalog_data = previous_data

    def write_catalog_container_to_path(self, catalog_path, container, folder=None, update_timestamp=False):
        """
        Écrit un catalogue vers un chemin arbitraire.
        Utilisé pour synchroniser la copie Data et la copie locale.
        """
        if not catalog_path:
            return
        os.makedirs(os.path.dirname(catalog_path), exist_ok=True)

        data = copy.deepcopy(container) if isinstance(container, dict) else self.empty_catalog_container()
        data.setdefault("catalogue_version", 3)
        data.setdefault("schema_version", "0.2-local")
        data.setdefault("project", {})

        if folder:
            data["project"]["source_folder"] = folder
        else:
            data["project"].setdefault("source_folder", "")

        if update_timestamp or not data["project"].get("updated_at"):
            data["project"]["updated_at"] = datetime.now().isoformat(timespec="seconds")

        info = data.setdefault("bisse_info", self.default_bisse_info())
        title = info.get("title") or data["project"].get("title") or data["project"].get("bisse_name") or (os.path.basename(folder) if folder else "")
        slug = info.get("slug") or self.slugify(title)
        info["title"] = title
        info["slug"] = slug
        data["project"]["bisse_name"] = title
        data["project"]["title"] = title

        with open(catalog_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    def project_id_from_container_or_folder(self, folder=None, container=None):
        info = {}
        project = {}
        if isinstance(container, dict):
            info = container.get("bisse_info", {}) or {}
            project = container.get("project", {}) or {}

        folder_slug = self.slugify(os.path.basename(folder or ""))
        candidates = [
            info.get("slug", ""),
            self.slugify(info.get("title", "")),
            self.slugify(project.get("title", "")),
            self.slugify(project.get("bisse_name", "")),
            folder_slug
        ]

        for candidate in candidates:
            candidate = self.slugify(candidate)
            if candidate and not self.is_generic_project_id(candidate):
                return candidate

        return folder_slug or "bisse"

    def read_project_record(self, project_id):
        paths = self.get_project_paths(project_id)
        record = self.read_json_file_safe(paths["project"], {})
        if not isinstance(record, dict):
            record = {}
        record.setdefault("schema_version", "0.1")
        record.setdefault("project_id", self.slugify(project_id))
        record.setdefault("title", "")
        record.setdefault("slug", record.get("project_id", ""))
        record.setdefault("linked_folder", "")
        record.setdefault("linked_folder_status", "unknown")
        record.setdefault("created_at", datetime.now().isoformat(timespec="seconds"))
        record.setdefault("updated_at", None)
        record.setdefault("last_synced_at", None)
        record.setdefault("hidden_from_workspace", False)
        return record

    def write_project_record(self, project_id, record):
        paths = self.get_project_paths(project_id)
        os.makedirs(paths["dir"], exist_ok=True)
        record["project_id"] = self.slugify(project_id)
        record["updated_at"] = datetime.now().isoformat(timespec="seconds")
        record["data_catalogue_path"] = paths["catalogue"]
        self.write_json_file_safe(paths["project"], record)

    def list_project_records(self):
        root = self.get_projects_root()
        records = []
        if not os.path.isdir(root):
            return records
        for name in sorted(os.listdir(root)):
            project_dir = os.path.join(root, name)
            if not os.path.isdir(project_dir):
                continue
            project_path = os.path.join(project_dir, "project.json")
            if not os.path.exists(project_path):
                continue
            record = self.read_project_record(name)
            records.append(record)
        return records

    def find_project_by_linked_folder(self, folder):
        if not folder:
            return None
        folder_abs = os.path.abspath(folder)
        for record in self.list_project_records():
            linked = record.get("linked_folder", "")
            if linked and os.path.abspath(linked) == folder_abs:
                return record
        return None

    def resolve_project_id_for_folder(self, folder, candidate_id):
        """
        Renvoie un project_id stable. Si un projet existant porte déjà ce slug
        mais pointe vers un autre dossier, on ajoute un suffixe court au lieu
        d'écraser silencieusement.
        """
        folder_abs = os.path.abspath(folder) if folder else ""
        linked = self.find_project_by_linked_folder(folder_abs)
        if linked:
            return linked.get("project_id")

        base = self.slugify(candidate_id or os.path.basename(folder_abs) or "bisse")
        paths = self.get_project_paths(base)
        if not os.path.exists(paths["project"]):
            return base

        record = self.read_project_record(base)
        linked_folder = record.get("linked_folder", "")
        if not linked_folder or not os.path.isdir(linked_folder):
            return base
        if folder_abs and os.path.abspath(linked_folder) == folder_abs:
            return base

        suffix = uuid.uuid5(uuid.NAMESPACE_URL, folder_abs or base).hex[:6]
        return self.slugify(f"{base}-{suffix}")

    def sync_project_catalogue_for_folder(self, project_id, folder, allow_data_to_local=False):
        """
        Synchronisation sécurisée post-réparation.

        Règle : le catalogue local du dossier bisse est prioritaire quand il existe.
        Data reçoit une copie portable. Data ne réécrit le local que dans un cas
        explicite : relocalisation d'un projet vers un dossier qui n'a pas encore
        de catalogue local.
        """
        project_id = self.slugify(project_id or os.path.basename(folder or "bisse"))
        paths = self.get_project_paths(project_id)
        os.makedirs(paths["dir"], exist_ok=True)

        data_catalog = paths["catalogue"]
        local_catalog = os.path.join(folder, "catalogue.json") if folder else ""

        data_exists = os.path.exists(data_catalog)
        local_exists = bool(local_catalog and os.path.exists(local_catalog))

        data_container = self.read_catalog_container_from_path(data_catalog, folder) if data_exists else None
        local_container = self.read_catalog_container_from_path(local_catalog, folder) if local_exists else None

        chosen = None
        source = ""

        if local_container is not None:
            chosen = local_container
            source = "local actif → Data copy"

            if data_container is None:
                self.write_catalog_container_to_path(data_catalog, chosen, folder, update_timestamp=False)
                self.log(f"📦 Copie Data créée depuis le catalogue local : {project_id}")
            elif not self.json_data_equal(data_container, local_container):
                self.backup_catalog_file(project_id, data_catalog, "data_avant_copie_locale")
                self.write_catalog_container_to_path(data_catalog, chosen, folder, update_timestamp=False)
                self.log(f"📦 Copie Data mise à jour depuis le catalogue local : {project_id}")

        elif data_container is not None:
            chosen = data_container
            source = "Data disponible"
            if allow_data_to_local and local_catalog:
                self.write_catalog_container_to_path(local_catalog, chosen, folder, update_timestamp=False)
                self.log(f"🔗 Catalogue Data copié vers le nouveau dossier local : {project_id}")
            else:
                self.log(
                    f"🛡️ Catalogue local absent pour {project_id}. "
                    "Data n'a pas été copié automatiquement vers le dossier local."
                )

        else:
            chosen = self.empty_catalog_container()
            source = "nouveau local + Data"
            self.write_catalog_container_to_path(local_catalog, chosen, folder, update_timestamp=True)
            self.write_catalog_container_to_path(data_catalog, chosen, folder, update_timestamp=False)
            self.log(f"📦 Nouveau catalogue local + copie Data créés : {project_id}")

        record = self.read_project_record(project_id)
        identity = self.catalog_identity(chosen)
        title = identity.get("title") or os.path.basename(folder or "") or record.get("title") or project_id
        slug = identity.get("slug") or self.slugify(title)

        record["title"] = title
        record["slug"] = slug
        record["linked_folder"] = os.path.abspath(folder) if folder else record.get("linked_folder", "")
        record["linked_folder_status"] = "ok" if folder and os.path.isdir(folder) else "missing"
        record["local_catalogue_path"] = local_catalog
        record["last_sync_direction"] = source
        record["last_synced_at"] = datetime.now().isoformat(timespec="seconds")
        record["hidden_from_workspace"] = False
        self.write_project_record(project_id, record)

        return chosen, record

    def ensure_project_for_folder(self, folder):
        folder = os.path.abspath(folder)
        local_catalog = os.path.join(folder, "catalogue.json")
        local_container = None

        if os.path.exists(local_catalog):
            try:
                local_container = self.read_catalog_container_from_path(local_catalog, folder)
            except Exception as exc:
                self.log(f"⚠️ Catalogue local illisible pendant l'import Data : {exc}")

        candidate_id = self.project_id_from_container_or_folder(folder, local_container)
        project_id = self.resolve_project_id_for_folder(folder, candidate_id)

        # Si l'ancien système retrouve un project_id générique, on préfère
        # immédiatement l'identité réelle du catalogue local.
        if self.is_generic_project_id(project_id) and not self.is_generic_project_id(candidate_id):
            project_id = candidate_id

        _container, _record = self.sync_project_catalogue_for_folder(
            project_id,
            folder,
            allow_data_to_local=False
        )

        self.current_project_id = project_id
        paths = self.get_project_paths(project_id)
        self.current_project_dir = paths["dir"]
        self.current_project_catalog_path = paths["catalogue"]
        self.local_catalog_path = local_catalog

        # Mode sécurisé : le catalogue actif est local.
        self.catalog_path = local_catalog

        return project_id

    def set_project_hidden_from_workspace(self, project_id, hidden=True):
        if not project_id:
            return
        record = self.read_project_record(project_id)
        record["hidden_from_workspace"] = bool(hidden)
        self.write_project_record(project_id, record)

    def relink_project_to_folder_dialog(self, project_id):
        project_id = self.resolve_project_id_reference(project_id)

        if not project_id:
            messagebox.showwarning(
                "Projet Data introuvable",
                (
                    "Impossible d'identifier le projet Data correspondant à cette ligne.\n\n"
                    "Essayez d'abord d'actualiser Mes bisses. Si le problème persiste, "
                    "ouvrez le dossier bisse une fois pour créer/importer son projet Data."
                )
            )
            return

        folder = filedialog.askdirectory(title="Relier ce projet à un dossier bisse local")
        if not folder:
            return

        self.relink_project_to_folder(project_id, folder)

    def relink_project_to_folder(self, project_id, folder):
        """
        Associe un projet Data existant à un dossier local sur cet ordinateur.

        v44 sécurité : si le dossier local possède déjà un catalogue, il gagne
        et Data reçoit une copie. Si le dossier local n'a pas de catalogue,
        Data peut être copié vers ce dossier, car l'action est explicite.
        """
        folder = os.path.abspath(folder)
        if not os.path.isdir(folder):
            messagebox.showerror("Dossier introuvable", f"Le dossier n'existe pas :\n{folder}")
            return

        project_id = self.resolve_project_id_reference(project_id)
        if not project_id:
            messagebox.showerror("Projet introuvable", "Impossible d'identifier le projet Data à relier.")
            return

        paths = self.get_project_paths(project_id)
        if not os.path.exists(paths["project"]):
            messagebox.showerror("Projet introuvable", f"Projet Data introuvable : {project_id}")
            return

        self.sync_project_catalogue_for_folder(
            project_id,
            folder,
            allow_data_to_local=True
        )

        workspace = self.read_workspace()
        bisses = workspace.setdefault("bisses", [])
        now = datetime.now().isoformat(timespec="seconds")
        found = False
        for entry in bisses:
            if entry.get("project_id") == project_id:
                entry["folder"] = folder
                entry["last_opened_at"] = now
                found = True
                break
        if not found:
            bisses.append({
                "project_id": project_id,
                "folder": folder,
                "first_opened_at": now,
                "last_opened_at": now
            })
        self.write_workspace(workspace)

        self.log(f"🔗 Projet relié à un dossier local : {project_id} → {folder}")
        messagebox.showinfo("Dossier relié", "Le projet a été relié au dossier local.")
        self.show_workspace_home()

    def open_app_data_folder(self):
        """
        Ouvre le dossier des données globales dans l'explorateur.
        """
        os.makedirs(self.app_data_folder, exist_ok=True)
        try:
            if os.name == "nt":
                os.startfile(self.app_data_folder)
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", self.app_data_folder])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", self.app_data_folder])
        except Exception as exc:
            messagebox.showerror(
                "Ouverture impossible",
                f"Impossible d'ouvrir le dossier :\\n{self.app_data_folder}\\n\\n{exc}"
            )


    def return_to_active_bisse_or_home(self):
        """
        Navigation sûre depuis les modules globaux.
        Si un bisse est ouvert, on revient à son tableau de bord.
        Sinon, on revient à l'accueil Mes bisses.
        """
        if self.base_folder and os.path.isdir(self.base_folder):
            self.load_folder(self.base_folder)
        else:
            self.show_workspace_home()

    def is_folder_in_publication_collection(self, folder):
        """
        Indique si un dossier de travail est déjà inclus dans la collection
        de publication globale.
        """
        if not folder:
            return False
        folder_abs = os.path.abspath(folder)
        try:
            collection = self.read_publication_collection()
            for entry in collection.get("bisses", []):
                if os.path.abspath(entry.get("folder", "")) == folder_abs:
                    return True
        except Exception:
            pass
        return False

    def show_library_home_panel(self):
        """
        Petit panneau d'accueil pour les actions globales de bibliothèque.
        La bibliothèque reste un cache global de l'Excel consolidé.
        """
        library = self.read_bisses_library()
        rows = library.get("rows", [])
        source = library.get("source_file") or "—"
        updated = library.get("updated_at") or "—"

        messagebox.showinfo(
            "Bibliothèque des bisses",
            (
                "Bibliothèque globale des informations générales.\\n\\n"
                f"Entrées : {len(rows)}\\n"
                f"Dernière mise à jour : {updated}\\n"
                f"Source : {source}\\n\\n"
                "Utilisez “Mettre à jour depuis Excel / CSV” pour remplacer "
                "la bibliothèque par le fichier consolidé. Pour appliquer une "
                "fiche à un bisse, ouvrez d'abord ce bisse puis utilisez "
                "“Remplir depuis bibliothèque” dans ses informations générales."
            )
        )


    def show_bisses_library_browser(self):
        """
        Navigateur de la bibliothèque générale des bisses.

        Ce module est global : il ne remplit pas le bisse actif.
        Il sert à consulter la référence issue de l'Excel consolidé et,
        si nécessaire, à créer/ouvrir un dossier de travail depuis une entrée.
        """
        self.clear_main_frame()
        self.status_header.config(text="Bibliothèque des bisses", fg="#6c3483")

        outer = self.make_scrollable_page(padx=14, pady=12)

        toolbar = tk.Frame(outer)
        toolbar.pack(fill="x", pady=(0, 10))

        tk.Button(
            toolbar,
            text="🏠 Mes bisses",
            command=self.show_workspace_home
        ).pack(side="left")

        tk.Label(
            toolbar,
            text="📚 Bibliothèque des bisses",
            font=("Arial", 16, "bold")
        ).pack(side="left", padx=14)

        tk.Button(
            toolbar,
            text="🔄 Mettre à jour depuis Excel / CSV",
            command=self.update_bisses_library_from_file_dialog
        ).pack(side="right")

        library = self.read_bisses_library()
        rows = library.get("rows", [])
        source = library.get("source_file") or "—"
        updated = library.get("updated_at") or "—"

        info = tk.LabelFrame(outer, text="Source", padx=10, pady=8)
        info.pack(fill="x", pady=(0, 10))

        tk.Label(
            info,
            text=(
                f"Entrées : {len(rows)} · Dernière mise à jour : {updated}\n"
                f"Source : {source}"
            ),
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=1320
        ).pack(fill="x")

        search_frame = tk.Frame(outer)
        search_frame.pack(fill="x", pady=(0, 8))

        tk.Label(search_frame, text="Recherche :").pack(side="left")
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=6)

        columns = ("nom", "slug", "region", "communes", "etat", "longueur", "sdt", "importance")

        table_frame = tk.Frame(outer)
        table_frame.pack(fill="both", expand=True)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse", height=18)

        headings = {
            "nom": "Nom",
            "slug": "Slug",
            "region": "Région",
            "communes": "Communes",
            "etat": "État",
            "longueur": "Longueur km",
            "sdt": "SDT",
            "importance": "Importance"
        }
        widths = {
            "nom": 250,
            "slug": 190,
            "region": 160,
            "communes": 220,
            "etat": 120,
            "longueur": 95,
            "sdt": 80,
            "importance": 130
        }

        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")

        tree.column("longueur", anchor="center")
        tree.column("sdt", anchor="center")

        tree_y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree_x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=tree_y_scroll.set, xscrollcommand=tree_x_scroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        tree_y_scroll.grid(row=0, column=1, sticky="ns")
        tree_x_scroll.grid(row=1, column=0, sticky="ew")

        filtered_rows = []

        def row_text(row):
            keys = (
                "nom", "slug", "region", "communes", "etat", "sentier",
                "cotation", "tags", "description", "itineraire",
                "sdt_numero", "sdt_importance"
            )
            return " ".join(str(row.get(key, "") or "") for key in keys).lower()

        def refresh_library_tree(*_args):
            nonlocal filtered_rows
            query = (search_var.get() or "").strip().lower()
            tokens = [tok for tok in query.split() if tok]

            tree.delete(*tree.get_children())
            filtered_rows = []

            for row in rows:
                haystack = row_text(row)
                if tokens and not all(tok in haystack for tok in tokens):
                    continue
                filtered_rows.append(row)
                iid = str(len(filtered_rows) - 1)
                tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        row.get("nom", ""),
                        row.get("slug", ""),
                        row.get("region", ""),
                        row.get("communes", ""),
                        row.get("etat", ""),
                        row.get("longueur_km", ""),
                        row.get("sdt_numero", ""),
                        row.get("sdt_importance", "")
                    )
                )

            status_var.set(f"{len(filtered_rows)} entrée(s) affichée(s) sur {len(rows)}.")

            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])

        status_var = tk.StringVar(value="")
        tk.Label(
            outer,
            textvariable=status_var,
            anchor="w",
            fg="#555555"
        ).pack(fill="x", pady=(6, 0))

        actions = tk.LabelFrame(outer, text="Actions sur l'entrée sélectionnée", padx=10, pady=8)
        actions.pack(fill="x", pady=(8, 0))

        def selected_row():
            selection = tree.selection()
            if not selection:
                return None
            try:
                return filtered_rows[int(selection[0])]
            except Exception:
                return None

        def show_selected_details():
            row = selected_row()
            if not row:
                messagebox.showwarning("Aucune sélection", "Sélectionnez une entrée de bibliothèque.")
                return
            self.show_library_entry_details(row)

        def create_or_open_selected_workspace_folder():
            row = selected_row()
            if not row:
                messagebox.showwarning("Aucune sélection", "Sélectionnez une entrée de bibliothèque.")
                return
            self.create_or_open_bisse_folder_from_library_entry(row)

        tk.Button(
            actions,
            text="🔎 Voir la fiche source",
            command=show_selected_details
        ).pack(side="left")

        tk.Button(
            actions,
            text="📂 Créer / ouvrir un dossier de travail",
            command=create_or_open_selected_workspace_folder,
            bg="#2c3e50",
            fg="white"
        ).pack(side="left", padx=6)

        tk.Button(
            actions,
            text="🔄 Actualiser",
            command=refresh_library_tree
        ).pack(side="right")

        tree.bind("<Double-1>", lambda _event: show_selected_details())
        search_var.trace_add("write", refresh_library_tree)
        refresh_library_tree()

    def show_library_entry_details(self, row):
        """
        Affiche toutes les colonnes utiles d'une entrée de bibliothèque.
        """
        window = tk.Toplevel(self.root)
        window.title(row.get("nom") or "Fiche bibliothèque")
        window.geometry("900x700")
        window.transient(self.root)

        title = row.get("nom") or "Entrée bibliothèque"
        slug = row.get("slug") or self.slugify(title)

        tk.Label(
            window,
            text=f"📚 {title} · {slug}",
            font=("Arial", 14, "bold"),
            anchor="w"
        ).pack(fill="x", padx=12, pady=(12, 6))

        frame = tk.Frame(window)
        frame.pack(fill="both", expand=True, padx=12, pady=6)

        text_widget = tk.Text(frame, wrap="word")
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        preferred_order = [
            "nom", "slug", "url", "region", "communes", "etat", "sentier",
            "cotation", "tags", "longueur_km", "altitude_haut_m",
            "altitude_bas_m", "prise_eau", "zones_irriguees", "autres_noms",
            "classement_ad", "sdt_numero", "sdt_importance",
            "description", "itineraire"
        ]

        already = set()
        for key in preferred_order:
            if key in row:
                text_widget.insert("end", f"{key}\n", ("label",))
                text_widget.insert("end", f"{row.get(key, '')}\n\n")
                already.add(key)

        for key in sorted(row.keys()):
            if key in already or key.startswith("_"):
                continue
            text_widget.insert("end", f"{key}\n", ("label",))
            text_widget.insert("end", f"{row.get(key, '')}\n\n")

        text_widget.tag_configure("label", font=("Arial", 10, "bold"))
        text_widget.configure(state="disabled")

        buttons = tk.Frame(window, padx=12, pady=10)
        buttons.pack(fill="x")

        tk.Button(
            buttons,
            text="📂 Créer / ouvrir un dossier de travail",
            command=lambda: self.create_or_open_bisse_folder_from_library_entry(row, parent_window=window),
            bg="#2c3e50",
            fg="white"
        ).pack(side="left")

        tk.Button(
            buttons,
            text="Fermer",
            command=window.destroy
        ).pack(side="right")

    def create_or_open_bisse_folder_from_library_entry(self, row, parent_window=None):
        """
        Crée ou ouvre un dossier de travail à partir d'une entrée de bibliothèque.

        Cette action ne dépend pas d'un bisse actif. Elle crée un catalogue local
        minimal dans le nouveau dossier, rempli avec les informations générales
        de l'entrée de bibliothèque.
        """
        title = str(row.get("nom") or "").strip()
        if not title:
            messagebox.showwarning("Entrée incomplète", "Cette entrée n'a pas de nom.")
            return

        slug = self.slugify(row.get("slug") or title)

        parent = filedialog.askdirectory(
            title="Choisir le dossier parent où créer/ouvrir ce bisse",
            initialdir=self.get_default_collection_root()
        )
        if not parent:
            return

        folder = os.path.join(parent, slug)
        already_exists = os.path.exists(folder)
        os.makedirs(folder, exist_ok=True)

        previous_base = self.base_folder
        previous_catalog_path = self.catalog_path
        previous_container = self.catalog_container
        previous_data = self.catalog_data

        try:
            self.base_folder = folder
            self.catalog_path = os.path.join(folder, "catalogue.json")
            self.catalog_container = self.empty_catalog_container()
            self.catalog_container["inventory_info"] = self.inventory_row_normalized(
                row,
                row.get("_source_file", ""),
                row.get("_source_sheet", "")
            )
            self.catalog_container["bisse_info"] = self.inventory_to_bisse_info_fields(
                self.catalog_container["inventory_info"]
            )
            self.catalog_data = []
            if not os.path.exists(self.catalog_path):
                self.save_catalog()
                self.log(f"💾 Catalogue créé depuis la bibliothèque : {self.catalog_path}")
            else:
                # Ne pas écraser un dossier de travail existant.
                self.log(f"ℹ️ Catalogue existant conservé : {self.catalog_path}")
        finally:
            self.base_folder = previous_base
            self.catalog_path = previous_catalog_path
            self.catalog_container = previous_container
            self.catalog_data = previous_data

        self.add_folder_to_workspace(folder)

        if parent_window is not None:
            try:
                parent_window.destroy()
            except Exception:
                pass

        if already_exists:
            message = "Dossier existant ajouté/ouvert depuis Mes bisses."
        else:
            message = "Nouveau dossier de travail créé et ajouté à Mes bisses."

        messagebox.showinfo("Mes bisses", f"{message}\n\n{folder}")
        self.load_folder(folder)


    # ============================================================
    # MISE À JOUR EN MASSE DEPUIS LA BIBLIOTHÈQUE
    # ============================================================

    def make_library_link_from_row(self, row):
        """
        Identifiant persistant d'une entrée de bibliothèque.
        Le slug est l'identifiant principal ; le nom reste en secours humain.
        """
        nom = str(row.get("nom") or "").strip()
        slug = self.slugify(row.get("slug") or nom)
        return {
            "slug": slug,
            "nom": nom,
            "linked_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": row.get("_source_file", ""),
            "source_sheet": row.get("_source_sheet", "")
        }

    def build_bisses_library_lookup(self):
        """
        Construit des index simples pour retrouver une ligne de bibliothèque.
        """
        library = self.read_bisses_library()
        rows = library.get("rows", [])

        by_slug = {}
        by_name = {}

        for row in rows:
            nom = str(row.get("nom") or "").strip()
            slug = self.slugify(row.get("slug") or nom)
            if slug and slug not in by_slug:
                by_slug[slug] = row

            name_key = self.slugify(nom)
            if name_key and name_key not in by_name:
                by_name[name_key] = row

        return rows, by_slug, by_name

    def read_catalog_container_for_folder(self, folder):
        """
        Lecture d'un catalogue sans changer durablement le bisse actif.

        v44 sécurité : lit d'abord le catalogue local du dossier. Data ne doit
        pas remplacer silencieusement le catalogue local.
        """
        folder = os.path.abspath(folder)
        local_catalog_path = os.path.join(folder, "catalogue.json")

        previous_base = self.base_folder
        previous_catalog_path = self.catalog_path
        previous_local_catalog_path = getattr(self, "local_catalog_path", "")
        previous_project_id = getattr(self, "current_project_id", "")
        previous_project_dir = getattr(self, "current_project_dir", "")
        previous_project_catalog_path = getattr(self, "current_project_catalog_path", "")
        previous_container = self.catalog_container
        previous_data = self.catalog_data

        try:
            self.base_folder = folder
            self.local_catalog_path = local_catalog_path
            self.catalog_path = local_catalog_path

            if not os.path.exists(local_catalog_path):
                # Lecture Data possible uniquement comme repli, sans écriture locale.
                project_id = self.resolve_project_id_for_folder(
                    folder,
                    self.slugify(os.path.basename(folder))
                )
                data_path = self.get_project_paths(project_id)["catalogue"]
                if os.path.exists(data_path):
                    self.catalog_path = data_path

            container = self.read_catalog_container()
            photos = container.get("photos", [])
            return container, photos
        finally:
            self.base_folder = previous_base
            self.catalog_path = previous_catalog_path
            self.local_catalog_path = previous_local_catalog_path
            self.current_project_id = previous_project_id
            self.current_project_dir = previous_project_dir
            self.current_project_catalog_path = previous_project_catalog_path
            self.catalog_container = previous_container
            self.catalog_data = previous_data

    def write_catalog_container_for_folder(self, folder, container):
        """
        Écrit un catalogue sans changer durablement le bisse actif.

        v44 sécurité : écrit d'abord dans le catalogue local du dossier,
        puis crée/met à jour la copie portable Data.
        """
        folder = os.path.abspath(folder)
        os.makedirs(folder, exist_ok=True)

        previous_base = self.base_folder
        previous_catalog_path = self.catalog_path
        previous_local_catalog_path = getattr(self, "local_catalog_path", "")
        previous_project_id = getattr(self, "current_project_id", "")
        previous_project_dir = getattr(self, "current_project_dir", "")
        previous_project_catalog_path = getattr(self, "current_project_catalog_path", "")
        previous_container = self.catalog_container
        previous_data = self.catalog_data

        try:
            self.base_folder = folder
            self.local_catalog_path = os.path.join(folder, "catalogue.json")
            self.catalog_path = self.local_catalog_path

            self.write_catalog_container_to_path(
                self.local_catalog_path,
                container,
                folder,
                update_timestamp=True
            )

            project_id = self.project_id_from_container_or_folder(folder, container)
            paths = self.get_project_paths(project_id)
            if os.path.exists(paths["catalogue"]):
                try:
                    self.backup_catalog_file(project_id, paths["catalogue"], "data_avant_ecriture_locale")
                except Exception:
                    pass
            self.write_catalog_container_to_path(
                paths["catalogue"],
                container,
                folder,
                update_timestamp=False
            )

            record = self.read_project_record(project_id)
            identity = self.catalog_identity(container)
            record["title"] = identity.get("title") or os.path.basename(folder)
            record["slug"] = identity.get("slug") or self.slugify(record["title"])
            record["linked_folder"] = folder
            record["linked_folder_status"] = "ok"
            record["last_synced_at"] = datetime.now().isoformat(timespec="seconds")
            record["last_sync_direction"] = "write local + Data copy"
            self.write_project_record(project_id, record)
        finally:
            self.base_folder = previous_base
            self.catalog_path = previous_catalog_path
            self.local_catalog_path = previous_local_catalog_path
            self.current_project_id = previous_project_id
            self.current_project_dir = previous_project_dir
            self.current_project_catalog_path = previous_project_catalog_path
            self.catalog_container = previous_container
            self.catalog_data = previous_data

    def find_library_row_for_catalog(self, folder, container, by_slug, by_name):
        """
        Retrouve la ligne de bibliothèque liée ou propose une correspondance.
        Retourne (row, link_state, candidate_key)
        link_state :
        - linked : lien déjà présent ;
        - proposed : correspondance proposée à confirmer ;
        - missing : aucune correspondance.
        """
        folder_name = os.path.basename(folder)
        info = container.get("bisse_info", {}) or {}
        inventory = container.get("inventory_info", {}) or {}
        link = container.get("library_link", {}) or {}

        # 1. Lien explicite déjà confirmé.
        linked_slug = self.slugify(link.get("slug") or "")
        if linked_slug and linked_slug in by_slug:
            return by_slug[linked_slug], "linked", linked_slug

        # 2. Ancien import inventaire : peut servir de lien implicite.
        inventory_slug = self.slugify(inventory.get("slug") or inventory.get("nom") or "")
        if inventory_slug and inventory_slug in by_slug:
            return by_slug[inventory_slug], "linked", inventory_slug

        # 3. Propositions à confirmer.
        candidates = [
            info.get("slug"),
            info.get("title"),
            folder_name
        ]
        for value in candidates:
            key = self.slugify(value or "")
            if key in by_slug:
                return by_slug[key], "proposed", key
            if key in by_name:
                return by_name[key], "proposed", key

        return None, "missing", ""

    def build_library_update_plan(self, folders):
        """
        Prépare la mise à jour : liens, champs vides à remplir, conflits.
        """
        rows, by_slug, by_name = self.build_bisses_library_lookup()
        labels = self.bisse_info_field_labels()

        plan = []
        for folder in folders:
            folder = os.path.abspath(folder)
            item = {
                "folder": folder,
                "summary": self.get_bisse_summary_from_folder(folder),
                "container": None,
                "row": None,
                "link_state": "missing",
                "auto_fill": [],
                "conflicts": [],
                "error": ""
            }

            try:
                container, _photos = self.read_catalog_container_for_folder(folder)
                item["container"] = container
                row, link_state, _key = self.find_library_row_for_catalog(folder, container, by_slug, by_name)
                item["row"] = row
                item["link_state"] = link_state

                if row:
                    inventory = self.inventory_row_normalized(
                        row,
                        row.get("_source_file", ""),
                        row.get("_source_sheet", "")
                    )
                    imported_fields = self.inventory_to_bisse_info_fields(inventory)
                    info = container.setdefault("bisse_info", self.default_bisse_info())

                    canonical_imported_fields = {}
                    for raw_key, imported_value in imported_fields.items():
                        key = self.canonical_bisse_info_update_key(raw_key)
                        if self.comparable_value(imported_value) == "":
                            continue
                        if key not in canonical_imported_fields:
                            canonical_imported_fields[key] = imported_value

                    self.synchronize_bisse_info_aliases(info)

                    for key, imported_value in canonical_imported_fields.items():
                        local_value = info.get(key)

                        if self.comparable_value(local_value) == "":
                            item["auto_fill"].append((key, imported_value))
                            continue

                        if self.values_equivalent_for_update(local_value, imported_value):
                            continue

                        item["conflicts"].append((key, local_value, imported_value, labels.get(key, key)))

            except Exception as exc:
                item["error"] = str(exc)

            plan.append(item)

        return plan

    def show_update_selected_bisses_from_library_dialog(self, folders):
        """
        Met à jour les informations générales des bisses sélectionnés depuis
        la bibliothèque globale, avec lien persistant et résolution prudente.
        """
        if not folders:
            messagebox.showwarning("Aucune sélection", "Sélectionnez un ou plusieurs bisses.")
            return

        library = self.read_bisses_library()
        if not library.get("rows"):
            messagebox.showwarning(
                "Bibliothèque vide",
                "La bibliothèque est vide. Mettez-la d'abord à jour depuis l'Excel consolidé."
            )
            return

        plan = self.build_library_update_plan(folders)
        usable = [item for item in plan if item.get("row") and not item.get("error")]

        if not usable:
            messagebox.showwarning(
                "Aucune correspondance",
                "Aucun des bisses sélectionnés n'a pu être relié à une entrée de bibliothèque."
            )
            return

        window = tk.Toplevel(self.root)
        window.title("Mettre à jour les infos depuis la bibliothèque")
        window.geometry("1250x720")
        window.transient(self.root)
        window.grab_set()

        tk.Label(
            window,
            text=(
                "Vérifiez les correspondances entre les dossiers de travail et la bibliothèque.\\n"
                "Les liens déjà confirmés seront réutilisés aux prochaines mises à jour. "
                "Les champs vides seront remplis automatiquement ; les conflits seront confirmés ensuite."
            ),
            justify="left",
            anchor="w"
        ).pack(fill="x", padx=12, pady=(10, 8))

        columns = ("bisse", "library", "link", "auto", "conflicts", "status", "folder")
        tree = ttk.Treeview(window, columns=columns, show="headings", selectmode="extended", height=18)
        headings = {
            "bisse": "Bisse local",
            "library": "Entrée bibliothèque",
            "link": "Lien",
            "auto": "Champs vides",
            "conflicts": "Conflits",
            "status": "Statut",
            "folder": "Dossier"
        }
        widths = {
            "bisse": 220,
            "library": 240,
            "link": 110,
            "auto": 90,
            "conflicts": 80,
            "status": 170,
            "folder": 390
        }
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        tree.column("auto", anchor="center")
        tree.column("conflicts", anchor="center")
        tree.pack(fill="both", expand=True, padx=12, pady=6)

        iid_to_item = {}
        default_selection = []

        for idx, item in enumerate(plan):
            row = item.get("row") or {}
            summary = item.get("summary", {})
            link_state = item.get("link_state", "missing")
            if item.get("error"):
                status = f"Erreur : {item['error']}"
            elif not row:
                status = "Aucune correspondance"
            elif link_state == "linked":
                status = "Lien existant"
            else:
                status = "Lien proposé à confirmer"

            iid = str(idx)
            iid_to_item[iid] = item
            tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    summary.get("title") or os.path.basename(item["folder"]),
                    row.get("nom", "") if row else "—",
                    "existant" if link_state == "linked" else ("à confirmer" if row else "—"),
                    len(item.get("auto_fill", [])),
                    len(item.get("conflicts", [])),
                    status,
                    item.get("folder", "")
                )
            )

            if row and not item.get("error"):
                default_selection.append(iid)

        if default_selection:
            tree.selection_set(default_selection)

        buttons = tk.Frame(window, padx=12, pady=10)
        buttons.pack(fill="x")

        def select_all():
            tree.selection_set(default_selection)

        def clear_selection():
            tree.selection_remove(tree.selection())

        def selected_items():
            return [iid_to_item[iid] for iid in tree.selection() if iid_to_item.get(iid, {}).get("row")]

        def continue_to_conflicts():
            selected = selected_items()
            if not selected:
                messagebox.showwarning("Aucune sélection", "Aucun bisse utilisable n'est sélectionné.")
                return

            proposed = [item for item in selected if item.get("link_state") == "proposed"]
            if proposed:
                if not messagebox.askyesno(
                    "Confirmer les liens proposés",
                    (
                        f"{len(proposed)} lien(s) bibliothèque sont proposés automatiquement.\\n\\n"
                        "Confirmer ces liens et les enregistrer pour les prochaines mises à jour ?"
                    )
                ):
                    return

            window.destroy()
            self.resolve_library_update_conflicts_and_apply(selected)

        tk.Button(buttons, text="Tout sélectionner", command=select_all).pack(side="left")
        tk.Button(buttons, text="Tout désélectionner", command=clear_selection).pack(side="left", padx=6)
        tk.Button(buttons, text="Annuler", command=window.destroy).pack(side="right")
        tk.Button(
            buttons,
            text="Continuer",
            command=continue_to_conflicts,
            bg="#2c3e50",
            fg="white"
        ).pack(side="right", padx=(0, 8))

    def resolve_library_update_conflicts_and_apply(self, selected_items):
        """
        Affiche les conflits de champs, puis applique la mise à jour.
        Par défaut, les corrections locales sont conservées.

        v35 :
        - les faux conflits équivalents après normalisation sont supprimés et commune/communes est dédoublé ;
        - un seul scroll global ;
        - chaque zone texte adapte sa hauteur à son contenu réel et à sa largeur réelle.
        """
        conflict_rows = []
        cleaned_conflicts_by_folder = {}

        for item in selected_items:
            cleaned = []
            seen_keys = set()
            for key, local_value, imported_value, label in item.get("conflicts", []):
                key = self.canonical_bisse_info_update_key(key)
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                if self.values_equivalent_for_update(local_value, imported_value):
                    continue

                cleaned.append((key, local_value, imported_value, label))
                conflict_rows.append({
                    "item": item,
                    "key": key,
                    "label": label,
                    "local": local_value,
                    "imported": imported_value
                })

            item["conflicts"] = cleaned
            cleaned_conflicts_by_folder[item["folder"]] = cleaned

        if not conflict_rows:
            self.apply_library_update_items(selected_items, {})
            return

        window = tk.Toplevel(self.root)
        window.title("Résoudre les conflits de mise à jour")
        window.geometry("1250x760")
        window.minsize(980, 620)
        window.transient(self.root)
        window.grab_set()

        header = tk.Frame(window, padx=12, pady=10)
        header.pack(fill="x")

        tk.Label(
            header,
            text=(
                "Les champs vides seront remplis automatiquement.\n"
                "Les différences purement typographiques sont ignorées.\n"
                "Pour les vrais conflits, choisissez quoi faire. Par défaut, les corrections locales sont conservées.\n"
                "Les différences significatives sont surlignées en jaune/orange."
            ),
            justify="left",
            anchor="w",
            wraplength=1120
        ).pack(side="left", fill="x", expand=True)

        # Zone unique de défilement : pas de mini-scroll dans les textes.
        scroll_outer = tk.Frame(window)
        scroll_outer.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        canvas = tk.Canvas(scroll_outer, highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(scroll_outer, orient="vertical", command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(scroll_outer, orient="horizontal", command=canvas.xview)

        scroll = tk.Frame(canvas, padx=4, pady=4)
        inner = canvas.create_window((0, 0), window=scroll, anchor="nw")

        canvas.configure(
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )

        canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        scroll_outer.grid_rowconfigure(0, weight=1)
        scroll_outer.grid_columnconfigure(0, weight=1)

        def on_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            width = max(event.width, 1120)
            canvas.itemconfigure(inner, width=width)
            window.after_idle(adjust_text_heights)

        scroll.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        def _on_mousewheel(event):
            try:
                if getattr(event, "num", None) == 4:
                    canvas.yview_scroll(-3, "units")
                elif getattr(event, "num", None) == 5:
                    canvas.yview_scroll(3, "units")
                else:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
            return "break"

        def _bind_wheel(_event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbind_wheel(_event=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        scroll_outer.bind("<Enter>", _bind_wheel)
        scroll_outer.bind("<Leave>", _unbind_wheel)

        choices = {}
        text_pairs = []
        adjust_pending = {"running": False}

        def display_lines(widget):
            """
            Nombre réel de lignes visuelles après wrap, selon la largeur réelle.
            """
            try:
                widget.update_idletasks()
                result = widget.count("1.0", "end-1c", "displaylines")
                if result and result[0]:
                    return int(result[0])
            except Exception:
                pass

            # Repli si count(displaylines) n'est pas disponible.
            try:
                return max(3, int(float(widget.index("end-1c").split(".")[0])))
            except Exception:
                return 3

        def adjust_text_heights():
            """
            Ajuste chaque paire local/bibliothèque à la hauteur réelle du texte
            le plus haut. Pas d'estimation fixe : la largeur réelle est prise
            en compte par Text.count(..., "displaylines").
            """
            if adjust_pending["running"]:
                return
            adjust_pending["running"] = True

            try:
                window.update_idletasks()
                changed = False

                for local_text, import_text in text_pairs:
                    local_lines = display_lines(local_text)
                    import_lines = display_lines(import_text)
                    height = max(3, local_lines, import_lines) + 1

                    current_local = int(str(local_text.cget("height")))
                    current_import = int(str(import_text.cget("height")))

                    if current_local != height:
                        local_text.configure(height=height)
                        changed = True
                    if current_import != height:
                        import_text.configure(height=height)
                        changed = True

                if changed:
                    window.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
            finally:
                adjust_pending["running"] = False

        for idx, conflict in enumerate(conflict_rows):
            item = conflict["item"]
            summary = item.get("summary", {})
            label = f"{summary.get('title') or os.path.basename(item['folder'])} — {conflict['label']}"

            frame = tk.LabelFrame(scroll, text=label, padx=8, pady=8)
            frame.grid(row=idx, column=0, sticky="ew", pady=6)
            scroll.grid_columnconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1, minsize=500)
            frame.grid_columnconfigure(1, weight=1, minsize=500)

            var = tk.StringVar(value="local")
            choices[(item["folder"], conflict["key"])] = var

            local_box = tk.Frame(frame)
            import_box = tk.Frame(frame)
            local_box.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
            import_box.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

            tk.Radiobutton(
                local_box,
                text="Garder la version locale",
                variable=var,
                value="local"
            ).pack(anchor="w")

            local_text = self.make_diff_text_box(
                local_box,
                conflict["local"],
                conflict["imported"],
                side="local",
                wheel_scroll_callback=_on_mousewheel
            )

            tk.Radiobutton(
                import_box,
                text="Utiliser la bibliothèque",
                variable=var,
                value="import"
            ).pack(anchor="w")

            import_text = self.make_diff_text_box(
                import_box,
                conflict["imported"],
                conflict["local"],
                side="import",
                wheel_scroll_callback=_on_mousewheel
            )

            text_pairs.append((local_text, import_text))

        buttons = tk.Frame(window, padx=12, pady=10)
        buttons.pack(fill="x", side="bottom")

        def set_all(value):
            for var in choices.values():
                var.set(value)

        def cleanup_wheel():
            try:
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")
            except Exception:
                pass

        def apply():
            cleanup_wheel()
            window.destroy()
            self.apply_library_update_items(selected_items, choices)

        def cancel():
            cleanup_wheel()
            window.destroy()

        tk.Button(buttons, text="Tout garder local", command=lambda: set_all("local")).pack(side="left")
        tk.Button(buttons, text="Tout prendre bibliothèque", command=lambda: set_all("import")).pack(side="left", padx=6)
        tk.Button(buttons, text="Annuler", command=cancel).pack(side="right")
        tk.Button(
            buttons,
            text="Appliquer la mise à jour",
            command=apply,
            bg="#2c3e50",
            fg="white"
        ).pack(side="right", padx=(0, 8))

        window.protocol("WM_DELETE_WINDOW", cancel)

        # Première mesure une fois que Tkinter connaît les vraies largeurs.
        window.after_idle(adjust_text_heights)
        window.after(250, adjust_text_heights)

    def apply_library_update_items(self, selected_items, conflict_choices):
        """
        Applique la mise à jour sur les catalogues locaux sélectionnés.
        """
        updated = 0
        auto_filled = 0
        conflicts_replaced = 0
        links_saved = 0
        errors = []

        for item in selected_items:
            folder = item["folder"]
            row = item.get("row")
            if not row:
                continue

            try:
                container = item.get("container")
                if not isinstance(container, dict):
                    container, _photos = self.read_catalog_container_for_folder(folder)

                inventory = self.inventory_row_normalized(
                    row,
                    row.get("_source_file", ""),
                    row.get("_source_sheet", "")
                )
                imported_fields_raw = self.inventory_to_bisse_info_fields(inventory)
                imported_fields = {}
                for raw_key, value in imported_fields_raw.items():
                    key = self.canonical_bisse_info_update_key(raw_key)
                    if key not in imported_fields:
                        imported_fields[key] = value

                info = container.setdefault("bisse_info", self.default_bisse_info())
                self.synchronize_bisse_info_aliases(info)

                # Champs vides : remplissage automatique.
                for key, imported_value in item.get("auto_fill", []):
                    if self.comparable_value(imported_value) != "" and self.comparable_value(info.get(key)) == "":
                        info[key] = imported_value
                        auto_filled += 1

                # Conflits : ne remplacer que si demandé.
                for key, _local_value, imported_value, _label in item.get("conflicts", []):
                    choice = conflict_choices.get((folder, key))
                    if choice is not None and choice.get() == "import":
                        info[key] = imported_value
                        conflicts_replaced += 1

                self.synchronize_bisse_info_aliases(info)
                container["bisse_info"] = info
                container["inventory_info"] = inventory
                container["library_link"] = self.make_library_link_from_row(row)
                links_saved += 1

                self.write_catalog_container_for_folder(folder, container)
                self.add_folder_to_workspace(folder)
                updated += 1

            except Exception as exc:
                errors.append(f"{os.path.basename(folder)} : {exc}")

        message = (
            f"{updated} bisse(s) mis à jour.\\n"
            f"Liens bibliothèque enregistrés : {links_saved}\\n"
            f"Champs vides remplis : {auto_filled}\\n"
            f"Conflits remplacés par la bibliothèque : {conflicts_replaced}"
        )
        if errors:
            message += "\\n\\nErreurs :\\n" + "\\n".join(errors[:8])
            if len(errors) > 8:
                message += f"\\n… et {len(errors) - 8} autre(s)."

        messagebox.showinfo("Mise à jour depuis la bibliothèque", message)
        self.show_workspace_home()

    def migrate_legacy_global_files_if_needed(self):
        """
        Migration prudente depuis les anciennes versions :
        si bisses_workspace.json, bisses_library.json ou bisses_collection.json
        existent dans l'ancien emplacement parent du bisse ouvert, et que le
        fichier global n'existe pas encore dans Gestion_Bisses_Data/, on le copie.

        On ne fusionne pas et on n'écrase jamais un fichier global existant.
        """
        if not self.base_folder:
            return

        try:
            legacy_root = os.path.dirname(os.path.abspath(self.base_folder))
        except Exception:
            return

        for filename in ("bisses_workspace.json", "bisses_library.json", "bisses_collection.json"):
            old_path = os.path.join(legacy_root, filename)
            new_path = self.get_app_data_path(filename)
            if os.path.exists(new_path):
                continue
            if os.path.exists(old_path):
                try:
                    shutil.copy2(old_path, new_path)
                    self.log(f"📦 Ancien fichier global copié vers Gestion_Bisses_Data : {filename}")
                except Exception as exc:
                    self.log(f"⚠️ Migration impossible pour {filename} : {exc}")

    def relative_to_base(self, path):
        try:
            rel = os.path.relpath(path, self.base_folder)
            return rel.replace("\\", "/")
        except Exception:
            return path.replace("\\", "/")

    def abs_from_base(self, relative_path):
        return os.path.join(self.base_folder, relative_path.replace("/", os.sep))

    def sanitize_filename_part(self, text):
        text = text.strip()
        text = re.sub(r"\s+", "", text)
        text = re.sub(r"[^A-Za-z0-9_\-]", "", text)
        return text or "Bisse"

    def default_gpx_categories(self):
        """
        Catégories initiales pour l'atelier GPX.
        Elles restent entièrement extensibles par l'utilisateur.
        """
        return [
            {
                "id": "non_classe",
                "label": "Non classé",
                "file_code": "non_classe",
                "color": "#8e44ad"
            },
            {
                "id": "ciel_ouvert",
                "label": "À ciel ouvert",
                "file_code": "ciel_ouvert",
                "color": "#1e88e5"
            },
            {
                "id": "canalise",
                "label": "Canalisé",
                "file_code": "canalise",
                "color": "#111111"
            },
            {
                "id": "abandonne",
                "label": "Abandonné",
                "file_code": "abandonne",
                "color": "#ef6c00"
            }
        ]


    # ============================================================
    # CATÉGORIES GLOBALES DE SEGMENTS
    # ============================================================

    def normalize_segment_category_record(self, record, fallback_order=100):
        record = copy.deepcopy(record) if isinstance(record, dict) else {}

        label = str(
            record.get("label")
            or record.get("name")
            or record.get("id")
            or "Catégorie"
        ).strip()

        category_id = str(record.get("id") or "").strip()
        if not category_id:
            category_id = self.slugify(label).replace("-", "_") or "categorie"

        file_code = str(record.get("file_code") or "").strip()
        if not file_code:
            file_code = category_id

        color = str(record.get("color") or "#666666").strip()
        if not re.fullmatch(r"#[0-9A-Fa-f]{6}", color):
            color = "#666666"

        try:
            order = int(record.get("order", fallback_order))
        except Exception:
            order = int(fallback_order)

        normalized = {
            "id": category_id,
            "label": label,
            "file_code": file_code,
            "color": color,
            "order": order,
            "active": bool(record.get("active", True)),
            "system": bool(record.get("system", False))
        }

        if record.get("scope"):
            normalized["scope"] = record.get("scope")
        if record.get("needs_integration"):
            normalized["needs_integration"] = True

        return normalized

    def default_global_segment_categories(self):
        defaults = []
        for order, category in enumerate(self.default_gpx_categories(), start=1):
            record = copy.deepcopy(category)
            record["order"] = order * 10
            record["active"] = True
            record["system"] = record.get("id") == "non_classe"
            defaults.append(self.normalize_segment_category_record(record, order * 10))
        return defaults

    def ensure_global_segment_categories_file(self):
        if os.path.exists(self.segment_categories_file):
            self.global_segment_categories = self.read_global_segment_categories(force_reload=True)
            return

        self.write_global_segment_categories(
            self.default_global_segment_categories(),
            create_backup=False
        )

    def read_global_segment_categories(self, force_reload=False):
        if self.global_segment_categories and not force_reload:
            return copy.deepcopy(self.global_segment_categories)

        data = self.read_json_file_safe(
            self.segment_categories_file,
            {"schema_version": "1.0", "categories": []}
        )

        raw_categories = data.get("categories", []) if isinstance(data, dict) else []
        categories = []
        seen = set()

        for index, raw in enumerate(raw_categories, start=1):
            category = self.normalize_segment_category_record(raw, index * 10)
            if category["id"] in seen:
                continue
            seen.add(category["id"])
            categories.append(category)

        # Garantit toujours la présence de « Non classé ».
        if "non_classe" not in seen:
            non_classe = self.default_global_segment_categories()[0]
            categories.insert(0, non_classe)

        categories.sort(key=lambda c: (int(c.get("order", 9999)), c.get("label", "").lower()))
        self.global_segment_categories = categories
        return copy.deepcopy(categories)

    def write_global_segment_categories(self, categories, create_backup=True):
        normalized = []
        seen = set()

        for index, raw in enumerate(categories or [], start=1):
            category = self.normalize_segment_category_record(raw, index * 10)
            if category["id"] in seen:
                continue
            seen.add(category["id"])
            normalized.append(category)

        if "non_classe" not in seen:
            normalized.insert(0, self.default_global_segment_categories()[0])

        # « Non classé » est structurel : toujours actif et protégé.
        for category in normalized:
            if category.get("id") == "non_classe":
                category["active"] = True
                category["system"] = True

        normalized.sort(key=lambda c: (int(c.get("order", 9999)), c.get("label", "").lower()))

        if create_backup and os.path.exists(self.segment_categories_file):
            backup_dir = os.path.join(self.app_data_folder, "segment_categories_backups")
            os.makedirs(backup_dir, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            shutil.copy2(
                self.segment_categories_file,
                os.path.join(backup_dir, f"segment_categories_{stamp}.json")
            )

        payload = {
            "schema_version": "1.0",
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "categories": normalized
        }

        os.makedirs(os.path.dirname(self.segment_categories_file), exist_ok=True)
        temp_path = self.segment_categories_file + ".tmp"

        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=4, ensure_ascii=False)

        os.replace(temp_path, self.segment_categories_file)
        self.global_segment_categories = normalized

    def get_global_segment_category_by_id(self, category_id):
        for category in self.read_global_segment_categories():
            if category.get("id") == category_id:
                return copy.deepcopy(category)
        return None

    def get_local_segment_categories_from_container(self, container=None):
        if container is None:
            container = self.catalog_container

        if not isinstance(container, dict):
            return []

        workshop = container.get("gpx_workshop", {}) or {}
        categories = workshop.get("categories", []) or []
        result = []

        for index, raw in enumerate(categories, start=1):
            result.append(self.normalize_segment_category_record(raw, 1000 + index * 10))

        return result

    def get_effective_gpx_categories(self, container=None, include_inactive=True, include_local=True):
        """
        Fusionne le référentiel global et les anciennes catégories locales.

        Une catégorie globale gagne toujours sur une catégorie locale portant
        le même identifiant : renommer ou recolorer dans Paramètres se reflète
        donc immédiatement dans tous les bisses, sans réécriture massive.
        """
        global_categories = self.read_global_segment_categories()
        local_categories = self.get_local_segment_categories_from_container(container)

        result = []
        global_ids = set()

        for category in global_categories:
            if not include_inactive and not category.get("active", True):
                continue
            record = copy.deepcopy(category)
            record["scope"] = "global"
            record["needs_integration"] = False
            result.append(record)
            global_ids.add(record.get("id"))

        if include_local:
            for local in local_categories:
                if local.get("id") in global_ids:
                    continue
                local["scope"] = "local"
                local["active"] = True
                local["needs_integration"] = True
                local["order"] = max(int(local.get("order", 1000)), 1000)
                result.append(local)

        result.sort(
            key=lambda c: (
                0 if c.get("scope") == "global" else 1,
                int(c.get("order", 9999)),
                c.get("label", "").lower()
            )
        )
        return result

    def refresh_global_category_interfaces(self):
        self.global_segment_categories = self.read_global_segment_categories(force_reload=True)

        try:
            self.refresh_segment_categories_manager_tree()
        except Exception:
            pass

        try:
            self.refresh_gpx_category_tree()
            self.refresh_gpx_category_combo()
            self.refresh_gpx_segment_tree()
        except Exception:
            pass

        if self.gpx_workshop_active:
            try:
                self.draw_gpx_workshop_map()
            except Exception:
                pass

    def iter_known_catalogues_for_category_usage(self):
        """
        Retourne un catalogue par projet connu, en privilégiant le catalogue
        local du dossier bisse lorsqu'il existe.
        """
        candidates = []

        if self.catalog_path and os.path.exists(self.catalog_path):
            candidates.append(self.catalog_path)

        for record in self.list_project_records():
            linked = record.get("linked_folder", "")
            local_path = os.path.join(linked, "catalogue.json") if linked else ""
            data_path = self.get_project_paths(record.get("project_id", ""))["catalogue"]

            if local_path and os.path.exists(local_path):
                candidates.append(local_path)
            elif os.path.exists(data_path):
                candidates.append(data_path)

        unique = []
        seen = set()
        for path in candidates:
            try:
                key = os.path.normcase(os.path.abspath(path))
            except Exception:
                key = str(path)
            if key in seen:
                continue
            seen.add(key)
            unique.append(path)

        return unique

    def scan_segment_category_usage(self):
        usage = {}

        for path in self.iter_known_catalogues_for_category_usage():
            raw = self.read_json_file_safe(path, {})
            if not isinstance(raw, dict):
                continue

            info = raw.get("bisse_info", {}) or {}
            project = raw.get("project", {}) or {}
            title = (
                info.get("title")
                or project.get("title")
                or project.get("bisse_name")
                or os.path.basename(os.path.dirname(path))
            )

            workshop = raw.get("gpx_workshop", {}) or {}
            for segment in workshop.get("segments", []) or []:
                ids = [segment.get("category_id", "non_classe")]
                bicolor = segment.get("bicolor_categories")
                if isinstance(bicolor, list):
                    ids.extend(bicolor[:2])

                # Une même catégorie ne compte qu'une fois par segment,
                # même si elle est à la fois principale et côté A/B.
                for category_id in dict.fromkeys(ids):
                    if not category_id:
                        continue
                    record = usage.setdefault(
                        category_id,
                        {"segments": 0, "bisses": set(), "catalogues": set()}
                    )
                    record["segments"] += 1
                    record["bisses"].add(str(title))
                    record["catalogues"].add(path)

        return usage

    def show_settings_dialog(self):
        if self.settings_window and self.settings_window.winfo_exists():
            self.settings_window.deiconify()
            self.settings_window.lift()
            self.settings_window.focus_force()
            return

        window = tk.Toplevel(self.root)
        self.settings_window = window
        window.title("Paramètres")
        window.geometry("940x650")
        window.minsize(820, 560)
        window.transient(self.root)

        notebook = ttk.Notebook(window)
        notebook.pack(fill="both", expand=True, padx=12, pady=12)

        categories_tab = tk.Frame(notebook, padx=12, pady=12)
        notebook.add(categories_tab, text="Catégories de segments")
        self.build_global_segment_categories_manager(categories_tab)

        info_tab = tk.Frame(notebook, padx=18, pady=18)
        notebook.add(info_tab, text="Données du logiciel")

        tk.Label(
            info_tab,
            text="Dossier des données globales",
            font=("Arial", 12, "bold")
        ).pack(anchor="w")

        tk.Label(
            info_tab,
            text=self.app_data_folder,
            justify="left",
            anchor="w",
            wraplength=820,
            fg="#555555"
        ).pack(fill="x", pady=(6, 12))

        tk.Button(
            info_tab,
            text="📂 Ouvrir Gestion_Bisses_Data",
            command=self.open_app_data_folder
        ).pack(anchor="w")

        tk.Label(
            info_tab,
            text=(
                "Les futurs paramètres NAS / mode réseau direct / mode local synchronisé "
                "seront ajoutés dans cette section."
            ),
            justify="left",
            anchor="w",
            wraplength=820,
            fg="#666666"
        ).pack(fill="x", pady=(24, 0))

        def close():
            self.segment_category_manager_tree = None
            self.settings_window = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", close)

    def build_global_segment_categories_manager(self, parent):
        tk.Label(
            parent,
            text=(
                "Référentiel global utilisé par tous les bisses. "
                "Les catégories locales anciennes restent lisibles et seront nettoyées "
                "dans l’étape de migration assistée."
            ),
            justify="left",
            anchor="w",
            wraplength=860,
            fg="#555555"
        ).pack(fill="x", pady=(0, 10))

        columns = ("etat", "label", "code", "couleur", "ordre", "usage")
        tree = ttk.Treeview(
            parent,
            columns=columns,
            show="headings",
            selectmode="browse",
            height=16
        )
        self.segment_category_manager_tree = tree

        tree.heading("etat", text="État")
        tree.heading("label", text="Nom")
        tree.heading("code", text="Code fichier")
        tree.heading("couleur", text="Couleur")
        tree.heading("ordre", text="Ordre")
        tree.heading("usage", text="Utilisation")

        tree.column("etat", width=100, anchor="center")
        tree.column("label", width=190)
        tree.column("code", width=150)
        tree.column("couleur", width=90, anchor="center")
        tree.column("ordre", width=65, anchor="center")
        tree.column("usage", width=180)

        tree.pack(fill="both", expand=True, pady=(0, 10))
        tree.bind("<Double-1>", lambda _event: self.edit_selected_global_segment_category())

        buttons = tk.Frame(parent)
        buttons.pack(fill="x")

        tk.Button(
            buttons,
            text="➕ Ajouter",
            command=self.open_global_segment_category_editor,
            bg="#8e44ad",
            fg="white"
        ).pack(side="left", padx=(0, 5))

        tk.Button(
            buttons,
            text="✏️ Modifier",
            command=self.edit_selected_global_segment_category
        ).pack(side="left", padx=5)

        tk.Button(
            buttons,
            text="⏯ Activer / désactiver",
            command=self.toggle_selected_global_segment_category
        ).pack(side="left", padx=5)

        tk.Button(
            buttons,
            text="🔎 Voir les utilisations",
            command=self.show_selected_global_segment_category_usage
        ).pack(side="left", padx=5)

        tk.Button(
            buttons,
            text="🗑️ Supprimer si inutilisée",
            command=self.delete_selected_global_segment_category,
            bg="#c0392b",
            fg="white"
        ).pack(side="left", padx=5)

        tk.Button(
            buttons,
            text="Actualiser",
            command=self.refresh_segment_categories_manager_tree
        ).pack(side="right")

        tk.Button(
            parent,
            text="🧹 Analyser, fusionner et nettoyer les catégories existantes",
            command=self.show_category_cleanup_assistant,
            bg="#d68910",
            fg="white"
        ).pack(fill="x", pady=(10, 0))

        self.refresh_segment_categories_manager_tree()

    def refresh_segment_categories_manager_tree(self):
        tree = self.segment_category_manager_tree
        if not tree or not tree.winfo_exists():
            return

        tree.delete(*tree.get_children())
        usage = self.scan_segment_category_usage()
        self.segment_category_manager_usage = usage

        for category in self.read_global_segment_categories(force_reload=True):
            record = usage.get(category.get("id"), {})
            segment_count = int(record.get("segments", 0))
            bisse_count = len(record.get("bisses", set()))
            usage_text = f"{segment_count} segment(s) · {bisse_count} bisse(s)"

            if category.get("system"):
                state = "Système"
            elif category.get("active", True):
                state = "Active"
            else:
                state = "Inactive"

            tree.insert(
                "",
                "end",
                iid=category.get("id"),
                values=(
                    state,
                    category.get("label", ""),
                    category.get("file_code", ""),
                    category.get("color", ""),
                    category.get("order", ""),
                    usage_text
                )
            )

    def get_selected_global_segment_category_id(self):
        tree = self.segment_category_manager_tree
        if not tree or not tree.winfo_exists():
            return None
        selection = tree.selection()
        return selection[0] if selection else None

    def open_global_segment_category_editor(self, category_id=None, preset=None):
        categories = self.read_global_segment_categories()
        existing = next(
            (copy.deepcopy(c) for c in categories if c.get("id") == category_id),
            None
        )

        if existing is None and isinstance(preset, dict):
            existing = self.normalize_segment_category_record(preset, 100)

        editing = category_id is not None and self.get_global_segment_category_by_id(category_id) is not None

        dialog = tk.Toplevel(self.root)
        dialog.title("Modifier une catégorie globale" if editing else "Ajouter une catégorie globale")
        dialog.geometry("520x390")
        dialog.minsize(480, 350)
        dialog.transient(self.root)
        dialog.grab_set()

        form = tk.Frame(dialog, padx=16, pady=14)
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        tk.Label(
            form,
            text="Catégorie globale de segment",
            font=("Arial", 14, "bold")
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))

        label_var = tk.StringVar(value=(existing or {}).get("label", ""))
        code_var = tk.StringVar(value=(existing or {}).get("file_code", ""))
        color_var = tk.StringVar(value=(existing or {}).get("color", "#2ecc71"))
        order_var = tk.IntVar(value=int((existing or {}).get("order", 100)))
        active_var = tk.BooleanVar(value=bool((existing or {}).get("active", True)))

        tk.Label(form, text="Nom affiché").grid(row=1, column=0, sticky="w", pady=5)
        tk.Entry(form, textvariable=label_var).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=5)

        tk.Label(form, text="Code fichier").grid(row=2, column=0, sticky="w", pady=5)
        tk.Entry(form, textvariable=code_var).grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=5)

        tk.Label(form, text="Couleur").grid(row=3, column=0, sticky="w", pady=5)
        color_row = tk.Frame(form)
        color_row.grid(row=3, column=1, sticky="ew", padx=(8, 0), pady=5)

        color_preview = tk.Label(
            color_row,
            textvariable=color_var,
            bg=color_var.get(),
            fg="white",
            width=16
        )
        color_preview.pack(side="left")

        def choose_color():
            chosen = colorchooser.askcolor(color=color_var.get(), parent=dialog)
            if chosen and chosen[1]:
                color_var.set(chosen[1])
                color_preview.config(bg=chosen[1])

        tk.Button(color_row, text="Choisir…", command=choose_color).pack(side="left", padx=8)

        tk.Label(form, text="Ordre").grid(row=4, column=0, sticky="w", pady=5)
        tk.Spinbox(form, from_=0, to=9999, textvariable=order_var, width=10).grid(
            row=4, column=1, sticky="w", padx=(8, 0), pady=5
        )

        tk.Checkbutton(
            form,
            text="Catégorie active et disponible pour les nouveaux segments",
            variable=active_var
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(8, 5))

        if editing:
            tk.Label(
                form,
                text=f"Identifiant stable : {category_id}",
                fg="#666666"
            ).grid(row=6, column=0, columnspan=2, sticky="w", pady=(5, 0))

        buttons = tk.Frame(form)
        buttons.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(22, 0))

        def validate():
            label = label_var.get().strip()
            if not label:
                messagebox.showerror("Catégorie", "Le nom affiché ne peut pas être vide.", parent=dialog)
                return

            file_code = self.sanitize_filename_part(
                code_var.get().strip().lower() or label.lower()
            )
            color = color_var.get().strip() or "#666666"

            if not re.fullmatch(r"#[0-9A-Fa-f]{6}", color):
                messagebox.showerror("Catégorie", "La couleur doit être au format #RRGGBB.", parent=dialog)
                return

            current = self.read_global_segment_categories()

            if editing:
                new_id = category_id
            else:
                base_id = self.slugify(label).replace("-", "_") or "categorie"
                new_id = base_id
                existing_ids = {c.get("id") for c in current}
                counter = 2
                while new_id in existing_ids:
                    new_id = f"{base_id}_{counter}"
                    counter += 1

            updated = []
            found = False

            for category in current:
                if category.get("id") != new_id:
                    updated.append(category)
                    continue

                found = True
                category = copy.deepcopy(category)
                category["label"] = label
                category["file_code"] = file_code
                category["color"] = color
                category["order"] = int(order_var.get())
                category["active"] = bool(active_var.get())
                updated.append(category)

            if not found:
                updated.append({
                    "id": new_id,
                    "label": label,
                    "file_code": file_code,
                    "color": color,
                    "order": int(order_var.get()),
                    "active": bool(active_var.get()),
                    "system": False
                })

            self.write_global_segment_categories(updated)
            self.refresh_global_category_interfaces()
            dialog.destroy()

            if self.gpx_workshop_active:
                self.gpx_workshop_status_var.set(
                    f"Catégorie globale {'modifiée' if editing else 'ajoutée'} : {label}."
                )

        tk.Button(buttons, text="Annuler", command=dialog.destroy).pack(side="right", padx=4)
        tk.Button(
            buttons,
            text="Enregistrer",
            command=validate,
            bg="#27ae60",
            fg="white"
        ).pack(side="right", padx=4)

    def edit_selected_global_segment_category(self):
        category_id = self.get_selected_global_segment_category_id()
        if not category_id:
            messagebox.showwarning("Catégorie", "Sélectionnez une catégorie globale.")
            return
        self.open_global_segment_category_editor(category_id=category_id)

    def toggle_selected_global_segment_category(self):
        category_id = self.get_selected_global_segment_category_id()
        if not category_id:
            messagebox.showwarning("Catégorie", "Sélectionnez une catégorie globale.")
            return

        categories = self.read_global_segment_categories()
        category = next((c for c in categories if c.get("id") == category_id), None)
        if not category:
            return

        if category.get("system"):
            messagebox.showinfo(
                "Catégorie protégée",
                "La catégorie « Non classé » est nécessaire au fonctionnement du logiciel."
            )
            return

        usage = self.scan_segment_category_usage().get(category_id, {})
        used_segments = int(usage.get("segments", 0))

        new_state = not category.get("active", True)
        if not new_state and used_segments:
            if not messagebox.askyesno(
                "Désactiver une catégorie utilisée ?",
                (
                    f"« {category.get('label')} » est encore utilisée par "
                    f"{used_segments} référence(s) de segment.\n\n"
                    "Elle ne sera plus proposée pour de nouveaux classements, "
                    "mais restera lisible et exportable tant que des segments l’utilisent.\n\n"
                    "Continuer ?"
                )
            ):
                return

        for item in categories:
            if item.get("id") == category_id:
                item["active"] = new_state

        self.write_global_segment_categories(categories)
        self.refresh_global_category_interfaces()

    def show_selected_global_segment_category_usage(self):
        category_id = self.get_selected_global_segment_category_id()
        if not category_id:
            messagebox.showwarning("Catégorie", "Sélectionnez une catégorie globale.")
            return

        category = self.get_global_segment_category_by_id(category_id) or {"label": category_id}
        usage = self.scan_segment_category_usage().get(
            category_id,
            {"segments": 0, "bisses": set(), "catalogues": set()}
        )

        bisses = sorted(usage.get("bisses", set()))
        if bisses:
            details = "\n".join(f"• {name}" for name in bisses)
        else:
            details = "Aucun bisse connu."

        messagebox.showinfo(
            f"Utilisation — {category.get('label')}",
            (
                f"Références de segments : {usage.get('segments', 0)}\n"
                f"Bisses concernés : {len(bisses)}\n\n"
                f"{details}"
            )
        )

    def delete_selected_global_segment_category(self):
        category_id = self.get_selected_global_segment_category_id()
        if not category_id:
            messagebox.showwarning("Catégorie", "Sélectionnez une catégorie globale.")
            return

        category = self.get_global_segment_category_by_id(category_id)
        if not category:
            return

        if category.get("system"):
            messagebox.showerror(
                "Suppression impossible",
                "La catégorie « Non classé » est protégée."
            )
            return

        usage = self.scan_segment_category_usage().get(category_id, {})
        used_segments = int(usage.get("segments", 0))

        if used_segments:
            messagebox.showerror(
                "Suppression bloquée",
                (
                    f"« {category.get('label')} » est encore utilisée par "
                    f"{used_segments} référence(s) de segment dans "
                    f"{len(usage.get('bisses', set()))} bisse(s).\n\n"
                    "La future migration assistée permettra de la remplacer partout "
                    "avant suppression."
                )
            )
            return

        if not messagebox.askyesno(
            "Supprimer la catégorie ?",
            (
                f"Supprimer définitivement la catégorie globale "
                f"« {category.get('label')} » ?\n\n"
                "Une sauvegarde automatique du référentiel sera créée."
            )
        ):
            return

        categories = [
            item for item in self.read_global_segment_categories()
            if item.get("id") != category_id
        ]
        self.write_global_segment_categories(categories)
        self.refresh_global_category_interfaces()


    # ============================================================
    # V49 — NETTOYAGE / MIGRATION ASSISTÉS DES CATÉGORIES
    # ============================================================

    def normalize_category_match_key(self, value):
        value = str(value or "").strip().lower()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = value.replace("&", " et ")
        value = re.sub(r"[^a-z0-9]+", "", value)
        return value

    def category_cleanup_hash_file(self, path):
        try:
            digest = hashlib.sha256()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(1024 * 1024), b""):
                    digest.update(chunk)
            return digest.hexdigest()
        except Exception:
            return ""

    def category_cleanup_atomic_write_json(self, path, data):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        temp_path = f"{path}.category_migration_tmp_{uuid.uuid4().hex}"
        try:
            with open(temp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            os.replace(temp_path, path)
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    def collect_category_cleanup_catalogues(self):
        """
        Regroupe les catalogues connus par projet.

        Pour chaque projet, le catalogue local est prioritaire pour l'analyse,
        mais la migration mettra à jour toutes les copies existantes :
        catalogue local + copie Gestion_Bisses_Data.
        """
        descriptors = {}
        path_to_key = {}

        def add_descriptor(key, title, project_id, folder, paths):
            clean_paths = []
            seen_paths = set()

            for path in paths:
                if not path or not os.path.exists(path):
                    continue
                absolute = os.path.abspath(path)
                norm = os.path.normcase(absolute)
                if norm in seen_paths:
                    continue
                seen_paths.add(norm)
                clean_paths.append(absolute)

            if not clean_paths:
                return

            # Évite qu'un même catalogue soit attribué à deux descripteurs.
            existing_key = None
            for path in clean_paths:
                norm = os.path.normcase(path)
                if norm in path_to_key:
                    existing_key = path_to_key[norm]
                    break

            if existing_key:
                existing = descriptors[existing_key]
                for path in clean_paths:
                    norm = os.path.normcase(path)
                    if norm not in {
                        os.path.normcase(p) for p in existing["paths"]
                    }:
                        existing["paths"].append(path)
                        path_to_key[norm] = existing_key
                return

            descriptor_key = str(key or project_id or clean_paths[0])
            suffix = 2
            base_key = descriptor_key
            while descriptor_key in descriptors:
                descriptor_key = f"{base_key}_{suffix}"
                suffix += 1

            primary = None
            if folder:
                local = os.path.abspath(os.path.join(folder, "catalogue.json"))
                for path in clean_paths:
                    if os.path.normcase(path) == os.path.normcase(local):
                        primary = path
                        break
            if primary is None:
                primary = clean_paths[0]

            descriptors[descriptor_key] = {
                "key": descriptor_key,
                "title": title or os.path.basename(folder or os.path.dirname(primary)),
                "project_id": project_id or "",
                "folder": folder or "",
                "primary_path": primary,
                "paths": clean_paths
            }

            for path in clean_paths:
                path_to_key[os.path.normcase(path)] = descriptor_key

        for record in self.list_project_records():
            project_id = record.get("project_id", "")
            linked_folder = record.get("linked_folder", "")
            data_path = self.get_project_paths(project_id)["catalogue"]
            local_path = (
                os.path.join(linked_folder, "catalogue.json")
                if linked_folder
                else ""
            )

            add_descriptor(
                project_id,
                record.get("title", ""),
                project_id,
                linked_folder,
                [local_path, data_path]
            )

        if self.catalog_path and os.path.exists(self.catalog_path):
            active_title = ""
            try:
                active_raw = self.read_json_file_safe(self.catalog_path, {})
                active_title = (
                    (active_raw.get("bisse_info", {}) or {}).get("title")
                    or (active_raw.get("project", {}) or {}).get("title")
                    or os.path.basename(self.base_folder or os.path.dirname(self.catalog_path))
                )
            except Exception:
                active_title = os.path.basename(self.base_folder or os.path.dirname(self.catalog_path))

            add_descriptor(
                self.current_project_id or self.catalog_path,
                active_title,
                self.current_project_id,
                self.base_folder,
                [
                    self.catalog_path,
                    self.current_project_catalog_path
                ]
            )

        # Périmètre supplémentaire choisi manuellement.
        extra_root = self.category_cleanup_extra_root
        if extra_root and os.path.isdir(extra_root):
            for current, dirs, files in os.walk(extra_root):
                dirs[:] = [
                    d for d in dirs
                    if d not in {
                        "Export_Platform",
                        "Export_JPG",
                        "Gestion_Bisses_Data",
                        "_A_VERIFIER_AVANT_SUPPRESSION",
                        "_RAPPORTS_NETTOYAGE_GESTION_BISSES",
                        "__pycache__"
                    }
                ]
                if "catalogue.json" not in files:
                    continue

                path = os.path.join(current, "catalogue.json")
                raw = self.read_json_file_safe(path, {})
                title = (
                    (raw.get("bisse_info", {}) or {}).get("title")
                    or (raw.get("project", {}) or {}).get("title")
                    or os.path.basename(current)
                )
                slug = (
                    (raw.get("bisse_info", {}) or {}).get("slug")
                    or self.slugify(title)
                )

                add_descriptor(
                    f"extra_{slug}_{len(descriptors)}",
                    title,
                    "",
                    current,
                    [path]
                )

        result = list(descriptors.values())
        result.sort(key=lambda item: item.get("title", "").lower())
        return result

    def build_category_cleanup_inventory(self):
        global_categories = self.read_global_segment_categories(force_reload=True)
        global_by_id = {
            category.get("id"): category
            for category in global_categories
        }

        descriptors = self.collect_category_cleanup_catalogues()
        self.category_cleanup_catalogues = descriptors

        inventory = {}

        def ensure_row(category_id):
            category_id = str(category_id or "non_classe")
            if category_id not in inventory:
                inventory[category_id] = {
                    "id": category_id,
                    "global": False,
                    "global_record": None,
                    "labels": {},
                    "colors": {},
                    "file_codes": {},
                    "local_definition_count": 0,
                    "segment_count": 0,
                    "manual_segment_count": 0,
                    "bisses": set(),
                    "catalogues": set(),
                    "suggested_target": None,
                    "suggestion_level": "",
                    "suggestion_reason": ""
                }
            return inventory[category_id]

        for category in global_categories:
            row = ensure_row(category.get("id"))
            row["global"] = True
            row["global_record"] = copy.deepcopy(category)
            label = category.get("label") or category.get("id")
            row["labels"][label] = row["labels"].get(label, 0) + 100000
            color = category.get("color")
            if color:
                row["colors"][color] = row["colors"].get(color, 0) + 100000
            file_code = category.get("file_code")
            if file_code:
                row["file_codes"][file_code] = row["file_codes"].get(file_code, 0) + 100000

        for descriptor in descriptors:
            path = descriptor.get("primary_path")
            raw = self.read_json_file_safe(path, {})
            if not isinstance(raw, dict):
                continue

            title = descriptor.get("title") or os.path.basename(os.path.dirname(path))
            workshop = raw.get("gpx_workshop", {}) or {}

            for raw_category in workshop.get("categories", []) or []:
                category = self.normalize_segment_category_record(raw_category, 1000)
                row = ensure_row(category.get("id"))
                row["local_definition_count"] += 1
                row["bisses"].add(title)
                row["catalogues"].add(path)

                label = category.get("label") or category.get("id")
                row["labels"][label] = row["labels"].get(label, 0) + 1

                color = category.get("color")
                if color:
                    row["colors"][color] = row["colors"].get(color, 0) + 1

                file_code = category.get("file_code")
                if file_code:
                    row["file_codes"][file_code] = row["file_codes"].get(file_code, 0) + 1

            for segment in workshop.get("segments", []) or []:
                ids = [segment.get("category_id", "non_classe")]
                bicolor = segment.get("bicolor_categories")
                if isinstance(bicolor, list):
                    ids.extend(bicolor[:2])

                for category_id in dict.fromkeys(
                    category_id for category_id in ids if category_id
                ):
                    row = ensure_row(category_id)
                    row["segment_count"] += 1
                    row["bisses"].add(title)
                    row["catalogues"].add(path)

            traces = raw.get("gpx_traces", {}) or {}
            for record in traces.get("manual_segments", []) or []:
                category_id = record.get("category")
                if not category_id:
                    continue
                row = ensure_row(category_id)
                row["manual_segment_count"] += 1
                row["bisses"].add(title)
                row["catalogues"].add(path)

                label = record.get("label")
                if label:
                    row["labels"][label] = row["labels"].get(label, 0) + 1
                color = record.get("color")
                if color:
                    row["colors"][color] = row["colors"].get(color, 0) + 1

        def best_value(mapping, fallback=""):
            if not mapping:
                return fallback
            return sorted(
                mapping.items(),
                key=lambda item: (-item[1], str(item[0]).lower())
            )[0][0]

        global_keys = {}
        for category in global_categories:
            keys = {
                self.normalize_category_match_key(category.get("id")),
                self.normalize_category_match_key(category.get("label")),
                self.normalize_category_match_key(category.get("file_code"))
            }
            for key in keys:
                if key:
                    global_keys.setdefault(key, set()).add(category.get("id"))

        for category_id, row in inventory.items():
            global_record = row.get("global_record") or {}
            row["label"] = (
                global_record.get("label")
                or best_value(row.get("labels"), category_id)
                or category_id
            )
            row["color"] = (
                global_record.get("color")
                or best_value(row.get("colors"), "#666666")
                or "#666666"
            )
            row["file_code"] = (
                global_record.get("file_code")
                or best_value(row.get("file_codes"), category_id)
                or category_id
            )
            row["usage_count"] = (
                int(row.get("segment_count", 0))
                + int(row.get("manual_segment_count", 0))
            )

            if category_id in global_by_id:
                if row.get("local_definition_count", 0):
                    row["suggested_target"] = category_id
                    row["suggestion_level"] = "safe"
                    row["suggestion_reason"] = "Même identifiant que la catégorie globale"
                continue

            exact_candidates = set()
            source_keys = {
                self.normalize_category_match_key(category_id),
                self.normalize_category_match_key(row.get("label")),
                self.normalize_category_match_key(row.get("file_code"))
            }
            for key in source_keys:
                exact_candidates.update(global_keys.get(key, set()))

            if len(exact_candidates) == 1:
                target_id = next(iter(exact_candidates))
                row["suggested_target"] = target_id
                row["suggestion_level"] = "safe"
                row["suggestion_reason"] = "Nom, code ou identifiant équivalent"
                continue

            # Proposition floue, jamais ajoutée automatiquement au plan.
            source_key = self.normalize_category_match_key(row.get("label") or category_id)
            scored = []
            if source_key:
                for global_category in global_categories:
                    target_key = self.normalize_category_match_key(
                        global_category.get("label") or global_category.get("id")
                    )
                    if not target_key:
                        continue
                    ratio = difflib.SequenceMatcher(None, source_key, target_key).ratio()
                    scored.append((ratio, global_category.get("id")))

            scored.sort(reverse=True)
            if scored and scored[0][0] >= 0.86:
                if len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.08:
                    row["suggested_target"] = scored[0][1]
                    row["suggestion_level"] = "review"
                    row["suggestion_reason"] = f"Ressemblance probable ({scored[0][0]*100:.0f} %)"

        self.category_cleanup_inventory = inventory
        return inventory

    def category_cleanup_target_label(self, category_id):
        category = self.get_global_segment_category_by_id(category_id)
        if category:
            return f"{category.get('label')} [{category_id}]"
        return str(category_id or "")

    def refresh_category_cleanup_target_options(self):
        if self.category_cleanup_target_var is None:
            return []

        options = []
        lookup = {}
        for category in self.read_global_segment_categories(force_reload=True):
            label = f"{category.get('label')} [{category.get('id')}]"
            options.append(label)
            lookup[label] = category.get("id")

        self.category_cleanup_target_lookup = lookup
        return options

    def category_cleanup_plan_text(self, category_id):
        action = self.category_cleanup_plan.get(category_id)
        if not action:
            return ""

        action_type = action.get("action")
        if action_type == "merge":
            target_id = action.get("target_id")
            if target_id == category_id:
                return "Nettoyer le doublon local"
            return f"Fusionner → {self.category_cleanup_target_label(target_id)}"
        if action_type == "delete_local":
            return "Retirer la définition locale inutilisée"
        return ""

    def refresh_category_cleanup_tree(self, rescan=True):
        tree = self.category_cleanup_tree
        if not tree or not tree.winfo_exists():
            return

        selected = set(tree.selection())
        if rescan:
            inventory = self.build_category_cleanup_inventory()
        else:
            inventory = self.category_cleanup_inventory

        tree.delete(*tree.get_children())

        for category_id, row in sorted(
            inventory.items(),
            key=lambda item: (
                0 if item[1].get("global") else 1,
                str(item[1].get("label", "")).lower(),
                item[0]
            )
        ):
            if row.get("global"):
                scope = "Globale"
                if row.get("local_definition_count"):
                    scope += " + doublon local"
            elif row.get("local_definition_count"):
                scope = "Locale"
            else:
                scope = "Utilisée sans définition"

            usage = (
                f"{row.get('segment_count', 0)} atelier"
                f" + {row.get('manual_segment_count', 0)} exporté"
            )
            bisses = str(len(row.get("bisses", set())))

            suggestion = ""
            if row.get("suggested_target"):
                prefix = "Sûre" if row.get("suggestion_level") == "safe" else "À vérifier"
                suggestion = (
                    f"{prefix} → "
                    f"{self.category_cleanup_target_label(row.get('suggested_target'))}"
                )
            elif not row.get("usage_count") and row.get("local_definition_count"):
                suggestion = "Locale inutilisée"

            plan_text = self.category_cleanup_plan_text(category_id)

            tags = []
            if plan_text:
                tags.append("planned")
            elif row.get("suggestion_level") == "safe":
                tags.append("safe")
            elif row.get("suggestion_level") == "review":
                tags.append("review")
            elif not row.get("usage_count") and row.get("local_definition_count"):
                tags.append("unused")

            tree.insert(
                "",
                "end",
                iid=category_id,
                values=(
                    scope,
                    category_id,
                    row.get("label", ""),
                    usage,
                    bisses,
                    suggestion,
                    plan_text
                ),
                tags=tuple(tags)
            )

        tree.tag_configure("planned", background="#d5f5e3")
        tree.tag_configure("safe", background="#eafaf1")
        tree.tag_configure("review", background="#fcf3cf")
        tree.tag_configure("unused", background="#f2f3f4")

        for iid in selected:
            if tree.exists(iid):
                tree.selection_add(iid)

        options = self.refresh_category_cleanup_target_options()
        if hasattr(self, "category_cleanup_target_combo"):
            self.category_cleanup_target_combo["values"] = options
            current = self.category_cleanup_target_var.get()
            if current not in options:
                self.category_cleanup_target_var.set(options[0] if options else "")

        if hasattr(self, "category_cleanup_scope_label"):
            extra = (
                f" + {self.category_cleanup_extra_root}"
                if self.category_cleanup_extra_root
                else ""
            )
            self.category_cleanup_scope_label.config(
                text=(
                    f"Périmètre : {len(self.category_cleanup_catalogues)} projet(s) connu(s)"
                    f"{extra}"
                )
            )

    def show_category_cleanup_assistant(self):
        if self.category_cleanup_window and self.category_cleanup_window.winfo_exists():
            self.category_cleanup_window.deiconify()
            self.category_cleanup_window.lift()
            self.category_cleanup_window.focus_force()
            return

        window = tk.Toplevel(self.root)
        self.category_cleanup_window = window
        window.title("Nettoyage assisté des catégories")
        window.geometry("1260x760")
        window.minsize(1040, 650)
        window.transient(self.root)

        frame = tk.Frame(window, padx=12, pady=12)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text="Nettoyage et migration assistés des catégories",
            font=("Arial", 15, "bold")
        ).pack(anchor="w")

        tk.Label(
            frame,
            text=(
                "Le logiciel recense les catégories globales, les anciennes définitions "
                "locales et leurs utilisations. Rien n’est modifié avant le bouton "
                "« Appliquer le plan ». Les propositions floues restent toujours à valider."
            ),
            justify="left",
            anchor="w",
            wraplength=1160,
            fg="#555555"
        ).pack(fill="x", pady=(4, 8))

        scope_row = tk.Frame(frame)
        scope_row.pack(fill="x", pady=(0, 8))

        self.category_cleanup_scope_label = tk.Label(
            scope_row,
            text="Périmètre : projets connus",
            anchor="w",
            fg="#555555"
        )
        self.category_cleanup_scope_label.pack(side="left", fill="x", expand=True)

        tk.Button(
            scope_row,
            text="📂 Ajouter / changer un dossier racine à scanner",
            command=self.choose_category_cleanup_extra_root
        ).pack(side="right", padx=(5, 0))

        tk.Button(
            scope_row,
            text="Retirer le dossier supplémentaire",
            command=self.clear_category_cleanup_extra_root
        ).pack(side="right", padx=5)

        table_frame = tk.Frame(frame)
        table_frame.pack(fill="both", expand=True)

        columns = (
            "scope", "id", "label", "usage", "bisses", "suggestion", "plan"
        )
        tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="extended"
        )
        self.category_cleanup_tree = tree

        headings = {
            "scope": "Portée",
            "id": "Identifiant",
            "label": "Nom observé",
            "usage": "Utilisation",
            "bisses": "Bisses",
            "suggestion": "Proposition",
            "plan": "Plan validé"
        }
        widths = {
            "scope": 150,
            "id": 155,
            "label": 170,
            "usage": 145,
            "bisses": 60,
            "suggestion": 245,
            "plan": 270
        }

        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(
                column,
                width=widths[column],
                anchor="center" if column in {"bisses"} else "w"
            )

        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        safe_row = tk.Frame(frame)
        safe_row.pack(fill="x", pady=(8, 4))

        tk.Button(
            safe_row,
            text="✅ Ajouter toutes les suggestions sûres au plan",
            command=self.add_safe_category_cleanup_suggestions_to_plan,
            bg="#27ae60",
            fg="white"
        ).pack(side="left")

        tk.Button(
            safe_row,
            text="🗑️ Planifier le retrait des locales inutilisées sélectionnées",
            command=self.plan_delete_unused_local_categories
        ).pack(side="left", padx=6)

        tk.Button(
            safe_row,
            text="➕ Créer une catégorie globale depuis la sélection",
            command=self.create_global_category_from_cleanup_selection,
            bg="#8e44ad",
            fg="white"
        ).pack(side="left", padx=6)

        merge_row = tk.Frame(frame)
        merge_row.pack(fill="x", pady=4)

        tk.Label(merge_row, text="Fusionner les lignes sélectionnées vers :").pack(side="left")

        self.category_cleanup_target_var = tk.StringVar()
        self.category_cleanup_target_combo = ttk.Combobox(
            merge_row,
            textvariable=self.category_cleanup_target_var,
            state="readonly",
            width=42
        )
        self.category_cleanup_target_combo.pack(side="left", padx=6)

        tk.Button(
            merge_row,
            text="➡ Ajouter cette fusion au plan",
            command=self.plan_selected_category_cleanup_merge,
            bg="#2980b9",
            fg="white"
        ).pack(side="left")

        tk.Button(
            merge_row,
            text="Retirer la sélection du plan",
            command=self.remove_selected_categories_from_cleanup_plan
        ).pack(side="left", padx=6)

        tk.Button(
            merge_row,
            text="Réinitialiser tout le plan",
            command=self.reset_category_cleanup_plan
        ).pack(side="left", padx=6)

        bottom = tk.Frame(frame)
        bottom.pack(fill="x", pady=(10, 0))

        tk.Button(
            bottom,
            text="🔄 Rescanner",
            command=lambda: self.refresh_category_cleanup_tree(rescan=True)
        ).pack(side="left")

        tk.Button(
            bottom,
            text="👁 Aperçu du plan",
            command=self.preview_category_cleanup_plan_dialog
        ).pack(side="right", padx=5)

        tk.Button(
            bottom,
            text="✅ Appliquer le plan avec sauvegardes",
            command=self.apply_category_cleanup_plan,
            bg="#c56a00",
            fg="white"
        ).pack(side="right", padx=5)

        def close():
            self.category_cleanup_tree = None
            self.category_cleanup_window = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", close)
        self.refresh_category_cleanup_tree(rescan=True)

    def choose_category_cleanup_extra_root(self):
        folder = filedialog.askdirectory(
            title="Choisir un dossier racine contenant des bisses",
            initialdir=(
                self.category_cleanup_extra_root
                or (os.path.dirname(self.base_folder) if self.base_folder else self.app_folder)
            )
        )
        if not folder:
            return
        self.category_cleanup_extra_root = os.path.abspath(folder)
        self.refresh_category_cleanup_tree(rescan=True)

    def clear_category_cleanup_extra_root(self):
        self.category_cleanup_extra_root = ""
        self.refresh_category_cleanup_tree(rescan=True)

    def selected_category_cleanup_ids(self):
        tree = self.category_cleanup_tree
        if not tree or not tree.winfo_exists():
            return []
        return list(tree.selection())

    def add_safe_category_cleanup_suggestions_to_plan(self):
        self.build_category_cleanup_inventory()
        added = 0

        for category_id, row in self.category_cleanup_inventory.items():
            if row.get("suggestion_level") != "safe":
                continue

            target_id = row.get("suggested_target")
            if not target_id:
                continue

            # Un pur global sans doublon local n'a rien à nettoyer.
            if category_id == target_id and not row.get("local_definition_count"):
                continue

            self.category_cleanup_plan[category_id] = {
                "action": "merge",
                "target_id": target_id,
                "reason": row.get("suggestion_reason", "")
            }
            added += 1

        self.refresh_category_cleanup_tree(rescan=False)
        messagebox.showinfo(
            "Suggestions ajoutées",
            f"{added} proposition(s) sûre(s) ajoutée(s) au plan."
        )

    def plan_selected_category_cleanup_merge(self):
        selected = self.selected_category_cleanup_ids()
        if not selected:
            messagebox.showwarning("Nettoyage", "Sélectionnez une ou plusieurs catégories.")
            return

        display_value = self.category_cleanup_target_var.get()
        target_id = self.category_cleanup_target_lookup.get(display_value)
        if not target_id:
            messagebox.showwarning("Nettoyage", "Choisissez une catégorie globale cible.")
            return

        for source_id in selected:
            source_global = self.get_global_segment_category_by_id(source_id)
            if source_global and source_global.get("system") and source_id != target_id:
                messagebox.showerror(
                    "Fusion impossible",
                    "La catégorie système « Non classé » ne peut pas être fusionnée vers une autre catégorie."
                )
                return

        for source_id in selected:
            self.category_cleanup_plan[source_id] = {
                "action": "merge",
                "target_id": target_id,
                "reason": "Choix manuel"
            }

        self.refresh_category_cleanup_tree(rescan=False)

    def plan_delete_unused_local_categories(self):
        selected = self.selected_category_cleanup_ids()
        if not selected:
            messagebox.showwarning(
                "Nettoyage",
                "Sélectionnez une ou plusieurs catégories locales inutilisées."
            )
            return

        errors = []
        planned = 0

        for category_id in selected:
            row = self.category_cleanup_inventory.get(category_id, {})
            if row.get("global"):
                errors.append(f"{category_id} est globale")
                continue
            if not row.get("local_definition_count"):
                errors.append(f"{category_id} n'a pas de définition locale")
                continue
            if row.get("usage_count"):
                errors.append(f"{category_id} est encore utilisée")
                continue

            self.category_cleanup_plan[category_id] = {
                "action": "delete_local",
                "target_id": None,
                "reason": "Définition locale inutilisée"
            }
            planned += 1

        self.refresh_category_cleanup_tree(rescan=False)

        if errors:
            messagebox.showwarning(
                "Certaines lignes ont été ignorées",
                "\n".join(f"• {item}" for item in errors)
            )
        elif planned:
            messagebox.showinfo(
                "Plan mis à jour",
                f"{planned} définition(s) locale(s) inutilisée(s) ajoutée(s) au plan."
            )

    def remove_selected_categories_from_cleanup_plan(self):
        selected = self.selected_category_cleanup_ids()
        for category_id in selected:
            self.category_cleanup_plan.pop(category_id, None)
        self.refresh_category_cleanup_tree(rescan=False)

    def reset_category_cleanup_plan(self):
        if self.category_cleanup_plan:
            if not messagebox.askyesno(
                "Réinitialiser le plan",
                "Retirer toutes les actions actuellement planifiées ?"
            ):
                return
        self.category_cleanup_plan = {}
        self.refresh_category_cleanup_tree(rescan=False)

    def create_global_category_from_cleanup_selection(self):
        selected = self.selected_category_cleanup_ids()
        if len(selected) != 1:
            messagebox.showwarning(
                "Créer une catégorie globale",
                "Sélectionnez exactement une catégorie locale."
            )
            return

        category_id = selected[0]
        row = self.category_cleanup_inventory.get(category_id, {})

        if row.get("global"):
            messagebox.showinfo(
                "Catégorie déjà globale",
                "Cette catégorie existe déjà dans le référentiel global."
            )
            return

        label = row.get("label") or category_id
        color = row.get("color") or "#666666"
        file_code = row.get("file_code") or category_id

        if not messagebox.askyesno(
            "Créer une catégorie globale",
            (
                f"Créer la catégorie globale suivante ?\n\n"
                f"Identifiant : {category_id}\n"
                f"Nom : {label}\n"
                f"Couleur : {color}\n\n"
                "Elle sera ensuite ajoutée au plan pour retirer ses anciennes "
                "définitions locales redondantes."
            )
        ):
            return

        categories = self.read_global_segment_categories()
        max_order = max(
            [int(category.get("order", 0)) for category in categories] + [0]
        )
        categories.append({
            "id": category_id,
            "label": label,
            "file_code": file_code,
            "color": color,
            "order": max_order + 10,
            "active": True,
            "system": False
        })
        self.write_global_segment_categories(categories)

        self.category_cleanup_plan[category_id] = {
            "action": "merge",
            "target_id": category_id,
            "reason": "Nouvelle catégorie globale créée depuis une locale"
        }

        self.refresh_global_category_interfaces()
        self.refresh_category_cleanup_tree(rescan=True)

    def apply_category_cleanup_to_container(self, container, plan, migration_id=None):
        data = copy.deepcopy(container)
        stats = {
            "local_definitions_removed": 0,
            "segment_references_changed": 0,
            "bicolor_references_changed": 0,
            "manual_references_changed": 0,
            "changed": False
        }

        if not isinstance(data, dict):
            return data, stats

        workshop = data.get("gpx_workshop")
        if isinstance(workshop, dict):
            old_categories = workshop.get("categories", []) or []
            new_categories = []

            for raw_category in old_categories:
                category_id = str((raw_category or {}).get("id") or "")
                action = plan.get(category_id)

                if action and action.get("action") in {"merge", "delete_local"}:
                    stats["local_definitions_removed"] += 1
                    stats["changed"] = True
                    continue

                new_categories.append(raw_category)

            workshop["categories"] = new_categories

            for segment in workshop.get("segments", []) or []:
                old_id = segment.get("category_id", "non_classe")
                action = plan.get(old_id)
                if action and action.get("action") == "merge":
                    target_id = action.get("target_id")
                    if target_id and target_id != old_id:
                        segment["category_id"] = target_id
                        stats["segment_references_changed"] += 1
                        stats["changed"] = True

                bicolor = segment.get("bicolor_categories")
                if isinstance(bicolor, list):
                    changed_here = False
                    new_bicolor = []
                    for category_id in bicolor:
                        action = plan.get(category_id)
                        if action and action.get("action") == "merge":
                            target_id = action.get("target_id") or category_id
                            if target_id != category_id:
                                changed_here = True
                            new_bicolor.append(target_id)
                        else:
                            new_bicolor.append(category_id)

                    if changed_here:
                        segment["bicolor_categories"] = new_bicolor
                        stats["bicolor_references_changed"] += 1
                        stats["changed"] = True

        traces = data.get("gpx_traces")
        if isinstance(traces, dict):
            for record in traces.get("manual_segments", []) or []:
                old_id = record.get("category")
                action = plan.get(old_id)
                if not action or action.get("action") != "merge":
                    continue

                target_id = action.get("target_id")
                if not target_id or target_id == old_id:
                    continue

                record["category"] = target_id
                target_category = self.get_global_segment_category_by_id(target_id)
                if target_category:
                    record["label"] = target_category.get("label")
                    record["color"] = target_category.get("color")

                stats["manual_references_changed"] += 1
                stats["changed"] = True

        if stats["changed"] and migration_id:
            data["last_category_migration"] = {
                "migration_id": migration_id,
                "applied_at": datetime.now().isoformat(timespec="seconds")
            }

        return data, stats

    def build_global_categories_after_cleanup_plan(self, plan):
        categories = self.read_global_segment_categories()
        remove_ids = set()

        for source_id, action in plan.items():
            if action.get("action") != "merge":
                continue
            target_id = action.get("target_id")
            if not target_id or target_id == source_id:
                continue

            source = self.get_global_segment_category_by_id(source_id)
            if source and not source.get("system"):
                remove_ids.add(source_id)

        result = [
            category for category in categories
            if category.get("id") not in remove_ids
        ]
        return result, sorted(remove_ids)

    def calculate_category_cleanup_preview(self):
        if not self.category_cleanup_plan:
            return {
                "actions": 0,
                "catalogues": 0,
                "files": 0,
                "local_definitions_removed": 0,
                "segment_references_changed": 0,
                "bicolor_references_changed": 0,
                "manual_references_changed": 0,
                "global_categories_removed": [],
                "invalid": []
            }

        descriptors = self.collect_category_cleanup_catalogues()
        totals = {
            "actions": len(self.category_cleanup_plan),
            "catalogues": 0,
            "files": 0,
            "local_definitions_removed": 0,
            "segment_references_changed": 0,
            "bicolor_references_changed": 0,
            "manual_references_changed": 0,
            "global_categories_removed": [],
            "invalid": []
        }

        for source_id, action in self.category_cleanup_plan.items():
            if action.get("action") == "delete_local":
                row = self.category_cleanup_inventory.get(source_id, {})
                if row.get("usage_count"):
                    totals["invalid"].append(
                        f"{source_id} est encore utilisée et ne peut pas être retirée sans remplacement."
                    )
            elif action.get("action") == "merge":
                target_id = action.get("target_id")
                if not self.get_global_segment_category_by_id(target_id):
                    totals["invalid"].append(
                        f"La cible globale {target_id} n'existe pas."
                    )

        touched_catalogues = 0
        touched_files = 0

        for descriptor in descriptors:
            raw = self.read_json_file_safe(descriptor.get("primary_path"), {})
            _, stats = self.apply_category_cleanup_to_container(
                raw,
                self.category_cleanup_plan
            )
            if not stats.get("changed"):
                continue

            touched_catalogues += 1
            touched_files += len(descriptor.get("paths", []))

            for key in (
                "local_definitions_removed",
                "segment_references_changed",
                "bicolor_references_changed",
                "manual_references_changed"
            ):
                totals[key] += int(stats.get(key, 0))

        _, removed_global = self.build_global_categories_after_cleanup_plan(
            self.category_cleanup_plan
        )

        totals["catalogues"] = touched_catalogues
        totals["files"] = touched_files
        totals["global_categories_removed"] = removed_global
        return totals

    def category_cleanup_preview_text(self, preview):
        lines = [
            f"Actions planifiées : {preview.get('actions', 0)}",
            f"Projets / catalogues concernés : {preview.get('catalogues', 0)}",
            f"Fichiers catalogue à sauvegarder et modifier : {preview.get('files', 0)}",
            "",
            f"Définitions locales retirées : {preview.get('local_definitions_removed', 0)}",
            f"Références de segments modifiées : {preview.get('segment_references_changed', 0)}",
            f"Références bicolores modifiées : {preview.get('bicolor_references_changed', 0)}",
            f"GPX déjà exportés / importés mis à jour : {preview.get('manual_references_changed', 0)}",
        ]

        removed = preview.get("global_categories_removed", [])
        if removed:
            lines.extend([
                "",
                "Catégories globales fusionnées puis retirées du référentiel :",
                *[f"• {category_id}" for category_id in removed]
            ])

        invalid = preview.get("invalid", [])
        if invalid:
            lines.extend([
                "",
                "Blocages détectés :",
                *[f"• {message}" for message in invalid]
            ])

        return "\n".join(lines)

    def preview_category_cleanup_plan_dialog(self):
        self.build_category_cleanup_inventory()
        preview = self.calculate_category_cleanup_preview()

        if not self.category_cleanup_plan:
            messagebox.showinfo(
                "Aperçu du nettoyage",
                "Aucune action n'est encore planifiée."
            )
            return

        messagebox.showinfo(
            "Aperçu du nettoyage",
            self.category_cleanup_preview_text(preview)
        )

    def apply_category_cleanup_plan(self):
        self.build_category_cleanup_inventory()
        preview = self.calculate_category_cleanup_preview()

        if not self.category_cleanup_plan:
            messagebox.showwarning(
                "Nettoyage",
                "Aucune action n'est planifiée."
            )
            return

        if preview.get("invalid"):
            messagebox.showerror(
                "Plan non applicable",
                self.category_cleanup_preview_text(preview)
            )
            return

        confirmation_text = (
            self.category_cleanup_preview_text(preview)
            + "\n\n"
            + "Avant toute modification, une copie de chaque fichier concerné "
              "et du référentiel global sera créée dans Gestion_Bisses_Data/"
              "category_migrations.\n\n"
            + "Appliquer ce plan ?"
        )

        if not messagebox.askyesno(
            "Appliquer la migration des catégories",
            confirmation_text
        ):
            return

        migration_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        migration_root = os.path.join(
            self.app_data_folder,
            "category_migrations",
            migration_id
        )
        backup_root = os.path.join(migration_root, "backups")
        os.makedirs(backup_root, exist_ok=True)

        descriptors = self.collect_category_cleanup_catalogues()
        prepared_writes = []
        backup_records = []
        aggregated_stats = {
            "local_definitions_removed": 0,
            "segment_references_changed": 0,
            "bicolor_references_changed": 0,
            "manual_references_changed": 0
        }

        seen_paths = set()

        try:
            for descriptor_index, descriptor in enumerate(descriptors, start=1):
                for path_index, path in enumerate(descriptor.get("paths", []), start=1):
                    norm = os.path.normcase(os.path.abspath(path))
                    if norm in seen_paths or not os.path.exists(path):
                        continue
                    seen_paths.add(norm)

                    raw = self.read_json_file_safe(path, {})
                    migrated, stats = self.apply_category_cleanup_to_container(
                        raw,
                        self.category_cleanup_plan,
                        migration_id=migration_id
                    )

                    if not stats.get("changed"):
                        continue

                    backup_name = (
                        f"{descriptor_index:03d}_"
                        f"{self.slugify(descriptor.get('project_id') or descriptor.get('title') or 'bisse')}_"
                        f"{path_index}_{os.path.basename(path)}"
                    )
                    backup_path = os.path.join(backup_root, backup_name)
                    shutil.copy2(path, backup_path)

                    backup_records.append({
                        "original_path": path,
                        "backup_path": backup_path,
                        "sha256_before": self.category_cleanup_hash_file(path),
                        "project_id": descriptor.get("project_id", ""),
                        "title": descriptor.get("title", "")
                    })
                    prepared_writes.append({
                        "path": path,
                        "data": migrated
                    })

                    for key in aggregated_stats:
                        aggregated_stats[key] += int(stats.get(key, 0))

            old_global = self.read_json_file_safe(
                self.segment_categories_file,
                {"schema_version": "1.0", "categories": []}
            )
            new_global_categories, removed_global_ids = (
                self.build_global_categories_after_cleanup_plan(
                    self.category_cleanup_plan
                )
            )
            new_global = {
                "schema_version": "1.0",
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "last_migration_id": migration_id,
                "categories": new_global_categories
            }

            global_changed = (
                json.dumps(old_global, sort_keys=True, ensure_ascii=False)
                != json.dumps(new_global, sort_keys=True, ensure_ascii=False)
            )

            if global_changed:
                global_backup = os.path.join(
                    backup_root,
                    "segment_categories_global_before.json"
                )
                shutil.copy2(self.segment_categories_file, global_backup)
                backup_records.append({
                    "original_path": self.segment_categories_file,
                    "backup_path": global_backup,
                    "sha256_before": self.category_cleanup_hash_file(
                        self.segment_categories_file
                    ),
                    "project_id": "__global__",
                    "title": "Référentiel global"
                })
                prepared_writes.append({
                    "path": self.segment_categories_file,
                    "data": new_global
                })

            manifest = {
                "migration_id": migration_id,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "status": "prepared",
                "plan": copy.deepcopy(self.category_cleanup_plan),
                "preview": preview,
                "removed_global_categories": removed_global_ids,
                "backups": backup_records,
                "writes": [item["path"] for item in prepared_writes],
                "stats": aggregated_stats
            }

            manifest_path = os.path.join(migration_root, "manifest.json")
            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(manifest, f, indent=4, ensure_ascii=False)

            written_paths = []
            try:
                for item in prepared_writes:
                    self.category_cleanup_atomic_write_json(
                        item["path"],
                        item["data"]
                    )
                    written_paths.append(item["path"])
            except Exception:
                # Retour automatique à l'état précédent pour tous les fichiers
                # qui ont déjà été écrits.
                for record in reversed(backup_records):
                    original_path = record.get("original_path")
                    backup_path = record.get("backup_path")
                    if (
                        original_path in written_paths
                        and backup_path
                        and os.path.exists(backup_path)
                    ):
                        shutil.copy2(backup_path, original_path)
                raise

            for record in backup_records:
                record["sha256_after"] = self.category_cleanup_hash_file(
                    record.get("original_path")
                )

            manifest["status"] = "applied"
            manifest["applied_at"] = datetime.now().isoformat(timespec="seconds")
            manifest["backups"] = backup_records

            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(manifest, f, indent=4, ensure_ascii=False)

            # Recharge le catalogue actif s'il fait partie des fichiers modifiés.
            active_norm = (
                os.path.normcase(os.path.abspath(self.catalog_path))
                if self.catalog_path
                else ""
            )
            written_norms = {
                os.path.normcase(os.path.abspath(path))
                for path in written_paths
            }

            if active_norm and active_norm in written_norms:
                self.catalog_container = self.read_catalog_container_from_path(
                    self.catalog_path,
                    self.base_folder
                )
                self.catalog_data = self.catalog_container.get("photos", [])

            self.global_segment_categories = self.read_global_segment_categories(
                force_reload=True
            )
            self.category_cleanup_plan = {}
            self.refresh_global_category_interfaces()
            self.refresh_category_cleanup_tree(rescan=True)
            self.refresh_segment_categories_manager_tree()

            if self.gpx_workshop_active:
                try:
                    self.load_gpx_workshop_state()
                    self.refresh_gpx_category_tree()
                    self.refresh_gpx_category_combo()
                    self.refresh_gpx_segment_tree()
                    self.draw_gpx_workshop_map()
                except Exception as exc:
                    self.log(
                        f"⚠️ Migration appliquée, mais rafraîchissement de l’atelier incomplet : {exc}"
                    )

            messagebox.showinfo(
                "Migration terminée",
                (
                    "Le nettoyage des catégories a été appliqué.\n\n"
                    f"Catalogues concernés : {preview.get('catalogues', 0)}\n"
                    f"Fichiers modifiés : {len(written_paths)}\n"
                    f"Définitions locales retirées : "
                    f"{aggregated_stats.get('local_definitions_removed', 0)}\n"
                    f"Références modifiées : "
                    f"{aggregated_stats.get('segment_references_changed', 0) + aggregated_stats.get('bicolor_references_changed', 0) + aggregated_stats.get('manual_references_changed', 0)}\n\n"
                    f"Manifest et sauvegardes :\n{migration_root}"
                )
            )

        except Exception as exc:
            self.log(f"❌ Erreur migration catégories : {exc}")
            messagebox.showerror(
                "Migration interrompue",
                (
                    "La migration n'a pas pu être terminée.\n\n"
                    "Les fichiers déjà écrits ont été restaurés depuis leurs sauvegardes "
                    "lorsque cela était nécessaire.\n\n"
                    f"Erreur : {exc}\n\n"
                    f"Dossier de diagnostic :\n{migration_root}"
                )
            )

    def edit_selected_gpx_category(self):
        if not self.gpx_category_tree:
            return

        selection = self.gpx_category_tree.selection()
        if not selection:
            messagebox.showwarning("Catégorie", "Sélectionnez une catégorie.")
            return

        category_id = selection[0]
        global_category = self.get_global_segment_category_by_id(category_id)

        if global_category:
            self.open_global_segment_category_editor(category_id=category_id)
            return

        local = self.get_gpx_category_by_id(category_id)
        if not local:
            return

        if not messagebox.askyesno(
            "Catégorie locale",
            (
                f"« {local.get('label', category_id)} » est une ancienne catégorie locale "
                "de ce bisse.\n\n"
                "L’intégrer maintenant au référentiel global avec le même identifiant ?\n\n"
                "Cela ne modifie aucun segment : la catégorie deviendra simplement "
                "disponible pour tous les bisses."
            )
        ):
            return

        categories = self.read_global_segment_categories()
        preset = copy.deepcopy(local)
        preset["active"] = True
        preset["system"] = False
        preset["order"] = max(
            [int(c.get("order", 0)) for c in categories] + [0]
        ) + 10
        categories.append(preset)
        self.write_global_segment_categories(categories)
        self.refresh_global_category_interfaces()
        self.open_global_segment_category_editor(category_id=category_id)

    def empty_gpx_workshop(self):
        return {
            "categories": self.default_gpx_categories(),
            "sources": [],
            "segments": [],
            "last_export_at": None
        }

    def default_bisse_info(self):
        title = os.path.basename(self.base_folder) if self.base_folder else ""
        return {
            "slug": self.slugify(title) if title else "",
            "title": title,
            "region": "Valais",
            "commune": "",
            "description": "",
            "itinerary": "",
            "length_km": None,
            "altitude_min_m": None,
            "altitude_max_m": None,
            "difficulty": "",
            "marked_trail": None,
            "state": "",
            "tags": []
        }

    def default_platform_export_state(self):
        return {
            "last_export_at": None,
            "last_export_folder": "",
            "target_repo_name": "Bisses"
        }

    def slugify(self, value):
        value = str(value or "").strip().lower()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = re.sub(r"[^a-z0-9]+", "-", value)
        value = value.strip("-")
        return value or "bisse"

    def parse_optional_float(self, value):
        value = str(value or "").strip().replace(",", ".")
        if not value:
            return None
        try:
            return float(value)
        except Exception:
            return None

    def parse_optional_int(self, value):
        value = str(value or "").strip()
        if not value:
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    def parse_tags_text(self, value):
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
        return [
            tag.strip()
            for tag in str(value or "").replace(";", ",").split(",")
            if tag.strip()
        ]

    def inventory_expected_columns(self):
        """
        Colonnes attendues dans la source d'inventaire bisses-valais/MVB.
        Toutes sont conservées dans catalogue.json, même si la plateforme
        n'en affiche qu'une partie.
        """
        return [
            "nom",
            "slug",
            "url",
            "description",
            "itineraire",
            "region",
            "communes",
            "etat",
            "sentier",
            "cotation",
            "tags",
            "longueur_km",
            "altitude_haut_m",
            "altitude_bas_m",
            "prise_eau",
            "zones_irriguees",
            "autres_noms",
            "classement_ad",
            "sdt_numero",
            "sdt_importance",
            "post_modified"
        ]

    def empty_inventory_info(self):
        data = {key: "" for key in self.inventory_expected_columns()}
        data["_source_file"] = ""
        data["_source_sheet"] = ""
        data["_imported_at"] = None
        return data

    def normalize_inventory_header(self, header):
        value = str(header or "").strip().lower()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")

        aliases = {
            "nom_lien_vers_site": "nom",
            "nom_du_bisse": "nom",
            "name": "nom",
            "itinerary": "itineraire",
            "itineraire": "itineraire",
            "région": "region",
            "commune": "communes",
            "communes": "communes",
            "etat_du_bisse": "etat",
            "état": "etat",
            "etat": "etat",
            "sentier_balise": "sentier",
            "sentier_balisé": "sentier",
            "cote": "cotation",
            "cotes": "cotation",
            "cotation": "cotation",
            "longueur": "longueur_km",
            "longueur_km": "longueur_km",
            "altitude_haut": "altitude_haut_m",
            "altitude_haute": "altitude_haut_m",
            "altitude_max": "altitude_haut_m",
            "altitude_max_m": "altitude_haut_m",
            "altitude_bas": "altitude_bas_m",
            "altitude_basse": "altitude_bas_m",
            "altitude_min": "altitude_bas_m",
            "altitude_min_m": "altitude_bas_m",
            "prise_d_eau": "prise_eau",
            "prise_deau": "prise_eau",
            "zones_irriguees": "zones_irriguees",
            "zones_irriguees_": "zones_irriguees",
            "autres_nom": "autres_noms",
            "autres_noms": "autres_noms",
            "classement_ad": "classement_ad",
            "sdt_no": "sdt_numero",
            "sdt_numero": "sdt_numero",
            "sdt_num": "sdt_numero",
            "sdt_importance": "sdt_importance",
            "date_modification": "post_modified",
            "post_modified": "post_modified"
        }
        return aliases.get(value, value)

    def clean_inventory_cell(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def inventory_row_normalized(self, row, source_file="", source_sheet=""):
        normalized = self.empty_inventory_info()
        for key in self.inventory_expected_columns():
            normalized[key] = self.clean_inventory_cell(row.get(key, ""))
        normalized["_source_file"] = os.path.basename(source_file) if source_file else ""
        normalized["_source_sheet"] = source_sheet or ""
        normalized["_imported_at"] = datetime.now().isoformat(timespec="seconds")
        return normalized

    def ensure_openpyxl_available(self):
        """
        Rend l'import Excel autonome : si openpyxl manque, le logiciel tente
        de l'installer automatiquement dans l'environnement Python courant.

        Si l'installation échoue, on remonte une erreur claire.
        """
        try:
            import openpyxl  # noqa: F401
            return True
        except Exception:
            pass

        self.log("📦 openpyxl absent : tentative d'installation automatique...")
        try:
            subprocess.check_call([
                sys.executable,
                "-m",
                "pip",
                "install",
                "openpyxl"
            ])
            import openpyxl  # noqa: F401
            self.log("✅ openpyxl installé automatiquement.")
            return True
        except Exception as exc:
            raise RuntimeError(
                "Impossible d'installer automatiquement openpyxl.\n\n"
                "Essayez manuellement dans l'invite de commande :\n"
                f"{sys.executable} -m pip install openpyxl\n\n"
                f"Détail : {exc}"
            ) from exc

    def inventory_to_bisse_info_fields(self, inventory):
        """
        Convertit une ligne d'inventaire en champs actifs de bisse_info.
        C'est cette fiche unifiée qui est utilisée par le logiciel et la plateforme.
        """
        nom = str(inventory.get("nom") or "").strip()
        slug = str(inventory.get("slug") or "").strip()

        result = {
            "title": nom,
            "slug": self.slugify(slug or nom) if (slug or nom) else "",
            "region": str(inventory.get("region") or "").strip(),
            "commune": str(inventory.get("communes") or "").strip(),
            "communes": str(inventory.get("communes") or "").strip(),
            "description": str(inventory.get("description") or "").strip(),
            "itinerary": str(inventory.get("itineraire") or "").strip(),
            "length_km": self.parse_optional_float(inventory.get("longueur_km")),
            "altitude_min_m": self.parse_optional_int(inventory.get("altitude_bas_m")),
            "altitude_max_m": self.parse_optional_int(inventory.get("altitude_haut_m")),
            "difficulty": str(inventory.get("cotation") or "").strip(),
            "state": str(inventory.get("etat") or "").strip(),
            "tags": self.parse_tags_text(inventory.get("tags")),
            "marked_trail": self.infer_marked_trail_from_inventory(inventory.get("sentier")),
        }
        return result

    def comparable_value(self, value):
        if value is None:
            return ""
        if isinstance(value, list):
            return "|".join(str(v).strip().lower() for v in value if str(v).strip())
        if isinstance(value, bool):
            return "true" if value else "false"
        return str(value).strip().lower()

    def display_value_for_conflict(self, value):
        if value is None:
            return "—"
        if isinstance(value, list):
            return ", ".join(str(v) for v in value) or "—"
        if isinstance(value, bool):
            return "oui" if value else "non"
        text = str(value)
        return text if text.strip() else "—"


    def normalize_text_for_diff(self, value):
        """
        Normalisation prudente pour décider si deux morceaux de texte sont
        réellement différents.

        On n'altère pas le texte affiché ; on calme seulement la comparaison :
        - apostrophes/guillemets/tirets typographiques ;
        - espaces insécables ;
        - espaces multiples ;
        - accents composés.
        """
        value = str(value or "")
        value = unicodedata.normalize("NFKC", value)

        replacements = {
            "\u00a0": " ",
            "\u202f": " ",
            "\u2009": " ",
            "\u2007": " ",
            "\u2018": "'",
            "\u2019": "'",
            "\u201a": "'",
            "\u201b": "'",
            "\u2032": "'",
            "\u00b4": "'",
            "\u0060": "'",
            "\u201c": '"',
            "\u201d": '"',
            "\u201e": '"',
            "\u201f": '"',
            "\u2033": '"',
            "\u2010": "-",
            "\u2011": "-",
            "\u2012": "-",
            "\u2013": "-",
            "\u2014": "-",
            "\u2212": "-",
            "\r\n": "\n",
            "\r": "\n",
        }
        for old, new in replacements.items():
            value = value.replace(old, new)

        value = re.sub(r"[ \t\f\v]+", " ", value)
        value = re.sub(r" *\n *", "\n", value)
        return value.strip().lower()

    def values_equivalent_for_update(self, local_value, imported_value):
        """
        Équivalence prudente pour la mise à jour depuis la bibliothèque.

        Si deux valeurs sont équivalentes après normalisation, elles ne doivent
        pas être présentées comme conflit, même si leur encodage ou typographie
        diffère légèrement.
        """
        return self.normalize_text_for_diff(local_value) == self.normalize_text_for_diff(imported_value)


    def canonical_bisse_info_update_key(self, key):
        """
        Clé canonique pour la mise à jour depuis bibliothèque.
        commune/communes représentent la même information côté utilisateur.
        """
        if key == "communes":
            return "commune"
        return key

    def synchronize_bisse_info_aliases(self, info):
        """
        Maintient les alias historiques commune/communes cohérents sans
        demander deux fois la même décision à l'utilisateur.
        """
        if not isinstance(info, dict):
            return info

        commune = info.get("commune")
        communes = info.get("communes")

        if self.comparable_value(commune) == "" and self.comparable_value(communes) != "":
            info["commune"] = communes
        elif self.comparable_value(communes) == "" and self.comparable_value(commune) != "":
            info["communes"] = commune
        elif self.comparable_value(commune) != "":
            # commune est le champ actif principal ; communes reste alias export/compatibilité.
            info["communes"] = commune

        return info

    def split_text_for_diff(self, value):
        """
        Découpe le texte en unités stables : lignes/paragraphes conservés,
        sans découpage caractère par caractère.
        """
        value = self.display_value_for_conflict(value)
        if not value:
            return []

        # Conserver les retours à la ligne comme séparateurs visibles.
        raw_lines = value.splitlines(keepends=True)
        if not raw_lines:
            return [value]

        units = []
        for line in raw_lines:
            # Si une ligne est très longue, on découpe en phrases pour éviter
            # qu'un paragraphe entier soit marqué à cause d'une seule phrase.
            if len(line) > 420:
                parts = re.split(r"([.!?;:]\s+)", line)
                current = ""
                for part in parts:
                    current += part
                    if re.match(r"[.!?;:]\s+$", part):
                        units.append(current)
                        current = ""
                if current:
                    units.append(current)
            else:
                units.append(line)

        return units

    def tokenize_diff_unit(self, value):
        """
        Découpe une unité différente en mots / espaces / ponctuation.
        Les espaces et ponctuations cosmétiques ne sont pas le cœur du diff.
        """
        value = str(value or "")
        if not value:
            return []
        return re.findall(r"\w+(?:['’]\w+)?|\s+|[^\w\s]", value, flags=re.UNICODE)

    def normalized_token_for_diff(self, token):
        token = self.normalize_text_for_diff(token)
        # Neutraliser les tokens purement espaces.
        if token.isspace() or token == "":
            return ""
        return token

    def estimate_text_widget_height(self, value, chars_per_line=86, min_lines=3, max_lines=80):
        """
        Hauteur estimée d'une zone de comparaison non scrollable.
        Le but est d'éviter le double scroll : le texte est affiché en entier
        et seule la fenêtre globale défile.
        """
        value = str(value or "")
        if not value:
            return min_lines

        total = 0
        for paragraph in value.splitlines() or [value]:
            total += max(1, (len(paragraph) // chars_per_line) + 1)
        return max(min_lines, min(max_lines, total + 1))

    def insert_diff_highlighted_text(self, widget, own_value, other_value, side=None):
        """
        Insère own_value dans le widget, en comparant à other_value.

        v35 : correction fiable.
        Le premier argument affiché est toujours celui qui apparaît dans la colonne.
        Il n'y a plus d'inversion selon side, ce qui évite d'afficher la valeur
        locale dans la colonne bibliothèque.
        """
        own_text = self.display_value_for_conflict(own_value)
        other_text = self.display_value_for_conflict(other_value)

        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.tag_configure("diff", background="#fff2a8")
        widget.tag_configure("diff_strong", background="#ffd27f")
        widget.tag_configure("empty", foreground="#777777", font=("Arial", 9, "italic"))

        if own_text == "":
            widget.insert("end", "—", ("empty",))
            widget.configure(state="disabled")
            return

        own_units = self.split_text_for_diff(own_text)
        other_units = self.split_text_for_diff(other_text)

        own_norm = [self.normalize_text_for_diff(unit) for unit in own_units]
        other_norm = [self.normalize_text_for_diff(unit) for unit in other_units]

        line_matcher = difflib.SequenceMatcher(None, own_norm, other_norm, autojunk=False)

        def insert_token_diff(own_unit, other_unit):
            own_tokens = self.tokenize_diff_unit(own_unit)
            other_tokens = self.tokenize_diff_unit(other_unit)
            own_tok_norm = [self.normalized_token_for_diff(tok) for tok in own_tokens]
            other_tok_norm = [self.normalized_token_for_diff(tok) for tok in other_tokens]

            tok_matcher = difflib.SequenceMatcher(None, own_tok_norm, other_tok_norm, autojunk=False)
            for opcode, i1, i2, _j1, _j2 in tok_matcher.get_opcodes():
                chunk = "".join(own_tokens[i1:i2])
                if not chunk:
                    continue

                normalized_chunk = self.normalize_text_for_diff(chunk)
                meaningful = bool(re.search(r"[\w\d]", normalized_chunk, flags=re.UNICODE))

                if opcode == "equal" or not meaningful:
                    widget.insert("end", chunk)
                elif opcode == "replace":
                    widget.insert("end", chunk, ("diff_strong",))
                else:
                    widget.insert("end", chunk, ("diff",))

        for opcode, i1, i2, j1, j2 in line_matcher.get_opcodes():
            if opcode == "equal":
                for unit in own_units[i1:i2]:
                    widget.insert("end", unit)
                continue

            own_slice = own_units[i1:i2]
            other_slice = other_units[j1:j2]
            other_norm_set = {self.normalize_text_for_diff(unit) for unit in other_slice}

            if opcode == "replace" and len(own_slice) == len(other_slice):
                for own_unit, other_unit in zip(own_slice, other_slice):
                    if self.normalize_text_for_diff(own_unit) == self.normalize_text_for_diff(other_unit):
                        widget.insert("end", own_unit)
                    else:
                        insert_token_diff(own_unit, other_unit)
                continue

            for own_unit in own_slice:
                norm = self.normalize_text_for_diff(own_unit)
                meaningful = bool(re.search(r"[\w\d]", norm, flags=re.UNICODE))
                if norm in other_norm_set or not meaningful:
                    widget.insert("end", own_unit)
                else:
                    widget.insert("end", own_unit, ("diff",))

        widget.configure(state="disabled")

    def make_diff_text_box(self, parent, own_value, other_value, side=None, wheel_scroll_callback=None):
        """
        Crée une zone de texte complète, non scrollable, avec différences surlignées.

        own_value est toujours la valeur affichée dans cette colonne.
        other_value est seulement la valeur de comparaison.
        """
        box = tk.Text(
            parent,
            height=3,
            wrap="word",
            relief="sunken",
            borderwidth=1,
            padx=6,
            pady=5,
            takefocus=False
        )
        box.pack(fill="both", expand=True)

        self.insert_diff_highlighted_text(box, own_value, other_value, side=side)

        def _redirect_mousewheel(event):
            if callable(wheel_scroll_callback):
                wheel_scroll_callback(event)
            return "break"

        box.bind("<MouseWheel>", _redirect_mousewheel)
        box.bind("<Button-4>", _redirect_mousewheel)
        box.bind("<Button-5>", _redirect_mousewheel)

        for sequence in ("<Prior>", "<Next>", "<Up>", "<Down>", "<Home>", "<End>"):
            box.bind(sequence, lambda _event: "break")

        return box

    def bisse_info_field_labels(self):
        return {
            "title": "Nom",
            "slug": "Identifiant web / slug",
            "region": "Région",
            "commune": "Communes",
            "communes": "Communes",
            "description": "Description",
            "itinerary": "Itinéraire",
            "length_km": "Longueur km",
            "altitude_min_m": "Altitude min",
            "altitude_max_m": "Altitude max",
            "difficulty": "Cotation / difficulté",
            "state": "État du bisse",
            "tags": "Tags",
            "marked_trail": "Sentier balisé"
        }

    def read_inventory_table(self, file_path):
        """
        Lit un inventaire Excel (.xlsx/.xlsm) ou CSV.

        Excel :
        - openpyxl est importé ou installé automatiquement ;
        - la feuille Fusion_FR est utilisée si elle existe.

        CSV :
        - plusieurs encodages sont testés pour éviter les erreurs classiques
          d'export Excel Windows.
        """
        ext = os.path.splitext(file_path)[1].lower()
        rows = []

        if ext in (".xlsx", ".xlsm"):
            self.ensure_openpyxl_available()
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_name = "Fusion_FR" if "Fusion_FR" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]

            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                return []

            headers = [self.normalize_inventory_header(h) for h in raw_rows[0]]
            for values in raw_rows[1:]:
                item = {}
                for header, value in zip(headers, values):
                    if not header:
                        continue
                    item[header] = self.clean_inventory_cell(value)
                if any(str(v or "").strip() for v in item.values()):
                    rows.append(self.inventory_row_normalized(item, file_path, sheet_name))

            return rows

        if ext == ".csv":
            errors = []
            for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    rows = []
                    with open(file_path, "r", encoding=encoding, newline="") as f:
                        sample = f.read(8192)
                        f.seek(0)
                        try:
                            dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
                        except Exception:
                            dialect = csv.excel
                            # Excel francophone exporte souvent en point-virgule.
                            if ";" in sample:
                                dialect.delimiter = ";"
                        reader = csv.DictReader(f, dialect=dialect)
                        for raw in reader:
                            item = {}
                            for header, value in raw.items():
                                item[self.normalize_inventory_header(header)] = self.clean_inventory_cell(value)
                            if any(str(v or "").strip() for v in item.values()):
                                rows.append(self.inventory_row_normalized(item, file_path, ""))
                    self.log(f"📥 CSV lu avec l'encodage {encoding}.")
                    return rows
                except UnicodeDecodeError as exc:
                    errors.append(f"{encoding}: {exc}")
                    continue
                except Exception as exc:
                    errors.append(f"{encoding}: {exc}")
                    continue

            raise RuntimeError(
                "Impossible de lire le CSV avec les encodages testés : "
                "utf-8-sig, utf-8, cp1252, latin-1.\n\n"
                + "\n".join(errors[:4])
            )

        raise ValueError("Format non reconnu. Utilisez un fichier .xlsx, .xlsm ou .csv.")


    def inventory_match_score(self, row):
        title_candidates = []
        try:
            container = self.read_catalog_container()
            info = container.get("bisse_info", {})
            project = container.get("project", {})
            title_candidates.extend([
                info.get("slug", ""),
                info.get("title", ""),
                project.get("title", ""),
                project.get("bisse_name", "")
            ])
        except Exception:
            pass

        if self.base_folder:
            title_candidates.append(os.path.basename(self.base_folder))

        target_slugs = {
            self.slugify(value)
            for value in title_candidates
            if str(value or "").strip()
        }

        row_slug = self.slugify(row.get("slug") or row.get("nom") or "")
        row_nom_slug = self.slugify(row.get("nom") or "")

        score = 0
        if row_slug in target_slugs:
            score += 100
        if row_nom_slug in target_slugs:
            score += 90
        for target in target_slugs:
            if target and (target in row_slug or row_slug in target):
                score += 35
            if target and (target in row_nom_slug or row_nom_slug in target):
                score += 35

        return score

    def infer_marked_trail_from_inventory(self, sentier):
        value = str(sentier or "").strip().lower()
        if not value:
            return None
        if "pas de sentier" in value or "aucun" in value:
            return False
        if "balis" in value or "sentier" in value:
            return True
        return None

    def apply_inventory_row_to_current_catalog(self, row, field_choices=None):
        """
        Stocke toutes les colonnes dans inventory_info et alimente la fiche active bisse_info.

        field_choices :
        - champ -> "import" : utiliser la valeur importée ;
        - champ -> "local" : conserver la valeur locale.
        Les champs vides localement sont remplis automatiquement.
        """
        field_choices = field_choices or {}

        self.ensure_catalog_file_exists()
        container = self.read_catalog_container()
        inventory = self.inventory_row_normalized(row, row.get("_source_file", ""), row.get("_source_sheet", ""))

        # Préserver les métadonnées d'import si la ligne est déjà normalisée.
        for meta_key in ("_source_file", "_source_sheet", "_imported_at"):
            if row.get(meta_key):
                inventory[meta_key] = row.get(meta_key)

        info = container.setdefault("bisse_info", self.default_bisse_info())
        new_info = dict(info)
        imported_fields = self.inventory_to_bisse_info_fields(inventory)

        for key, imported_value in imported_fields.items():
            if self.comparable_value(imported_value) == "":
                continue

            local_value = new_info.get(key)
            choice = field_choices.get(key)

            if choice == "local":
                continue

            if choice == "import":
                new_info[key] = imported_value
                continue

            # Comportement par défaut : remplir les champs vides ou identiques,
            # ne pas écraser un champ local différent sans choix explicite.
            if self.comparable_value(local_value) == "":
                new_info[key] = imported_value
            elif self.comparable_value(local_value) == self.comparable_value(imported_value):
                new_info[key] = imported_value

        container["bisse_info"] = new_info
        container["inventory_info"] = inventory
        self.catalog_container = container
        self.catalog_data = container.get("photos", [])
        self.save_catalog()

        self.log(f"📥 Informations inventaire importées : {new_info.get('title', '')}")
        return new_info

    def show_inventory_conflict_dialog(self, row, after_import=None):
        """
        Affiche un écran unique de résolution des conflits.
        Les champs vides localement sont remplis automatiquement ;
        seuls les vrais conflits sont présentés.
        """
        self.ensure_catalog_file_exists()
        container = self.read_catalog_container()
        info = container.setdefault("bisse_info", self.default_bisse_info())

        inventory = self.inventory_row_normalized(row, row.get("_source_file", ""), row.get("_source_sheet", ""))
        for meta_key in ("_source_file", "_source_sheet", "_imported_at"):
            if row.get(meta_key):
                inventory[meta_key] = row.get(meta_key)

        imported_fields = self.inventory_to_bisse_info_fields(inventory)
        labels = self.bisse_info_field_labels()

        conflicts = []
        auto_fill = []

        for key, imported_value in imported_fields.items():
            if self.comparable_value(imported_value) == "":
                continue

            local_value = info.get(key)
            if self.comparable_value(local_value) == "":
                auto_fill.append(key)
            elif self.comparable_value(local_value) != self.comparable_value(imported_value):
                conflicts.append((key, local_value, imported_value))

        if not conflicts:
            self.apply_inventory_row_to_current_catalog(row, field_choices={key: "import" for key in auto_fill})
            messagebox.showinfo(
                "Import terminé",
                (
                    "Aucun conflit détecté.\n\n"
                    f"Champs remplis automatiquement : {len(auto_fill)}"
                )
            )
            if callable(after_import):
                after_import()
            else:
                self.show_bisse_info_editor()
            return

        window = tk.Toplevel(self.root)
        window.title("Résoudre les conflits d’import")
        window.geometry("1180x720")
        window.transient(self.root)
        window.grab_set()

        header = tk.Frame(window, padx=12, pady=10)
        header.pack(fill="x")

        tk.Label(
            header,
            text=(
                "L’import va remplir automatiquement les champs vides.\n"
                "Pour les champs déjà remplis différemment, choisissez la version à conserver."
            ),
            justify="left",
            anchor="w"
        ).pack(side="left", fill="x", expand=True)

        tk.Label(
            header,
            text=f"Automatiques : {len(auto_fill)} · Conflits : {len(conflicts)}",
            fg="#555555"
        ).pack(side="right")

        canvas = tk.Canvas(window, highlightthickness=0)
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=canvas.yview)
        scroll = tk.Frame(canvas, padx=12, pady=8)
        inner = canvas.create_window((0, 0), window=scroll, anchor="nw")

        def on_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfigure(inner, width=event.width)

        scroll.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        choices = {}

        def make_row(parent, row_index, key, local_value, imported_value):
            frame = tk.LabelFrame(parent, text=labels.get(key, key), padx=8, pady=8)
            frame.grid(row=row_index, column=0, sticky="ew", pady=6)
            frame.grid_columnconfigure(0, weight=1)
            frame.grid_columnconfigure(1, weight=1)

            var = tk.StringVar(value="local")
            choices[key] = var

            local_box = tk.Frame(frame)
            local_box.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
            import_box = tk.Frame(frame)
            import_box.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

            tk.Radiobutton(local_box, text="Garder la version actuelle", variable=var, value="local").pack(anchor="w")
            local_text = tk.Text(local_box, height=4, wrap="word")
            local_text.pack(fill="both", expand=True)
            local_text.insert("1.0", self.display_value_for_conflict(local_value))
            local_text.config(state="disabled")

            tk.Radiobutton(import_box, text="Utiliser la version importée", variable=var, value="import").pack(anchor="w")
            import_text = tk.Text(import_box, height=4, wrap="word")
            import_text.pack(fill="both", expand=True)
            import_text.insert("1.0", self.display_value_for_conflict(imported_value))
            import_text.config(state="disabled")

        scroll.grid_columnconfigure(0, weight=1)
        for i, (key, local_value, imported_value) in enumerate(conflicts):
            make_row(scroll, i, key, local_value, imported_value)

        buttons = tk.Frame(window, padx=12, pady=10)
        buttons.pack(fill="x")

        def set_all(value):
            for var in choices.values():
                var.set(value)

        def apply_choices():
            field_choices = {key: "import" for key in auto_fill}
            for key, var in choices.items():
                field_choices[key] = var.get()

            self.apply_inventory_row_to_current_catalog(row, field_choices=field_choices)
            window.destroy()
            messagebox.showinfo("Import terminé", "Les informations d’inventaire ont été appliquées.")
            if callable(after_import):
                after_import()
            else:
                self.show_bisse_info_editor()

        tk.Button(buttons, text="Tout garder local", command=lambda: set_all("local")).pack(side="left")
        tk.Button(buttons, text="Tout remplacer par l’import", command=lambda: set_all("import")).pack(side="left", padx=8)
        tk.Button(buttons, text="Annuler", command=window.destroy).pack(side="right", padx=(8, 0))
        tk.Button(
            buttons,
            text="Appliquer les choix",
            command=apply_choices,
            bg="#27ae60",
            fg="white"
        ).pack(side="right")


    def import_inventory_for_current_bisse_dialog(self, after_import=None):
        """
        Import direct depuis un Excel/CSV, conservé comme option ponctuelle.
        Pour l'usage courant, préférer la bibliothèque permanente v24.
        """
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return

        initial = os.path.dirname(self.base_folder) if self.base_folder else os.getcwd()
        file_path = filedialog.askopenfilename(
            title="Importer une source d'inventaire Excel / CSV",
            initialdir=initial,
            filetypes=[
                ("Inventaire Excel ou CSV", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if not file_path:
            return

        try:
            rows = self.read_inventory_table(file_path)
        except Exception as exc:
            messagebox.showerror("Import impossible", str(exc))
            return

        if not rows:
            messagebox.showwarning("Inventaire vide", "Aucune ligne lisible trouvée dans ce fichier.")
            return

        self.choose_inventory_row_from_rows(
            rows=rows,
            title="Choisir le bisse à importer",
            intro=(
                "Sélectionnez la ligne d’inventaire à importer. "
                "La fiche active sera remplie après résolution des conflits."
            ),
            after_select=lambda row: self.show_inventory_conflict_dialog(row, after_import=after_import)
        )

    def empty_catalog_container(self):
        """
        Structure moderne de catalogue local.

        Le fichier local reste riche : il contient les photos complètes,
        l'atelier GPX, les ressources de travail et les champs nécessaires
        à l'export vers la plateforme GitHub Pages « Bisses ».
        """
        title = os.path.basename(self.base_folder) if self.base_folder else ""
        return {
            "catalogue_version": 3,
            "schema_version": "0.2-local",
            "project": {
                "bisse_name": title,
                "title": title,
                "year": datetime.now().year,
                "source_folder": self.base_folder or "",
                "updated_at": datetime.now().isoformat(timespec="seconds")
            },
            "bisse_info": self.default_bisse_info(),
            "photos": [],
            "gpx_traces": {
                "manual_segments": [],
                "live_topo": []
            },
            "gpx_workshop": self.empty_gpx_workshop(),
            "external_resources": [],
            "inventory_info": self.empty_inventory_info(),
            "platform_export": self.default_platform_export_state()
        }

    def read_catalog_container(self):
        if not os.path.exists(self.catalog_path):
            self.catalog_container = self.empty_catalog_container()
            return self.catalog_container

        with open(self.catalog_path, "r", encoding="utf-8") as f:
            raw = json.load(f)

        # Compatibilité avec les anciennes versions : catalogue = simple liste de photos.
        if isinstance(raw, list):
            container = self.empty_catalog_container()
            container["photos"] = raw
            self.catalog_container = container
            return container

        if not isinstance(raw, dict):
            raise ValueError("Format de catalogue.json non reconnu.")

        raw.setdefault("catalogue_version", 3)
        raw.setdefault("schema_version", "0.2-local")
        raw.setdefault("project", {})
        raw["project"].setdefault("bisse_name", os.path.basename(self.base_folder) if self.base_folder else "")
        raw["project"].setdefault("title", raw["project"].get("bisse_name", ""))
        raw["project"].setdefault("year", datetime.now().year)
        raw["project"].setdefault("source_folder", self.base_folder or "")
        raw["project"].setdefault("updated_at", datetime.now().isoformat(timespec="seconds"))

        raw.setdefault("bisse_info", self.default_bisse_info())
        default_info = self.default_bisse_info()
        for key, value in default_info.items():
            raw["bisse_info"].setdefault(key, value)

        if not raw["bisse_info"].get("title"):
            raw["bisse_info"]["title"] = raw["project"].get("title") or raw["project"].get("bisse_name", "")
        if not raw["bisse_info"].get("slug"):
            raw["bisse_info"]["slug"] = self.slugify(raw["bisse_info"].get("title") or raw["project"].get("bisse_name", ""))

        raw.setdefault("photos", [])
        for index, entry in enumerate(raw.get("photos", []), start=1):
            if isinstance(entry, dict):
                entry.setdefault("platform_selected", False)
                entry.setdefault("platform_order", 0)
                entry.setdefault("platform_caption", "")

        raw.setdefault("gpx_traces", {})
        raw["gpx_traces"].setdefault("manual_segments", [])
        raw["gpx_traces"].setdefault("live_topo", [])

        raw.setdefault("gpx_workshop", self.empty_gpx_workshop())
        raw["gpx_workshop"].setdefault("categories", self.default_gpx_categories())
        raw["gpx_workshop"].setdefault("sources", [])
        raw["gpx_workshop"].setdefault("segments", [])
        raw["gpx_workshop"].setdefault("last_export_at", None)

        raw.setdefault("external_resources", [])
        raw.setdefault("inventory_info", self.empty_inventory_info())
        default_inventory = self.empty_inventory_info()
        for key, value in default_inventory.items():
            raw["inventory_info"].setdefault(key, value)
        raw.setdefault("platform_export", self.default_platform_export_state())
        raw["platform_export"].setdefault("last_export_at", None)
        raw["platform_export"].setdefault("last_export_folder", "")
        raw["platform_export"].setdefault("target_repo_name", "Bisses")

        self.catalog_container = raw
        return raw

    def save_catalog(self):
        if not isinstance(self.catalog_container, dict):
            self.catalog_container = self.empty_catalog_container()

        self.ensure_safe_before_save(interactive=True)

        self.catalog_container.setdefault("schema_version", "0.2-local")
        self.catalog_container.setdefault("project", {})
        self.catalog_container["project"]["bisse_name"] = os.path.basename(self.base_folder) if self.base_folder else ""
        self.catalog_container["project"].setdefault("title", self.catalog_container["project"].get("bisse_name", ""))
        self.catalog_container["project"].setdefault("year", datetime.now().year)
        self.catalog_container["project"]["source_folder"] = self.base_folder or ""
        self.catalog_container["project"]["updated_at"] = datetime.now().isoformat(timespec="seconds")

        self.catalog_container.setdefault("bisse_info", self.default_bisse_info())
        self.catalog_container["bisse_info"].setdefault(
            "slug",
            self.slugify(
                self.catalog_container["bisse_info"].get("title")
                or self.catalog_container["project"].get("bisse_name", "")
            )
        )
        self.catalog_container["photos"] = self.catalog_data

        self.catalog_container.setdefault("gpx_traces", {})
        self.catalog_container["gpx_traces"].setdefault("manual_segments", [])
        self.catalog_container["gpx_traces"].setdefault("live_topo", [])

        self.catalog_container.setdefault("gpx_workshop", self.empty_gpx_workshop())
        self.catalog_container["gpx_workshop"].setdefault("categories", self.default_gpx_categories())
        self.catalog_container["gpx_workshop"].setdefault("sources", [])
        self.catalog_container["gpx_workshop"].setdefault("segments", [])
        self.catalog_container["gpx_workshop"].setdefault("last_export_at", None)

        self.catalog_container.setdefault("external_resources", [])
        self.catalog_container.setdefault("inventory_info", self.empty_inventory_info())
        self.catalog_container.setdefault("platform_export", self.default_platform_export_state())

        # Mode sécurisé : écriture principale dans le dossier local actif.
        self.reset_active_catalog_paths_to_local()
        with open(self.catalog_path, "w", encoding="utf-8") as f:
            json.dump(self.catalog_container, f, indent=4, ensure_ascii=False)

        # Copie portable dans Data, sans jamais réécrire le local depuis Data.
        try:
            self.write_portable_data_copy_for_active_project()
        except Exception as exc:
            self.log(f"⚠️ Copie Data portable impossible : {exc}")

    def read_catalog(self):
        container = self.read_catalog_container()
        return container.get("photos", [])

    def ensure_catalog_file_exists(self):
        """
        Crée un catalogue.json vide si le dossier n'en a pas encore.
        Cela permet aux modules non-photo, notamment l'atelier GPX,
        de fonctionner indépendamment de la présence de photos.
        """
        self.catalog_container = self.read_catalog_container()
        self.catalog_data = self.catalog_container.get("photos", [])

        if not os.path.exists(self.catalog_path):
            self.save_catalog()
            self.log(f"💾 Catalogue vide créé pour le dossier : {self.catalog_path}")

        return self.catalog_container


    # ============================================================
    # MES BISSES / ESPACE DE TRAVAIL
    # ============================================================

    def get_workspace_path(self):
        """
        Liste persistante des dossiers bisses ouverts dans le logiciel.

        Fichier global du logiciel, toujours lu au démarrage depuis :
        Gestion_Bisses_Data/bisses_workspace.json
        """
        return self.get_app_data_path("bisses_workspace.json")

    def default_workspace(self):
        return {
            "schema_version": "0.1",
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "bisses": []
        }

    def read_workspace(self):
        path = self.get_workspace_path()
        if not os.path.exists(path):
            return self.default_workspace()
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return self.default_workspace()
            data.setdefault("schema_version", "0.1")
            data.setdefault("updated_at", None)
            data.setdefault("bisses", [])
            if not isinstance(data.get("bisses"), list):
                data["bisses"] = []
            return data
        except Exception:
            return self.default_workspace()

    def write_workspace(self, workspace):
        path = self.get_workspace_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        workspace["updated_at"] = datetime.now().isoformat(timespec="seconds")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(workspace, f, indent=2, ensure_ascii=False)
        self.log(f"💾 Liste Mes bisses sauvegardée : {path}")

    def add_folder_to_workspace(self, folder):
        if not folder:
            return

        folder_abs = os.path.abspath(folder)
        try:
            project_id = self.ensure_project_for_folder(folder_abs)
        except Exception:
            project_id = ""

        workspace = self.read_workspace()
        bisses = workspace.setdefault("bisses", [])

        now = datetime.now().isoformat(timespec="seconds")
        found = False

        for entry in bisses:
            same_project = project_id and entry.get("project_id") == project_id
            same_folder = os.path.abspath(entry.get("folder", "")) == folder_abs if entry.get("folder") else False
            if same_project or same_folder:
                entry["project_id"] = project_id or entry.get("project_id", "")
                entry["folder"] = folder_abs
                entry["last_opened_at"] = now
                found = True
                break

        if not found:
            bisses.append({
                "project_id": project_id,
                "folder": folder_abs,
                "first_opened_at": now,
                "last_opened_at": now
            })

        if project_id:
            self.set_project_hidden_from_workspace(project_id, False)

        bisses.sort(key=lambda entry: self.get_bisse_summary_from_project(entry.get("project_id", ""), entry.get("folder", "")).get("title", "").lower())
        self.write_workspace(workspace)

    def looks_like_bisse_folder(self, folder):
        """
        Détection volontairement souple d'un dossier bisse.
        Un dossier est accepté s'il contient au moins :
        - catalogue.json, ou
        - Photos/, ou
        - Fichiers GPX/, ou
        - Export_JPG/, ou
        - Export_Platform/.
        """
        if not folder or not os.path.isdir(folder):
            return False
        markers = [
            "catalogue.json",
            "Photos",
            "Fichiers GPX",
            "Export_JPG",
            "Export_Platform"
        ]
        return any(os.path.exists(os.path.join(folder, marker)) for marker in markers)

    def find_bisse_folders_in_parent(self, parent_folder):
        """
        Recherche les dossiers bisses directement sous un dossier parent.
        On reste volontairement au niveau immédiat pour éviter d'importer
        des sous-dossiers techniques comme Photos/ ou Fichiers GPX/.
        """
        if not parent_folder or not os.path.isdir(parent_folder):
            return []

        candidates = []
        for name in sorted(os.listdir(parent_folder), key=str.lower):
            path = os.path.join(parent_folder, name)
            if not os.path.isdir(path):
                continue
            if self.looks_like_bisse_folder(path):
                candidates.append(os.path.abspath(path))
        return candidates

    def add_multiple_bisse_folders_to_workspace_dialog(self):
        """
        Ajoute en une fois plusieurs dossiers bisses à Mes bisses.
        L'utilisateur choisit un dossier parent ; le logiciel propose ensuite
        les sous-dossiers détectés.
        """
        parent = filedialog.askdirectory(
            title="Choisir le dossier parent contenant plusieurs dossiers bisses",
            initialdir=self.get_default_collection_root()
        )
        if not parent:
            return

        candidates = self.find_bisse_folders_in_parent(parent)
        if not candidates:
            messagebox.showwarning(
                "Aucun dossier bisse détecté",
                (
                    "Aucun sous-dossier reconnu comme dossier bisse n'a été trouvé.\n\n"
                    "Un dossier est reconnu s'il contient catalogue.json, Photos/, "
                    "Fichiers GPX/, Export_JPG/ ou Export_Platform/."
                )
            )
            return

        window = tk.Toplevel(self.root)
        window.title("Ajouter plusieurs bisses à Mes bisses")
        window.geometry("980x620")
        window.transient(self.root)
        window.grab_set()

        tk.Label(
            window,
            text=(
                f"{len(candidates)} dossier(s) bisse détecté(s) dans :\n{parent}\n\n"
                "Décochez ceux que vous ne voulez pas ajouter."
            ),
            justify="left",
            anchor="w"
        ).pack(fill="x", padx=12, pady=(10, 8))

        canvas = tk.Canvas(window, highlightthickness=0)
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=canvas.yview)
        scroll = tk.Frame(canvas, padx=12, pady=8)
        inner = canvas.create_window((0, 0), window=scroll, anchor="nw")

        def on_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfigure(inner, width=event.width)

        scroll.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        choices = []

        for folder in candidates:
            summary = self.get_bisse_summary_from_folder(folder)
            var = tk.BooleanVar(value=True)
            choices.append((var, folder))

            row = tk.Frame(scroll)
            row.pack(fill="x", pady=3)

            tk.Checkbutton(row, variable=var).pack(side="left")

            label = (
                f"{summary.get('title') or os.path.basename(folder)}"
                f" · {summary.get('id') or self.slugify(os.path.basename(folder))}"
                f" · {'catalogue OK' if summary.get('has_catalog') else 'sans catalogue'}"
                f"\n{folder}"
            )
            tk.Label(row, text=label, justify="left", anchor="w").pack(side="left", fill="x", expand=True)

        buttons = tk.Frame(window, padx=12, pady=10)
        buttons.pack(fill="x")

        def set_all(value):
            for var, _folder in choices:
                var.set(value)

        def apply_import():
            selected = [folder for var, folder in choices if var.get()]
            if not selected:
                messagebox.showwarning("Aucune sélection", "Aucun dossier n'est sélectionné.")
                return

            for folder in selected:
                self.add_folder_to_workspace(folder)

            window.destroy()
            self.show_workspace_home()
            messagebox.showinfo(
                "Mes bisses",
                f"{len(selected)} dossier(s) ajouté(s) à Mes bisses."
            )

        tk.Button(buttons, text="Tout sélectionner", command=lambda: set_all(True)).pack(side="left")
        tk.Button(buttons, text="Tout désélectionner", command=lambda: set_all(False)).pack(side="left", padx=6)
        tk.Button(buttons, text="Annuler", command=window.destroy).pack(side="right")
        tk.Button(
            buttons,
            text="Ajouter les dossiers sélectionnés",
            command=apply_import,
            bg="#2c3e50",
            fg="white"
        ).pack(side="right", padx=(0, 8))

    def get_workspace_entries(self):
        """
        Retourne les projets visibles dans Mes bisses.

        v43 : la liste n'est plus dépendante uniquement des chemins locaux.
        Les projets présents dans Gestion_Bisses_Data/projects restent visibles
        même si leur dossier source n'existe pas sur cet ordinateur.
        """
        workspace = self.read_workspace()
        entries = []
        seen_projects = set()

        for entry in workspace.get("bisses", []):
            project_id = entry.get("project_id", "")
            folder = entry.get("folder", "")

            if not project_id and folder:
                try:
                    record = self.find_project_by_linked_folder(folder)
                    if record:
                        project_id = record.get("project_id", "")
                except Exception:
                    project_id = ""

            if project_id:
                seen_projects.add(project_id)

            summary = self.get_bisse_summary_from_project(project_id, folder) if project_id else self.get_bisse_summary_from_folder(folder)
            summary["last_opened_at"] = entry.get("last_opened_at", "")
            summary["first_opened_at"] = entry.get("first_opened_at", "")
            entries.append(summary)

        for record in self.list_project_records():
            project_id = record.get("project_id", "")
            if not project_id or project_id in seen_projects:
                continue
            if record.get("hidden_from_workspace"):
                continue
            summary = self.get_bisse_summary_from_project(project_id, record.get("linked_folder", ""))
            summary["last_opened_at"] = record.get("updated_at") or ""
            summary["first_opened_at"] = record.get("created_at") or ""
            entries.append(summary)

        entries.sort(key=lambda item: (item.get("title", "").lower(), item.get("id", "")))
        return entries

    def remove_folder_from_workspace(self, folder):
        if not folder:
            return
        folder_abs = os.path.abspath(folder)
        workspace = self.read_workspace()
        before = len(workspace.get("bisses", []))
        workspace["bisses"] = [
            entry for entry in workspace.get("bisses", [])
            if os.path.abspath(entry.get("folder", "")) != folder_abs
        ]
        after = len(workspace.get("bisses", []))
        self.write_workspace(workspace)
        self.log(f"🗑️ {before - after} bisse(s) retiré(s) de Mes bisses.")


    def remove_workspace_entry(self, project_id="", folder=""):
        workspace = self.read_workspace()
        before = len(workspace.get("bisses", []))
        folder_abs = os.path.abspath(folder) if folder else ""

        workspace["bisses"] = [
            entry for entry in workspace.get("bisses", [])
            if not (
                (project_id and entry.get("project_id") == project_id)
                or (folder_abs and entry.get("folder") and os.path.abspath(entry.get("folder", "")) == folder_abs)
            )
        ]

        if project_id:
            self.set_project_hidden_from_workspace(project_id, True)

        self.write_workspace(workspace)
        self.log(f"🗑️ {before - len(workspace.get('bisses', []))} entrée(s) retirée(s) de Mes bisses.")

    def get_bisse_summary_from_project(self, project_id, folder=""):
        project_id = self.slugify(project_id or "")
        folder = os.path.abspath(folder) if folder else ""

        record = self.read_project_record(project_id) if project_id else {}
        if not folder:
            folder = record.get("linked_folder", "")

        paths = self.get_project_paths(project_id) if project_id else {}
        data_catalog_path = paths.get("catalogue", "")
        local_catalog_path = os.path.join(folder, "catalogue.json") if folder else ""

        source_exists = bool(folder and os.path.isdir(folder))
        title = record.get("title") or (os.path.basename(folder) if folder else project_id)
        slug = record.get("slug") or self.slugify(title)
        region = ""
        commune = ""
        photos_selected = 0
        segments_count = 0
        has_catalog = os.path.exists(data_catalog_path) or os.path.exists(local_catalog_path)

        raw = None
        catalog_source = data_catalog_path if os.path.exists(data_catalog_path) else local_catalog_path
        if catalog_source and os.path.exists(catalog_source):
            try:
                with open(catalog_source, "r", encoding="utf-8") as f:
                    raw = json.load(f)

                if isinstance(raw, list):
                    photos = raw
                    info = {}
                    workshop = {}
                else:
                    photos = raw.get("photos", [])
                    info = raw.get("bisse_info", {}) or {}
                    workshop = raw.get("gpx_workshop", {}) or {}

                title = info.get("title") or (raw.get("project", {}).get("title") if isinstance(raw, dict) else "") or title
                slug = self.slugify(info.get("slug") or title)
                region = info.get("region", "")
                commune = info.get("commune", "")
                photos_selected = sum(1 for p in photos if isinstance(p, dict) and p.get("platform_selected"))
                segments_count = len(workshop.get("segments", []))
            except Exception as exc:
                return {
                    "project_id": project_id,
                    "id": slug,
                    "title": title,
                    "folder": folder,
                    "source_exists": source_exists,
                    "region": region,
                    "commune": commune,
                    "photos_selected": photos_selected,
                    "segments_count": segments_count,
                    "has_catalog": has_catalog,
                    "exportable": False,
                    "status": f"catalogue illisible : {exc}"
                }

        if not source_exists:
            status = "dossier source à relier"
            exportable = False
        elif has_catalog:
            status = "prêt"
            exportable = True
        else:
            status = "catalogue manquant"
            exportable = False

        return {
            "project_id": project_id,
            "id": slug,
            "title": title,
            "folder": folder,
            "source_exists": source_exists,
            "region": region,
            "commune": commune,
            "photos_selected": photos_selected,
            "segments_count": segments_count,
            "has_catalog": has_catalog,
            "exportable": exportable,
            "status": status
        }


    def resolve_project_id_reference(self, reference):
        """
        Résout une référence approximative vers un vrai project_id Data.

        Utile quand une ligne de Mes bisses vient d'un ancien workspace sans
        project_id, mais contient encore un slug, un titre ou un iid de Treeview.
        """
        if not reference:
            return ""

        ref_slug = self.slugify(str(reference))
        if not ref_slug:
            return ""

        try:
            paths = self.get_project_paths(ref_slug)
            if os.path.exists(paths["project"]):
                return ref_slug
        except Exception:
            pass

        for record in self.list_project_records():
            project_id = record.get("project_id", "")
            if not project_id:
                continue
            candidates = [
                project_id,
                record.get("slug", ""),
                record.get("title", "")
            ]
            if ref_slug in {self.slugify(value) for value in candidates if value}:
                return project_id

        return ""

    def resolve_project_id_for_workspace_entry(self, entry, iid=None):
        """
        Retrouve le project_id associé à une ligne de Mes bisses.
        Ne dépend pas du dossier source local : fonctionne aussi quand le
        dossier est justement à relier.
        """
        if not isinstance(entry, dict):
            entry = {}

        direct_candidates = [
            entry.get("project_id", ""),
            entry.get("id", ""),
            entry.get("slug", ""),
            iid or ""
        ]

        for candidate in direct_candidates:
            project_id = self.resolve_project_id_reference(candidate)
            if project_id:
                return project_id

        folder = entry.get("folder", "")
        if folder:
            try:
                record = self.find_project_by_linked_folder(folder)
                if record and record.get("project_id"):
                    return record.get("project_id")
            except Exception:
                pass

        title_slug = self.slugify(entry.get("title", ""))
        if title_slug:
            for record in self.list_project_records():
                if self.slugify(record.get("title", "")) == title_slug:
                    return record.get("project_id", "")

        return ""


    # ============================================================
    # SÉCURITÉ POST-RÉPARATION : identité, chemins, sauvegarde
    # ============================================================

    def is_generic_project_id(self, value):
        slug = self.slugify(value or "")
        return slug in {"", "bisse", "bisses", "projet", "project", "nouveau-bisse", "catalogue"}

    def catalog_identity(self, container):
        if not isinstance(container, dict):
            return {"title": "", "slug": ""}
        info = container.get("bisse_info", {}) or {}
        project = container.get("project", {}) or {}
        title = (
            info.get("title")
            or project.get("title")
            or project.get("bisse_name")
            or ""
        )
        slug = (
            info.get("slug")
            or self.slugify(title)
            or ""
        )
        return {"title": str(title or ""), "slug": self.slugify(slug)}

    def identity_tokens(self, value):
        slug = self.slugify(value or "")
        ignored = {
            "bisse", "bisses", "de", "du", "des", "d", "le", "la", "les",
            "local", "adrien", "musee", "inventaire", "test"
        }
        return {token for token in slug.split("-") if token and token not in ignored}

    def catalogue_seems_compatible_with_folder(self, container, folder):
        """
        Garde-fou contre les mélanges évidents de catalogues.

        v44b : corrige les faux positifs dus aux apostrophes et aux variantes
        de séparateurs :
        - Bisse d'Ergisch ↔ bisse-dergisch
        - Fiescherwyssa ↔ fiescher-wyssa
        """
        identity = self.catalog_identity(container)
        cat_slug = self.slugify(identity.get("slug") or identity.get("title") or "")
        folder_slug = self.slugify(os.path.basename(folder or ""))

        if not cat_slug or not folder_slug:
            return True

        def compact(value):
            return self.slugify(value).replace("-", "")

        cat_compact = compact(cat_slug)
        folder_compact = compact(folder_slug)

        if cat_slug == folder_slug or cat_slug in folder_slug or folder_slug in cat_slug:
            return True

        if cat_compact and folder_compact:
            if cat_compact == folder_compact or cat_compact in folder_compact or folder_compact in cat_compact:
                return True

        cat_tokens = self.identity_tokens(cat_slug)
        folder_tokens = self.identity_tokens(folder_slug)

        if not cat_tokens or not folder_tokens:
            return True

        if cat_tokens.intersection(folder_tokens):
            return True

        # Dernier filet pour les composés collés/décollés.
        cat_joined = "".join(sorted(cat_tokens))
        folder_joined = "".join(sorted(folder_tokens))
        if cat_joined and folder_joined and (
            cat_joined == folder_joined
            or cat_joined in folder_joined
            or folder_joined in cat_joined
        ):
            return True

        return False

    def reset_active_catalog_paths_to_local(self):
        """
        En mode sécurisé, le catalogue actif est le catalogue local du dossier bisse.
        Data garde une copie portable, mais ne peut plus écraser silencieusement le local.
        """
        if not self.base_folder:
            return
        expected_local = os.path.join(self.base_folder, "catalogue.json")
        self.local_catalog_path = expected_local
        self.catalog_path = expected_local

    def ensure_safe_before_save(self, interactive=True):
        """
        Vérifie qu'une sauvegarde ne va pas écrire le catalogue d'un bisse
        dans le dossier d'un autre bisse.
        """
        if not self.base_folder:
            raise RuntimeError("Aucun dossier bisse actif.")

        expected_local = os.path.abspath(os.path.join(self.base_folder, "catalogue.json"))
        current_catalog_path = os.path.abspath(self.catalog_path or "")
        current_local_path = os.path.abspath(getattr(self, "local_catalog_path", "") or expected_local)

        if current_catalog_path != expected_local:
            self.log(
                "🛡️ Sécurité : catalog_path ne pointait pas vers le catalogue local actif. "
                "Chemin corrigé avant sauvegarde."
            )
            self.catalog_path = expected_local

        if current_local_path != expected_local:
            self.log(
                "🛡️ Sécurité : local_catalog_path ne pointait pas vers le dossier actif. "
                "Chemin corrigé avant sauvegarde."
            )
            self.local_catalog_path = expected_local

        if not self.catalogue_seems_compatible_with_folder(self.catalog_container, self.base_folder):
            identity = self.catalog_identity(self.catalog_container)
            message = (
                "Le catalogue actif ne semble pas correspondre au dossier ouvert.\n\n"
                f"Dossier actif :\n{self.base_folder}\n\n"
                f"Catalogue : {identity.get('title') or identity.get('slug')}\n\n"
                "Par sécurité, la sauvegarde est bloquée pour éviter un nouvel écrasement."
            )
            self.log("⛔ Sauvegarde bloquée : identité catalogue/dossier incohérente.")
            if interactive:
                messagebox.showerror("Sauvegarde bloquée", message)
            raise RuntimeError("Sauvegarde bloquée : identité catalogue/dossier incohérente.")

        return True

    def write_portable_data_copy_for_active_project(self):
        """
        Écrit une copie portable dans Gestion_Bisses_Data/projects/<project_id>/catalogue.json.
        Ne modifie jamais le catalogue local.
        """
        if not self.base_folder or not isinstance(self.catalog_container, dict):
            return

        project_id = getattr(self, "current_project_id", "") or self.project_id_from_container_or_folder(
            self.base_folder,
            self.catalog_container
        )
        project_id = self.slugify(project_id or os.path.basename(self.base_folder))

        paths = self.get_project_paths(project_id)
        os.makedirs(paths["dir"], exist_ok=True)

        if os.path.exists(paths["catalogue"]):
            try:
                existing = self.read_catalog_container_from_path(paths["catalogue"], self.base_folder)
                if not self.json_data_equal(existing, self.catalog_container):
                    self.backup_catalog_file(project_id, paths["catalogue"], "data_avant_copie_locale")
            except Exception:
                pass

        self.write_catalog_container_to_path(
            paths["catalogue"],
            self.catalog_container,
            self.base_folder,
            update_timestamp=False
        )

        record = self.read_project_record(project_id)
        identity = self.catalog_identity(self.catalog_container)
        record["title"] = identity.get("title") or os.path.basename(self.base_folder)
        record["slug"] = identity.get("slug") or self.slugify(record["title"])
        record["linked_folder"] = os.path.abspath(self.base_folder)
        record["linked_folder_status"] = "ok"
        record["local_catalogue_path"] = os.path.join(self.base_folder, "catalogue.json")
        record["last_sync_direction"] = "local → Data copy"
        record["last_synced_at"] = datetime.now().isoformat(timespec="seconds")
        record["hidden_from_workspace"] = False
        self.write_project_record(project_id, record)

        self.current_project_id = project_id
        self.current_project_dir = paths["dir"]
        self.current_project_catalog_path = paths["catalogue"]

    def audit_data_integrity(self):
        """
        Produit un petit rapport texte dans Gestion_Bisses_Data pour repérer
        les projets Data suspects après réparation.
        """
        lines = []
        lines.append("Audit Gestion_Bisses_Data")
        lines.append(datetime.now().isoformat(timespec="seconds"))
        lines.append("")

        for record in self.list_project_records():
            project_id = record.get("project_id", "")
            paths = self.get_project_paths(project_id)
            linked = record.get("linked_folder", "")
            data_exists = os.path.exists(paths["catalogue"])
            local_path = os.path.join(linked, "catalogue.json") if linked else ""
            local_exists = bool(local_path and os.path.exists(local_path))

            lines.append(f"Projet : {project_id}")
            if self.is_generic_project_id(project_id):
                lines.append("  ⚠️ project_id générique : à mettre en quarantaine si un vrai projet existe déjà")
            lines.append(f"  titre record : {record.get('title', '')}")
            lines.append(f"  linked_folder : {linked}")
            lines.append(f"  dossier existe : {'oui' if linked and os.path.isdir(linked) else 'non'}")
            lines.append(f"  catalogue Data : {'oui' if data_exists else 'non'}")
            lines.append(f"  catalogue local : {'oui' if local_exists else 'non'}")

            try:
                if data_exists:
                    data_container = self.read_catalog_container_from_path(paths["catalogue"], linked or self.base_folder)
                    data_id = self.catalog_identity(data_container)
                    lines.append(f"  identité Data : {data_id.get('title')} / {data_id.get('slug')}")
                    lines.append(f"  Data compatible dossier : {'oui' if self.catalogue_seems_compatible_with_folder(data_container, linked) else 'NON'}")
                if local_exists:
                    local_container = self.read_catalog_container_from_path(local_path, linked)
                    local_id = self.catalog_identity(local_container)
                    lines.append(f"  identité locale : {local_id.get('title')} / {local_id.get('slug')}")
                    lines.append(f"  local compatible dossier : {'oui' if self.catalogue_seems_compatible_with_folder(local_container, linked) else 'NON'}")
            except Exception as exc:
                lines.append(f"  erreur lecture : {exc}")

            lines.append("")

        report_path = os.path.join(
            self.app_data_folder,
            f"audit_integrite_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        self.log(f"🛡️ Audit d'intégrité créé : {report_path}")
        return report_path

    def audit_data_integrity_dialog(self):
        try:
            report = self.audit_data_integrity()
            messagebox.showinfo(
                "Audit terminé",
                f"Rapport créé :\n{report}"
            )
        except Exception as exc:
            messagebox.showerror("Audit impossible", str(exc))

    def show_workspace_home(self):
        """
        Accueil compact : tout doit rester visible sans scroll vertical.

        v43 : Mes bisses affiche les projets Data, même si leur dossier source
        doit être relié sur cet ordinateur.
        """
        self.clear_main_frame()
        self.status_header.config(text="Mes bisses", fg="#34495e")

        outer = tk.Frame(self.main_frame)
        outer.pack(fill="both", expand=True)

        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(3, weight=1)

        header = tk.Frame(outer)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        header.grid_columnconfigure(1, weight=1)

        tk.Label(
            header,
            text="🏠 Mes bisses",
            font=("Arial", 18, "bold")
        ).grid(row=0, column=0, sticky="w")

        tk.Label(
            header,
            text=f"Données : {self.app_data_folder}",
            justify="left",
            anchor="w",
            fg="#666666"
        ).grid(row=0, column=1, sticky="ew", padx=12)

        tk.Button(
            header,
            text="🔄 Actualiser",
            command=self.show_workspace_home
        ).grid(row=0, column=2, sticky="e")

        action_area = tk.Frame(outer)
        action_area.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        for col in range(3):
            action_area.grid_columnconfigure(col, weight=1)

        work_actions = tk.LabelFrame(action_area, text="Mes bisses", padx=8, pady=6)
        work_actions.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        work_actions.grid_columnconfigure(0, weight=1)
        work_actions.grid_columnconfigure(1, weight=1)

        tk.Button(
            work_actions,
            text="📂 Ouvrir un dossier bisse",
            command=self.select_base_folder,
            bg="#2c3e50",
            fg="white"
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))

        tk.Button(
            work_actions,
            text="📁 Ajouter plusieurs",
            command=self.add_multiple_bisse_folders_to_workspace_dialog
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))

        publication_actions = tk.LabelFrame(action_area, text="Lot plateforme", padx=8, pady=6)
        publication_actions.grid(row=0, column=1, sticky="nsew", padx=4)
        publication_actions.grid_columnconfigure(0, weight=1)

        tk.Button(
            publication_actions,
            text="🌐 Publication / export",
            command=self.show_publication_module,
            bg="#1f618d",
            fg="white"
        ).grid(row=0, column=0, sticky="ew")

        library_actions = tk.LabelFrame(action_area, text="Bibliothèque", padx=8, pady=6)
        library_actions.grid(row=0, column=2, sticky="nsew", padx=(4, 0))
        library_actions.grid_columnconfigure(0, weight=1)
        library_actions.grid_columnconfigure(1, weight=1)

        tk.Button(
            library_actions,
            text="📚 Ouvrir",
            command=self.show_bisses_library_browser,
            bg="#6c3483",
            fg="white"
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))

        tk.Button(
            library_actions,
            text="🔄 Maj Excel/CSV",
            command=self.update_bisses_library_from_file_dialog
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))

        data_row = tk.Frame(outer)
        data_row.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        data_row.grid_columnconfigure(0, weight=1)

        tk.Label(
            data_row,
            text=(
                "Les catalogues de travail sont maintenant conservés dans "
                "Gestion_Bisses_Data/projects/. Les dossiers locaux restent des sources liées."
            ),
            justify="left",
            anchor="w",
            fg="#666666"
        ).grid(row=0, column=0, sticky="ew")

        tk.Button(
            data_row,
            text="📂 Données globales",
            command=self.open_app_data_folder
        ).grid(row=0, column=1, sticky="e", padx=(8, 0))

        entries = self.get_workspace_entries()
        entry_by_iid = {}

        center = tk.Frame(outer)
        center.grid(row=3, column=0, sticky="nsew")
        center.grid_rowconfigure(0, weight=1)
        center.grid_columnconfigure(0, weight=1)

        if not entries:
            empty = tk.LabelFrame(center, text="Aucun bisse dans la liste de travail", padx=14, pady=12)
            empty.grid(row=0, column=0, sticky="nsew")
            empty.grid_columnconfigure(0, weight=1)
            empty.grid_rowconfigure(0, weight=1)

            tk.Label(
                empty,
                text=(
                    "Ouvrez un dossier bisse ou copiez un dossier Gestion_Bisses_Data existant.\n"
                    "Les projets portables apparaîtront ici."
                ),
                justify="center",
                fg="#555555"
            ).grid(row=0, column=0)
        else:
            publication_folders = set()
            try:
                for item in self.read_publication_collection().get("bisses", []):
                    folder = item.get("folder", "")
                    if folder:
                        publication_folders.add(os.path.abspath(folder))
            except Exception:
                publication_folders = set()

            table_frame = tk.Frame(center)
            table_frame.grid(row=0, column=0, sticky="nsew")
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)

            columns = (
                "title", "slug", "region", "commune",
                "photos", "segments", "publication", "source", "status",
                "last_opened", "folder"
            )
            tree = ttk.Treeview(
                table_frame,
                columns=columns,
                show="headings",
                selectmode="extended",
                height=8
            )

            headings = {
                "title": "Bisse",
                "slug": "Slug",
                "region": "Région",
                "commune": "Communes",
                "photos": "Photos",
                "segments": "Segments",
                "publication": "Lot",
                "source": "Dossier source",
                "status": "Catalogue",
                "last_opened": "Dernière ouverture",
                "folder": "Dossier"
            }
            widths = {
                "title": 210,
                "slug": 140,
                "region": 115,
                "commune": 140,
                "photos": 70,
                "segments": 75,
                "publication": 55,
                "source": 115,
                "status": 120,
                "last_opened": 135,
                "folder": 360
            }

            for col in columns:
                tree.heading(col, text=headings[col])
                tree.column(col, width=widths[col], anchor="w")

            for col in ("photos", "segments", "publication", "source", "status"):
                tree.column(col, anchor="center")

            yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

            tree.grid(row=0, column=0, sticky="nsew")
            yscroll.grid(row=0, column=1, sticky="ns")
            xscroll.grid(row=1, column=0, sticky="ew")

            for entry in entries:
                folder = entry.get("folder", "")
                project_id = entry.get("project_id", "") or entry.get("id", "")
                iid = project_id or folder or f"entry-{len(entry_by_iid)}"
                if iid in entry_by_iid:
                    iid = f"{iid}-{len(entry_by_iid)}"

                source_ok = bool(entry.get("source_exists"))
                in_publication = os.path.abspath(folder) in publication_folders if folder else False
                entry_by_iid[iid] = entry

                tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        entry.get("title", ""),
                        entry.get("id", ""),
                        entry.get("region", ""),
                        entry.get("commune", ""),
                        entry.get("photos_selected", 0),
                        entry.get("segments_count", 0),
                        "✅" if in_publication else "—",
                        "✅" if source_ok else "🔗 à relier",
                        "✅" if entry.get("has_catalog") else "⚪",
                        entry.get("last_opened_at", ""),
                        folder if source_ok else "(dossier source à relier)"
                    )
                )

            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])

            def selected_entries():
                result = []
                for iid in tree.selection():
                    if iid not in entry_by_iid:
                        continue
                    entry = dict(entry_by_iid[iid])
                    entry["_iid"] = iid
                    if not entry.get("project_id"):
                        resolved = self.resolve_project_id_for_workspace_entry(entry, iid)
                        if resolved:
                            entry["project_id"] = resolved
                    result.append(entry)
                return result

            def selected_existing_folders():
                folders = []
                for entry in selected_entries():
                    folder = entry.get("folder", "")
                    if folder and os.path.isdir(folder):
                        folders.append(folder)
                return folders

            def open_selected():
                entries_sel = selected_entries()
                if not entries_sel:
                    messagebox.showwarning("Aucune sélection", "Sélectionnez un bisse à ouvrir.")
                    return

                entry = entries_sel[0]
                folder = entry.get("folder", "")
                if folder and os.path.isdir(folder):
                    self.load_folder(folder)
                    return

                project_id = self.resolve_project_id_for_workspace_entry(entry, entry.get("_iid", ""))
                if messagebox.askyesno(
                    "Dossier source à relier",
                    "Le dossier source de ce projet n'existe pas sur cet ordinateur.\n\nVoulez-vous le relier à un dossier local ?"
                ):
                    self.relink_project_to_folder_dialog(project_id)

            def relink_selected():
                entries_sel = selected_entries()
                if not entries_sel:
                    messagebox.showwarning("Aucune sélection", "Sélectionnez un projet à relier.")
                    return

                entry = entries_sel[0]
                project_id = self.resolve_project_id_for_workspace_entry(entry, entry.get("_iid", ""))
                self.relink_project_to_folder_dialog(project_id)

            def remove_selected():
                entries_sel = selected_entries()
                if not entries_sel:
                    messagebox.showwarning("Aucune sélection", "Sélectionnez un ou plusieurs bisses à retirer.")
                    return
                if messagebox.askyesno(
                    "Retirer de Mes bisses",
                    (
                        f"Retirer {len(entries_sel)} bisse(s) de la liste de travail ?\n\n"
                        "Les projets Data et les dossiers ne seront pas supprimés."
                    )
                ):
                    for entry in entries_sel:
                        self.remove_workspace_entry(entry.get("project_id", ""), entry.get("folder", ""))
                    self.show_workspace_home()

            def add_selected_to_publication():
                folders = selected_existing_folders()
                if not folders:
                    messagebox.showwarning(
                        "Dossier source à relier",
                        "Aucun des bisses sélectionnés n'a de dossier source local disponible."
                    )
                    return
                for folder in folders:
                    self.add_folder_to_publication_collection(folder)
                self.show_workspace_home()
                messagebox.showinfo(
                    "Lot plateforme",
                    f"{len(folders)} bisse(s) inclus dans le lot plateforme."
                )

            def update_selected_from_library():
                folders = selected_existing_folders()
                if not folders:
                    messagebox.showwarning(
                        "Dossier source à relier",
                        "Aucun des bisses sélectionnés n'a de dossier source local disponible."
                    )
                    return
                self.show_update_selected_bisses_from_library_dialog(folders)

            tree.bind("<Double-1>", lambda _event: open_selected())

        actions = tk.LabelFrame(outer, text="Actions sur la sélection", padx=8, pady=6)
        actions.grid(row=4, column=0, sticky="ew", pady=(6, 0))

        if not entries:
            tk.Label(
                actions,
                text="Aucun bisse sélectionnable pour le moment.",
                fg="#666666"
            ).pack(side="left")
            return

        tk.Button(
            actions,
            text="📂 Ouvrir la sélection",
            command=open_selected,
            bg="#2c3e50",
            fg="white"
        ).pack(side="left")

        tk.Button(
            actions,
            text="🔗 Relier à un dossier local",
            command=relink_selected
        ).pack(side="left", padx=6)

        tk.Button(
            actions,
            text="🌐 Inclure dans le lot",
            command=add_selected_to_publication
        ).pack(side="left", padx=6)

        tk.Button(
            actions,
            text="📚 Maj infos bibliothèque",
            command=update_selected_from_library
        ).pack(side="left", padx=6)

        tk.Button(
            actions,
            text="🗑️ Retirer",
            command=remove_selected
        ).pack(side="left", padx=6)

        tk.Button(
            actions,
            text="Tout sélectionner",
            command=lambda: tree.selection_set(tree.get_children())
        ).pack(side="right")

    def get_entry_image_path(self, entry):
        """
        Chemin réel de l'image utilisée par le logiciel.

        Logique :
        - si image_relative_path existe, on l'utilise ;
        - sinon compatibilité avec les anciens catalogues :
          Export_JPG/filename puis Photos/filename.
        """
        rel = entry.get("image_relative_path")
        if rel:
            return self.abs_from_base(rel)

        filename = entry.get("filename", "")

        candidate_export = os.path.join(self.export_folder, filename)
        if os.path.exists(candidate_export):
            return candidate_export

        candidate_photos = os.path.join(self.photos_folder, filename)
        if os.path.exists(candidate_photos):
            return candidate_photos

        return candidate_export

    def set_entry_image_path(self, entry, image_path):
        entry["image_relative_path"] = self.relative_to_base(image_path)
        entry["filename"] = os.path.basename(image_path)

    def get_capture_datetime_for_sort(self, image_path, entry=None):
        """
        Date de prise de vue pour le tri.
        Priorité :
        1. EXIF DateTimeOriginal
        2. EXIF DateTime
        3. date_taken du catalogue
        4. date de modification du fichier
        """
        try:
            img = Image.open(image_path)
            info = img._getexif()
            if info:
                for tag in (36867, 306):
                    date_str = info.get(tag)
                    if date_str:
                        return datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S")
        except Exception:
            pass

        if entry:
            date_taken = entry.get("date_taken")
            if date_taken:
                try:
                    return datetime.fromisoformat(date_taken.replace("Z", "+00:00")).replace(tzinfo=None)
                except Exception:
                    pass

        try:
            return datetime.fromtimestamp(os.path.getmtime(image_path))
        except Exception:
            return datetime.max

    # ============================================================
    # ANALYSE DU DOSSIER
    # ============================================================

    def select_base_folder(self):
        folder = filedialog.askdirectory(title="Sélectionnez le dossier du Bisse")
        if not folder:
            return

        if os.path.abspath(folder) != os.path.abspath(self.base_folder or folder):
            self.manual_photos_folder = ""

        self.load_folder(folder)

    def load_folder(self, folder):
        self.base_folder = os.path.abspath(folder)

        try:
            self.migrate_legacy_global_files_if_needed()
        except Exception as exc:
            self.log(f"⚠️ Migration des fichiers globaux impossible : {exc}")

        candidate_photos_folder = os.path.join(self.base_folder, "Photos")

        manual_is_valid = False
        if self.manual_photos_folder and os.path.isdir(self.manual_photos_folder):
            try:
                manual_is_valid = os.path.commonpath([
                    os.path.abspath(self.base_folder),
                    os.path.abspath(self.manual_photos_folder)
                ]) == os.path.abspath(self.base_folder)
            except Exception:
                manual_is_valid = False

        if manual_is_valid:
            self.photos_folder = self.manual_photos_folder
        elif os.path.exists(candidate_photos_folder) and os.path.isdir(candidate_photos_folder):
            self.photos_folder = candidate_photos_folder
        else:
            self.photos_folder = self.base_folder

        self.export_folder = os.path.join(self.base_folder, "Export_JPG")
        self.gpx_folder = os.path.join(self.base_folder, "Fichiers GPX")
        self.local_catalog_path = os.path.join(self.base_folder, "catalogue.json")
        self.catalog_path = self.local_catalog_path

        try:
            project_id = self.ensure_project_for_folder(self.base_folder)
            self.reset_active_catalog_paths_to_local()
            self.log(f"📦 Projet Data associé : {project_id} — catalogue actif local sécurisé")
        except Exception as exc:
            self.current_project_id = ""
            self.current_project_dir = ""
            self.current_project_catalog_path = ""
            self.catalog_path = self.local_catalog_path
            self.log(f"⚠️ Projet Data indisponible, utilisation du catalogue local : {exc}")

        try:
            self.add_folder_to_workspace(self.base_folder)
        except Exception as exc:
            self.log(f"⚠️ Impossible d'ajouter ce dossier à Mes bisses : {exc}")

        self.status_header.config(text=f"Dossier actif : {self.base_folder}", fg="green")
        self.log(f"📂 Analyse du dossier : {self.base_folder}")

        has_raw_photos = self.folder_has_images(self.photos_folder)

        has_export_photos = (
            os.path.exists(self.export_folder)
            and self.folder_has_images(self.export_folder)
        )

        has_catalog = os.path.exists(self.catalog_path)
        is_geolocated = False
        self.catalog_data = []

        if has_catalog:
            try:
                self.catalog_data = self.read_catalog()
                if any(str(entry.get("gps_sync", "")).startswith("OK") for entry in self.catalog_data):
                    is_geolocated = True
            except Exception as e:
                self.log(f"⚠️ Catalogue illisible : {e}")
                has_catalog = False
                self.catalog_data = []

        self.show_contextual_interface(
            has_raw=has_raw_photos,
            has_export=has_export_photos,
            has_cat=has_catalog,
            is_geo=is_geolocated
        )

    def folder_has_images(self, folder):
        if not folder or not os.path.exists(folder):
            return False

        valid_ext = (".heic", ".heif", ".jpg", ".jpeg")
        return any(
            f.lower().endswith(valid_ext)
            for f in os.listdir(folder)
            if os.path.isfile(os.path.join(folder, f))
        )

    def show_contextual_interface(self, has_raw, has_export, has_cat, is_geo):
        """
        Tableau de bord du dossier.

        v41 prudente :
        - le tableau de bord devient surtout un hub de synthèse et d'accès aux ateliers ;
        - aucune fonction n'est supprimée ;
        - les commandes techniques restent disponibles dans "Actions avancées / maintenance",
          replié par défaut ;
        - la logique GPX n'est pas changée ici, la détection automatique Fichiers GPX
          reste prévue pour v42.
        """
        self.clear_main_frame()

        outer = tk.Frame(self.main_frame)
        outer.pack(fill="both", expand=True)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(2, weight=1)

        # ------------------------------------------------------------------
        # Calculs d'état
        # ------------------------------------------------------------------
        photo_ok = 0
        photo_discarded = 0
        photo_geolocated = 0
        photo_titles = 0
        photo_descriptions = 0
        platform_selected_count = 0
        platform_missing_title = 0
        platform_missing_description = 0

        if has_cat:
            try:
                self.catalog_data = self.read_catalog()
                for entry in self.catalog_data:
                    status = entry.get("status")
                    if status == "OK":
                        photo_ok += 1
                        if entry.get("gps_coordinates"):
                            photo_geolocated += 1
                        if (entry.get("title") or "").strip():
                            photo_titles += 1
                        if (entry.get("description") or "").strip():
                            photo_descriptions += 1
                        if entry.get("platform_selected"):
                            platform_selected_count += 1
                            if not (entry.get("title") or "").strip():
                                platform_missing_title += 1
                            if not (entry.get("description") or "").strip():
                                platform_missing_description += 1
                    elif status == "SUPPRIMEE":
                        photo_discarded += 1
            except Exception:
                pass

        trace_summary = "Atelier GPX disponible même sans catalogue photo."
        workshop_sources = 0
        workshop_segments = 0
        categorized_segments = 0
        exported_segments = 0

        try:
            self.read_catalog_container()
            trace_summary = self.get_trace_summary_text()
            workshop = self.get_gpx_workshop_state()
            workshop_sources = len(workshop.get("sources", []))
            workshop_segments = len(workshop.get("segments", []))
            categorized_segments = sum(
                1 for seg in workshop.get("segments", [])
                if seg.get("category_id", "non_classe") != "non_classe"
            )
            traces = self.get_trace_sections()
            exported_segments = len(traces.get("manual_segments", []))
        except Exception as e:
            trace_summary = f"Résumé des tracés indisponible : {e}"

        bisse_info = {}
        try:
            bisse_info = self.read_catalog_container().get("bisse_info", {})
        except Exception:
            bisse_info = {}

        platform_title = bisse_info.get("title") or os.path.basename(self.base_folder or "")
        platform_slug = bisse_info.get("slug") or self.slugify(platform_title)

        # ------------------------------------------------------------------
        # En-tête synthétique
        # ------------------------------------------------------------------
        header = tk.Frame(outer)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        header.grid_columnconfigure(0, weight=1)

        title_text = platform_title or "Bisse actif"
        tk.Label(
            header,
            text=f"📁 {title_text}",
            font=("Arial", 17, "bold"),
            anchor="w"
        ).grid(row=0, column=0, sticky="ew")

        tk.Button(
            header,
            text="🏠 Mes bisses",
            command=self.show_workspace_home
        ).grid(row=0, column=1, sticky="e", padx=(8, 0))

        tk.Label(
            header,
            text=f"Dossier : {self.base_folder or '—'}",
            justify="left",
            anchor="w",
            fg="#666666",
            wraplength=1300
        ).grid(row=1, column=0, columnspan=2, sticky="ew", pady=(3, 0))

        # ------------------------------------------------------------------
        # Bandeau état court
        # ------------------------------------------------------------------
        status = tk.Frame(outer)
        status.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        for col in range(4):
            status.grid_columnconfigure(col, weight=1)

        def status_card(col, title, lines, accent=None):
            card = tk.LabelFrame(status, text=title, padx=8, pady=6)
            card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 4, 0 if col == 3 else 4))
            tk.Label(
                card,
                text="\n".join(lines),
                justify="left",
                anchor="w",
                fg=accent or "#333333",
                wraplength=310
            ).pack(fill="x")

        status_card(
            0,
            "Dossier",
            [
                f"Photos source : {'oui' if has_raw else 'non'}",
                f"Export JPG : {'oui' if has_export else 'non'}",
                f"Catalogue : {'oui' if has_cat else 'non'}"
            ]
        )

        status_card(
            1,
            "Photos",
            [
                f"Actives : {photo_ok}" if has_cat else "Actives : —",
                f"GPS : {photo_geolocated}/{photo_ok}" if has_cat else "GPS : —",
                f"Titre/desc. : {photo_titles}/{photo_descriptions}" if has_cat else "Titre/desc. : —"
            ]
        )

        status_card(
            2,
            "GPX",
            [
                f"Sources : {workshop_sources}",
                f"Segments : {workshop_segments}",
                f"Classés/exportés : {categorized_segments}/{exported_segments}"
            ]
        )

        status_card(
            3,
            "Plateforme",
            [
                f"Photos choisies : {platform_selected_count}",
                f"Sans titre : {platform_missing_title}",
                f"Sans description : {platform_missing_description}"
            ],
            accent="#7d3c98" if platform_missing_title or platform_missing_description else None
        )

        # ------------------------------------------------------------------
        # Cartes principales : accès aux ateliers / modules
        # ------------------------------------------------------------------
        modules = tk.Frame(outer)
        modules.grid(row=2, column=0, sticky="nsew")
        modules.grid_columnconfigure(0, weight=1)
        modules.grid_columnconfigure(1, weight=1)
        modules.grid_columnconfigure(2, weight=1)
        modules.grid_rowconfigure(0, weight=1)

        photos_frame = tk.LabelFrame(modules, text="📷 Photos", padx=10, pady=8)
        photos_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        gpx_frame = tk.LabelFrame(modules, text="🧭 Tracés GPX", padx=10, pady=8)
        gpx_frame.grid(row=0, column=1, sticky="nsew", padx=5)

        info_frame = tk.LabelFrame(modules, text="📝 Fiche / publication", padx=10, pady=8)
        info_frame.grid(row=0, column=2, sticky="nsew", padx=(5, 0))

        # Photos : accès principal + amorçage si nécessaire.
        tk.Label(
            photos_frame,
            text=(
                f"{photo_ok} photo(s) active(s), {photo_geolocated} géolocalisée(s).\n"
                "Travail courant : carte, visionneuse, titres, descriptions, sélection plateforme."
                if has_cat else
                "Aucun catalogue photo utilisable pour l’instant.\n"
                "L’atelier Photos peut s’ouvrir en état vide, mais il faut créer ou compléter le catalogue pour travailler sur les images."
            ),
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=410
        ).pack(fill="x", pady=(0, 8))

        tk.Button(
            photos_frame,
            text="⚙️ Préparation photos",
            command=self.show_photo_preparation_dialog,
            bg="#2c3e50" if has_raw and not has_cat else "#ecf0f1",
            fg="white" if has_raw and not has_cat else "black",
            height=2 if has_raw and not has_cat else 1
        ).pack(fill="x", pady=(0, 6))

        tk.Button(
            photos_frame,
            text="🗺️ Ouvrir l’atelier Photos",
            command=self.show_map_interface,
            bg="#2980b9",
            fg="white",
            height=2
        ).pack(fill="x", pady=4)

        # GPX : on garde seulement la porte d'entrée principale ici.
        tk.Label(
            gpx_frame,
            text=(
                f"{workshop_sources} source(s), {workshop_segments} segment(s), "
                f"{categorized_segments} segment(s) classé(s).\n"
                "Travail courant : import brut, sens amont → aval, segmentation, catégories, export."
            ),
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=410
        ).pack(fill="x", pady=(0, 8))

        tk.Button(
            gpx_frame,
            text="🧭 Ouvrir l’atelier Tracés GPX",
            command=self.show_gpx_workshop,
            bg="#d35400",
            fg="white",
            height=2
        ).pack(fill="x", pady=4)

        tk.Label(
            gpx_frame,
            text=trace_summary,
            justify="left",
            anchor="w",
            fg="#666666",
            wraplength=410
        ).pack(fill="x", pady=(8, 0))

        # Fiche / publication : portes d'entrée principales.
        tk.Label(
            info_frame,
            text=(
                f"Fiche : {platform_title or 'à compléter'}\n"
                f"Slug : {platform_slug or '—'}\n"
                f"Photos plateforme : {platform_selected_count}"
            ),
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=410
        ).pack(fill="x", pady=(0, 8))

        tk.Button(
            info_frame,
            text="📝 Informations générales",
            command=self.show_bisse_info_editor,
            bg="#34495e",
            fg="white",
            height=2
        ).pack(fill="x", pady=4)

        tk.Button(
            info_frame,
            text="🌐 Publication / export",
            command=self.show_publication_module,
            bg="#1f618d",
            fg="white",
            height=2
        ).pack(fill="x", pady=4)

        # ------------------------------------------------------------------
        # Maintenance : seule la maintenance transversale reste sur le dashboard.
        # Les fonctions Photos et GPX ont été replacées dans leurs modules.
        # ------------------------------------------------------------------
        maintenance_shell = tk.Frame(outer)
        maintenance_shell.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        maintenance_shell.grid_columnconfigure(0, weight=1)

        maintenance_visible = tk.BooleanVar(value=False)
        maintenance_content = tk.LabelFrame(
            maintenance_shell,
            text="Maintenance du dossier",
            padx=10,
            pady=8
        )

        maintenance_button = tk.Button(
            maintenance_shell,
            text="▶ Maintenance du dossier",
            anchor="w"
        )
        maintenance_button.grid(row=0, column=0, sticky="ew")

        def toggle_maintenance():
            if maintenance_visible.get():
                maintenance_content.grid_remove()
                maintenance_button.config(text="▶ Maintenance du dossier")
                maintenance_visible.set(False)
            else:
                maintenance_content.grid(row=1, column=0, sticky="ew", pady=(6, 0))
                maintenance_button.config(text="▼ Maintenance du dossier")
                maintenance_visible.set(True)

        maintenance_button.config(command=toggle_maintenance)

        tk.Button(
            maintenance_content,
            text="✅ Vérifier la structure du dossier",
            command=self.check_local_folder_structure
        ).pack(side="left", fill="x", expand=True, padx=(0, 6))

        tk.Button(
            maintenance_content,
            text="🛡️ Audit Data",
            command=self.audit_data_integrity_dialog
        ).pack(side="left", fill="x", expand=True, padx=(0, 6))

        tk.Label(
            maintenance_content,
            text=(
                "Les commandes Photos sont maintenant dans “Préparation photos”. "
                "Les commandes GPX sont dans l’atelier GPX, onglet “Import / branches”."
            ),
            justify="left",
            anchor="w",
            fg="#666666",
            wraplength=900
        ).pack(side="left", fill="x", expand=True, padx=(6, 0))

    def check_local_folder_structure(self):
        """
        Contrôle léger de la structure locale attendue.
        Ne déplace rien automatiquement : il crée seulement les dossiers
        standards manquants après confirmation.
        """
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return

        expected = [
            ("Photos", os.path.join(self.base_folder, "Photos")),
            ("Fichiers GPX", self.gpx_folder or os.path.join(self.base_folder, "Fichiers GPX")),
            ("Export_Platform", os.path.join(self.base_folder, "Export_Platform"))
        ]

        lines = [f"Dossier bisse : {self.base_folder}", ""]
        missing = []

        for label, path in expected:
            if os.path.isdir(path):
                lines.append(f"✅ {label}/")
            else:
                lines.append(f"⚪ {label}/ manquant")
                missing.append(path)

        if os.path.exists(self.catalog_path):
            lines.append("✅ catalogue.json")
        else:
            lines.append("⚪ catalogue.json manquant")

        if missing:
            if messagebox.askyesno(
                "Structure du dossier",
                "\n".join(lines) + "\n\nCréer les dossiers manquants ?"
            ):
                for path in missing:
                    os.makedirs(path, exist_ok=True)
                self.log("✅ Dossiers manquants créés.")
                self.load_folder(self.base_folder)
            else:
                self.log("Structure vérifiée, aucun dossier créé.")
        else:
            messagebox.showinfo("Structure du dossier", "\n".join(lines))

    def get_bisses_library_path(self):
        """
        Bibliothèque permanente issue du fichier consolidé
        bisses_valais_consolidation_api_mvb.xlsx.

        Fichier global du logiciel, toujours stocké dans :
        Gestion_Bisses_Data/bisses_library.json
        """
        return self.get_app_data_path("bisses_library.json")

    def default_bisses_library(self):
        return {
            "schema_version": "0.1",
            "source_file": "",
            "source_updated_at": None,
            "updated_at": None,
            "rows": []
        }

    def read_bisses_library(self):
        path = self.get_bisses_library_path()
        if not os.path.exists(path):
            return self.default_bisses_library()
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return self.default_bisses_library()
            data.setdefault("schema_version", "0.1")
            data.setdefault("source_file", "")
            data.setdefault("source_updated_at", None)
            data.setdefault("updated_at", None)
            data.setdefault("rows", [])
            if not isinstance(data.get("rows"), list):
                data["rows"] = []
            return data
        except Exception:
            return self.default_bisses_library()

    def write_bisses_library(self, library):
        path = self.get_bisses_library_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        library["updated_at"] = datetime.now().isoformat(timespec="seconds")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(library, f, indent=2, ensure_ascii=False)
        self.log(f"💾 Bibliothèque des bisses sauvegardée : {path}")

    def update_bisses_library_from_file(self, file_path):
        rows = self.read_inventory_table(file_path)
        if not rows:
            raise ValueError("Aucune ligne lisible trouvée dans ce fichier.")

        rows.sort(key=lambda r: (str(r.get("nom") or "").lower(), str(r.get("slug") or "").lower()))

        library = self.default_bisses_library()
        library["source_file"] = os.path.abspath(file_path)
        try:
            library["source_updated_at"] = datetime.fromtimestamp(
                os.path.getmtime(file_path)
            ).isoformat(timespec="seconds")
        except Exception:
            library["source_updated_at"] = None
        library["rows"] = rows
        self.write_bisses_library(library)
        return library

    def update_bisses_library_from_file_dialog(self, after_update=None):
        initial = self.get_default_collection_root()
        file_path = filedialog.askopenfilename(
            title="Mettre à jour la bibliothèque depuis le fichier consolidé",
            initialdir=initial,
            filetypes=[
                ("Inventaire consolidé Excel ou CSV", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if not file_path:
            return

        try:
            library = self.update_bisses_library_from_file(file_path)
        except Exception as exc:
            messagebox.showerror("Mise à jour impossible", str(exc))
            return

        messagebox.showinfo(
            "Bibliothèque mise à jour",
            (
                f"Bibliothèque mise à jour :\n{self.get_bisses_library_path()}\n\n"
                f"Lignes importées : {len(library.get('rows', []))}"
            )
        )
        if callable(after_update):
            after_update()

    def choose_inventory_row_from_rows(self, rows, title, intro, after_select):
        """
        Fenêtre commune de sélection d'une ligne d'inventaire, utilisée
        à la fois pour l'import direct Excel/CSV et pour la bibliothèque locale.
        """
        if not rows:
            messagebox.showwarning("Inventaire vide", "Aucune ligne disponible.")
            return

        rows = sorted(rows, key=lambda r: (-self.inventory_match_score(r), str(r.get("nom") or "").lower()))

        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("950x520")
        window.transient(self.root)
        window.grab_set()

        tk.Label(
            window,
            text=intro,
            justify="left"
        ).pack(fill="x", padx=12, pady=(10, 6))

        search_var = tk.StringVar(value="")
        search_row = tk.Frame(window)
        search_row.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(search_row, text="Filtrer :").pack(side="left")
        search_entry = tk.Entry(search_row, textvariable=search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        columns = ("score", "nom", "slug", "region", "communes", "etat")
        tree = ttk.Treeview(window, columns=columns, show="headings", selectmode="browse", height=14)
        for col, label, width in (
            ("score", "Score", 60),
            ("nom", "Nom", 230),
            ("slug", "Slug", 170),
            ("region", "Région", 150),
            ("communes", "Communes", 190),
            ("etat", "État", 120)
        ):
            tree.heading(col, text=label)
            tree.column(col, width=width, anchor="w")
        tree.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        current_rows = []

        def populate():
            nonlocal current_rows
            query = self.slugify(search_var.get())
            tree.delete(*tree.get_children())
            current_rows = []
            for row in rows:
                haystack = " ".join(str(row.get(key, "")) for key in ("nom", "slug", "region", "communes", "etat", "autres_noms"))
                if query and query not in self.slugify(haystack):
                    continue
                current_rows.append(row)
                item_id = str(len(current_rows) - 1)
                tree.insert(
                    "",
                    "end",
                    iid=item_id,
                    values=(
                        self.inventory_match_score(row),
                        row.get("nom", ""),
                        row.get("slug", ""),
                        row.get("region", ""),
                        row.get("communes", ""),
                        row.get("etat", "")
                    )
                )
            if current_rows:
                tree.selection_set("0")
                tree.focus("0")

        def import_selected():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Aucune sélection", "Sélectionnez une ligne d’inventaire.")
                return
            row = current_rows[int(selection[0])]
            window.destroy()
            after_select(row)

        search_var.trace_add("write", lambda *_: populate())
        tree.bind("<Double-1>", lambda _event: import_selected())

        buttons = tk.Frame(window)
        buttons.pack(fill="x", padx=12, pady=(0, 12))
        tk.Button(buttons, text="Annuler", command=window.destroy).pack(side="left")
        tk.Button(
            buttons,
            text="📥 Utiliser la ligne sélectionnée",
            command=import_selected,
            bg="#1f618d",
            fg="white"
        ).pack(side="right")

        populate()
        search_entry.focus_set()

    def fill_current_bisse_from_library_dialog(self, after_import=None):
        """
        Remplit la fiche du bisse ouvert depuis la bibliothèque permanente.
        Si la bibliothèque n'existe pas encore, propose de la créer depuis Excel/CSV.
        """
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return

        library = self.read_bisses_library()
        rows = library.get("rows", [])

        if not rows:
            if messagebox.askyesno(
                "Bibliothèque absente",
                (
                    "Aucune bibliothèque locale n'est encore disponible.\n\n"
                    "Voulez-vous l'importer maintenant depuis le fichier Excel consolidé ?"
                )
            ):
                self.update_bisses_library_from_file_dialog(
                    after_update=lambda: self.fill_current_bisse_from_library_dialog(after_import=after_import)
                )
            return

        source = library.get("source_file") or "source inconnue"
        updated = library.get("updated_at") or "—"

        self.choose_inventory_row_from_rows(
            rows=rows,
            title="Remplir depuis la bibliothèque des bisses",
            intro=(
                "Sélectionnez le bisse de référence dans la bibliothèque locale.\n"
                f"Source : {source}\n"
                f"Mise à jour bibliothèque : {updated}"
            ),
            after_select=lambda row: self.show_inventory_conflict_dialog(row, after_import=after_import)
        )

    def show_bisse_info_editor(self):
        """
        Éditeur des informations générales du bisse.

        v24 :
        - bibliothèque permanente des bisses dans bisses_library.json ;
        - mise à jour de la bibliothèque depuis le fichier Excel/CSV consolidé ;
        - remplissage de la fiche active depuis cette bibliothèque ;
        - import direct Excel/CSV conservé comme option ponctuelle.
        """
        self.ensure_catalog_file_exists()
        container = self.read_catalog_container()
        info = container.setdefault("bisse_info", self.default_bisse_info())
        resources = container.setdefault("external_resources", [])
        inventory = container.setdefault("inventory_info", self.empty_inventory_info())
        default_inventory = self.empty_inventory_info()
        for key, value in default_inventory.items():
            inventory.setdefault(key, value)

        self.clear_main_frame()
        self.status_header.config(text="Informations générales du bisse", fg="#34495e")

        outer = tk.Frame(self.main_frame)
        outer.pack(fill="both", expand=True)

        toolbar = tk.Frame(outer, padx=14, pady=10, bg="#f4f4f4")
        toolbar.pack(fill="x")

        tk.Button(
            toolbar,
            text="↩️ Tableau de bord",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(side="left")

        tk.Label(
            toolbar,
            text="📝 Informations générales du bisse",
            font=("Arial", 16, "bold"),
            bg="#f4f4f4"
        ).pack(side="left", padx=16)

        vars_text = {}

        marked_var = tk.StringVar(
            value="inconnu" if info.get("marked_trail") is None else ("oui" if info.get("marked_trail") else "non")
        )

        description_text = None
        itinerary_text = None
        resources_text = None

        def save_info():
            nonlocal description_text, itinerary_text, resources_text

            try:
                external_resources = json.loads(resources_text.get("1.0", tk.END).strip() or "[]")
                if not isinstance(external_resources, list):
                    raise ValueError("external_resources doit être une liste JSON.")
            except Exception as exc:
                messagebox.showerror("Ressources externes invalides", str(exc))
                return

            current_container = self.read_catalog_container()
            current_info = current_container.setdefault("bisse_info", self.default_bisse_info())

            new_info = dict(current_info)
            new_info["title"] = vars_text["title"].get().strip()
            new_info["slug"] = self.slugify(vars_text["slug"].get().strip() or new_info["title"])
            new_info["region"] = vars_text["region"].get().strip()
            new_info["commune"] = vars_text["commune"].get().strip()
            new_info["communes"] = vars_text["commune"].get().strip()
            new_info["description"] = description_text.get("1.0", tk.END).strip()
            new_info["itinerary"] = itinerary_text.get("1.0", tk.END).strip()
            new_info["length_km"] = self.parse_optional_float(vars_text["length_km"].get())
            new_info["altitude_min_m"] = self.parse_optional_int(vars_text["altitude_min_m"].get())
            new_info["altitude_max_m"] = self.parse_optional_int(vars_text["altitude_max_m"].get())
            new_info["difficulty"] = vars_text["difficulty"].get().strip()
            new_info["state"] = vars_text["state"].get().strip()
            new_info["tags"] = self.parse_tags_text(vars_text["tags"].get())

            if marked_var.get() == "oui":
                new_info["marked_trail"] = True
            elif marked_var.get() == "non":
                new_info["marked_trail"] = False
            else:
                new_info["marked_trail"] = None

            # inventory_info reste conservé en arrière-plan comme trace complète
            # de la dernière ligne importée. Il n'est plus édité dans cet écran :
            # la fiche active est bisse_info.
            current_container.setdefault("inventory_info", self.empty_inventory_info())

            current_container["bisse_info"] = new_info
            current_container["external_resources"] = external_resources
            self.catalog_container = current_container
            self.catalog_data = current_container.get("photos", [])
            self.save_catalog()

            self.log("💾 Informations générales / inventaire du bisse sauvegardés.")
            self.status_header.config(
                text=f"Informations sauvegardées : {new_info.get('title', '')}",
                fg="#27ae60"
            )
            messagebox.showinfo("Sauvegardé", "Informations du bisse sauvegardées.")

        tk.Button(
            toolbar,
            text="📚 Remplir depuis bibliothèque",
            command=lambda: self.fill_current_bisse_from_library_dialog(after_import=self.show_bisse_info_editor),
            bg="#1f618d",
            fg="white"
        ).pack(side="right", padx=(6, 0))

        tk.Button(
            toolbar,
            text="🔄 Mettre à jour bibliothèque",
            command=lambda: self.update_bisses_library_from_file_dialog(after_update=self.show_bisse_info_editor)
        ).pack(side="right", padx=(6, 0))

        tk.Button(
            toolbar,
            text="📥 Import direct Excel / CSV",
            command=lambda: self.import_inventory_for_current_bisse_dialog(after_import=self.show_bisse_info_editor)
        ).pack(side="right", padx=(6, 0))

        tk.Button(
            toolbar,
            text="💾 Sauvegarder",
            command=save_info,
            bg="#27ae60",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="right")

        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, padx=18, pady=14)
        inner_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")

        def on_inner_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfigure(inner_window, width=event.width)

        def on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        scroll_frame.bind("<Configure>", on_inner_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.bind("<Enter>", lambda _event: canvas.bind_all("<MouseWheel>", on_mousewheel))
        canvas.bind("<Leave>", lambda _event: canvas.unbind_all("<MouseWheel>"))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        library = self.read_bisses_library()
        library_count = len(library.get("rows", []))
        library_source = library.get("source_file") or "—"
        library_updated = library.get("updated_at") or "—"

        intro = (
            "Ces informations alimentent le catalogue local et l’export vers "
            "la plateforme GitHub Pages “Bisses”. La bibliothèque permanente "
            "permet de remplir cette fiche sans recharger l’Excel à chaque fois.\n"
            f"Bibliothèque : {library_count} bisse(s) · source : {library_source} · mise à jour : {library_updated}"
        )
        tk.Label(
            scroll_frame,
            text=intro,
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=1120
        ).pack(fill="x", pady=(0, 14))

        main_frame = tk.LabelFrame(scroll_frame, text="Informations principales", padx=10, pady=10)
        main_frame.pack(fill="x", pady=(0, 12))
        main_frame.grid_columnconfigure(1, weight=1)

        fields = [
            ("title", "Nom du bisse"),
            ("slug", "Identifiant web / slug"),
            ("region", "Région"),
            ("commune", "Communes"),
            ("length_km", "Longueur km"),
            ("altitude_min_m", "Altitude min m"),
            ("altitude_max_m", "Altitude max m"),
            ("difficulty", "Cotation / difficulté"),
            ("state", "État du bisse"),
            ("tags", "Tags (séparés par virgules)")
        ]

        for row, (key, label) in enumerate(fields):
            tk.Label(main_frame, text=label).grid(row=row, column=0, sticky="w", pady=4, padx=(0, 8))
            if key == "tags":
                value = ", ".join(info.get("tags", [])) if isinstance(info.get("tags"), list) else str(info.get("tags") or "")
            else:
                value = "" if info.get(key) is None else str(info.get(key, ""))
            var = tk.StringVar(value=value)
            vars_text[key] = var
            tk.Entry(main_frame, textvariable=var).grid(row=row, column=1, sticky="ew", pady=4)

        row = len(fields)
        tk.Label(main_frame, text="Sentier balisé").grid(row=row, column=0, sticky="w", pady=4, padx=(0, 8))
        ttk.Combobox(
            main_frame,
            textvariable=marked_var,
            values=["inconnu", "oui", "non"],
            state="readonly"
        ).grid(row=row, column=1, sticky="ew", pady=4)

        texts_frame = tk.LabelFrame(scroll_frame, text="Textes", padx=10, pady=10)
        texts_frame.pack(fill="both", expand=True, pady=(0, 12))
        texts_frame.grid_columnconfigure(1, weight=1)

        tk.Label(texts_frame, text="Description").grid(row=0, column=0, sticky="nw", pady=4, padx=(0, 8))
        description_text = tk.Text(texts_frame, height=7, wrap="word")
        description_text.grid(row=0, column=1, sticky="nsew", pady=4)
        description_text.insert("1.0", info.get("description", "") or "")

        tk.Label(texts_frame, text="Itinéraire").grid(row=1, column=0, sticky="nw", pady=4, padx=(0, 8))
        itinerary_text = tk.Text(texts_frame, height=6, wrap="word")
        itinerary_text.grid(row=1, column=1, sticky="nsew", pady=4)
        itinerary_text.insert("1.0", info.get("itinerary", "") or "")
        texts_frame.grid_rowconfigure(0, weight=1)
        texts_frame.grid_rowconfigure(1, weight=1)

        # L'inventaire importé est conservé intégralement dans catalogue.json
        # sous inventory_info, mais il n'est plus affiché comme un grand module
        # éditable. L'écran reste centré sur la fiche active du bisse.
        source_file = inventory.get("_source_file") or ""
        imported_at = inventory.get("_imported_at") or ""
        if source_file or imported_at:
            inventory_note = f"Données importées depuis : {source_file or 'source inconnue'}"
            if imported_at:
                inventory_note += f" · import : {imported_at}"
        else:
            inventory_note = (
                "Aucune source d’inventaire importée pour l’instant. "
                "Utilisez le bouton d’import pour remplir automatiquement la fiche."
            )

        tk.Label(
            scroll_frame,
            text=inventory_note,
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=1120
        ).pack(fill="x", pady=(0, 12))

        resources_frame = tk.LabelFrame(scroll_frame, text="Ressources externes", padx=10, pady=10)
        resources_frame.pack(fill="both", expand=True, pady=(0, 12))
        resources_frame.grid_columnconfigure(0, weight=1)
        tk.Label(
            resources_frame,
            text="Format JSON, liste de ressources externes. Ce bloc reste séparé du contenu principal.",
            fg="#555555",
            justify="left",
            anchor="w"
        ).grid(row=0, column=0, sticky="ew")
        resources_text = tk.Text(resources_frame, height=8, wrap="word", font=("Consolas", 9))
        resources_text.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        resources_text.insert("1.0", json.dumps(resources, indent=2, ensure_ascii=False))
        resources_frame.grid_rowconfigure(1, weight=1)


    def platform_category_id(self, local_id):
        mapping = {
            "ciel_ouvert": "open",
            "open": "open",
            "canalise": "canalized",
            "canalized": "canalized",
            "abandonne": "abandoned",
            "abandoned": "abandoned",
            "non_classe": "unknown",
            "inconnu": "unknown",
            "mixed": "mixed"
        }
        return mapping.get(local_id, self.slugify(local_id))

    def platform_category_record(self, category):
        """
        Prépare une catégorie de segment pour l'export plateforme.

        Important :
        - on peut traduire l'identifiant local vers l'identifiant web ;
        - on ne remplace jamais la couleur choisie dans l'atelier GPX.
        La plateforme doit lire les couleurs décidées localement.
        """
        local_id = category.get("id", "")
        platform_id = self.platform_category_id(local_id)

        default_names = {
            "open": "À ciel ouvert",
            "canalized": "Canalisé",
            "abandoned": "Abandonné",
            "unknown": "Non classé"
        }

        default_codes = {
            "open": "OPEN",
            "canalized": "CAN",
            "abandoned": "ABD",
            "unknown": "UNK"
        }

        return {
            "id": platform_id,
            "name": (
                category.get("label")
                or category.get("name")
                or default_names.get(platform_id, platform_id)
            ),
            "file_code": (
                category.get("file_code")
                or default_codes.get(platform_id, platform_id)
            ).upper(),
            "color": category.get("color") or "#666666"
        }


    def segment_part_to_lonlat(self, points):
        coordinates = []
        for point in points:
            if point and len(point) >= 2:
                lat = float(point[0])
                lon = float(point[1])
                coordinates.append([lon, lat])
        return coordinates

    def build_platform_geojson(self, container):
        workshop = container.get("gpx_workshop", {})
        segments = workshop.get("segments", [])
        features = []

        ordered_segments = sorted(segments, key=self.gpx_segment_sort_key)

        for order, segment in enumerate(ordered_segments, start=1):
            parts = [
                self.segment_part_to_lonlat(part.get("points", []))
                for part in segment.get("parts", [])
                if len(part.get("points", [])) >= 2
            ]
            parts = [part for part in parts if len(part) >= 2]

            if not parts:
                continue

            category_id = self.platform_category_id(segment.get("category_id", "non_classe"))

            display_mode = segment.get("display_mode", "single")
            bicolor_structure_types = []
            bicolor_colors = []
            if display_mode == "bicolor":
                a_id, b_id = self.get_segment_bicolor_category_ids(segment)
                bicolor_structure_types = [
                    self.platform_category_id(a_id),
                    self.platform_category_id(b_id)
                ]
                bicolor_colors = [
                    self.get_gpx_category_by_id(a_id, container=container).get("color", "#666666"),
                    self.get_gpx_category_by_id(b_id, container=container).get("color", "#666666")
                ]

            if len(parts) == 1:
                geometry = {
                    "type": "LineString",
                    "coordinates": parts[0]
                }
                continuous = True
            else:
                geometry = {
                    "type": "MultiLineString",
                    "coordinates": parts
                }
                continuous = False

            features.append({
                "type": "Feature",
                "properties": {
                    "id": segment.get("id") or f"seg-{order:03d}",
                    "name": segment.get("name") or f"Segment {order:03d}",
                    "order": order,
                    "structure_type": "mixed" if display_mode == "bicolor" else category_id,
                    "structure_types": bicolor_structure_types if display_mode == "bicolor" else [category_id],
                    "display_mode": display_mode,
                    "bicolor_colors": bicolor_colors if display_mode == "bicolor" else [],
                    "water_status": segment.get("water_status", "unknown"),
                    "continuous": continuous
                },
                "geometry": geometry
            })

        # Repli : segments GPX déjà exportés/importés dans gpx_traces.
        if not features:
            manual_segments = container.get("gpx_traces", {}).get("manual_segments", [])
            for order, record in enumerate(manual_segments, start=1):
                parts = [
                    self.segment_part_to_lonlat(part)
                    for part in record.get("segments", [])
                    if len(part) >= 2
                ]
                parts = [part for part in parts if len(part) >= 2]
                if not parts:
                    continue

                category_id = self.platform_category_id(record.get("category", "unknown"))
                if len(parts) == 1:
                    geometry = {"type": "LineString", "coordinates": parts[0]}
                    continuous = True
                else:
                    geometry = {"type": "MultiLineString", "coordinates": parts}
                    continuous = False

                features.append({
                    "type": "Feature",
                    "properties": {
                        "id": record.get("id") or f"seg-{order:03d}",
                        "name": record.get("label") or f"Segment {order:03d}",
                        "order": order,
                        "structure_type": category_id,
                        "water_status": record.get("water_status", "unknown"),
                        "continuous": continuous
                    },
                    "geometry": geometry
                })

        return {
            "type": "FeatureCollection",
            "features": features
        }

    def selected_platform_photos(self, container):
        photos = []
        for index, entry in enumerate(container.get("photos", [])):
            if entry.get("status") != "OK":
                continue
            if not entry.get("platform_selected"):
                continue
            image_path = self.get_entry_image_path(entry)
            coords = entry.get("gps_coordinates") or {}
            photos.append((index, entry, image_path, coords))

        photos.sort(
            key=lambda item: (
                int(item[1].get("platform_order") or 999999),
                item[1].get("date_taken") or "",
                item[1].get("filename") or ""
            )
        )
        return photos

    def export_platform_photos(self, container, media_dir, slug):
        os.makedirs(media_dir, exist_ok=True)
        exported = []
        selected = self.selected_platform_photos(container)

        for number, (_idx, entry, image_path, coords) in enumerate(selected, start=1):
            if not image_path or not os.path.exists(image_path):
                self.log(f"⚠️ Photo plateforme introuvable : {entry.get('filename')}")
                continue

            output_name = f"photo_{number:03d}_web.jpg"
            output_path = os.path.join(media_dir, output_name)

            try:
                img = Image.open(image_path)
                img = ImageOps.exif_transpose(img)
                if img.mode != "RGB":
                    img = img.convert("RGB")
                img.thumbnail((1600, 1600), Image.Resampling.LANCZOS)
                img.save(output_path, "JPEG", quality=85, optimize=True)
            except Exception as exc:
                self.log(f"❌ Export photo web impossible {entry.get('filename')}: {exc}")
                continue

            exported.append({
                "filename_web": f"media/{slug}/{output_name}",
                "title": entry.get("title", "") or "",
                "description": entry.get("description", "") or "",
                "date": entry.get("date_taken", "") or "",
                "lat": coords.get("lat"),
                "lon": coords.get("lon"),
                "platform_order": int(entry.get("platform_order") or number)
            })

        return exported

    def build_platform_catalogue(self, container, slug, exported_photos):
        info = copy.deepcopy(container.get("bisse_info", self.default_bisse_info()))
        info["slug"] = slug
        info.setdefault("title", os.path.basename(self.base_folder))
        info.setdefault("region", "Valais")
        info.setdefault("tags", [])

        project = container.get("project", {})
        workshop = container.get("gpx_workshop", {}) or {}

        used_ids = []
        for segment in workshop.get("segments", []) or []:
            main_id = segment.get("category_id", "non_classe")
            if main_id and main_id not in used_ids:
                used_ids.append(main_id)

            bicolor = segment.get("bicolor_categories")
            if isinstance(bicolor, list):
                for category_id in bicolor[:2]:
                    if category_id and category_id not in used_ids:
                        used_ids.append(category_id)

        # Repli pour les anciens GPX exportés sans segments d'atelier.
        if not used_ids:
            for record in container.get("gpx_traces", {}).get("manual_segments", []) or []:
                category_id = record.get("category")
                if category_id and category_id not in used_ids:
                    used_ids.append(category_id)

        categories = []
        seen = set()

        for category_id in used_ids:
            category = self.get_gpx_category_by_id(category_id, container=container)
            record = self.platform_category_record(category)
            if record["id"] in seen:
                continue
            seen.add(record["id"])
            categories.append(record)

        return {
            "schema_version": "0.1",
            "project": {
                "id": f"{slug}-{project.get('year', datetime.now().year)}",
                "title": info.get("title", ""),
                "year": project.get("year", datetime.now().year),
                "source_folder": ""
            },
            "bisse_info": info,
            "segment_categories": categories,
            "photos": exported_photos,
            "external_resources": copy.deepcopy(container.get("external_resources", [])),
            "inventory_info": copy.deepcopy(container.get("inventory_info", self.empty_inventory_info()))
        }

    def configure_paths_for_bisse_folder(self, folder):
        """
        Configure les chemins internes sans reconstruire l'interface.

        v44 sécurité : configure un contexte local complet et cohérent pour
        éviter qu'un export multi-bisses conserve des chemins Data/local d'un
        autre bisse.
        """
        folder = os.path.abspath(folder)
        self.base_folder = folder
        candidate_photos = os.path.join(folder, "Photos")
        self.photos_folder = candidate_photos if os.path.isdir(candidate_photos) else folder
        self.export_folder = os.path.join(folder, "Export_JPG")
        self.gpx_folder = os.path.join(folder, "Fichiers GPX")
        self.local_catalog_path = os.path.join(folder, "catalogue.json")
        self.catalog_path = self.local_catalog_path

        try:
            project_id = self.ensure_project_for_folder(folder)
            self.reset_active_catalog_paths_to_local()
            self.log(f"🛡️ Contexte export sécurisé : {project_id} → {folder}")
        except Exception as exc:
            self.current_project_id = ""
            self.current_project_dir = ""
            self.current_project_catalog_path = ""
            self.catalog_path = self.local_catalog_path
            self.log(f"⚠️ Contexte Data non disponible pour export : {exc}")

    def get_default_collection_root(self):
        if self.base_folder:
            return os.path.dirname(os.path.abspath(self.base_folder))
        return os.getcwd()

    def get_publication_collection_path(self):
        """
        Collection globale des bisses inclus dans l'export plateforme.
        Elle est volontairement séparée de Mes bisses, mais stockée au
        même endroit global que les autres fichiers du logiciel.
        """
        return self.get_app_data_path("bisses_collection.json")

    def default_publication_collection(self):
        return {
            "schema_version": "0.1",
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "bisses": []
        }

    def read_publication_collection(self):
        path = self.get_publication_collection_path()
        if not os.path.exists(path):
            return self.default_publication_collection()
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return self.default_publication_collection()
            data.setdefault("schema_version", "0.1")
            data.setdefault("bisses", [])
            return data
        except Exception:
            return self.default_publication_collection()

    def write_publication_collection(self, collection):
        path = self.get_publication_collection_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        collection["updated_at"] = datetime.now().isoformat(timespec="seconds")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(collection, f, indent=2, ensure_ascii=False)
        self.log(f"💾 Collection de bisses sauvegardée : {path}")

    def get_bisse_summary_from_folder(self, folder):
        folder = os.path.abspath(folder)
        try:
            record = self.find_project_by_linked_folder(folder)
            if record:
                return self.get_bisse_summary_from_project(record.get("project_id", ""), folder)
        except Exception:
            pass
        catalog_path = os.path.join(folder, "catalogue.json")
        title = os.path.basename(folder)
        slug = self.slugify(title)
        region = ""
        commune = ""
        photos_selected = 0
        segments_count = 0
        has_catalog = os.path.exists(catalog_path)

        if has_catalog:
            try:
                with open(catalog_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)

                if isinstance(raw, list):
                    photos = raw
                    info = {}
                    workshop = {}
                else:
                    photos = raw.get("photos", [])
                    info = raw.get("bisse_info", {}) or {}
                    workshop = raw.get("gpx_workshop", {}) or {}

                title = info.get("title") or raw.get("project", {}).get("title") if isinstance(raw, dict) else title
                title = title or os.path.basename(folder)
                slug = self.slugify(info.get("slug") or title)
                region = info.get("region", "")
                commune = info.get("commune", "")
                photos_selected = sum(1 for p in photos if isinstance(p, dict) and p.get("platform_selected"))
                segments_count = len(workshop.get("segments", []))
            except Exception as exc:
                return {
                    "id": slug,
                    "title": title,
                    "folder": folder,
                    "region": region,
                    "commune": commune,
                    "photos_selected": photos_selected,
                    "segments_count": segments_count,
                    "has_catalog": has_catalog,
                    "exportable": False,
                    "status": f"catalogue illisible : {exc}"
                }

        exportable = has_catalog
        status = "prêt" if exportable else "catalogue manquant"

        return {
            "id": slug,
            "title": title,
            "folder": folder,
            "region": region,
            "commune": commune,
            "photos_selected": photos_selected,
            "segments_count": segments_count,
            "has_catalog": has_catalog,
            "exportable": exportable,
            "status": status
        }

    def refresh_publication_tree(self):
        if not self.publication_tree:
            return

        for item in self.publication_tree.get_children():
            self.publication_tree.delete(item)

        collection = self.read_publication_collection()
        for entry in collection.get("bisses", []):
            folder = entry.get("folder", "")
            summary = self.get_bisse_summary_from_folder(folder) if folder else entry
            item_id = summary.get("folder") or summary.get("id")
            self.publication_tree.insert(
                "",
                "end",
                iid=item_id,
                values=(
                    summary.get("title", ""),
                    summary.get("id", ""),
                    summary.get("photos_selected", 0),
                    summary.get("segments_count", 0),
                    "✅" if summary.get("exportable") else "⚠️",
                    summary.get("folder", "")
                )
            )

        self.publication_status_var.set(
            f"Collection : {self.get_publication_collection_path()} · {len(collection.get('bisses', []))} bisse(s)"
        )

    def add_folder_to_publication_collection(self, folder):
        if not folder:
            return
        summary = self.get_bisse_summary_from_folder(folder)
        collection = self.read_publication_collection()
        bisses = collection.setdefault("bisses", [])
        folder_abs = os.path.abspath(folder)

        for index, entry in enumerate(bisses):
            if os.path.abspath(entry.get("folder", "")) == folder_abs:
                bisses[index] = summary
                break
        else:
            bisses.append(summary)

        bisses.sort(key=lambda e: (e.get("title", "").lower(), e.get("id", "")))
        self.write_publication_collection(collection)
        self.refresh_publication_tree()

    def add_current_bisse_to_publication_collection(self):
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return
        self.add_folder_to_publication_collection(self.base_folder)

    def add_existing_bisse_folder_to_publication_collection(self):
        folder = filedialog.askdirectory(
            title="Ajouter un dossier bisse à la collection",
            initialdir=self.get_default_collection_root()
        )
        if folder:
            self.add_folder_to_publication_collection(folder)

    def remove_selected_bisse_from_publication_collection(self):
        if not self.publication_tree:
            return
        selection = self.publication_tree.selection()
        if not selection:
            messagebox.showwarning("Aucune sélection", "Sélectionnez un bisse à retirer de la collection.")
            return

        selected_folders = {os.path.abspath(item) for item in selection}
        collection = self.read_publication_collection()
        before = len(collection.get("bisses", []))
        collection["bisses"] = [
            entry for entry in collection.get("bisses", [])
            if os.path.abspath(entry.get("folder", "")) not in selected_folders
        ]
        after = len(collection.get("bisses", []))
        self.write_publication_collection(collection)
        self.refresh_publication_tree()
        self.log(f"🗑️ {before - after} bisse(s) retiré(s) de la lot plateforme.")

    def build_bisse_index_entry(self, slug, platform_catalogue):
        return {
            "id": slug,
            "title": platform_catalogue.get("bisse_info", {}).get("title", ""),
            "region": platform_catalogue.get("bisse_info", {}).get("region", ""),
            "commune": platform_catalogue.get("bisse_info", {}).get("commune", ""),
            "catalogue": f"data/bisses/{slug}/catalogue.json",
            "segments": f"data/bisses/{slug}/segments.geojson"
        }

    def export_current_bisse_to_platform_root(self, export_root):
        """
        Exporte le dossier bisse actuellement configuré dans export_root.
        Ne génère pas bisses_index.json : la fonction appelante décide si elle
        exporte un seul bisse ou une collection.

        v44 sécurité : l'export est en lecture seule vis-à-vis de catalogue.json.
        Il ne met plus à jour platform_export dans le catalogue de travail.
        """
        self.ensure_catalog_file_exists()
        container = self.read_catalog_container()

        if not self.catalogue_seems_compatible_with_folder(container, self.base_folder):
            identity = self.catalog_identity(container)
            raise RuntimeError(
                "Identité incohérente avant export : "
                f"dossier={self.base_folder} / catalogue={identity.get('title') or identity.get('slug')}"
            )

        info = container.setdefault("bisse_info", self.default_bisse_info())
        title = info.get("title") or container.get("project", {}).get("title") or os.path.basename(self.base_folder)
        slug = self.slugify(info.get("slug") or title)
        info["slug"] = slug
        if not info.get("title"):
            info["title"] = title

        data_root = os.path.join(export_root, "data")
        bisse_data_dir = os.path.join(data_root, "bisses", slug)
        media_dir = os.path.join(export_root, "media", slug)

        # Nettoyage ciblé du bisse exporté, pas de tout l'export global.
        if os.path.isdir(bisse_data_dir):
            shutil.rmtree(bisse_data_dir)
        if os.path.isdir(media_dir):
            shutil.rmtree(media_dir)

        os.makedirs(bisse_data_dir, exist_ok=True)
        os.makedirs(media_dir, exist_ok=True)

        exported_photos = self.export_platform_photos(container, media_dir, slug)
        platform_catalogue = self.build_platform_catalogue(container, slug, exported_photos)
        geojson = self.build_platform_geojson(container)

        with open(os.path.join(bisse_data_dir, "catalogue.json"), "w", encoding="utf-8") as f:
            json.dump(platform_catalogue, f, indent=2, ensure_ascii=False)

        with open(os.path.join(bisse_data_dir, "segments.geojson"), "w", encoding="utf-8") as f:
            json.dump(geojson, f, indent=2, ensure_ascii=False)

        return {
            "slug": slug,
            "catalogue": platform_catalogue,
            "geojson": geojson,
            "photos_count": len(exported_photos),
            "segments_count": len(geojson.get("features", [])),
            "index_entry": self.build_bisse_index_entry(slug, platform_catalogue)
        }

    def export_for_platform(self):
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return

        try:
            export_root = os.path.join(self.base_folder, "Export_Platform")
            result = self.export_current_bisse_to_platform_root(export_root)
            data_root = os.path.join(export_root, "data")
            os.makedirs(data_root, exist_ok=True)
            with open(os.path.join(data_root, "bisses_index.json"), "w", encoding="utf-8") as f:
                json.dump([result["index_entry"]], f, indent=2, ensure_ascii=False)
        except Exception as exc:
            messagebox.showerror("Erreur export", f"Impossible d'exporter pour la plateforme :\n{exc}")
            return

        self.last_platform_export_root = export_root
        self.log(f"🌐 Export plateforme créé : {export_root}")
        self.log(f"   Bisse : {result['slug']}")
        self.log(f"   Photos web : {result['photos_count']}")
        self.log(f"   Segments GeoJSON : {result['segments_count']}")

        messagebox.showinfo(
            "Export plateforme terminé",
            (
                f"Export créé dans :\n{export_root}\n\n"
                f"Photos web : {result['photos_count']}\n"
                f"Segments GeoJSON : {result['segments_count']}\n\n"
                "Vous pouvez copier data/ et media/ dans le dépôt GitHub Pages « Bisses »."
            )
        )

    def export_publication_collection(self):
        collection = self.read_publication_collection()
        bisses = collection.get("bisses", [])

        if not bisses:
            messagebox.showwarning(
                "Collection vide",
                "Ajoutez au moins un dossier bisse à la lot plateforme."
            )
            return

        export_root = os.path.join(self.get_default_collection_root(), "Export_Platform")

        if os.path.isdir(os.path.join(export_root, "data")) or os.path.isdir(os.path.join(export_root, "media")):
            if not messagebox.askyesno(
                "Remplacer l'export global",
                (
                    f"Le dossier d'export existe déjà :\n{export_root}\n\n"
                    "Remplacer les dossiers data/ et media/ générés ?"
                )
            ):
                return

        # Nettoyage global seulement pour un export collection assumé.
        for name in ("data", "media"):
            path = os.path.join(export_root, name)
            if os.path.isdir(path):
                shutil.rmtree(path)
        os.makedirs(os.path.join(export_root, "data"), exist_ok=True)
        os.makedirs(os.path.join(export_root, "media"), exist_ok=True)

        saved_state = {
            "base_folder": self.base_folder,
            "photos_folder": self.photos_folder,
            "manual_photos_folder": self.manual_photos_folder,
            "export_folder": self.export_folder,
            "gpx_folder": self.gpx_folder,
            "catalog_path": self.catalog_path,
            "local_catalog_path": getattr(self, "local_catalog_path", ""),
            "current_project_id": getattr(self, "current_project_id", ""),
            "current_project_dir": getattr(self, "current_project_dir", ""),
            "current_project_catalog_path": getattr(self, "current_project_catalog_path", ""),
            "catalog_data": self.catalog_data,
            "catalog_container": self.catalog_container,
        }

        index_entries = []
        errors = []
        exported = 0

        try:
            for entry in bisses:
                folder = entry.get("folder", "")
                if not folder or not os.path.isdir(folder):
                    errors.append(f"Dossier introuvable : {folder}")
                    continue
                if not os.path.exists(os.path.join(folder, "catalogue.json")):
                    errors.append(f"Catalogue manquant : {folder}")
                    continue

                try:
                    self.configure_paths_for_bisse_folder(folder)
                    result = self.export_current_bisse_to_platform_root(export_root)
                    index_entries.append(result["index_entry"])
                    exported += 1
                    self.log(f"✅ Exporté : {result['slug']} ({result['photos_count']} photos, {result['segments_count']} segments)")
                except Exception as exc:
                    errors.append(f"{folder}: {exc}")
        finally:
            self.base_folder = saved_state["base_folder"]
            self.photos_folder = saved_state["photos_folder"]
            self.manual_photos_folder = saved_state["manual_photos_folder"]
            self.export_folder = saved_state["export_folder"]
            self.gpx_folder = saved_state["gpx_folder"]
            self.catalog_path = saved_state["catalog_path"]
            self.local_catalog_path = saved_state["local_catalog_path"]
            self.current_project_id = saved_state["current_project_id"]
            self.current_project_dir = saved_state["current_project_dir"]
            self.current_project_catalog_path = saved_state["current_project_catalog_path"]
            self.catalog_data = saved_state["catalog_data"]
            self.catalog_container = saved_state["catalog_container"]

        index_entries.sort(key=lambda e: (e.get("title", "").lower(), e.get("id", "")))
        with open(os.path.join(export_root, "data", "bisses_index.json"), "w", encoding="utf-8") as f:
            json.dump(index_entries, f, indent=2, ensure_ascii=False)

        self.last_platform_export_root = export_root
        self.refresh_publication_tree()

        if errors:
            self.log("⚠️ Erreurs export collection :")
            for err in errors:
                self.log(f"   - {err}")

        messagebox.showinfo(
            "Export collection terminé",
            (
                f"Dossier export :\n{export_root}\n\n"
                f"Bisses exportés : {exported}\n"
                f"Erreurs : {len(errors)}\n\n"
                "data/bisses_index.json contient maintenant plusieurs bisses."
            )
        )

    def copy_last_export_to_github_repo(self):
        export_root = self.last_platform_export_root or os.path.join(self.get_default_collection_root(), "Export_Platform")
        data_src = os.path.join(export_root, "data")
        media_src = os.path.join(export_root, "media")

        if not os.path.isdir(data_src) or not os.path.isdir(media_src):
            messagebox.showwarning(
                "Export absent",
                "Aucun export data/ + media/ complet n'a été trouvé. Lancez d'abord un export plateforme."
            )
            return

        repo = filedialog.askdirectory(
            title="Choisir le dossier du dépôt GitHub Pages Bisses",
            initialdir=self.get_default_collection_root()
        )
        if not repo:
            return

        # Petite vérification non bloquante.
        looks_like_repo = os.path.exists(os.path.join(repo, "index.html")) or os.path.isdir(os.path.join(repo, ".git"))
        if not looks_like_repo:
            if not messagebox.askyesno(
                "Dossier inhabituel",
                (
                    "Le dossier choisi ne ressemble pas clairement au dépôt Bisses "
                    "(pas de index.html ni de .git détecté).\n\nContinuer quand même ?"
                )
            ):
                return

        if not messagebox.askyesno(
            "Copier vers Bisses",
            (
                f"Remplacer dans :\n{repo}\n\n"
                "Les dossiers data/ et media/ seront remplacés.\n"
                "index.html, assets/ et .nojekyll ne seront pas touchés.\n\n"
                "Continuer ?"
            )
        ):
            return

        for name, src in (("data", data_src), ("media", media_src)):
            dst = os.path.join(repo, name)
            if os.path.isdir(dst):
                shutil.rmtree(dst)
            shutil.copytree(src, dst)

        self.log(f"📤 Export copié vers le dépôt Bisses : {repo}")
        messagebox.showinfo(
            "Copie terminée",
            "Les dossiers data/ et media/ ont été copiés dans le dépôt GitHub Pages Bisses."
        )

    def show_publication_module(self):
        self.clear_main_frame()
        self.status_header.config(text="Publication / Export plateforme Bisses", fg="#1f618d")

        outer = self.make_scrollable_page(padx=14, pady=12)

        toolbar = tk.Frame(outer)
        toolbar.pack(fill="x", pady=(0, 10))

        tk.Button(
            toolbar,
            text="🏠 Mes bisses",
            command=self.show_workspace_home
        ).pack(side="left")

        tk.Button(
            toolbar,
            text="↩️ Bisse actif",
            command=self.return_to_active_bisse_or_home
        ).pack(side="left", padx=(6, 0))

        tk.Label(
            toolbar,
            text="🌐 Publication / export plateforme",
            font=("Arial", 16, "bold")
        ).pack(side="left", padx=14)

        tk.Button(
            toolbar,
            text="📤 Copier dernier export vers GitHub Bisses",
            command=self.copy_last_export_to_github_repo,
            bg="#117864",
            fg="white"
        ).pack(side="right")

        info = tk.LabelFrame(outer, text="Principe", padx=10, pady=8)
        info.pack(fill="x", pady=(0, 10))
        tk.Label(
            info,
            text=(
                "Ce module gère l'export statique lu par GitHub Pages. "
                "Il peut exporter le bisse ouvert ou une collection de plusieurs dossiers bisses, "
                "puis faciliter la copie de data/ et media/ vers le dépôt local “Bisses”."
            ),
            justify="left",
            anchor="w",
            wraplength=1320
        ).pack(fill="x")

        actions = tk.LabelFrame(outer, text="Actions rapides", padx=10, pady=8)
        actions.pack(fill="x", pady=(0, 10))

        tk.Button(
            actions,
            text="🌐 Exporter le bisse ouvert",
            command=self.export_for_platform,
            bg="#1f618d",
            fg="white"
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            actions,
            text="➕ Inclure le bisse ouvert dans le lot",
            command=self.add_current_bisse_to_publication_collection
        ).pack(side="left", fill="x", expand=True, padx=4)

        tk.Button(
            actions,
            text="📁 Ajouter un autre dossier bisse",
            command=self.add_existing_bisse_folder_to_publication_collection
        ).pack(side="left", fill="x", expand=True, padx=4)

        tk.Button(
            actions,
            text="🌍 Exporter le lot multi-bisses",
            command=self.export_publication_collection,
            bg="#7d3c98",
            fg="white"
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        collection_frame = tk.LabelFrame(outer, text="Lot plateforme", padx=10, pady=8)
        collection_frame.pack(fill="both", expand=True)

        tk.Label(
            collection_frame,
            textvariable=self.publication_status_var,
            anchor="w",
            justify="left",
            fg="#555555"
        ).pack(fill="x", pady=(0, 6))

        columns = ("title", "slug", "photos", "segments", "exportable", "folder")

        publication_table_frame = tk.Frame(collection_frame)
        publication_table_frame.pack(fill="both", expand=True)
        publication_table_frame.grid_rowconfigure(0, weight=1)
        publication_table_frame.grid_columnconfigure(0, weight=1)

        self.publication_tree = ttk.Treeview(
            publication_table_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
            height=12
        )
        self.publication_tree.heading("title", text="Bisse")
        self.publication_tree.heading("slug", text="Slug")
        self.publication_tree.heading("photos", text="Photos plateforme")
        self.publication_tree.heading("segments", text="Segments")
        self.publication_tree.heading("exportable", text="État")
        self.publication_tree.heading("folder", text="Dossier")
        self.publication_tree.column("title", width=240)
        self.publication_tree.column("slug", width=150)
        self.publication_tree.column("photos", width=130, anchor="center")
        self.publication_tree.column("segments", width=90, anchor="center")
        self.publication_tree.column("exportable", width=70, anchor="center")
        self.publication_tree.column("folder", width=650)

        publication_y_scroll = ttk.Scrollbar(publication_table_frame, orient="vertical", command=self.publication_tree.yview)
        publication_x_scroll = ttk.Scrollbar(publication_table_frame, orient="horizontal", command=self.publication_tree.xview)
        self.publication_tree.configure(
            yscrollcommand=publication_y_scroll.set,
            xscrollcommand=publication_x_scroll.set
        )

        self.publication_tree.grid(row=0, column=0, sticky="nsew")
        publication_y_scroll.grid(row=0, column=1, sticky="ns")
        publication_x_scroll.grid(row=1, column=0, sticky="ew")

        bottom = tk.Frame(collection_frame)
        bottom.pack(fill="x", pady=(8, 0))

        tk.Button(
            bottom,
            text="🗑️ Retirer la sélection du lot",
            command=self.remove_selected_bisse_from_publication_collection
        ).pack(side="left")

        tk.Button(
            bottom,
            text="🔄 Rafraîchir",
            command=self.refresh_publication_tree
        ).pack(side="left", padx=8)

        self.refresh_publication_tree()

    def show_gpx_sync_selector(self):
        """
        Ouvre le flux de sélection d'une ou plusieurs traces GPX horodatées
        puis lance la synchronisation photos.
        """
        self.select_gpx()
        if self.gpx_files:
            self.run_sync()

    def show_resync_interface(self):
        self.clear_main_frame()

        tk.Label(
            self.main_frame,
            text="🔄 Re-synchronisation GPS",
            font=("Arial", 14, "bold")
        ).pack(pady=10)

        tk.Label(
            self.main_frame,
            text=(
                "Choisissez une ou plusieurs traces GPX horodatées.\n"
                "Utile quand la sortie est divisée en deux parties : le logiciel "
                "choisira la trace pertinente selon l'heure de chaque photo."
            ),
            fg="gray",
            justify="center",
            wraplength=760
        ).pack(pady=5)

        gpx_frame = tk.Frame(self.main_frame)
        gpx_frame.pack(fill="x", pady=15)

        tk.Button(
            gpx_frame,
            text="🛰️ 1. Choisir une ou plusieurs traces GPX",
            command=self.select_gpx
        ).pack(side="left", padx=5)

        self.lbl_gpx = tk.Label(gpx_frame, text="Aucune trace", fg="gray")
        self.lbl_gpx.pack(side="left", padx=10)

        self.btn_sync = tk.Button(
            self.main_frame,
            text="📍 2. Re-synchroniser les photos",
            command=self.run_sync,
            bg="#27ae60",
            fg="white",
            height=2,
            state="disabled"
        )
        self.btn_sync.pack(pady=10, fill="x")

        tk.Button(
            self.main_frame,
            text="↩️ Retour",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(pady=10)

    def choose_manual_photos_folder(self):
        """
        Force le dossier photos actif pour ce dossier de bisse.
        Utile si les images sont dans un sous-dossier non standard.
        """
        initial = self.photos_folder if self.photos_folder and os.path.isdir(self.photos_folder) else self.base_folder
        folder = filedialog.askdirectory(
            title="Choisir le dossier contenant les photos à utiliser",
            initialdir=initial
        )
        if not folder:
            return

        if not self.folder_has_images(folder):
            messagebox.showwarning(
                "Aucune image",
                (
                    "Le dossier choisi ne contient aucune image reconnue "
                    "(.jpg, .jpeg, .heic, .heif).\n\n"
                    "Le dossier photos actif n'a pas été modifié."
                )
            )
            return

        try:
            inside_base = os.path.commonpath([
                os.path.abspath(self.base_folder),
                os.path.abspath(folder)
            ]) == os.path.abspath(self.base_folder)
        except Exception:
            inside_base = True

        if not inside_base:
            if not messagebox.askyesno(
                "Dossier hors du bisse",
                (
                    "Le dossier choisi ne semble pas être à l'intérieur du dossier de bisse actif.\n\n"
                    "Continuer quand même ?"
                )
            ):
                return

        self.manual_photos_folder = folder
        self.photos_folder = folder
        self.log(f"📂 Dossier photos actif choisi manuellement : {folder}")
        self.load_folder(self.base_folder)

    # ============================================================
    # MODULE 1 : CATALOGUE + CONVERSION SI NÉCESSAIRE
    # ============================================================

    def run_conversion(self):
        if not messagebox.askyesno(
            "Confirmation",
            (
                "Le logiciel va :\n\n"
                "1. Utiliser directement les JPG/JPEG présents dans Photos\n"
                "2. Convertir les HEIC/HEIF en JPG dans Export_JPG\n"
                "3. Lire les coordonnées GPS déjà présentes dans les JPG si elles existent\n"
                "4. Créer ou recréer catalogue.json\n\n"
                "Continuer ?"
            )
        ):
            return

        self.clear_log()
        self.progress["value"] = 0
        self.log("🚀 [MODULE 1] Démarrage création catalogue / conversion...")
        self.process_conversion_logic()

    def process_conversion_logic(self):
        try:
            valid_ext = (".heic", ".heif", ".jpg", ".jpeg")

            if not os.path.exists(self.photos_folder):
                messagebox.showerror(
                    "Erreur",
                    f"Dossier photos introuvable :\n{self.photos_folder}"
                )
                return

            files = [
                f for f in os.listdir(self.photos_folder)
                if os.path.isfile(os.path.join(self.photos_folder, f))
                and f.lower().endswith(valid_ext)
            ]

            if not files:
                messagebox.showwarning("Vide", "Aucune image trouvée.")
                return

            catalog = []
            success_count = 0
            convert_count = 0
            direct_jpg_count = 0
            gps_found_count = 0
            error_count = 0
            total = len(files)

            self.log(f"📁 Source : {self.photos_folder}")
            self.log(f"🖼️ {total} image(s) trouvée(s).")
            self.log("ℹ️ Les JPG/JPEG sont traités directement dans Photos, sans copie.")

            for i, filename in enumerate(files):
                input_path = os.path.join(self.photos_folder, filename)
                base_name, ext = os.path.splitext(filename)
                ext_lower = ext.lower()

                is_heic = ext_lower in (".heic", ".heif")
                is_jpg = ext_lower in (".jpg", ".jpeg")

                try:
                    if is_heic:
                        os.makedirs(self.export_folder, exist_ok=True)

                        output_filename = base_name + ".jpg"
                        output_path = os.path.join(self.export_folder, output_filename)

                        self.log(f"🔄 Conversion HEIC/HEIF : {filename} -> {output_filename}")

                        img = Image.open(input_path)
                        exif_data = img.info.get("exif", b"")

                        if img.mode != "RGB":
                            img = img.convert("RGB")

                        if exif_data:
                            img.save(
                                output_path,
                                "JPEG",
                                quality=100,
                                subsampling=0,
                                exif=exif_data
                            )
                        else:
                            img.save(
                                output_path,
                                "JPEG",
                                quality=100,
                                subsampling=0
                            )

                        working_path = output_path
                        output_filename = os.path.basename(working_path)
                        convert_count += 1

                    elif is_jpg:
                        working_path = input_path
                        output_filename = os.path.basename(working_path)
                        direct_jpg_count += 1
                        self.log(f"📷 JPG original utilisé directement : {filename}")

                    else:
                        continue

                    existing_meta = self.read_text_metadata_from_jpg(working_path)
                    title = existing_meta.get("title", "") if existing_meta.get("ok") else ""
                    description = existing_meta.get("description", "") if existing_meta.get("ok") else ""

                    existing_gps = self.read_gps_metadata_from_jpg(working_path)

                    if existing_gps.get("ok"):
                        gps_sync = "OK_METADATA"
                        gps_source = "JPG_EXIF"
                        gps_coordinates = {
                            "lat": existing_gps["lat"],
                            "lon": existing_gps["lon"],
                            "ele": existing_gps["ele"]
                        }
                        gps_found_count += 1

                        try:
                            capture_dt = self.get_capture_datetime_for_sort(working_path)
                            date_taken = capture_dt.isoformat() if capture_dt != datetime.max else None
                        except Exception:
                            date_taken = None

                        self.log(
                            f"📍 GPS EXIF détecté : {output_filename} -> "
                            f"{existing_gps['lat']:.6f}, {existing_gps['lon']:.6f}"
                        )

                    else:
                        gps_sync = "NON_ENCORE_FAIT"
                        gps_source = None
                        gps_coordinates = None
                        date_taken = None

                    entry = {
                        "filename": output_filename,
                        "original_filename": filename,
                        "source_relative_path": self.relative_to_base(input_path),
                        "image_relative_path": self.relative_to_base(working_path),
                        "status": "OK",
                        "converted_from_heic": is_heic,
                        "uses_original_jpg": is_jpg,
                        "copied_from_jpg": False,
                        "gps_sync": gps_sync,
                        "gps_source": gps_source,
                        "date_taken": date_taken,
                        "gps_coordinates": gps_coordinates,
                        "title": title,
                        "description": description,
                        "platform_selected": False,
                        "platform_order": 0,
                        "platform_caption": ""
                    }

                    catalog.append(entry)
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    self.log(f"❌ Erreur sur {filename} : {e}")

                    catalog.append({
                        "filename": filename,
                        "original_filename": filename,
                        "source_relative_path": self.relative_to_base(input_path),
                        "image_relative_path": self.relative_to_base(input_path),
                        "status": "ERREUR",
                        "error": str(e),
                        "converted_from_heic": is_heic,
                        "uses_original_jpg": is_jpg,
                        "copied_from_jpg": False,
                        "gps_sync": "ERREUR_CONVERSION",
                        "gps_source": None,
                        "date_taken": None,
                        "gps_coordinates": None,
                        "title": "",
                        "description": "",
                        "platform_selected": False,
                        "platform_order": 0,
                        "platform_caption": ""
                    })

                self.progress["value"] = ((i + 1) / total) * 100
                self.root.update_idletasks()

            self.catalog_data = catalog
            self.save_catalog()

            self.log("-" * 40)
            self.log("✅ MODULE 1 TERMINÉ")
            self.log(f"🖼️ Photos traitées : {success_count}/{total}")
            self.log(f"📷 JPG originaux utilisés directement : {direct_jpg_count}")
            self.log(f"🔄 HEIC/HEIF convertis : {convert_count}")
            self.log(f"📍 GPS EXIF trouvés : {gps_found_count}")
            self.log(f"❌ Erreurs : {error_count}")
            self.log(f"💾 Catalogue créé : {self.catalog_path}")

            messagebox.showinfo(
                "Succès",
                (
                    "Module 1 terminé !\n\n"
                    f"Photos traitées : {success_count}/{total}\n"
                    f"JPG utilisés directement : {direct_jpg_count}\n"
                    f"HEIC/HEIF convertis : {convert_count}\n"
                    f"GPS EXIF trouvés : {gps_found_count}\n\n"
                    "Vous pouvez maintenant renommer, géolocaliser ou ouvrir la carte."
                )
            )

            self.load_folder(self.base_folder)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur générale : {e}")

    # ============================================================
    # MODULE RENOMMAGE
    # ============================================================

    def show_rename_interface(self):
        try:
            self.catalog_data = self.read_catalog()
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de lire le catalogue :\n{e}")
            return

        if not self.catalog_data:
            messagebox.showwarning(
                "Renommage indisponible",
                (
                    "Aucune photo active n’est disponible dans le catalogue.\n\n"
                    "Vous pouvez créer le catalogue photo depuis le tableau de bord "
                    "ou continuer à travailler dans l’atelier GPX."
                )
            )
            return

        self.clear_main_frame()
        self.status_header.config(text="Module renommage des photos", fg="#8e44ad")

        top = tk.Frame(self.main_frame)
        top.pack(fill="x", pady=(0, 8))

        tk.Button(
            top,
            text="↩️ Retour au dossier",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(side="left", padx=4)

        default_prefix = self.sanitize_filename_part(os.path.basename(self.base_folder))
        self.rename_prefix_var.set(default_prefix)
        self.rename_year_var.set(str(datetime.now().year))
        self.rename_start_var.set(1)

        controls = tk.LabelFrame(
            self.main_frame,
            text="Paramètres de renommage",
            padx=10,
            pady=10
        )
        controls.pack(fill="x", pady=8)

        tk.Label(controls, text="Préfixe").grid(row=0, column=0, sticky="w", padx=5)
        tk.Entry(
            controls,
            textvariable=self.rename_prefix_var,
            width=28
        ).grid(row=0, column=1, sticky="w", padx=5)

        tk.Label(controls, text="Année").grid(row=0, column=2, sticky="w", padx=5)
        tk.Entry(
            controls,
            textvariable=self.rename_year_var,
            width=10
        ).grid(row=0, column=3, sticky="w", padx=5)

        tk.Label(controls, text="Premier numéro").grid(row=0, column=4, sticky="w", padx=5)
        tk.Spinbox(
            controls,
            from_=1,
            to=99999,
            textvariable=self.rename_start_var,
            width=8
        ).grid(row=0, column=5, sticky="w", padx=5)

        tk.Button(
            controls,
            text="👁️ Prévisualiser",
            command=self.preview_rename_plan
        ).grid(row=0, column=6, padx=10)

        tk.Button(
            controls,
            text="✅ Appliquer le renommage",
            command=self.apply_rename_plan,
            bg="#8e44ad",
            fg="white"
        ).grid(row=0, column=7, padx=10)

        info = (
            "Les photos sont triées par date de prise de vue EXIF.\n"
            "Les JPG originaux dans Photos sont renommés directement.\n"
            "Les JPG issus de HEIC/HEIF sont renommés dans Export_JPG.\n"
            "Le catalogue est mis à jour automatiquement."
        )
        tk.Label(
            controls,
            text=info,
            fg="#555555",
            justify="left"
        ).grid(row=1, column=0, columnspan=8, sticky="w", pady=(10, 0))

        table_frame = tk.Frame(self.main_frame)
        table_frame.pack(fill="both", expand=True, pady=8)

        columns = ("order", "date", "current", "new", "folder", "status")
        self.rename_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings"
        )

        self.rename_tree.heading("order", text="#")
        self.rename_tree.heading("date", text="Date de prise de vue")
        self.rename_tree.heading("current", text="Nom actuel")
        self.rename_tree.heading("new", text="Nouveau nom")
        self.rename_tree.heading("folder", text="Dossier")
        self.rename_tree.heading("status", text="Statut")

        self.rename_tree.column("order", width=50, anchor="center")
        self.rename_tree.column("date", width=160)
        self.rename_tree.column("current", width=230)
        self.rename_tree.column("new", width=230)
        self.rename_tree.column("folder", width=280)
        self.rename_tree.column("status", width=150)

        y_scroll = tk.Scrollbar(table_frame, orient="vertical", command=self.rename_tree.yview)
        self.rename_tree.configure(yscrollcommand=y_scroll.set)

        self.rename_tree.pack(side="left", fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")

        self.preview_rename_plan()

    def build_rename_plan(self):
        prefix = self.sanitize_filename_part(self.rename_prefix_var.get())
        year = self.sanitize_filename_part(self.rename_year_var.get())
        start = int(self.rename_start_var.get())

        candidates = []

        for catalog_index, entry in enumerate(self.catalog_data):
            if entry.get("status") != "OK":
                continue

            image_path = self.get_entry_image_path(entry)

            if not os.path.exists(image_path):
                candidates.append({
                    "catalog_index": catalog_index,
                    "entry": entry,
                    "source_path": image_path,
                    "target_path": image_path,
                    "date": datetime.max,
                    "status": "Fichier introuvable",
                    "skip": True
                })
                continue

            if not image_path.lower().endswith((".jpg", ".jpeg")):
                candidates.append({
                    "catalog_index": catalog_index,
                    "entry": entry,
                    "source_path": image_path,
                    "target_path": image_path,
                    "date": datetime.max,
                    "status": "Non JPG",
                    "skip": True
                })
                continue

            dt = self.get_capture_datetime_for_sort(image_path, entry)
            candidates.append({
                "catalog_index": catalog_index,
                "entry": entry,
                "source_path": image_path,
                "date": dt,
                "skip": False
            })

        sortable = [p for p in candidates if not p.get("skip")]
        skipped = [p for p in candidates if p.get("skip")]

        sortable.sort(key=lambda p: (p["date"], os.path.basename(p["source_path"]).lower()))

        plan = []
        number = start

        for p in sortable:
            folder = os.path.dirname(p["source_path"])
            new_name = f"{prefix}_{year}_{number}.jpg"
            target_path = os.path.join(folder, new_name)

            status = "À renommer"
            if os.path.abspath(p["source_path"]) == os.path.abspath(target_path):
                status = "Déjà correct"

            p["target_path"] = target_path
            p["new_name"] = new_name
            p["status"] = status
            p["order_number"] = number

            plan.append(p)
            number += 1

        for p in skipped:
            p["new_name"] = os.path.basename(p["source_path"])
            p["order_number"] = ""
            plan.append(p)

        return plan

    def preview_rename_plan(self):
        if not self.rename_tree:
            return

        self.rename_plan = self.build_rename_plan()

        for item in self.rename_tree.get_children():
            self.rename_tree.delete(item)

        for i, p in enumerate(self.rename_plan, start=1):
            source_path = p.get("source_path", "")
            target_path = p.get("target_path", source_path)
            dt = p.get("date")

            if isinstance(dt, datetime) and dt != datetime.max:
                date_text = dt.strftime("%Y-%m-%d %H:%M:%S")
            else:
                date_text = "—"

            folder_text = self.relative_to_base(os.path.dirname(source_path))

            self.rename_tree.insert(
                "",
                "end",
                values=(
                    p.get("order_number", i),
                    date_text,
                    os.path.basename(source_path),
                    os.path.basename(target_path),
                    folder_text,
                    p.get("status", "")
                )
            )

        self.log(f"👁️ Aperçu renommage généré : {len(self.rename_plan)} entrée(s).")

    def apply_rename_plan(self):
        if not self.rename_plan:
            self.preview_rename_plan()

        if not self.rename_plan:
            messagebox.showwarning("Aucun renommage", "Aucun plan de renommage disponible.")
            return

        active_plan = [
            p for p in self.rename_plan
            if not p.get("skip")
            and os.path.abspath(p.get("source_path", "")) != os.path.abspath(p.get("target_path", ""))
        ]

        if not active_plan:
            messagebox.showinfo("Rien à faire", "Tous les noms sont déjà corrects.")
            return

        if not messagebox.askyesno(
            "Confirmation",
            (
                f"Renommer {len(active_plan)} photo(s) ?\n\n"
                "Cette opération modifiera les fichiers JPG et le catalogue.\n"
                "Elle n'altère pas les fichiers HEIC originaux."
            )
        ):
            return

        try:
            source_set = {os.path.abspath(p["source_path"]) for p in active_plan}

            for p in active_plan:
                target_abs = os.path.abspath(p["target_path"])
                if os.path.exists(p["target_path"]) and target_abs not in source_set:
                    messagebox.showerror(
                        "Collision de nom",
                        (
                            "Un fichier cible existe déjà :\n\n"
                            f"{p['target_path']}\n\n"
                            "Renommage annulé."
                        )
                    )
                    return

            temp_moves = []

            for p in active_plan:
                source_path = p["source_path"]
                folder = os.path.dirname(source_path)
                temp_name = f".tmp_rename_{uuid.uuid4().hex}_{os.path.basename(source_path)}"
                temp_path = os.path.join(folder, temp_name)

                os.rename(source_path, temp_path)
                temp_moves.append((p, temp_path))

            for p, temp_path in temp_moves:
                os.rename(temp_path, p["target_path"])

                entry = self.catalog_data[p["catalog_index"]]
                self.set_entry_image_path(entry, p["target_path"])

                if "original_filename_before_rename" not in entry:
                    entry["original_filename_before_rename"] = entry.get(
                        "original_filename",
                        os.path.basename(p["source_path"])
                    )

                entry["renamed"] = True
                entry["rename_date"] = datetime.now().isoformat(timespec="seconds")
                entry["previous_filename"] = os.path.basename(p["source_path"])

                self.log(
                    f"🔤 Renommé : {os.path.basename(p['source_path'])} -> {os.path.basename(p['target_path'])}"
                )

            self.save_catalog()

            messagebox.showinfo(
                "Renommage terminé",
                f"{len(active_plan)} photo(s) renommée(s)."
            )

            self.load_folder(self.base_folder)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur renommage : {e}")


    # ============================================================
    # MODULE TRACÉS DU BISSE : IMPORT DES GPX SUISSEMOBILE / TOPO
    # ============================================================

    def strip_accents(self, text):
        normalized = unicodedata.normalize("NFD", str(text))
        return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")

    def normalize_gpx_filename(self, filename):
        value = self.strip_accents(filename).lower()
        value = value.replace("-", "_").replace(" ", "_")
        value = re.sub(r"_+", "_", value)
        return value

    def get_trace_category_display(self, category):
        mapping = {
            "ciel_ouvert": ("À ciel ouvert", "#1e88e5"),
            "canalise": ("Canalisé", "#111111"),
            "abandonne": ("Abandonné", "#ef6c00"),
            "inconnu": ("Catégorie inconnue", "#8e44ad"),
            "topo_live": ("Trace live / topo", "#7f8c8d")
        }
        return mapping.get(category, ("Catégorie inconnue", "#8e44ad"))

    def detect_trace_category_from_filename(self, filename):
        """
        Détection par convention de nommage :
        - *_ciel_ouvert*.gpx
        - *_canalise*.gpx / *_canalisé*.gpx
        - *_abandonne*.gpx / *_abandonné*.gpx
        - *_Topo_Année*.gpx pour la trace live.
        """
        norm = self.normalize_gpx_filename(os.path.basename(filename))

        if "topo" in norm:
            return "topo_live", "live_topo"

        if "ciel" in norm and "ouvert" in norm:
            return "ciel_ouvert", "manual_segments"

        if "canalis" in norm:
            return "canalise", "manual_segments"

        if "abandon" in norm:
            return "abandonne", "manual_segments"

        return "inconnu", "manual_segments"

    def haversine_distance_m(self, lat1, lon1, lat2, lon2):
        r = 6371000.0
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)

        a = (
            math.sin(dphi / 2) ** 2
            + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
        )
        return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    def calculate_trace_length_m(self, segments):
        total = 0.0
        for segment in segments:
            if len(segment) < 2:
                continue
            for p1, p2 in zip(segment[:-1], segment[1:]):
                total += self.haversine_distance_m(
                    float(p1[0]), float(p1[1]),
                    float(p2[0]), float(p2[1])
                )
        return total

    def extract_segments_from_gpx(self, gpx):
        """
        Extrait les géométries utiles d'un GPX.
        Les exports peuvent contenir :
        - des tracks avec segments ;
        - des routes ;
        - plus rarement des waypoints uniquement.
        """
        segments = []

        # Tracks / track segments
        for track in gpx.tracks:
            for segment in track.segments:
                points = []
                for point in segment.points:
                    if point.latitude is None or point.longitude is None:
                        continue
                    ele = point.elevation if point.elevation is not None else None
                    points.append([float(point.latitude), float(point.longitude), ele])
                if len(points) >= 2:
                    segments.append(points)

        # Routes, si le fichier GPX n'est pas structuré en tracks.
        for route in gpx.routes:
            points = []
            for point in route.points:
                if point.latitude is None or point.longitude is None:
                    continue
                ele = point.elevation if point.elevation is not None else None
                points.append([float(point.latitude), float(point.longitude), ele])
            if len(points) >= 2:
                segments.append(points)

        # Si vraiment aucun segment, on ne transforme pas les waypoints en tracé.
        return segments

    def parse_gpx_trace_file(self, gpx_path):
        with open(gpx_path, "r", encoding="utf-8") as f:
            gpx = gpxpy.parse(f)

        segments = self.extract_segments_from_gpx(gpx)
        category, storage_group = self.detect_trace_category_from_filename(gpx_path)
        label, color = self.get_trace_category_display(category)

        point_count = sum(len(segment) for segment in segments)
        length_m = self.calculate_trace_length_m(segments)

        trace_record = {
            "id": uuid.uuid4().hex,
            "category": category,
            "label": label,
            "color": color,
            "source_filename": os.path.basename(gpx_path),
            "source_relative_path": self.relative_to_base(gpx_path),
            "imported_at": datetime.now().isoformat(timespec="seconds"),
            "point_count": point_count,
            "length_m": round(length_m, 2),
            "segments": segments
        }

        return storage_group, trace_record

    def get_trace_sections(self):
        if not isinstance(self.catalog_container, dict):
            try:
                self.read_catalog_container()
            except Exception:
                self.catalog_container = self.empty_catalog_container()

        traces = self.catalog_container.setdefault("gpx_traces", {})
        traces.setdefault("manual_segments", [])
        traces.setdefault("live_topo", [])
        return traces

    def get_trace_summary_text(self):
        try:
            traces = self.get_trace_sections()
            manual = traces.get("manual_segments", [])
            live = traces.get("live_topo", [])

            counts = {
                "ciel_ouvert": 0,
                "canalise": 0,
                "abandonne": 0,
                "inconnu": 0
            }

            for record in manual:
                category = record.get("category", "inconnu")
                counts[category] = counts.get(category, 0) + 1

            if not manual and not live:
                return (
                    "Aucun tracé manuel importé. "
                    "Placez les fichiers .gpx de SuisseMobile dans « Fichiers GPX », "
                    "puis utilisez le bouton d'import."
                )

            return (
                "Tronçons importés : "
                f"ciel ouvert {counts.get('ciel_ouvert', 0)} · "
                f"canalisé {counts.get('canalise', 0)} · "
                f"abandonné {counts.get('abandonne', 0)} · "
                f"inconnu {counts.get('inconnu', 0)}. "
                f"Trace live / topo : {len(live)}."
            )

        except Exception as e:
            return f"Résumé des tracés indisponible : {e}"

    def import_bisse_traces_from_gpx_folder(self):
        """
        Importe tous les GPX présents dans le dossier « Fichiers GPX ».

        Les tracés manuels dessinés sur SuisseMobile sont classés par le nom de fichier.
        Les fichiers contenant « topo » sont importés comme traces live / terrain.
        """
        self.ensure_catalog_file_exists()

        if not os.path.exists(self.gpx_folder):
            messagebox.showerror(
                "Dossier manquant",
                (
                    "Le dossier « Fichiers GPX » n'existe pas encore.\n\n"
                    f"Chemin attendu :\n{self.gpx_folder}"
                )
            )
            return

        gpx_files = [
            os.path.join(self.gpx_folder, filename)
            for filename in sorted(os.listdir(self.gpx_folder))
            if filename.lower().endswith(".gpx")
            and os.path.isfile(os.path.join(self.gpx_folder, filename))
        ]

        if not gpx_files:
            messagebox.showwarning(
                "Aucun GPX",
                "Aucun fichier .gpx trouvé dans le dossier « Fichiers GPX »."
            )
            return

        if not messagebox.askyesno(
            "Importer les tracés du bisse",
            (
                f"{len(gpx_files)} fichier(s) GPX détecté(s).\n\n"
                "L'import remplacera les tracés déjà importés dans catalogue.json,\n"
                "sans modifier les fichiers GPX sources.\n\n"
                "Continuer ?"
            )
        ):
            return

        try:
            self.catalog_data = self.read_catalog()
            traces = self.get_trace_sections()
            new_manual = []
            new_live = []

            imported_count = 0
            empty_count = 0
            error_count = 0

            self.log("🧭 Import des GPX de « Fichiers GPX »...")

            for gpx_path in gpx_files:
                filename = os.path.basename(gpx_path)
                try:
                    storage_group, record = self.parse_gpx_trace_file(gpx_path)

                    if not record.get("segments"):
                        empty_count += 1
                        self.log(f"⚠️ GPX sans tracé exploitable : {filename}")
                        continue

                    if storage_group == "live_topo":
                        new_live.append(record)
                    else:
                        new_manual.append(record)

                    imported_count += 1
                    length_km = record.get("length_m", 0) / 1000.0
                    self.log(
                        f"✅ GPX importé : {filename} · "
                        f"{record.get('label')} · "
                        f"{record.get('point_count', 0)} points · "
                        f"{length_km:.2f} km"
                    )

                except Exception as e:
                    error_count += 1
                    self.log(f"❌ Erreur import GPX {filename} : {e}")

            traces["manual_segments"] = new_manual
            traces["live_topo"] = new_live
            self.save_catalog()

            self.log("-" * 40)
            self.log("✅ Import des tracés terminé")
            self.log(f"🧭 GPX importés : {imported_count}")
            self.log(f"⚠️ GPX sans tracé exploitable : {empty_count}")
            self.log(f"❌ Erreurs : {error_count}")

            messagebox.showinfo(
                "Import des tracés terminé",
                (
                    f"GPX importés : {imported_count}\n"
                    f"GPX sans tracé exploitable : {empty_count}\n"
                    f"Erreurs : {error_count}\n\n"
                    "Les tracés peuvent maintenant être affichés sur la carte."
                )
            )

            self.load_folder(self.base_folder)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur globale lors de l'import des tracés : {e}")

    def iter_trace_records(self):
        traces = self.get_trace_sections()
        for record in traces.get("manual_segments", []):
            yield record
        for record in traces.get("live_topo", []):
            yield record

    def should_display_trace_record(self, record):
        category = record.get("category", "inconnu")

        if category == "ciel_ouvert":
            return self.show_trace_ciel_var.get()
        if category == "canalise":
            return self.show_trace_canalise_var.get()
        if category == "abandonne":
            return self.show_trace_abandonne_var.get()
        if category == "topo_live":
            return self.show_trace_topo_var.get()
        if category == "inconnu":
            return self.show_trace_inconnu_var.get()

        # Catégories personnalisées issues de l’atelier GPX :
        # elles restent visibles par défaut dans l’atelier Photos.
        return True

    def clear_map_trace_paths(self):
        for path in self.map_trace_paths:
            try:
                path.delete()
            except Exception:
                pass
        self.map_trace_paths = []

    def refresh_trace_display(self):
        self.draw_bisse_traces_on_map()

    def thin_positions_for_display(self, positions, max_points=1400):
        """
        Allège uniquement le rendu visuel des tracés très denses.
        Les géométries complètes restent intactes dans catalogue.json
        et sont toujours utilisées pour les exports, les découpes et les calculs.

        Ce filtrage évite de transmettre plusieurs milliers de sommets
        au widget de carte à chaque rafraîchissement, ce qui accélère
        l'affichage des cartes après la refonte GPX.
        """
        if not positions or len(positions) <= max_points:
            return positions

        step = max(1, math.ceil(len(positions) / max_points))
        thinned = positions[::step]

        if thinned[-1] != positions[-1]:
            thinned.append(positions[-1])

        return thinned

    def draw_bisse_traces_on_map(self):
        if not self.map_widget:
            return

        self.clear_map_trace_paths()

        for record in self.iter_trace_records():
            if not self.should_display_trace_record(record):
                continue

            color = record.get("color", "#8e44ad")
            category = record.get("category", "inconnu")

            # Tracés plus lisibles : un large contour blanc, puis le tracé coloré par-dessus.
            # C'est particulièrement utile sur les fonds swisstopo très détaillés.
            if category == "topo_live":
                outline_width = 6
                core_width = 3
            else:
                outline_width = 10
                core_width = 6

            for segment in record.get("segments", []):
                positions = [
                    (float(point[0]), float(point[1]))
                    for point in segment
                    if point and len(point) >= 2
                ]

                if len(positions) < 2:
                    continue

                # Rendu allégé uniquement à l'écran ; la géométrie complète
                # reste stockée et utilisée ailleurs.
                positions = self.thin_positions_for_display(positions)

                # 1) Contour blanc
                try:
                    outline = self.map_widget.set_path(
                        positions,
                        color="#ffffff",
                        width=outline_width,
                        name=record.get("source_filename", "")
                    )
                except TypeError:
                    outline = self.map_widget.set_path(
                        positions,
                        color="#ffffff",
                        width=outline_width
                    )

                self.map_trace_paths.append(outline)

                # 2) Ligne colorée au-dessus
                try:
                    core = self.map_widget.set_path(
                        positions,
                        color=color,
                        width=core_width,
                        name=record.get("source_filename", "")
                    )
                except TypeError:
                    core = self.map_widget.set_path(
                        positions,
                        color=color,
                        width=core_width
                    )

                self.map_trace_paths.append(core)

    def collect_displayed_trace_points(self):
        points = []
        for record in self.iter_trace_records():
            if not self.should_display_trace_record(record):
                continue
            for segment in record.get("segments", []):
                for point in segment:
                    if point and len(point) >= 2:
                        points.append((float(point[0]), float(point[1])))
        return points

    def fit_map_to_content(self):
        """
        Cadre la carte sur les photos géolocalisées ET les tracés affichés.
        """
        if not self.map_widget:
            return

        points = []
        if self.show_photos_on_map_var.get():
            points.extend((p["lat"], p["lon"]) for p in self.geolocated_photos)
        points.extend(self.collect_displayed_trace_points())

        if not points:
            return

        lats = [p[0] for p in points]
        lons = [p[1] for p in points]

        min_lat = min(lats)
        max_lat = max(lats)
        min_lon = min(lons)
        max_lon = max(lons)

        center_lat = (min_lat + max_lat) / 2
        center_lon = (min_lon + max_lon) / 2

        # Un seul cadrage global limite les changements successifs de tuiles.
        # fit_bounding_box est déjà utilisé dans l'atelier GPX ; on l'utilise ici aussi.
        try:
            if min_lat == max_lat and min_lon == max_lon:
                self.map_widget.set_position(center_lat, center_lon)
                self.map_widget.set_zoom(17)
            else:
                self.map_widget.fit_bounding_box(
                    (max_lat, min_lon),
                    (min_lat, max_lon)
                )
        except Exception:
            # Repli robuste si une version locale de TkinterMapView se comporte autrement.
            self.map_widget.set_position(center_lat, center_lon)

            span = max(max_lat - min_lat, max_lon - min_lon)

            if span < 0.002:
                zoom = 18
            elif span < 0.005:
                zoom = 17
            elif span < 0.01:
                zoom = 16
            elif span < 0.03:
                zoom = 15
            elif span < 0.08:
                zoom = 14
            elif span < 0.15:
                zoom = 13
            elif span < 0.3:
                zoom = 12
            else:
                zoom = 11

            self.map_widget.set_zoom(zoom)


    # ============================================================
    # RENOMMAGE DES FICHIERS GPX DEPUIS LA VISIONNEUSE CARTE
    # ============================================================

    def list_gpx_files_in_project_folder(self):
        """
        Retourne les fichiers .gpx présents dans « Fichiers GPX ».
        """
        if not self.gpx_folder or not os.path.exists(self.gpx_folder):
            return []

        return [
            os.path.join(self.gpx_folder, filename)
            for filename in sorted(os.listdir(self.gpx_folder), key=lambda name: name.lower())
            if filename.lower().endswith(".gpx")
            and os.path.isfile(os.path.join(self.gpx_folder, filename))
        ]

    def validate_gpx_new_filename(self, raw_name):
        """
        Valide et normalise un nouveau nom GPX saisi dans la boîte de dialogue.
        Le nom reste libre, mais doit rester un nom de fichier simple dans le dossier courant.
        """
        name = (raw_name or "").strip()

        if not name:
            return None, "Un nom de fichier GPX ne peut pas être vide."

        if os.path.basename(name) != name or "/" in name or "\\" in name:
            return None, f"Nom invalide : {name}"

        if not name.lower().endswith(".gpx"):
            name += ".gpx"

        forbidden_chars = '<>:"/\\|?*'
        if any(ch in name for ch in forbidden_chars):
            return None, f"Nom invalide : {name}"

        if name in (".gpx",):
            return None, "Nom invalide."

        return name, ""

    def update_imported_trace_records_after_gpx_rename(self, rename_map):
        """
        Met à jour les références GPX déjà importées dans catalogue.json.

        rename_map :
            {
                old_abs_path: new_abs_path,
                ...
            }

        Si le nouveau nom change la catégorie détectée
        (ciel_ouvert / canalise / abandonne / topo),
        la catégorie, la couleur et le rangement manuel/topo sont mis à jour.
        """
        if not rename_map:
            return

        traces = self.get_trace_sections()
        all_records = list(traces.get("manual_segments", [])) + list(traces.get("live_topo", []))

        old_rel_to_new_abs = {
            self.relative_to_base(old_abs): new_abs
            for old_abs, new_abs in rename_map.items()
        }

        old_name_to_new_abs = {
            os.path.basename(old_abs): new_abs
            for old_abs, new_abs in rename_map.items()
        }

        updated_records = []

        for record in all_records:
            old_rel = record.get("source_relative_path", "")
            old_name = record.get("source_filename", "")

            new_abs = None
            if old_rel in old_rel_to_new_abs:
                new_abs = old_rel_to_new_abs[old_rel]
            elif old_name in old_name_to_new_abs:
                new_abs = old_name_to_new_abs[old_name]

            if new_abs:
                new_name = os.path.basename(new_abs)
                category, storage_group = self.detect_trace_category_from_filename(new_name)
                label, color = self.get_trace_category_display(category)

                record["source_filename"] = new_name
                record["source_relative_path"] = self.relative_to_base(new_abs)
                record["category"] = category
                record["label"] = label
                record["color"] = color
                record["renamed_at"] = datetime.now().isoformat(timespec="seconds")
                record["storage_group_hint"] = storage_group

            updated_records.append(record)

        new_manual = []
        new_live = []

        for record in updated_records:
            category = record.get("category", "inconnu")
            if category == "topo_live":
                new_live.append(record)
            else:
                new_manual.append(record)

        traces["manual_segments"] = new_manual
        traces["live_topo"] = new_live

        # Mise à jour parallèle des GPX sources de l’atelier de segmentation.
        workshop = self.get_gpx_workshop_state()
        for source in workshop.get("sources", []):
            old_rel = source.get("source_relative_path", "")
            old_name = source.get("source_filename", "")
            new_abs = None
            if old_rel in old_rel_to_new_abs:
                new_abs = old_rel_to_new_abs[old_rel]
            elif old_name in old_name_to_new_abs:
                new_abs = old_name_to_new_abs[old_name]
            if new_abs:
                source["source_filename"] = os.path.basename(new_abs)
                source["source_relative_path"] = self.relative_to_base(new_abs)
                source["label"] = os.path.splitext(os.path.basename(new_abs))[0]
                source["renamed_at"] = datetime.now().isoformat(timespec="seconds")

    def show_gpx_rename_dialog(self):
        """
        Ouvre une boîte de dialogue depuis la visionneuse carte
        pour renommer directement les fichiers .gpx du dossier « Fichiers GPX ».
        """
        if not self.gpx_folder or not os.path.exists(self.gpx_folder):
            messagebox.showerror(
                "Dossier GPX introuvable",
                (
                    "Le dossier « Fichiers GPX » n'existe pas.\n\n"
                    f"Chemin attendu :\n{self.gpx_folder}"
                )
            )
            return

        gpx_paths = self.list_gpx_files_in_project_folder()

        if not gpx_paths:
            messagebox.showwarning(
                "Aucun GPX",
                "Aucun fichier .gpx trouvé dans le dossier « Fichiers GPX »."
            )
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Renommer les fichiers GPX")
        dialog.geometry("980x560")
        dialog.minsize(760, 380)
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(
            dialog,
            text="✏️ Renommage des fichiers GPX du dossier « Fichiers GPX »",
            font=("Arial", 14, "bold")
        ).pack(anchor="w", padx=16, pady=(14, 4))

        tk.Label(
            dialog,
            text=(
                "Modifiez les noms ci-dessous. Les noms peuvent être saisis avec ou sans l'extension .gpx.\n"
                "Après renommage, les références déjà importées dans catalogue.json sont mises à jour ; "
                "la catégorie/couleur est recalculée à partir du nouveau nom."
            ),
            justify="left",
            fg="#555555",
            wraplength=920
        ).pack(anchor="w", padx=16, pady=(0, 10))

        outer = tk.Frame(dialog)
        outer.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        rows_frame = tk.Frame(canvas)

        rows_frame.bind(
            "<Configure>",
            lambda event: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=rows_frame, anchor="nw")

        def _resize_rows(event):
            canvas.itemconfigure(canvas_window, width=event.width)

        canvas.bind("<Configure>", _resize_rows)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        headers = ["Nom actuel", "Catégorie détectée", "Nouveau nom"]
        for col, header in enumerate(headers):
            tk.Label(
                rows_frame,
                text=header,
                font=("Arial", 10, "bold"),
                anchor="w"
            ).grid(row=0, column=col, sticky="ew", padx=6, pady=(0, 6))

        rows_frame.grid_columnconfigure(0, weight=2)
        rows_frame.grid_columnconfigure(1, weight=1)
        rows_frame.grid_columnconfigure(2, weight=3)

        entry_rows = []

        for row_index, gpx_path in enumerate(gpx_paths, start=1):
            current_name = os.path.basename(gpx_path)
            category, _storage_group = self.detect_trace_category_from_filename(current_name)
            label, _color = self.get_trace_category_display(category)

            name_var = tk.StringVar(value=current_name)

            tk.Label(
                rows_frame,
                text=current_name,
                anchor="w",
                justify="left",
                wraplength=280
            ).grid(row=row_index, column=0, sticky="ew", padx=6, pady=3)

            tk.Label(
                rows_frame,
                text=label,
                anchor="w",
                justify="left",
                fg="#555555"
            ).grid(row=row_index, column=1, sticky="ew", padx=6, pady=3)

            tk.Entry(
                rows_frame,
                textvariable=name_var
            ).grid(row=row_index, column=2, sticky="ew", padx=6, pady=3)

            entry_rows.append({
                "old_path": gpx_path,
                "old_name": current_name,
                "new_name_var": name_var
            })

        footer = tk.Frame(dialog)
        footer.pack(fill="x", padx=16, pady=(4, 14))

        status_var = tk.StringVar(value="")
        tk.Label(
            footer,
            textvariable=status_var,
            fg="#8e44ad",
            justify="left",
            anchor="w"
        ).pack(side="left", fill="x", expand=True)

        def apply_renames():
            planned = []
            validation_errors = []
            seen_new_names = {}

            for item in entry_rows:
                old_path = item["old_path"]
                old_name = item["old_name"]
                raw_new_name = item["new_name_var"].get()

                new_name, error = self.validate_gpx_new_filename(raw_new_name)
                if error:
                    validation_errors.append(f"{old_name} → {error}")
                    continue

                new_name_key = new_name.lower()
                if new_name_key in seen_new_names:
                    validation_errors.append(
                        f"Nom cible en double : {new_name} "
                        f"(déjà utilisé pour {seen_new_names[new_name_key]})"
                    )
                    continue

                seen_new_names[new_name_key] = old_name

                new_path = os.path.join(self.gpx_folder, new_name)
                planned.append({
                    "old_path": old_path,
                    "old_name": old_name,
                    "new_path": new_path,
                    "new_name": new_name
                })

            if validation_errors:
                messagebox.showerror(
                    "Renommage impossible",
                    "\n".join(validation_errors[:12])
                )
                return

            changes = [
                item for item in planned
                if os.path.abspath(item["old_path"]) != os.path.abspath(item["new_path"])
            ]

            if not changes:
                status_var.set("Aucun changement de nom à appliquer.")
                return

            source_paths = {os.path.abspath(item["old_path"]) for item in changes}

            collisions = []
            for item in changes:
                target_abs = os.path.abspath(item["new_path"])
                if os.path.exists(item["new_path"]) and target_abs not in source_paths:
                    collisions.append(item["new_name"])

            if collisions:
                messagebox.showerror(
                    "Collision de noms",
                    (
                        "Ces fichiers existent déjà dans « Fichiers GPX » :\n\n"
                        + "\n".join(collisions[:12])
                    )
                )
                return

            if not messagebox.askyesno(
                "Confirmer le renommage GPX",
                (
                    f"Renommer {len(changes)} fichier(s) GPX ?\n\n"
                    "Le catalogue sera mis à jour pour conserver les liens avec les tracés importés."
                )
            ):
                return

            temp_moves = []
            rename_map = {}

            try:
                for item in changes:
                    old_path = item["old_path"]
                    folder = os.path.dirname(old_path)
                    temp_path = os.path.join(
                        folder,
                        f".tmp_gpx_rename_{uuid.uuid4().hex}.gpx"
                    )

                    os.rename(old_path, temp_path)
                    temp_moves.append((item, temp_path))

                for item, temp_path in temp_moves:
                    os.rename(temp_path, item["new_path"])
                    rename_map[item["old_path"]] = item["new_path"]
                    self.log(f"🧭 GPX renommé : {item['old_name']} → {item['new_name']}")

                # Relire le catalogue moderne, mettre à jour les traces importées,
                # puis sauvegarder.
                try:
                    self.read_catalog_container()
                except Exception:
                    pass

                self.update_imported_trace_records_after_gpx_rename(rename_map)
                self.save_catalog()

                # Si l’atelier GPX est actuellement ouvert, rafraîchir aussi ses listes.
                if self.gpx_source_tree:
                    self.refresh_gpx_source_tree()
                    self.draw_gpx_workshop_map()

                messagebox.showinfo(
                    "Renommage GPX terminé",
                    f"{len(changes)} fichier(s) GPX renommé(s)."
                )

                dialog.destroy()

                # Rafraîchissement léger de la carte et des traces importées.
                self.draw_bisse_traces_on_map()
                self.fit_map_to_content()

            except Exception as e:
                # Tentative de restauration si une erreur survient entre les deux phases.
                for item, temp_path in temp_moves:
                    try:
                        if os.path.exists(temp_path) and not os.path.exists(item["old_path"]):
                            os.rename(temp_path, item["old_path"])
                    except Exception:
                        pass

                messagebox.showerror("Erreur de renommage GPX", str(e))
                self.log(f"❌ Erreur lors du renommage GPX : {e}")

        tk.Button(
            footer,
            text="Annuler",
            command=dialog.destroy,
            width=14
        ).pack(side="right", padx=(8, 0))

        tk.Button(
            footer,
            text="✅ Appliquer le renommage",
            command=apply_renames,
            bg="#34495e",
            fg="white",
            width=24
        ).pack(side="right")

        self.root.wait_window(dialog)


    # ============================================================
    # ATELIER TRACÉS GPX : ORIENTATION, SEGMENTATION, FUSION,
    # CATÉGORIES LIBRES ET EXPORT
    # ============================================================

    def get_gpx_workshop_state(self):
        if not isinstance(self.catalog_container, dict):
            self.read_catalog_container()

        self.catalog_container.setdefault("gpx_workshop", self.empty_gpx_workshop())
        workshop = self.catalog_container["gpx_workshop"]
        workshop.setdefault("categories", self.default_gpx_categories())
        workshop.setdefault("sources", [])
        workshop.setdefault("segments", [])
        workshop.setdefault("last_export_at", None)
        return workshop

    def save_gpx_workshop_state(self):
        workshop = self.get_gpx_workshop_state()
        self.catalog_container["gpx_workshop"] = workshop
        self.save_catalog()

    def get_gpx_category_by_id(self, category_id, container=None):
        for category in self.get_effective_gpx_categories(
            container=container,
            include_inactive=True,
            include_local=True
        ):
            if category.get("id") == category_id:
                return category

        return {
            "id": "non_classe",
            "label": "Non classé",
            "file_code": "non_classe",
            "color": "#8e44ad",
            "active": True,
            "scope": "global"
        }

    def category_label_to_id(self, label):
        for category in self.get_effective_gpx_categories(
            include_inactive=True,
            include_local=True
        ):
            if category.get("label") == label:
                return category.get("id")
        return "non_classe"

    def gpx_workshop_categories_labels(self):
        """
        Les catégories globales inactives restent lisibles sur les anciens
        segments mais ne sont plus proposées pour un nouveau classement.
        Les anciennes catégories locales restent disponibles jusqu'à la migration.
        """
        labels = []
        for category in self.get_effective_gpx_categories(
            include_inactive=True,
            include_local=True
        ):
            if category.get("scope") == "global" and not category.get("active", True):
                continue
            label = category.get("label", category.get("id", ""))
            if label and label not in labels:
                labels.append(label)
        return labels

    def snapshot_gpx_segments(self, reason=""):
        workshop = self.get_gpx_workshop_state()
        self.gpx_workshop_undo_stack.append({
            "reason": reason,
            "segments": copy.deepcopy(workshop.get("segments", []))
        })
        self.gpx_workshop_redo_stack.clear()
        if len(self.gpx_workshop_undo_stack) > 80:
            self.gpx_workshop_undo_stack = self.gpx_workshop_undo_stack[-80:]

    def undo_gpx_segment_action(self):
        if not self.gpx_workshop_undo_stack:
            self.gpx_workshop_status_var.set("Aucune action de segment à annuler.")
            return

        workshop = self.get_gpx_workshop_state()
        current = copy.deepcopy(workshop.get("segments", []))
        previous = self.gpx_workshop_undo_stack.pop()
        self.gpx_workshop_redo_stack.append({
            "reason": previous.get("reason", ""),
            "segments": current
        })
        workshop["segments"] = previous.get("segments", [])
        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(f"↶ Annulé : {previous.get('reason') or 'dernière action'}")

    def redo_gpx_segment_action(self):
        if not self.gpx_workshop_redo_stack:
            self.gpx_workshop_status_var.set("Aucune action de segment à rétablir.")
            return

        workshop = self.get_gpx_workshop_state()
        current = copy.deepcopy(workshop.get("segments", []))
        future = self.gpx_workshop_redo_stack.pop()
        self.gpx_workshop_undo_stack.append({
            "reason": future.get("reason", ""),
            "segments": current
        })
        workshop["segments"] = future.get("segments", [])
        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(f"↷ Rétabli : {future.get('reason') or 'dernière action'}")

    def parse_gpx_source_file(self, gpx_path):
        with open(gpx_path, "r", encoding="utf-8") as f:
            gpx = gpxpy.parse(f)

        parts = self.extract_segments_from_gpx(gpx)
        if not parts:
            raise ValueError("Le GPX ne contient aucun tracé exploitable.")

        existing_sources = self.get_gpx_workshop_state().get("sources", [])
        branch_order = len(existing_sources) + 1

        return {
            "id": uuid.uuid4().hex,
            "label": os.path.splitext(os.path.basename(gpx_path))[0],
            "source_filename": os.path.basename(gpx_path),
            "source_relative_path": self.relative_to_base(gpx_path),
            "imported_at": datetime.now().isoformat(timespec="seconds"),
            "parts": parts,
            "orientation_defined": False,
            "orientation_label": "À définir",
            "upstream_endpoint": None,
            "branch_order": branch_order,
            "visible": True
        }

    def show_gpx_workshop(self):
        try:
            self.ensure_catalog_file_exists()
            self.get_gpx_workshop_state()
            self.gpx_workshop_photos = self.load_geolocated_photos()
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir ou de créer le catalogue :\n{e}")
            return

        self.clear_main_frame()
        self.gpx_workshop_active = True
        self.status_header.config(
            text="Atelier Tracés GPX · orientation · segmentation · classement · export",
            fg="#d35400"
        )

        toolbar = tk.Frame(self.main_frame)
        toolbar.pack(fill="x", pady=(0, 8))

        left_actions = tk.Frame(toolbar)
        left_actions.pack(side="left", fill="x", expand=True)

        tk.Button(
            left_actions,
            text="↩️ Tableau de bord",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(side="left", padx=(0, 4))

        tk.Button(
            left_actions,
            text="📥 Importer des GPX sources",
            command=self.import_gpx_sources_into_workshop,
            bg="#d35400",
            fg="white"
        ).pack(side="left", padx=4)

        tk.Button(
            left_actions,
            text="✏️ Renommer les GPX du dossier",
            command=self.show_gpx_rename_dialog
        ).pack(side="left", padx=4)

        tk.Button(
            left_actions,
            text="💾 Enregistrer l’atelier",
            command=self.save_gpx_workshop_state,
            bg="#27ae60",
            fg="white"
        ).pack(side="left", padx=4)

        tk.Button(
            left_actions,
            text="📤 Exporter les tronçons GPX",
            command=self.export_gpx_workshop_segments,
            bg="#2980b9",
            fg="white"
        ).pack(side="left", padx=4)

        photo_mode = tk.LabelFrame(toolbar, text="Photos", padx=5, pady=2)
        photo_mode.pack(side="right", padx=(8, 0))

        for label, value in (
            ("Visibles", "visible"),
            ("Discrètes", "discrete"),
            ("Masquées", "hidden"),
        ):
            tk.Radiobutton(
                photo_mode,
                text=label,
                variable=self.gpx_photo_display_mode_var,
                value=value,
                command=lambda v=value: self.set_gpx_photo_display_mode(v)
            ).pack(side="left", padx=2)

        tk.Checkbutton(
            photo_mode,
            text="Visionneuse intégrée",
            variable=self.gpx_photo_viewer_integrated_var,
            command=self.toggle_gpx_photo_viewer_mode
        ).pack(side="left", padx=(10, 2))

        main_paned = tk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL, sashwidth=7)
        main_paned.pack(fill="both", expand=True)

        map_panel = tk.Frame(main_paned, bg="#eeeeee")
        editor_panel = tk.Frame(main_paned, padx=10, pady=8)

        main_paned.add(map_panel, minsize=720, width=980)
        main_paned.add(editor_panel, minsize=420, width=510)

        self.build_gpx_workshop_map_panel(map_panel)
        self.build_gpx_workshop_editor_panel(editor_panel)
        self.apply_gpx_photo_viewer_mode(open_floating=False)

        self.set_gpx_editor_swisstopo_layer("color_auto")
        self.refresh_gpx_source_tree()
        self.refresh_gpx_segment_tree()
        self.refresh_gpx_category_tree()
        self.refresh_gpx_category_combo()
        self.draw_gpx_workshop_map()
        self.fit_gpx_workshop_map_to_content()
        self.start_photo_layer_watch("gpx")

        self.gpx_workshop_status_var.set(
            "Importez les GPX sources, préparez les segments, puis utilisez les photos visibles ou discrètes comme repères de terrain."
        )
        self.log("🧭 Atelier Tracés GPX ouvert.")

    def build_gpx_workshop_map_panel(self, parent):
        top = tk.Frame(parent, bg="#eeeeee")
        top.pack(fill="x", padx=8, pady=(6, 4))

        tk.Label(
            top,
            text="Carte de segmentation",
            font=("Arial", 12, "bold"),
            bg="#eeeeee"
        ).pack(side="left")

        tk.Button(
            top,
            text="Cadrer tout",
            command=self.fit_gpx_workshop_map_to_content
        ).pack(side="right", padx=2)

        tk.Button(top, text="Aérienne", command=lambda: self.set_gpx_editor_swisstopo_layer("image")).pack(side="right", padx=2)
        tk.Button(top, text="Grise", command=lambda: self.set_gpx_editor_swisstopo_layer("grey")).pack(side="right", padx=2)
        tk.Button(top, text="Standard", command=lambda: self.set_gpx_editor_swisstopo_layer("color")).pack(side="right", padx=2)
        tk.Button(top, text="25k", command=lambda: self.set_gpx_editor_swisstopo_layer("color_detail")).pack(side="right", padx=2)
        tk.Button(top, text="10k", command=lambda: self.set_gpx_editor_swisstopo_layer("color_10k")).pack(side="right", padx=2)
        tk.Button(top, text="Auto", command=lambda: self.set_gpx_editor_swisstopo_layer("color_auto")).pack(side="right", padx=2)

        tk.Label(
            parent,
            text=(
                "La sélection d’un segment peut aussi se faire en cliquant sur son tracé. "
                "En mode découpe, cliquez sur le tracé choisi pour créer une coupure."
            ),
            justify="left",
            anchor="w",
            wraplength=930,
            bg="#eeeeee",
            fg="#555555"
        ).pack(fill="x", padx=8, pady=(0, 4))

        # La carte et la visionneuse intégrée partagent un panneau horizontal.
        # La visionneuse peut être ajoutée/retirée sans recréer l'atelier.
        self.gpx_map_viewer_paned = tk.PanedWindow(
            parent,
            orient=tk.HORIZONTAL,
            sashwidth=7,
            bg="#d6dce1"
        )
        self.gpx_map_viewer_paned.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.gpx_map_holder = tk.Frame(self.gpx_map_viewer_paned, bg="#eeeeee")
        self.gpx_photo_integrated_frame = tk.Frame(self.gpx_map_viewer_paned, bg="#20252b")

        self.gpx_map_viewer_paned.add(self.gpx_map_holder, minsize=480, width=760)

        self.gpx_editor_map = tkintermapview.TkinterMapView(
            self.gpx_map_holder,
            width=760,
            height=700,
            corner_radius=0
        )
        self.gpx_editor_map.pack(fill="both", expand=True)
        self.gpx_editor_map.add_left_click_map_command(self.handle_gpx_workshop_map_click)

        self.build_gpx_integrated_photo_viewer(self.gpx_photo_integrated_frame)

    def build_gpx_integrated_photo_viewer(self, parent):
        top = tk.Frame(parent, bg="#20252b", padx=8, pady=7)
        top.pack(fill="x")

        tk.Button(top, text="◀", width=4, command=lambda: self.navigate_gpx_photo_viewer(-1)).pack(side="left")
        tk.Label(
            top,
            textvariable=self.gpx_photo_viewer_index_var,
            bg="#20252b",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left", padx=8)
        tk.Button(top, text="▶", width=4, command=lambda: self.navigate_gpx_photo_viewer(1)).pack(side="left")
        tk.Button(
            top,
            text="Détacher",
            command=self.switch_gpx_viewer_to_floating
        ).pack(side="right")

        tk.Label(
            parent,
            textvariable=self.gpx_photo_viewer_filename_var,
            bg="#20252b",
            fg="white",
            font=("Arial", 11, "bold"),
            anchor="w",
            justify="left",
            wraplength=360
        ).pack(fill="x", padx=10, pady=(2, 4))

        tk.Label(
            parent,
            textvariable=self.gpx_photo_viewer_meta_var,
            bg="#20252b",
            fg="#d8e4ec",
            anchor="w",
            justify="left",
            wraplength=360
        ).pack(fill="x", padx=10, pady=(0, 8))

        self.gpx_photo_integrated_canvas = tk.Canvas(
            parent,
            bg="#15191d",
            highlightthickness=0
        )
        self.gpx_photo_integrated_canvas.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.gpx_photo_integrated_canvas.bind(
            "<Configure>",
            lambda _event: self.render_gpx_photo_viewer_image()
        )

    def gpx_integrated_viewer_is_attached(self):
        if not self.gpx_map_viewer_paned or not self.gpx_photo_integrated_frame:
            return False
        try:
            return str(self.gpx_photo_integrated_frame) in {str(p) for p in self.gpx_map_viewer_paned.panes()}
        except Exception:
            return False

    def apply_gpx_photo_viewer_mode(self, open_floating=True):
        if not self.gpx_workshop_active or not self.gpx_map_viewer_paned:
            return

        integrated = bool(self.gpx_photo_viewer_integrated_var.get())

        if integrated:
            if not self.gpx_integrated_viewer_is_attached():
                self.gpx_map_viewer_paned.add(
                    self.gpx_photo_integrated_frame,
                    minsize=320,
                    width=380
                )
            if self.gpx_photo_viewer_window and self.gpx_photo_viewer_window.winfo_exists():
                self.gpx_photo_viewer_window.withdraw()
            self.render_gpx_photo_viewer_image()
        else:
            if self.gpx_integrated_viewer_is_attached():
                self.gpx_map_viewer_paned.forget(self.gpx_photo_integrated_frame)
            if open_floating and self.gpx_photo_viewer_current_photo:
                self.ensure_gpx_photo_viewer_window()
                self.gpx_photo_viewer_window.deiconify()
                self.gpx_photo_viewer_window.lift()
                self.render_gpx_photo_viewer_image()

    def toggle_gpx_photo_viewer_mode(self):
        self.apply_gpx_photo_viewer_mode(open_floating=True)

    def switch_gpx_viewer_to_integrated(self):
        self.gpx_photo_viewer_integrated_var.set(True)
        self.apply_gpx_photo_viewer_mode(open_floating=False)

    def switch_gpx_viewer_to_floating(self):
        self.gpx_photo_viewer_integrated_var.set(False)
        self.apply_gpx_photo_viewer_mode(open_floating=True)

    def get_active_gpx_photo_viewer_canvas(self):
        if self.gpx_photo_viewer_integrated_var.get() and self.gpx_integrated_viewer_is_attached():
            return self.gpx_photo_integrated_canvas
        return self.gpx_photo_viewer_canvas

    def create_scrollable_tab(self, notebook, title):
        """
        Crée un onglet à défilement vertical.
        Cela évite que les boutons du bas soient écrasés quand la fenêtre,
        le journal ou le panneau latéral manquent de hauteur.
        """
        outer = tk.Frame(notebook)
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, padx=8, pady=8)

        inner_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def on_inner_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            # Le contenu garde la largeur disponible ; seul le vertical scrolle.
            canvas.itemconfigure(inner_window, width=event.width)

        def on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        inner.bind("<Configure>", on_inner_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.bind("<Enter>", lambda _event: canvas.bind_all("<MouseWheel>", on_mousewheel))
        canvas.bind("<Leave>", lambda _event: canvas.unbind_all("<MouseWheel>"))

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        notebook.add(outer, text=title)
        return inner

    def build_gpx_workshop_editor_panel(self, parent):
        ttk.Label(
            parent,
            textvariable=self.gpx_workshop_status_var,
            wraplength=470,
            justify="left"
        ).pack(fill="x", pady=(0, 8))

        notebook = ttk.Notebook(parent)
        notebook.pack(fill="both", expand=True)

        sources_tab = self.create_scrollable_tab(notebook, "1. Import / branches")
        segments_tab = self.create_scrollable_tab(notebook, "2. Segments")
        categories_tab = self.create_scrollable_tab(notebook, "3. Catégories")

        self.build_gpx_sources_tab(sources_tab)
        self.build_gpx_segments_tab(segments_tab)
        self.build_gpx_categories_tab(categories_tab)

    def build_gpx_sources_tab(self, parent):
        tk.Label(
            parent,
            text=(
                "Zone d’import des GPX sources. Chaque fichier importé devient une branche de travail. "
                "Vous pouvez retirer une branche de l’atelier sans supprimer le fichier GPX du disque. "
                "Définissez ensuite le sens amont → aval avant de préparer les segments."
            ),
            justify="left",
            anchor="w",
            wraplength=470,
            fg="#555555"
        ).pack(fill="x", pady=(0, 6))

        columns = ("visible", "ordre", "branche", "sens")
        self.gpx_source_tree = ttk.Treeview(
            parent,
            columns=columns,
            show="headings",
            selectmode="browse",
            height=7
        )
        self.gpx_source_tree.heading("visible", text="👁")
        self.gpx_source_tree.heading("ordre", text="Ordre")
        self.gpx_source_tree.heading("branche", text="Branche / GPX")
        self.gpx_source_tree.heading("sens", text="Sens")
        self.gpx_source_tree.column("visible", width=42, anchor="center")
        self.gpx_source_tree.column("ordre", width=55, anchor="center")
        self.gpx_source_tree.column("branche", width=230)
        self.gpx_source_tree.column("sens", width=120)
        self.gpx_source_tree.pack(fill="x", pady=(0, 6))
        self.gpx_source_tree.bind("<<TreeviewSelect>>", self.on_gpx_source_selected)
        self.gpx_source_tree.bind("<Button-1>", self.on_gpx_source_tree_click)

        tk.Label(
            parent,
            textvariable=self.gpx_workshop_selected_source_var,
            justify="left",
            anchor="w",
            wraplength=470
        ).pack(fill="x", pady=(0, 6))

        prep_frame = tk.LabelFrame(parent, text="Import / préparation GPX", padx=8, pady=6)
        prep_frame.pack(fill="x", pady=4)

        import_row = tk.Frame(prep_frame)
        import_row.pack(fill="x", pady=3)

        tk.Button(
            import_row,
            text="📥 Importer des GPX sources",
            command=self.import_gpx_sources_into_workshop,
            bg="#d35400",
            fg="white"
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            import_row,
            text="📥 Importer GPX déjà catégorisés",
            command=self.import_bisse_traces_from_gpx_folder
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        gpx_prep_row = tk.Frame(prep_frame)
        gpx_prep_row.pack(fill="x", pady=3)

        tk.Button(
            gpx_prep_row,
            text="✏️ Renommer les GPX du dossier",
            command=self.show_gpx_rename_dialog
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            gpx_prep_row,
            text="🗑️ Retirer la branche sélectionnée",
            command=self.remove_selected_gpx_source_from_workshop
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        orientation_row = tk.Frame(parent)
        orientation_row.pack(fill="x", pady=4)

        tk.Button(
            orientation_row,
            text="A = amont",
            command=lambda: self.set_selected_source_orientation("A")
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            orientation_row,
            text="B = amont",
            command=lambda: self.set_selected_source_orientation("B")
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        order_row = tk.Frame(parent)
        order_row.pack(fill="x", pady=4)

        tk.Button(
            order_row,
            text="↑ Branche plus haute",
            command=lambda: self.move_selected_source_order(-1)
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            order_row,
            text="↓ Branche plus basse",
            command=lambda: self.move_selected_source_order(1)
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        visibility_row = tk.Frame(parent)
        visibility_row.pack(fill="x", pady=4)

        tk.Button(
            visibility_row,
            text="👁 Tout afficher les branches",
            command=self.show_all_gpx_sources
        ).pack(fill="x")

        endpoint_row = tk.Frame(parent)
        endpoint_row.pack(fill="x", pady=4)

        self.gpx_endpoint_toggle_button = tk.Button(
            endpoint_row,
            text="🙈 Masquer les repères A / B",
            command=self.toggle_gpx_endpoint_markers
        )
        self.gpx_endpoint_toggle_button.pack(fill="x")

        tk.Button(
            parent,
            text="✂️ Créer / recréer les segments depuis les branches importées",
            command=self.prepare_workshop_segments_from_sources,
            bg="#d35400",
            fg="white",
            height=2
        ).pack(fill="x", pady=(10, 4))

    def build_gpx_segments_tab(self, parent):
        tk.Label(
            parent,
            text=(
                "Les segments peuvent être découpés, catégorisés, fusionnés même s’ils sont discontinus "
                "ou issus de branches différentes. Une fusion crée un tronçon logique à plusieurs parties."
            ),
            justify="left",
            anchor="w",
            wraplength=470,
            fg="#555555"
        ).pack(fill="x", pady=(0, 6))

        columns = ("visible", "categorie", "parties", "ordre", "id")
        self.gpx_segment_tree = ttk.Treeview(
            parent,
            columns=columns,
            show="headings",
            selectmode="extended",
            height=7
        )
        self.gpx_segment_tree.heading("visible", text="👁")
        self.gpx_segment_tree.heading("categorie", text="Catégorie")
        self.gpx_segment_tree.heading("parties", text="Parties")
        self.gpx_segment_tree.heading("ordre", text="Ordre")
        self.gpx_segment_tree.heading("id", text="ID")
        self.gpx_segment_tree.column("visible", width=42, anchor="center")
        self.gpx_segment_tree.column("categorie", width=150)
        self.gpx_segment_tree.column("parties", width=60, anchor="center")
        self.gpx_segment_tree.column("ordre", width=70, anchor="center")
        self.gpx_segment_tree.column("id", width=90)
        self.gpx_segment_tree.pack(fill="x", pady=(0, 6))
        self.gpx_segment_tree.bind("<<TreeviewSelect>>", self.on_gpx_segment_selected)
        self.gpx_segment_tree.bind("<Button-1>", self.on_gpx_segment_tree_click)

        tk.Label(
            parent,
            textvariable=self.gpx_workshop_selected_segment_var,
            justify="left",
            anchor="w",
            wraplength=470
        ).pack(fill="x", pady=(0, 6))

        category_row = tk.Frame(parent)
        category_row.pack(fill="x", pady=4)

        tk.Label(category_row, text="Catégorie :").pack(side="left", padx=(0, 5))
        self.gpx_category_combo = ttk.Combobox(
            category_row,
            state="readonly",
            width=28
        )
        self.gpx_category_combo.pack(side="left", fill="x", expand=True, padx=(0, 5))

        tk.Button(
            category_row,
            text="Appliquer",
            command=self.apply_category_to_selected_segments
        ).pack(side="right")

        advanced_toggle_row = tk.Frame(parent)
        advanced_toggle_row.pack(fill="x", pady=(8, 2))

        tk.Button(
            advanced_toggle_row,
            text="Options avancées ▸ Segment bicolore",
            command=self.toggle_gpx_advanced_options
        ).pack(fill="x")

        self.gpx_advanced_options_container = tk.LabelFrame(
            parent,
            text="Options avancées",
            padx=8,
            pady=8
        )

        mode_row = tk.Frame(self.gpx_advanced_options_container)
        mode_row.pack(fill="x", pady=3)

        tk.Label(mode_row, text="Affichage du segment :").pack(side="left", padx=(0, 8))
        tk.Radiobutton(
            mode_row,
            text="Couleur simple",
            variable=self.gpx_display_mode_var,
            value="single"
        ).pack(side="left", padx=(0, 8))
        tk.Radiobutton(
            mode_row,
            text="Bicolore",
            variable=self.gpx_display_mode_var,
            value="bicolor"
        ).pack(side="left")

        bicolor_row_a = tk.Frame(self.gpx_advanced_options_container)
        bicolor_row_a.pack(fill="x", pady=3)
        tk.Label(bicolor_row_a, text="Catégorie côté A :", width=18, anchor="w").pack(side="left")
        self.gpx_bicolor_a_combo = ttk.Combobox(
            bicolor_row_a,
            state="readonly",
            width=26
        )
        self.gpx_bicolor_a_combo.pack(side="left", fill="x", expand=True)

        bicolor_row_b = tk.Frame(self.gpx_advanced_options_container)
        bicolor_row_b.pack(fill="x", pady=3)
        tk.Label(bicolor_row_b, text="Catégorie côté B :", width=18, anchor="w").pack(side="left")
        self.gpx_bicolor_b_combo = ttk.Combobox(
            bicolor_row_b,
            state="readonly",
            width=26
        )
        self.gpx_bicolor_b_combo.pack(side="left", fill="x", expand=True)

        advanced_apply_row = tk.Frame(self.gpx_advanced_options_container)
        advanced_apply_row.pack(fill="x", pady=(8, 0))
        tk.Button(
            advanced_apply_row,
            text="Appliquer les options avancées",
            command=self.apply_gpx_advanced_display_options
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))
        tk.Button(
            advanced_apply_row,
            text="Revenir à couleur simple",
            command=self.reset_selected_segments_to_single_color
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        action_row_1 = tk.Frame(parent)
        action_row_1.pack(fill="x", pady=4)

        tk.Button(
            action_row_1,
            text="✂️ Couper le segment sélectionné",
            command=self.activate_gpx_cut_mode
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            action_row_1,
            text="🔗 Fusionner la sélection",
            command=self.merge_selected_gpx_segments
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        action_row_2 = tk.Frame(parent)
        action_row_2.pack(fill="x", pady=4)

        tk.Button(
            action_row_2,
            text="↔ Dissocier un segment discontinu",
            command=self.split_selected_discontinuous_segment
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            action_row_2,
            text="🗑️ Retirer les segments sélectionnés",
            command=self.delete_selected_gpx_segments
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        action_row_3 = tk.Frame(parent)
        action_row_3.pack(fill="x", pady=4)

        tk.Button(
            action_row_3,
            text="🎯 Cadrer la sélection",
            command=self.fit_selected_gpx_segments
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            action_row_3,
            text="👁 Isoler la sélection",
            command=self.isolate_selected_gpx_segments
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        action_row_4 = tk.Frame(parent)
        action_row_4.pack(fill="x", pady=4)

        tk.Button(
            action_row_4,
            text="👁 Tout réafficher",
            command=self.show_all_gpx_segments
        ).pack(fill="x")

        history_row = tk.Frame(parent)
        history_row.pack(fill="x", pady=(10, 4))

        tk.Button(
            history_row,
            text="↶ Annuler",
            command=self.undo_gpx_segment_action
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            history_row,
            text="↷ Rétablir",
            command=self.redo_gpx_segment_action
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        tk.Button(
            parent,
            text="📤 Exporter les segments GPX classés",
            command=self.export_gpx_workshop_segments,
            bg="#2980b9",
            fg="white",
            height=2
        ).pack(fill="x", pady=(10, 0))

    def build_gpx_categories_tab(self, parent):
        tk.Label(
            parent,
            text=(
                "Les catégories sont désormais globales et partagées entre tous les bisses. "
                "Les anciennes catégories locales restent visibles avec l’indication "
                "« Locale à intégrer » jusqu’au nettoyage assisté."
            ),
            justify="left",
            anchor="w",
            wraplength=470,
            fg="#555555"
        ).pack(fill="x", pady=(0, 6))

        columns = ("statut", "label", "code", "couleur")
        self.gpx_category_tree = ttk.Treeview(
            parent,
            columns=columns,
            show="headings",
            selectmode="browse",
            height=11
        )
        self.gpx_category_tree.heading("statut", text="Portée / état")
        self.gpx_category_tree.heading("label", text="Nom")
        self.gpx_category_tree.heading("code", text="Code fichier")
        self.gpx_category_tree.heading("couleur", text="Couleur")
        self.gpx_category_tree.column("statut", width=125)
        self.gpx_category_tree.column("label", width=160)
        self.gpx_category_tree.column("code", width=130)
        self.gpx_category_tree.column("couleur", width=85, anchor="center")
        self.gpx_category_tree.pack(fill="x", pady=(0, 8))
        self.gpx_category_tree.bind(
            "<Double-1>",
            lambda _event: self.edit_selected_gpx_category()
        )

        row = tk.Frame(parent)
        row.pack(fill="x", pady=4)

        tk.Button(
            row,
            text="➕ Ajouter une catégorie globale",
            command=self.add_gpx_category,
            bg="#8e44ad",
            fg="white"
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))

        tk.Button(
            row,
            text="✏️ Modifier / intégrer",
            command=self.edit_selected_gpx_category
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))

        tk.Button(
            parent,
            text="⚙️ Ouvrir le gestionnaire global dans Paramètres",
            command=self.show_settings_dialog
        ).pack(fill="x", pady=4)

        tk.Label(
            parent,
            text=(
                "Une modification globale du nom ou de la couleur est immédiatement "
                "utilisée par les cartes et les exports, sans réécrire tous les catalogues."
            ),
            justify="left",
            anchor="w",
            wraplength=470,
            fg="#666666"
        ).pack(fill="x", pady=(10, 0))

    def get_swisstopo_tile_url(self, layer_name):
        """
        Fonds Swisstopo utilisables par TkinterMapView en EPSG:3857.

        "color_auto" n'est pas une couche WMTS : c'est une logique interne
        qui choisit une carte nationale fixe selon le zoom courant.
        """
        layers = {
            # Mode concret de repli pour l'auto.
            "color_auto": "ch.swisstopo.pixelkarte-farbe",

            # Cartes nationales fixes / noscale.
            "color_1000k": "ch.swisstopo.pixelkarte-farbe-pk1000.noscale",
            "color_500k": "ch.swisstopo.pixelkarte-farbe-pk500.noscale",
            "color_200k": "ch.swisstopo.pixelkarte-farbe-pk200.noscale",
            "color_100k": "ch.swisstopo.pixelkarte-farbe-pk100.noscale",
            "color_50k": "ch.swisstopo.pixelkarte-farbe-pk50.noscale",
            "color_detail": "ch.swisstopo.pixelkarte-farbe-pk25.noscale",

            # Carte nationale 1:10'000.
            "color_10k": "ch.swisstopo.landeskarte-farbe-10",

            # Fonds existants.
            "color": "ch.swisstopo.pixelkarte-farbe",
            "grey": "ch.swisstopo.pixelkarte-grau",
            "image": "ch.swisstopo.swissimage",
        }

        layer_id = layers.get(layer_name)
        if not layer_id:
            return None

        # Les pixelkarten et swissimage sont servies en JPEG ; la LK10 en PNG.
        extension = "png" if layer_name == "color_10k" else "jpeg"

        return f"https://wmts.geo.admin.ch/1.0.0/{layer_id}/default/current/3857/{{z}}/{{x}}/{{y}}.{extension}"


    def get_map_zoom_value(self, widget):
        """
        Récupère le niveau de zoom courant de TkinterMapView.
        Selon les versions, le zoom est exposé soit via un attribut,
        soit via une méthode.
        """
        if widget is None:
            return None

        try:
            zoom = getattr(widget, "zoom", None)
            if zoom is not None:
                return int(round(float(zoom)))
        except Exception:
            pass

        try:
            getter = getattr(widget, "get_zoom", None)
            if callable(getter):
                return int(round(float(getter())))
        except Exception:
            pass

        return None

    def get_swisstopo_auto_layer_for_zoom(self, zoom):
        """
        Choix automatique de l'échelle cartographique.

        Objectif :
        - vues larges : cartes moins détaillées, lisibles et professionnelles ;
        - vues locales : 25k plus tôt que le fond WebMercator standard ;
        - vues très rapprochées : 10k.

        Les seuils sont adaptés à TkinterMapView / WebMercator, pas une copie
        exacte du viewer map.geo.admin. Ils sont volontairement simples pour
        rester stables et ajustables.
        """
        if zoom is None:
            return "color"

        if zoom >= 17:
            return "color_10k"
        if zoom >= 14:
            return "color_detail"
        if zoom >= 13:
            return "color_50k"
        if zoom >= 12:
            return "color_100k"
        if zoom >= 11:
            return "color_200k"
        if zoom >= 10:
            return "color_500k"
        return "color_1000k"


    def describe_swisstopo_layer(self, layer_name):
        labels = {
            "color_auto": "Auto",
            "color": "Carte nationale standard",
            "color_1000k": "Carte nationale 1:1'000'000",
            "color_500k": "Carte nationale 1:500'000",
            "color_200k": "Carte nationale 1:200'000",
            "color_100k": "Carte nationale 1:100'000",
            "color_50k": "Carte nationale 1:50'000",
            "color_detail": "Carte nationale 1:25'000",
            "color_10k": "Carte nationale 1:10'000",
            "grey": "Carte grise",
            "image": "Photo aérienne"
        }
        return labels.get(layer_name, layer_name)


    def set_tile_server_for_widget(self, widget, context, layer_name, force=False):
        if widget is None:
            return

        if layer_name == "color_auto":
            zoom = self.get_map_zoom_value(widget)
            layer_name = self.get_swisstopo_auto_layer_for_zoom(zoom)

        url = self.get_swisstopo_tile_url(layer_name)
        if not url:
            return

        if not force and self.swisstopo_last_layer.get(context) == layer_name:
            return

        widget.set_tile_server(url, max_zoom=19)
        self.swisstopo_last_layer[context] = layer_name

        if context == "photo":
            self.log(f"🗺️ Fond swisstopo : {self.describe_swisstopo_layer(layer_name)}")
        else:
            self.log(f"🗺️ Fond swisstopo atelier GPX : {self.describe_swisstopo_layer(layer_name)}")

    def set_swisstopo_auto_mode(self, context, enabled=True):
        self.swisstopo_auto_enabled = bool(enabled)

        widget = self.map_widget if context == "photo" else self.gpx_editor_map
        if not widget:
            return

        if self.swisstopo_auto_enabled:
            self.apply_swisstopo_auto_layer(context, force=True)
            self.start_swisstopo_auto_watch(context)
        else:
            self.stop_swisstopo_auto_watch(context)

    def apply_swisstopo_auto_layer(self, context, force=False):
        if not self.swisstopo_auto_enabled:
            return

        widget = self.map_widget if context == "photo" else self.gpx_editor_map
        if not widget:
            return

        zoom = self.get_map_zoom_value(widget)
        layer_name = self.get_swisstopo_auto_layer_for_zoom(zoom)
        self.set_tile_server_for_widget(widget, context, layer_name, force=force)

    def start_swisstopo_auto_watch(self, context):
        """
        Surveille discrètement le zoom et change le fond seulement quand
        le seuil change. Cela évite de remplacer le serveur de tuiles à chaque
        mouvement de carte.
        """
        self.stop_swisstopo_auto_watch(context)

        def tick():
            widget = self.map_widget if context == "photo" else self.gpx_editor_map
            if not widget or not self.swisstopo_auto_enabled:
                self.swisstopo_auto_after_id[context] = None
                return

            self.apply_swisstopo_auto_layer(context, force=False)

            try:
                self.swisstopo_auto_after_id[context] = self.root.after(700, tick)
            except Exception:
                self.swisstopo_auto_after_id[context] = None

        try:
            self.swisstopo_auto_after_id[context] = self.root.after(700, tick)
        except Exception:
            self.swisstopo_auto_after_id[context] = None

    def stop_swisstopo_auto_watch(self, context):
        after_id = self.swisstopo_auto_after_id.get(context)
        if after_id:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass
        self.swisstopo_auto_after_id[context] = None

    def set_gpx_editor_swisstopo_layer(self, layer_name):
        if not self.gpx_editor_map:
            return

        if layer_name == "color_auto":
            self.set_swisstopo_auto_mode("gpx", True)
            return

        self.set_swisstopo_auto_mode("gpx", False)
        self.set_tile_server_for_widget(self.gpx_editor_map, "gpx", layer_name, force=True)

    def import_gpx_sources_into_workshop(self):
        if not os.path.exists(self.gpx_folder):
            os.makedirs(self.gpx_folder, exist_ok=True)

        selected_files = filedialog.askopenfilenames(
            title="Choisir les GPX sources à intégrer à l’atelier",
            initialdir=self.gpx_folder,
            filetypes=[("Fichiers GPX", "*.gpx")]
        )

        if not selected_files:
            return

        workshop = self.get_gpx_workshop_state()
        existing_rel = {
            source.get("source_relative_path")
            for source in workshop.get("sources", [])
        }

        imported = 0
        skipped = 0
        errors = 0

        for raw_path in selected_files:
            source_path = raw_path
            try:
                # Si le fichier vient d'ailleurs, on le copie dans le dossier GPX du projet.
                source_path = os.path.abspath(source_path)
                gpx_folder_abs = os.path.abspath(self.gpx_folder)
                if os.path.dirname(source_path) != gpx_folder_abs:
                    target_path = os.path.join(self.gpx_folder, os.path.basename(source_path))
                    if os.path.exists(target_path) and os.path.abspath(target_path) != source_path:
                        base, ext = os.path.splitext(os.path.basename(source_path))
                        counter = 2
                        while True:
                            candidate = os.path.join(self.gpx_folder, f"{base}_{counter}{ext}")
                            if not os.path.exists(candidate):
                                target_path = candidate
                                break
                            counter += 1
                    with open(source_path, "rb") as src, open(target_path, "wb") as dst:
                        dst.write(src.read())
                    source_path = target_path

                rel = self.relative_to_base(source_path)
                if rel in existing_rel:
                    skipped += 1
                    continue

                record = self.parse_gpx_source_file(source_path)
                workshop.setdefault("sources", []).append(record)
                existing_rel.add(rel)
                imported += 1
                self.log(f"✅ GPX source importé dans l’atelier : {record.get('source_filename')}")
            except Exception as e:
                errors += 1
                self.log(f"❌ Import GPX source impossible ({os.path.basename(raw_path)}): {e}")

        # Réordonne proprement les branches.
        for index, source in enumerate(workshop.get("sources", []), start=1):
            source["branch_order"] = index

        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        self.draw_gpx_workshop_map()
        self.fit_gpx_workshop_map_to_content()

        self.gpx_workshop_status_var.set(
            f"Import terminé : {imported} source(s), {skipped} déjà présente(s), {errors} erreur(s)."
        )

    def refresh_gpx_source_tree(self):
        if not self.gpx_source_tree:
            return

        self.gpx_source_tree.delete(*self.gpx_source_tree.get_children())
        sources = sorted(
            self.get_gpx_workshop_state().get("sources", []),
            key=lambda s: (s.get("branch_order", 10**9), s.get("source_filename", "").lower())
        )

        for source in sources:
            sens = source.get("orientation_label", "À définir")
            self.gpx_source_tree.insert(
                "",
                "end",
                iid=source.get("id"),
                values=(
                    "👁" if source.get("visible", True) else "🙈",
                    source.get("branch_order", ""),
                    source.get("source_filename", ""),
                    sens
                )
            )

    def refresh_gpx_segment_tree(self):
        if not self.gpx_segment_tree:
            return

        self.gpx_segment_tree.delete(*self.gpx_segment_tree.get_children())

        segments = sorted(
            self.get_gpx_workshop_state().get("segments", []),
            key=self.gpx_segment_sort_key
        )

        for index, segment in enumerate(segments, start=1):
            category = self.get_gpx_category_by_id(segment.get("category_id", "non_classe"))
            category_label = category.get("label", "Non classé")
            if segment.get("display_mode") == "bicolor":
                a_id, b_id = self.get_segment_bicolor_category_ids(segment)
                a_label = self.get_gpx_category_by_id(a_id).get("label", a_id)
                b_label = self.get_gpx_category_by_id(b_id).get("label", b_id)
                category_label = f"◐ {a_label} / {b_label}"
            parts = segment.get("parts", [])
            self.gpx_segment_tree.insert(
                "",
                "end",
                iid=segment.get("id"),
                values=(
                    "👁" if segment.get("visible", True) else "🙈",
                    category_label,
                    len(parts),
                    index,
                    str(segment.get("id", ""))[:8]
                )
            )

    def refresh_gpx_category_tree(self):
        if not self.gpx_category_tree:
            return

        try:
            if not self.gpx_category_tree.winfo_exists():
                return
        except Exception:
            return

        self.gpx_category_tree.delete(*self.gpx_category_tree.get_children())

        for category in self.get_effective_gpx_categories(
            include_inactive=True,
            include_local=True
        ):
            if category.get("scope") == "local":
                status = "Locale à intégrer"
            elif category.get("active", True):
                status = "Globale active"
            else:
                status = "Globale inactive"

            self.gpx_category_tree.insert(
                "",
                "end",
                iid=category.get("id"),
                values=(
                    status,
                    category.get("label", ""),
                    category.get("file_code", ""),
                    category.get("color", "")
                )
            )

    def toggle_gpx_advanced_options(self):
        if not self.gpx_advanced_options_container:
            return

        self.gpx_advanced_options_visible = not self.gpx_advanced_options_visible
        if self.gpx_advanced_options_visible:
            self.gpx_advanced_options_container.pack(fill="x", pady=(0, 8))
        else:
            self.gpx_advanced_options_container.pack_forget()

    def set_bicolor_combos_from_segment(self, segment):
        if not segment:
            return

        labels = self.gpx_workshop_categories_labels()
        if not labels:
            return

        mode = segment.get("display_mode", "single")
        self.gpx_display_mode_var.set("bicolor" if mode == "bicolor" else "single")

        a_id, b_id = self.get_segment_bicolor_category_ids(segment)

        a_category = self.get_gpx_category_by_id(a_id)
        b_category = self.get_gpx_category_by_id(b_id)

        if self.gpx_bicolor_a_combo:
            self.gpx_bicolor_a_combo.set(a_category.get("label", labels[0]))
        if self.gpx_bicolor_b_combo:
            self.gpx_bicolor_b_combo.set(b_category.get("label", labels[min(1, len(labels)-1)]))

    def get_segment_bicolor_category_ids(self, segment):
        """
        Retourne les deux catégories utilisées pour l'affichage bicolore.
        Repli : catégorie principale + même catégorie si rien n'est défini.
        """
        main = segment.get("category_id", "non_classe")
        cats = segment.get("bicolor_categories")

        if isinstance(cats, list) and len(cats) >= 2:
            return cats[0] or main, cats[1] or main

        return main, main

    def apply_gpx_advanced_display_options(self):
        ids = self.get_selected_gpx_segment_ids()
        if not ids:
            messagebox.showwarning("Aucun segment", "Sélectionnez au moins un segment.")
            return

        mode = self.gpx_display_mode_var.get() or "single"

        a_label = self.gpx_bicolor_a_combo.get() if self.gpx_bicolor_a_combo else ""
        b_label = self.gpx_bicolor_b_combo.get() if self.gpx_bicolor_b_combo else ""

        a_id = self.category_label_to_id(a_label)
        b_id = self.category_label_to_id(b_label)

        self.snapshot_gpx_segments("Options avancées d'affichage")
        workshop = self.get_gpx_workshop_state()
        changed = 0

        for segment in workshop.get("segments", []):
            if segment.get("id") not in ids:
                continue

            if mode == "bicolor":
                segment["display_mode"] = "bicolor"
                segment["bicolor_categories"] = [a_id, b_id]
                # Le segment garde sa catégorie principale pour le classement simple.
                # Si la catégorie principale n'était pas définie, on la rend cohérente
                # avec le côté B, généralement la structure principale.
                if segment.get("category_id", "non_classe") == "non_classe":
                    segment["category_id"] = b_id
            else:
                segment["display_mode"] = "single"
                segment.pop("bicolor_categories", None)

            changed += 1

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        if self.gpx_segment_tree:
            self.gpx_segment_tree.selection_set(ids)
        self.draw_gpx_workshop_map()

        if mode == "bicolor":
            self.gpx_workshop_status_var.set(
                f"{changed} segment(s) passé(s) en affichage bicolore : {a_label} / {b_label}."
            )
        else:
            self.gpx_workshop_status_var.set(f"{changed} segment(s) repassé(s) en couleur simple.")

    def reset_selected_segments_to_single_color(self):
        self.gpx_display_mode_var.set("single")
        self.apply_gpx_advanced_display_options()

    def refresh_gpx_category_combo(self):
        labels = self.gpx_workshop_categories_labels()

        if self.gpx_category_combo:
            self.gpx_category_combo["values"] = labels
            if labels and self.gpx_category_combo.get() not in labels:
                self.gpx_category_combo.set(labels[0])

        if self.gpx_bicolor_a_combo:
            self.gpx_bicolor_a_combo["values"] = labels
            if labels and self.gpx_bicolor_a_combo.get() not in labels:
                self.gpx_bicolor_a_combo.set(labels[0])

        if self.gpx_bicolor_b_combo:
            self.gpx_bicolor_b_combo["values"] = labels
            if labels and self.gpx_bicolor_b_combo.get() not in labels:
                self.gpx_bicolor_b_combo.set(labels[min(1, len(labels) - 1)])

    def on_gpx_source_selected(self, _event=None):
        if not self.gpx_source_tree:
            return

        selection = self.gpx_source_tree.selection()
        if not selection:
            self.gpx_workshop_selected_source_id = None
            self.gpx_workshop_selected_source_var.set("Aucune branche sélectionnée")
            self.draw_gpx_workshop_map()
            return

        source_id = selection[0]
        self.gpx_workshop_selected_source_id = source_id
        source = self.find_gpx_source(source_id)
        if source:
            self.gpx_workshop_selected_source_var.set(
                f"Branche sélectionnée : {source.get('source_filename')} · sens : {source.get('orientation_label', 'À définir')}"
            )
        self.draw_gpx_workshop_map()

    def on_gpx_segment_selected(self, _event=None):
        ids = self.get_selected_gpx_segment_ids()
        if not ids:
            self.gpx_workshop_selected_segment_var.set("Aucun segment sélectionné")
        elif len(ids) == 1:
            segment = self.find_gpx_segment(ids[0])
            category = self.get_gpx_category_by_id(segment.get("category_id", "non_classe")) if segment else {}
            display_note = ""
            if segment and segment.get("display_mode") == "bicolor":
                a_id, b_id = self.get_segment_bicolor_category_ids(segment)
                a_label = self.get_gpx_category_by_id(a_id).get("label", a_id)
                b_label = self.get_gpx_category_by_id(b_id).get("label", b_id)
                display_note = f" · bicolore : {a_label} / {b_label}"
            self.gpx_workshop_selected_segment_var.set(
                f"1 segment sélectionné · catégorie : {category.get('label', '—')} · parties : {len(segment.get('parts', [])) if segment else 0}{display_note}"
            )
            if self.gpx_category_combo and category:
                self.gpx_category_combo.set(category.get("label", "Non classé"))
            self.set_bicolor_combos_from_segment(segment)
        else:
            self.gpx_workshop_selected_segment_var.set(f"{len(ids)} segments sélectionnés")
        self.draw_gpx_workshop_map()

    def find_gpx_source(self, source_id):
        for source in self.get_gpx_workshop_state().get("sources", []):
            if source.get("id") == source_id:
                return source
        return None

    def find_gpx_segment(self, segment_id):
        for segment in self.get_gpx_workshop_state().get("segments", []):
            if segment.get("id") == segment_id:
                return segment
        return None

    def get_selected_gpx_segment_ids(self):
        if not self.gpx_segment_tree:
            return []
        return list(self.gpx_segment_tree.selection())

    def remove_selected_gpx_source_from_workshop(self):
        """
        Retire une branche de la zone d'import de l'atelier GPX.

        Important :
        - ne supprime jamais le fichier .gpx du dossier ;
        - si des segments existent déjà et proviennent de cette branche,
          on propose de retirer aussi les parties de segments concernées,
          afin d'éviter des segments orphelins ou incohérents.
        """
        if not self.gpx_source_tree:
            return

        selection = self.gpx_source_tree.selection()
        if not selection:
            messagebox.showwarning(
                "Aucune branche",
                "Sélectionnez d'abord une branche GPX à retirer de l'atelier."
            )
            return

        source_id = selection[0]
        source = self.find_gpx_source(source_id)
        if not source:
            return

        source_name = source.get("source_filename", "branche sélectionnée")
        workshop = self.get_gpx_workshop_state()
        segments = workshop.get("segments", [])

        dependent_segments = []
        dependent_parts_count = 0
        for segment in segments:
            parts = segment.get("parts", [])
            matching = [part for part in parts if part.get("source_id") == source_id]
            if matching:
                dependent_segments.append(segment)
                dependent_parts_count += len(matching)

        if dependent_segments:
            message = (
                f"Retirer « {source_name} » de l’atelier ?\n\n"
                "Le fichier GPX ne sera pas supprimé du disque.\n\n"
                f"Attention : {len(dependent_segments)} segment(s) utilisent encore "
                f"{dependent_parts_count} partie(s) provenant de cette branche.\n\n"
                "Si vous continuez, ces parties seront aussi retirées des segments. "
                "Les segments devenus vides seront supprimés."
            )
        else:
            message = (
                f"Retirer « {source_name} » de l’atelier ?\n\n"
                "Le fichier GPX ne sera pas supprimé du disque."
            )

        if not messagebox.askyesno("Retirer la branche", message):
            return

        self.snapshot_gpx_segments(f"Retrait branche {source_name}")

        # Retirer la source.
        workshop["sources"] = [
            item for item in workshop.get("sources", [])
            if item.get("id") != source_id
        ]

        # Retirer proprement les parties de segments qui venaient de cette source.
        if dependent_segments:
            cleaned_segments = []
            removed_empty_segments = 0

            for segment in workshop.get("segments", []):
                remaining_parts = [
                    part for part in segment.get("parts", [])
                    if part.get("source_id") != source_id
                ]

                if remaining_parts:
                    segment["parts"] = remaining_parts
                    cleaned_segments.append(segment)
                else:
                    removed_empty_segments += 1

            workshop["segments"] = cleaned_segments
        else:
            removed_empty_segments = 0

        self.gpx_workshop_selected_source_id = None
        self.gpx_workshop_selected_source_var.set("Aucune branche sélectionnée")

        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()

        if dependent_segments:
            self.gpx_workshop_status_var.set(
                f"Branche retirée : {source_name}. "
                f"Parties liées retirées ; segments vides supprimés : {removed_empty_segments}."
            )
        else:
            self.gpx_workshop_status_var.set(f"Branche retirée de l’atelier : {source_name}.")

        self.log(f"🗑️ Branche GPX retirée de l’atelier : {source_name}")

    def on_gpx_source_tree_click(self, event):
        if not self.gpx_source_tree:
            return

        region = self.gpx_source_tree.identify("region", event.x, event.y)
        column = self.gpx_source_tree.identify_column(event.x)
        item_id = self.gpx_source_tree.identify_row(event.y)

        if region == "cell" and column == "#1" and item_id:
            self.gpx_source_tree.selection_set(item_id)
            source = self.find_gpx_source(item_id)
            if source:
                source["visible"] = not source.get("visible", True)
                self.save_gpx_workshop_state()
                self.refresh_gpx_source_tree()
                try:
                    self.gpx_source_tree.selection_set(item_id)
                except Exception:
                    pass
                self.draw_gpx_workshop_map()
                self.gpx_workshop_status_var.set(
                    f"Branche {'affichée' if source.get('visible', True) else 'masquée'} : {source.get('source_filename', '')}"
                )
            return "break"

    def on_gpx_segment_tree_click(self, event):
        if not self.gpx_segment_tree:
            return

        region = self.gpx_segment_tree.identify("region", event.x, event.y)
        column = self.gpx_segment_tree.identify_column(event.x)
        item_id = self.gpx_segment_tree.identify_row(event.y)

        if region == "cell" and column == "#1" and item_id:
            current_selection = list(self.gpx_segment_tree.selection())
            segment = self.find_gpx_segment(item_id)
            if segment:
                segment["visible"] = not segment.get("visible", True)
                self.save_gpx_workshop_state()
                self.refresh_gpx_segment_tree()
                try:
                    for sid in current_selection:
                        self.gpx_segment_tree.selection_add(sid)
                except Exception:
                    pass
                self.draw_gpx_workshop_map()
            return "break"

    def toggle_selected_gpx_source_visibility(self):
        if not self.gpx_source_tree:
            return
        selection = self.gpx_source_tree.selection()
        if not selection:
            messagebox.showwarning("Aucune branche", "Sélectionnez une branche GPX.")
            return

        source = self.find_gpx_source(selection[0])
        if not source:
            return

        source["visible"] = not source.get("visible", True)
        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        self.gpx_source_tree.selection_set(source.get("id"))
        self.draw_gpx_workshop_map()

    def show_all_gpx_sources(self):
        for source in self.get_gpx_workshop_state().get("sources", []):
            source["visible"] = True
        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set("Toutes les branches sont réaffichées.")

    def toggle_selected_gpx_segments_visibility(self):
        ids = self.get_selected_gpx_segment_ids()
        if not ids:
            messagebox.showwarning("Aucun segment", "Sélectionnez au moins un segment.")
            return

        segments = [
            segment for segment in self.get_gpx_workshop_state().get("segments", [])
            if segment.get("id") in ids
        ]
        if not segments:
            return

        new_visible = not any(segment.get("visible", True) for segment in segments)
        for segment in segments:
            segment["visible"] = new_visible

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        if self.gpx_segment_tree:
            for sid in ids:
                try:
                    self.gpx_segment_tree.selection_add(sid)
                except Exception:
                    pass
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            "Segments sélectionnés réaffichés." if new_visible else "Segments sélectionnés masqués."
        )

    def show_all_gpx_segments(self):
        for segment in self.get_gpx_workshop_state().get("segments", []):
            segment["visible"] = True
        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set("Tous les segments sont réaffichés.")

    def isolate_selected_gpx_segments(self):
        ids = set(self.get_selected_gpx_segment_ids())
        if not ids:
            messagebox.showwarning("Aucun segment", "Sélectionnez au moins un segment à isoler.")
            return

        for segment in self.get_gpx_workshop_state().get("segments", []):
            segment["visible"] = segment.get("id") in ids

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        if self.gpx_segment_tree:
            for sid in ids:
                try:
                    self.gpx_segment_tree.selection_add(sid)
                except Exception:
                    pass
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set("Sélection isolée sur la carte. Utilisez « Tout réafficher » pour revenir.")

    def fit_selected_gpx_segments(self):
        ids = set(self.get_selected_gpx_segment_ids())
        if not ids or not self.gpx_editor_map:
            messagebox.showwarning("Aucune sélection", "Sélectionnez au moins un segment à cadrer.")
            return

        points = []
        for segment in self.get_gpx_workshop_state().get("segments", []):
            if segment.get("id") not in ids:
                continue
            for part in segment.get("parts", []):
                for point in part.get("points", []):
                    if point and len(point) >= 2:
                        points.append((float(point[0]), float(point[1])))

        if not points:
            self.gpx_workshop_status_var.set("Impossible de cadrer : aucun point trouvé dans la sélection.")
            return

        lats = [p[0] for p in points]
        lons = [p[1] for p in points]
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)

        lat_pad = max((max_lat - min_lat) * 0.15, 0.00015)
        lon_pad = max((max_lon - min_lon) * 0.15, 0.00015)

        try:
            self.gpx_editor_map.fit_bounding_box(
                (max_lat + lat_pad, min_lon - lon_pad),
                (min_lat - lat_pad, max_lon + lon_pad)
            )
        except Exception:
            self.gpx_editor_map.set_position((min_lat + max_lat) / 2, (min_lon + max_lon) / 2)

        self.gpx_workshop_status_var.set("🎯 Sélection cadrée.")

    def source_endpoint_positions(self, source):
        parts = source.get("parts", [])
        valid_parts = [part for part in parts if len(part) >= 2]
        if not valid_parts:
            return None, None
        first = valid_parts[0][0]
        last = valid_parts[-1][-1]
        return (float(first[0]), float(first[1])), (float(last[0]), float(last[1]))

    def set_selected_source_orientation(self, upstream_endpoint):
        source = self.find_gpx_source(self.gpx_workshop_selected_source_id)
        if not source:
            messagebox.showwarning("Aucune branche", "Sélectionnez d'abord une branche GPX.")
            return

        if upstream_endpoint not in ("A", "B"):
            return

        if self.get_gpx_workshop_state().get("segments"):
            if not messagebox.askyesno(
                "Segments déjà préparés",
                (
                    "Des segments existent déjà.\n\n"
                    "Changer le sens de cette branche ne recalculera pas automatiquement les segments existants. "
                    "Pour repartir proprement, utilisez ensuite « Préparer les segments depuis les branches orientées ».\n\n"
                    "Continuer ?"
                )
            ):
                return

        if upstream_endpoint == "B":
            reversed_parts = []
            for part in reversed(source.get("parts", [])):
                reversed_parts.append(list(reversed(part)))
            source["parts"] = reversed_parts
            source["orientation_label"] = "B = amont"
            source["upstream_endpoint"] = "B"
        else:
            source["orientation_label"] = "A = amont"
            source["upstream_endpoint"] = "A"

        source["orientation_defined"] = True
        # Une fois le sens déterminé, les repères A/B peuvent disparaître
        # pour ne pas encombrer la carte. L'utilisateur peut les réafficher
        # avec la case dédiée s'il veut contrôler une autre branche.
        self.gpx_workshop_show_endpoints_var.set(False)

        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        if self.gpx_source_tree:
            self.gpx_source_tree.selection_set(source.get("id"))
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            f"Sens amont → aval défini pour {source.get('source_filename')} : {source.get('orientation_label')}."
        )

    def move_selected_source_order(self, delta):
        source = self.find_gpx_source(self.gpx_workshop_selected_source_id)
        if not source:
            messagebox.showwarning("Aucune branche", "Sélectionnez d'abord une branche GPX.")
            return

        sources = sorted(
            self.get_gpx_workshop_state().get("sources", []),
            key=lambda s: s.get("branch_order", 10**9)
        )
        current_index = next((i for i, s in enumerate(sources) if s.get("id") == source.get("id")), None)
        if current_index is None:
            return

        new_index = max(0, min(len(sources) - 1, current_index + delta))
        if new_index == current_index:
            return

        sources.insert(new_index, sources.pop(current_index))
        for order, src in enumerate(sources, start=1):
            src["branch_order"] = order

        workshop = self.get_gpx_workshop_state()
        workshop["sources"] = sources
        self.save_gpx_workshop_state()
        self.refresh_gpx_source_tree()
        self.gpx_source_tree.selection_set(source.get("id"))
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            "Ordre des branches modifié. Cet ordre sert de référence pour la numérotation d’export entre branches."
        )

    def prepare_workshop_segments_from_sources(self):
        workshop = self.get_gpx_workshop_state()
        sources = sorted(
            workshop.get("sources", []),
            key=lambda s: s.get("branch_order", 10**9)
        )

        if not sources:
            messagebox.showwarning("Aucune branche", "Importez d'abord un ou plusieurs GPX sources.")
            return

        unoriented = [s for s in sources if not s.get("orientation_defined")]
        if unoriented:
            names = "\n".join(f"• {s.get('source_filename')}" for s in unoriented[:8])
            messagebox.showwarning(
                "Sens à définir",
                (
                    "Définissez d'abord manuellement le sens amont → aval des branches suivantes :\n\n"
                    f"{names}"
                )
            )
            return

        if workshop.get("segments"):
            if not messagebox.askyesno(
                "Recréer les segments ?",
                (
                    "Des segments existent déjà dans l’atelier.\n\n"
                    "Les recréer depuis les branches orientées effacera les coupes et fusions actuelles.\n\n"
                    "Continuer ?"
                )
            ):
                return

        segments = []
        for source in sources:
            for part_order, points in enumerate(source.get("parts", [])):
                if len(points) < 2:
                    continue

                part_record = {
                    "source_id": source.get("id"),
                    "source_filename": source.get("source_filename"),
                    "branch_order": source.get("branch_order", 10**9),
                    "part_order": part_order,
                    "start_fraction": 0.0,
                    "end_fraction": 1.0,
                    "points": copy.deepcopy(points)
                }

                segments.append({
                    "id": uuid.uuid4().hex,
                    "category_id": "non_classe",
                    "visible": True,
                    "parts": [part_record],
                    "created_at": datetime.now().isoformat(timespec="seconds")
                })

        self.snapshot_gpx_segments("Préparation des segments depuis les branches")
        workshop["segments"] = segments
        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.fit_gpx_workshop_map_to_content()
        self.gpx_workshop_status_var.set(
            f"{len(segments)} segment(s) initial(aux) préparé(s). Découpez, classez et fusionnez ensuite."
        )

    def gpx_part_sort_key(self, part):
        return (
            int(part.get("branch_order", 10**9)),
            int(part.get("part_order", 10**9)),
            float(part.get("start_fraction", 10**9))
        )

    def gpx_segment_sort_key(self, segment):
        parts = segment.get("parts", [])
        if not parts:
            return (10**9, 10**9, 10**9)
        return min(self.gpx_part_sort_key(part) for part in parts)

    def clear_gpx_editor_photo_markers(self):
        """
        Supprime uniquement la couche photo de l'atelier GPX.
        Les tracés et repères A/B restent intacts.
        """
        self.clear_photo_layer_objects("gpx")

    def sync_gpx_workshop_photo_markers(self):
        """
        Source unique et autoritaire pour les photos de l'atelier GPX.
        Toute reconstruction de la carte passe par cette méthode.
        """
        self.refresh_photo_layer("gpx", force=True)

    def toggle_gpx_workshop_photos(self):
        """
        Compatibilité avec les anciens appels booléens.
        """
        mode = "visible" if self.gpx_workshop_show_photos_var.get() else "hidden"
        self.set_gpx_photo_display_mode(mode)

    def clear_gpx_editor_drawings(self):
        for path in self.gpx_editor_paths:
            try:
                path.delete()
            except Exception:
                pass
        self.gpx_editor_paths = []

        for marker in self.gpx_editor_markers:
            try:
                marker.delete()
            except Exception:
                pass
        self.gpx_editor_markers = []

        self.clear_gpx_editor_photo_markers()

    def get_gpx_endpoint_icon(self, label):
        """
        Crée une pastille A/B lisible : la lettre est dans la bulle,
        pas affichée au-dessus du marqueur standard.
        """
        label = str(label).upper()
        cache_key = label
        if cache_key in self.gpx_endpoint_icon_cache:
            return self.gpx_endpoint_icon_cache[cache_key]

        size = 34
        image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)

        fill = "#f39c12" if label == "A" else "#8e44ad"
        edge = "#7f4f00" if label == "A" else "#4a235a"

        draw.ellipse((1, 1, size - 2, size - 2), fill="#ffffff", outline="#ffffff")
        draw.ellipse((3, 3, size - 4, size - 4), fill=fill, outline=edge, width=2)

        try:
            font = ImageFont.truetype("arialbd.ttf", 15)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", 15)
            except Exception:
                font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), label, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        text_x = (size - text_w) / 2
        text_y = (size - text_h) / 2 - 1
        draw.text((text_x, text_y), label, fill="#ffffff", font=font)

        icon = ImageTk.PhotoImage(image)
        self.gpx_endpoint_icon_cache[cache_key] = icon
        return icon

    def update_gpx_endpoint_toggle_button(self):
        if not self.gpx_endpoint_toggle_button:
            return

        if self.gpx_workshop_show_endpoints_var.get():
            self.gpx_endpoint_toggle_button.config(text="🙈 Masquer les repères A / B")
        else:
            self.gpx_endpoint_toggle_button.config(text="👁 Afficher les repères A / B")

    def toggle_gpx_endpoint_markers(self):
        self.gpx_workshop_show_endpoints_var.set(
            not self.gpx_workshop_show_endpoints_var.get()
        )
        self.update_gpx_endpoint_toggle_button()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            "Repères A / B affichés."
            if self.gpx_workshop_show_endpoints_var.get()
            else "Repères A / B masqués."
        )


    def add_gpx_endpoint_marker(self, lat, lon, label):
        if not self.gpx_editor_map:
            return

        try:
            icon = self.get_gpx_endpoint_icon(label)
            marker = self.gpx_editor_map.set_marker(
                lat,
                lon,
                icon=icon,
                icon_anchor="center"
            )
        except TypeError:
            # Repli si une ancienne version de tkintermapview ne supporte pas icon.
            marker = self.gpx_editor_map.set_marker(lat, lon, text=str(label))

        self.gpx_editor_markers.append(marker)

    def is_gpx_source_visible(self, source_id):
        """
        Visibilité d'une branche source.
        Utilisé aussi quand les segments existent : masquer une branche masque
        les parties de segment issues de cette branche.
        """
        if not source_id:
            return True

        source = self.find_gpx_source(source_id)
        if not source:
            return True

        return source.get("visible", True)

    def offset_positions_m(self, positions, offset_m):
        """
        Décale approximativement une polyligne de quelques mètres à gauche/droite.
        Suffisant pour un rendu visuel bicolore dans l'atelier.
        Les données exportées restent la géométrie centrale exacte.
        """
        if len(positions) < 2 or offset_m == 0:
            return positions

        # Approximation locale : 1 degré latitude ≈ 111'320 m.
        mean_lat = sum(p[0] for p in positions) / len(positions)
        meters_per_deg_lat = 111320.0
        meters_per_deg_lon = max(1.0, 111320.0 * math.cos(math.radians(mean_lat)))

        xy = [
            (lon * meters_per_deg_lon, lat * meters_per_deg_lat)
            for lat, lon in positions
        ]

        out = []
        n = len(xy)

        for i, (x, y) in enumerate(xy):
            if i == 0:
                dx = xy[1][0] - x
                dy = xy[1][1] - y
            elif i == n - 1:
                dx = x - xy[i - 1][0]
                dy = y - xy[i - 1][1]
            else:
                dx = xy[i + 1][0] - xy[i - 1][0]
                dy = xy[i + 1][1] - xy[i - 1][1]

            length = math.hypot(dx, dy)
            if length == 0:
                out.append(positions[i])
                continue

            nx = -dy / length
            ny = dx / length

            ox = x + nx * offset_m
            oy = y + ny * offset_m

            out_lat = oy / meters_per_deg_lat
            out_lon = ox / meters_per_deg_lon
            out.append((out_lat, out_lon))

        return out

    def draw_path_layers(self, positions, layers, command_segment_id=None):
        for layer_index, (layer_color, width) in enumerate(layers):
            try:
                path = self.gpx_editor_map.set_path(
                    positions,
                    color=layer_color,
                    width=width,
                    command=(
                        (lambda _path, sid=command_segment_id: self.select_gpx_segment_by_id(sid))
                        if command_segment_id and layer_index == len(layers) - 1 else None
                    )
                )
            except TypeError:
                path = self.gpx_editor_map.set_path(positions, color=layer_color)
            self.gpx_editor_paths.append(path)

    def bicolor_offset_m_for_current_zoom(self, positions, desired_px=2.0):
        """
        Calcule un décalage en mètres correspondant à un petit nombre de pixels
        à l'écran. Cela stabilise le rendu bicolore selon le niveau de zoom.

        Avant, le décalage était fixe en mètres : il paraissait presque nul à
        certains zooms et trop séparé à d'autres. Ici, on garde un écart visuel
        constant en pixels.
        """
        try:
            zoom = self.get_map_zoom_value(self.gpx_editor_map)
        except Exception:
            zoom = None

        if zoom is None:
            return 1.15

        try:
            mean_lat = sum(p[0] for p in positions) / len(positions)
            meters_per_pixel = 156543.03392 * math.cos(math.radians(mean_lat)) / (2 ** zoom)
            return max(0.15, meters_per_pixel * desired_px)
        except Exception:
            return 1.15

    def draw_gpx_bicolor_part_on_editor_map(self, segment, positions, selected=False):
        """
        Rendu bicolore compact et plus stable au zoom.

        Objectif visuel :
        - un seul segment cartographique épais ;
        - deux flancs colorés collés longitudinalement ;
        - halo externe commun ;
        - écart des deux flancs calculé en pixels, donc beaucoup moins variable
          quand on zoome/dézoome.
        """
        a_id, b_id = self.get_segment_bicolor_category_ids(segment)
        color_a = self.get_gpx_category_by_id(a_id).get("color", "#ef6c00")
        color_b = self.get_gpx_category_by_id(b_id).get("color", "#111111")

        # Décalage visuel quasi constant, exprimé d'abord en pixels puis converti
        # en mètres pour produire des coordonnées lat/lon.
        offset_m = self.bicolor_offset_m_for_current_zoom(positions, desired_px=1.9)
        positions_a = self.offset_positions_m(positions, offset_m)
        positions_b = self.offset_positions_m(positions, -offset_m)

        if selected:
            # Halo commun : le segment reste un seul objet visuel.
            self.draw_path_layers(
                positions,
                [
                    ("#2c3e50", 18),
                    ("#ffffff", 14)
                ],
                command_segment_id=None
            )
            colored_width = 8
        else:
            self.draw_path_layers(
                positions,
                [
                    ("#ffffff", 13)
                ],
                command_segment_id=None
            )
            colored_width = 7

        # Pas de halo individuel : les deux couleurs se touchent visuellement.
        self.draw_path_layers(
            positions_a,
            [(color_a, colored_width)],
            command_segment_id=None
        )
        self.draw_path_layers(
            positions_b,
            [(color_b, colored_width)],
            command_segment_id=segment.get("id")
        )


    def draw_gpx_segment_part_on_editor_map(self, segment, part, selected=False):
        if not self.is_gpx_source_visible(part.get("source_id")):
            return

        category = self.get_gpx_category_by_id(segment.get("category_id", "non_classe"))
        color = category.get("color", "#8e44ad")

        positions = [
            (float(point[0]), float(point[1]))
            for point in part.get("points", [])
            if point and len(point) >= 2
        ]
        if len(positions) < 2:
            return

        positions = self.thin_positions_for_display(positions)

        if segment.get("display_mode") == "bicolor":
            self.draw_gpx_bicolor_part_on_editor_map(segment, positions, selected=selected)
            return

        if selected:
            layers = [
                ("#2c3e50", 17),  # halo sombre discret
                ("#ffffff", 13),  # contour blanc renforcé
                (color, 8)        # couleur de catégorie conservée
            ]
        else:
            layers = [
                ("#ffffff", 10),
                (color, 5)
            ]

        self.draw_path_layers(positions, layers, command_segment_id=segment.get("id"))

    def draw_gpx_workshop_map(self):
        if not self.gpx_editor_map:
            return

        self.clear_gpx_editor_drawings()
        workshop = self.get_gpx_workshop_state()
        segments = workshop.get("segments", [])
        selected_segment_ids = set(self.get_selected_gpx_segment_ids())

        if segments:
            visible_segments = [
                segment for segment in segments
                if segment.get("visible", True)
            ]

            normal_segments = [
                segment for segment in visible_segments
                if segment.get("id") not in selected_segment_ids
            ]
            selected_segments = [
                segment for segment in visible_segments
                if segment.get("id") in selected_segment_ids
            ]

            # Les segments sélectionnés sont dessinés en dernier, donc au-dessus
            # des autres sans changer leur couleur de catégorie.
            for segment in normal_segments + selected_segments:
                selected = segment.get("id") in selected_segment_ids
                for part in segment.get("parts", []):
                    self.draw_gpx_segment_part_on_editor_map(segment, part, selected=selected)
        else:
            selected_source_id = self.gpx_workshop_selected_source_id
            for source in workshop.get("sources", []):
                if not source.get("visible", True):
                    continue

                selected = source.get("id") == selected_source_id
                line_color = "#d35400" if selected else "#7f8c8d"
                halo_width = 13 if selected else 9
                core_width = 7 if selected else 4

                for part in source.get("parts", []):
                    positions = [
                        (float(point[0]), float(point[1]))
                        for point in part
                        if point and len(point) >= 2
                    ]
                    if len(positions) < 2:
                        continue

                    positions = self.thin_positions_for_display(positions)

                    for layer_color, width in (("#ffffff", halo_width), (line_color, core_width)):
                        path = self.gpx_editor_map.set_path(
                            positions,
                            color=layer_color,
                            width=width
                        )
                        self.gpx_editor_paths.append(path)

            source = self.find_gpx_source(selected_source_id)
            if source and source.get("visible", True) and self.gpx_workshop_show_endpoints_var.get():
                endpoint_a, endpoint_b = self.source_endpoint_positions(source)
                if endpoint_a:
                    self.add_gpx_endpoint_marker(endpoint_a[0], endpoint_a[1], "A")
                if endpoint_b:
                    self.add_gpx_endpoint_marker(endpoint_b[0], endpoint_b[1], "B")

        self.update_gpx_endpoint_toggle_button()

        # Synchronisation stricte de la case « Afficher les photos » après
        # toute modification qui redessine la carte.
        self.sync_gpx_workshop_photo_markers()

    def select_gpx_segment_by_id(self, segment_id):
        if self.gpx_segment_tree:
            self.gpx_segment_tree.selection_set(segment_id)
            self.gpx_segment_tree.focus(segment_id)
            self.on_gpx_segment_selected()

    def draw_photos_on_gpx_workshop_map(self, clear_first=True):
        """
        Conservée pour compatibilité : la couche est désormais entièrement
        gérée par refresh_photo_layer().
        """
        self.refresh_photo_layer("gpx", force=True)

    def show_gpx_photo_quick_view(self, photo):
        """
        Actualise la visionneuse unique de l'atelier GPX.
        Selon l'option choisie, elle est intégrée à côté de la carte ou flottante.
        """
        if not self.gpx_workshop_photos:
            self.gpx_workshop_photos = self.load_geolocated_photos()

        if not self.gpx_workshop_photos:
            return

        index = 0
        target_id = photo.get("catalog_index")
        for i, candidate in enumerate(self.gpx_workshop_photos):
            if candidate.get("catalog_index") == target_id:
                index = i
                break

        if self.gpx_photo_viewer_integrated_var.get():
            self.apply_gpx_photo_viewer_mode(open_floating=False)
            self.show_gpx_viewer_photo_by_index(index, center_map=True)
            return

        self.ensure_gpx_photo_viewer_window()
        self.show_gpx_viewer_photo_by_index(index, center_map=True)

        try:
            self.gpx_photo_viewer_window.deiconify()
            self.gpx_photo_viewer_window.lift()
            self.gpx_photo_viewer_window.focus_force()
        except Exception:
            pass

    def ensure_gpx_photo_viewer_window(self):
        if self.gpx_photo_viewer_window and self.gpx_photo_viewer_window.winfo_exists():
            return

        window = tk.Toplevel(self.root)
        window.title("Visionneuse photos · Atelier GPX")
        window.geometry(self.gpx_photo_viewer_geometry or "820x720")
        window.minsize(620, 480)
        window.transient(self.root)
        window.protocol("WM_DELETE_WINDOW", lambda: self.close_gpx_photo_viewer(destroy=False))

        self.gpx_photo_viewer_window = window

        top = tk.Frame(window, padx=10, pady=8)
        top.pack(fill="x")

        tk.Button(top, text="◀ Précédente", command=lambda: self.navigate_gpx_photo_viewer(-1)).pack(side="left")
        tk.Label(
            top,
            textvariable=self.gpx_photo_viewer_index_var,
            font=("Arial", 11, "bold"),
            width=20
        ).pack(side="left", padx=10)
        tk.Button(top, text="Suivante ▶", command=lambda: self.navigate_gpx_photo_viewer(1)).pack(side="left")
        tk.Button(top, text="Intégrer", command=self.switch_gpx_viewer_to_integrated).pack(side="right", padx=(4, 0))
        tk.Button(top, text="Masquer", command=lambda: self.close_gpx_photo_viewer(destroy=False)).pack(side="right")

        tk.Label(
            window,
            textvariable=self.gpx_photo_viewer_filename_var,
            font=("Arial", 12, "bold"),
            anchor="w",
            justify="left"
        ).pack(fill="x", padx=12, pady=(0, 4))

        tk.Label(
            window,
            textvariable=self.gpx_photo_viewer_meta_var,
            anchor="w",
            justify="left",
            wraplength=780,
            fg="#444444"
        ).pack(fill="x", padx=12, pady=(0, 8))

        self.gpx_photo_viewer_canvas = tk.Canvas(window, bg="#20252b", highlightthickness=0)
        self.gpx_photo_viewer_canvas.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.gpx_photo_viewer_canvas.bind("<Configure>", lambda _event: self.render_gpx_photo_viewer_image())

        window.bind("<Left>", lambda _event: self.navigate_gpx_photo_viewer(-1))
        window.bind("<Right>", lambda _event: self.navigate_gpx_photo_viewer(1))
        window.bind("<Escape>", lambda _event: self.close_gpx_photo_viewer(destroy=False))

    def close_gpx_photo_viewer(self, destroy=False):
        window = self.gpx_photo_viewer_window
        if not window:
            return

        try:
            if window.winfo_exists():
                self.gpx_photo_viewer_geometry = window.geometry()
                if destroy:
                    window.destroy()
                else:
                    window.withdraw()
        except Exception:
            pass

        if destroy:
            self.gpx_photo_viewer_window = None
            self.gpx_photo_viewer_canvas = None
            self.gpx_photo_viewer_image = None
            self.gpx_photo_viewer_current_photo = None
            self.gpx_photo_viewer_index = -1
            self.gpx_photo_viewer_pil_cache = {}

    def load_gpx_viewer_pil_image(self, photo):
        image_path = photo.get("image_path", "")
        if not image_path or not os.path.exists(image_path):
            return None

        cache_key = os.path.abspath(image_path)
        if cache_key in self.gpx_photo_viewer_pil_cache:
            return self.gpx_photo_viewer_pil_cache[cache_key]

        try:
            image = Image.open(image_path)
            image = ImageOps.exif_transpose(image)
            image_copy = image.copy()
            self.gpx_photo_viewer_pil_cache[cache_key] = image_copy
            return image_copy
        except Exception:
            return None

    def preload_gpx_viewer_adjacent_images(self):
        photos = self.gpx_workshop_photos
        if not photos or self.gpx_photo_viewer_index < 0:
            return

        wanted_paths = set()
        for offset in (-1, 0, 1):
            index = (self.gpx_photo_viewer_index + offset) % len(photos)
            photo = photos[index]
            path = os.path.abspath(photo.get("image_path", ""))
            if path:
                wanted_paths.add(path)
                self.load_gpx_viewer_pil_image(photo)

        for cache_key in list(self.gpx_photo_viewer_pil_cache):
            if cache_key not in wanted_paths:
                self.gpx_photo_viewer_pil_cache.pop(cache_key, None)

    def show_gpx_viewer_photo_by_index(self, index, center_map=True):
        photos = self.gpx_workshop_photos
        if not photos:
            return

        index %= len(photos)
        photo = photos[index]

        self.gpx_photo_viewer_index = index
        self.gpx_photo_viewer_current_photo = photo
        self.gpx_photo_viewer_index_var.set(f"Photo {index + 1} / {len(photos)}")
        self.gpx_photo_viewer_filename_var.set(photo.get("filename", ""))

        meta = []
        if photo.get("title"):
            meta.append(f"Titre : {photo.get('title')}")
        if photo.get("description"):
            meta.append(f"Description : {photo.get('description')}")
        if photo.get("date_taken"):
            meta.append(f"Date : {photo.get('date_taken')}")
        if photo.get("lat") is not None and photo.get("lon") is not None:
            meta.append(f"Coordonnées : {photo.get('lat'):.6f}, {photo.get('lon'):.6f}")
        self.gpx_photo_viewer_meta_var.set("\n".join(meta) if meta else "Aucune métadonnée textuelle.")

        self.preload_gpx_viewer_adjacent_images()
        self.render_gpx_photo_viewer_image()
        self.refresh_photo_layer("gpx", force=True)

        if center_map:
            self.center_map_on_photo_if_needed(
                self.gpx_editor_map,
                photo.get("lat"),
                photo.get("lon")
            )

    def render_gpx_photo_viewer_image(self):
        canvas = self.get_active_gpx_photo_viewer_canvas()
        photo = self.gpx_photo_viewer_current_photo

        if not canvas:
            return

        canvas.delete("all")

        if not photo:
            canvas.create_text(
                max(10, canvas.winfo_width() // 2),
                max(10, canvas.winfo_height() // 2),
                text="Cliquez sur une pastille photo",
                fill="#d8e4ec",
                font=("Arial", 12, "bold")
            )
            return

        image = self.load_gpx_viewer_pil_image(photo)

        if image is None:
            canvas.create_text(
                max(10, canvas.winfo_width() // 2),
                max(10, canvas.winfo_height() // 2),
                text="Image introuvable ou illisible",
                fill="white"
            )
            return

        canvas_w = max(100, canvas.winfo_width())
        canvas_h = max(100, canvas.winfo_height())
        image_w, image_h = image.size
        scale = min(canvas_w / image_w, canvas_h / image_h, 1.0)
        target = (max(1, int(image_w * scale)), max(1, int(image_h * scale)))
        rendered = image.resize(target, Image.Resampling.LANCZOS)
        self.gpx_photo_viewer_image = ImageTk.PhotoImage(rendered)
        canvas.create_image(
            canvas_w // 2,
            canvas_h // 2,
            image=self.gpx_photo_viewer_image,
            anchor="center"
        )

    def navigate_gpx_photo_viewer(self, delta):
        if not self.gpx_workshop_photos:
            return "break"

        if self.gpx_photo_viewer_index < 0:
            new_index = 0
        else:
            new_index = (self.gpx_photo_viewer_index + int(delta)) % len(self.gpx_workshop_photos)

        self.show_gpx_viewer_photo_by_index(new_index, center_map=True)
        return "break"

    def focus_is_text_editor(self):
        try:
            widget = self.root.focus_get()
            if widget is None:
                return False
            return widget.winfo_class() in {
                "Entry", "Text", "TEntry", "TCombobox", "Spinbox", "TSpinbox"
            }
        except Exception:
            return False

    def handle_photo_navigation_key(self, event):
        if self.focus_is_text_editor():
            return None

        direction = -1 if str(event.keysym) == "Left" else 1

        if self.gpx_workshop_active and self.gpx_photo_viewer_current_photo:
            integrated_visible = (
                self.gpx_photo_viewer_integrated_var.get()
                and self.gpx_integrated_viewer_is_attached()
            )
            floating_visible = (
                self.gpx_photo_viewer_window
                and self.gpx_photo_viewer_window.winfo_exists()
                and str(self.gpx_photo_viewer_window.state()) != "withdrawn"
            )
            if integrated_visible or floating_visible:
                return self.navigate_gpx_photo_viewer(direction)

        if self.viewer_canvas is not None and self.current_photo is not None:
            if direction < 0:
                self.select_previous_photo()
            else:
                self.select_next_photo()
            return "break"

        return None

    def handle_global_escape_key(self, _event=None):
        if self.gpx_workshop_click_mode == "cut":
            self.cancel_gpx_cut_mode(silent=False)
            return "break"

        if (
            self.gpx_photo_viewer_window
            and self.gpx_photo_viewer_window.winfo_exists()
            and str(self.gpx_photo_viewer_window.state()) != "withdrawn"
        ):
            self.close_gpx_photo_viewer(destroy=False)
            return "break"

        return None

    def fit_gpx_workshop_map_to_content(self):
        if not self.gpx_editor_map:
            return

        points = []
        workshop = self.get_gpx_workshop_state()

        if workshop.get("segments"):
            for segment in workshop.get("segments", []):
                if not segment.get("visible", True):
                    continue
                for part in segment.get("parts", []):
                    if not self.is_gpx_source_visible(part.get("source_id")):
                        continue
                    for point in part.get("points", []):
                        if point and len(point) >= 2:
                            points.append((float(point[0]), float(point[1])))
        else:
            for source in workshop.get("sources", []):
                if not source.get("visible", True):
                    continue
                for part in source.get("parts", []):
                    for point in part:
                        if point and len(point) >= 2:
                            points.append((float(point[0]), float(point[1])))

        if self.gpx_photo_display_mode_var.get() != "hidden":
            points.extend(
                (p["lat"], p["lon"])
                for p in self.gpx_workshop_photos
                if p.get("lat") is not None and p.get("lon") is not None
            )

        if not points:
            return

        lats = [p[0] for p in points]
        lons = [p[1] for p in points]
        self.gpx_editor_map.fit_bounding_box(
            (max(lats), min(lons)),
            (min(lats), max(lons))
        )

    def handle_gpx_workshop_map_click(self, coords):
        if self.consume_photo_map_click_guard("gpx"):
            return

        if self.photo_spider_state.get("gpx") is not None:
            self.close_photo_spider("gpx", redraw=True)

        if self.gpx_workshop_click_mode == "cut":
            self.cut_selected_gpx_segment_at_coords(coords)

    def activate_gpx_cut_mode(self):
        selected = self.get_selected_gpx_segment_ids()
        if len(selected) != 1:
            messagebox.showwarning(
                "Sélection requise",
                "Sélectionnez exactement un segment avant d'activer le mode découpe."
            )
            return

        segment = self.find_gpx_segment(selected[0])
        if not segment:
            return

        if len(segment.get("parts", [])) != 1:
            messagebox.showwarning(
                "Segment discontinu",
                (
                    "Ce segment contient plusieurs parties discontinues.\n\n"
                    "Utilisez d’abord « Dissocier un segment discontinu », "
                    "puis coupez la partie voulue."
                )
            )
            return

        self.gpx_workshop_click_mode = "cut"
        self.gpx_workshop_pending_segment_id = selected[0]
        self.enter_gpx_segment_edit_photo_mode()
        self.gpx_workshop_status_var.set(
            "Mode découpe actif : les photos restent visibles en mode discret et les segments sont prioritaires. Cliquez près du tracé."
        )

    def set_gpx_photo_display_mode(self, mode, automatic=False):
        if mode not in {"visible", "discrete", "hidden"}:
            return

        if not automatic and self.gpx_workshop_click_mode == "cut":
            # Un choix manuel pendant la découpe devient le nouvel état voulu :
            # il ne sera pas écrasé à la fin du mode.
            self.gpx_photo_mode_before_edit = None

        self.gpx_photo_display_mode_var.set(mode)
        self.gpx_workshop_show_photos_var.set(mode != "hidden")
        self.refresh_photo_layer("gpx", force=True)

    def enter_gpx_segment_edit_photo_mode(self):
        if self.gpx_photo_mode_before_edit is None:
            self.gpx_photo_mode_before_edit = self.gpx_photo_display_mode_var.get()

        if self.gpx_photo_display_mode_var.get() == "visible":
            self.set_gpx_photo_display_mode("discrete", automatic=True)

    def exit_gpx_segment_edit_photo_mode(self):
        previous = self.gpx_photo_mode_before_edit
        self.gpx_photo_mode_before_edit = None

        if previous in {"visible", "discrete", "hidden"}:
            self.set_gpx_photo_display_mode(previous, automatic=True)

    def cancel_gpx_cut_mode(self, silent=False):
        if self.gpx_workshop_click_mode != "cut":
            return

        self.gpx_workshop_click_mode = None
        self.gpx_workshop_pending_segment_id = None
        self.exit_gpx_segment_edit_photo_mode()

        if not silent:
            self.gpx_workshop_status_var.set("Mode découpe annulé.")

    def cut_selected_gpx_segment_at_coords(self, coords):
        segment_id = self.gpx_workshop_pending_segment_id
        segment = self.find_gpx_segment(segment_id)
        if not segment:
            self.gpx_workshop_click_mode = None
            self.gpx_workshop_pending_segment_id = None
            self.exit_gpx_segment_edit_photo_mode()
            return

        parts = segment.get("parts", [])
        if len(parts) != 1:
            self.gpx_workshop_status_var.set("Découpe annulée : le segment n'est plus simple.")
            self.gpx_workshop_click_mode = None
            self.gpx_workshop_pending_segment_id = None
            self.exit_gpx_segment_edit_photo_mode()
            return

        part = parts[0]
        points = part.get("points", [])
        if len(points) < 3:
            self.gpx_workshop_status_var.set("Découpe impossible : segment trop court.")
            self.gpx_workshop_click_mode = None
            self.gpx_workshop_pending_segment_id = None
            self.exit_gpx_segment_edit_photo_mode()
            return

        click_lat, click_lon = float(coords[0]), float(coords[1])
        nearest_index = min(
            range(len(points)),
            key=lambda i: self.haversine_distance_m(
                click_lat, click_lon,
                float(points[i][0]), float(points[i][1])
            )
        )

        if nearest_index <= 0 or nearest_index >= len(points) - 1:
            self.gpx_workshop_status_var.set(
                "Découpe trop proche d'une extrémité : cliquez un peu plus à l'intérieur du segment."
            )
            return

        left_points = copy.deepcopy(points[:nearest_index + 1])
        right_points = copy.deepcopy(points[nearest_index:])

        n = max(1, len(points) - 1)
        split_ratio = nearest_index / n
        start_fraction = float(part.get("start_fraction", 0.0))
        end_fraction = float(part.get("end_fraction", 1.0))
        mid_fraction = start_fraction + (end_fraction - start_fraction) * split_ratio

        left_part = copy.deepcopy(part)
        left_part["points"] = left_points
        left_part["end_fraction"] = mid_fraction

        right_part = copy.deepcopy(part)
        right_part["points"] = right_points
        right_part["start_fraction"] = mid_fraction

        left_segment = {
            "id": uuid.uuid4().hex,
            "category_id": segment.get("category_id", "non_classe"),
            "visible": segment.get("visible", True),
            "display_mode": segment.get("display_mode", "single"),
            "bicolor_categories": copy.deepcopy(segment.get("bicolor_categories", [])),
            "parts": [left_part],
            "created_at": datetime.now().isoformat(timespec="seconds")
        }

        right_segment = {
            "id": uuid.uuid4().hex,
            "category_id": segment.get("category_id", "non_classe"),
            "visible": segment.get("visible", True),
            "display_mode": segment.get("display_mode", "single"),
            "bicolor_categories": copy.deepcopy(segment.get("bicolor_categories", [])),
            "parts": [right_part],
            "created_at": datetime.now().isoformat(timespec="seconds")
        }

        self.snapshot_gpx_segments("Découpe d’un segment")
        workshop = self.get_gpx_workshop_state()
        new_segments = []
        for existing in workshop.get("segments", []):
            if existing.get("id") == segment_id:
                new_segments.extend([left_segment, right_segment])
            else:
                new_segments.append(existing)
        workshop["segments"] = new_segments

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_click_mode = None
        self.gpx_workshop_pending_segment_id = None
        self.exit_gpx_segment_edit_photo_mode()
        self.gpx_workshop_status_var.set("✂️ Segment découpé.")

    def point_distance_m(self, a, b):
        return self.haversine_distance_m(float(a[0]), float(a[1]), float(b[0]), float(b[1]))

    def parts_touch(self, points_a, points_b, tolerance_m=2.0):
        if not points_a or not points_b:
            return None

        combos = [
            ("a_end_b_start", points_a[-1], points_b[0]),
            ("a_start_b_end", points_a[0], points_b[-1]),
            ("a_start_b_start", points_a[0], points_b[0]),
            ("a_end_b_end", points_a[-1], points_b[-1])
        ]

        best_mode = None
        best_distance = None
        for mode, pa, pb in combos:
            distance = self.point_distance_m(pa, pb)
            if best_distance is None or distance < best_distance:
                best_mode = mode
                best_distance = distance

        if best_distance is not None and best_distance <= tolerance_m:
            return best_mode
        return None

    def join_two_gpx_parts_if_touching(self, part_a, part_b, tolerance_m=2.0):
        points_a = copy.deepcopy(part_a.get("points", []))
        points_b = copy.deepcopy(part_b.get("points", []))

        if len(points_a) < 2 or len(points_b) < 2:
            return None

        mode = self.parts_touch(points_a, points_b, tolerance_m=tolerance_m)
        if not mode:
            return None

        if mode == "a_end_b_start":
            joined_points = points_a + points_b[1:]
        elif mode == "a_start_b_end":
            joined_points = points_b + points_a[1:]
        elif mode == "a_start_b_start":
            joined_points = list(reversed(points_a)) + points_b[1:]
        elif mode == "a_end_b_end":
            joined_points = points_a + list(reversed(points_b))[1:]
        else:
            return None

        new_part = copy.deepcopy(part_a)
        new_part["points"] = joined_points
        new_part["joined_parts"] = True

        same_origin = (
            part_a.get("source_id") == part_b.get("source_id")
            and part_a.get("part_order") == part_b.get("part_order")
        )

        try:
            new_part["branch_order"] = min(int(part_a.get("branch_order", 10**9)), int(part_b.get("branch_order", 10**9)))
            new_part["part_order"] = min(int(part_a.get("part_order", 10**9)), int(part_b.get("part_order", 10**9)))
            new_part["start_fraction"] = min(float(part_a.get("start_fraction", 0.0)), float(part_b.get("start_fraction", 0.0)))
            new_part["end_fraction"] = max(float(part_a.get("end_fraction", 1.0)), float(part_b.get("end_fraction", 1.0)))
        except Exception:
            pass

        if not same_origin:
            new_part["mixed_origin_join"] = True

        return new_part

    def normalize_merged_gpx_parts(self, parts, tolerance_m=2.0):
        """
        Recollage intelligent :
        - parties jointives → une seule polyligne continue ;
        - parties disjointes → plusieurs parties dans le même segment logique.
        """
        remaining = [
            copy.deepcopy(part)
            for part in parts
            if len(part.get("points", [])) >= 2
        ]
        remaining.sort(key=self.gpx_part_sort_key)

        changed = True
        while changed and len(remaining) > 1:
            changed = False
            for i in range(len(remaining)):
                if changed:
                    break
                for j in range(i + 1, len(remaining)):
                    joined = self.join_two_gpx_parts_if_touching(remaining[i], remaining[j], tolerance_m=tolerance_m)
                    if joined:
                        rebuilt = []
                        for k, part in enumerate(remaining):
                            if k not in (i, j):
                                rebuilt.append(part)
                        rebuilt.append(joined)
                        rebuilt.sort(key=self.gpx_part_sort_key)
                        remaining = rebuilt
                        changed = True
                        break

        remaining.sort(key=self.gpx_part_sort_key)
        return remaining

    def merge_selected_gpx_segments(self):
        ids = self.get_selected_gpx_segment_ids()
        if len(ids) < 2:
            messagebox.showwarning("Fusion", "Sélectionnez au moins deux segments pour les fusionner.")
            return

        workshop = self.get_gpx_workshop_state()
        selected_segments = [
            segment for segment in workshop.get("segments", [])
            if segment.get("id") in ids
        ]
        if len(selected_segments) < 2:
            return

        categories = {seg.get("category_id", "non_classe") for seg in selected_segments}
        mixed_categories = len(categories) != 1
        category_id = next(iter(categories)) if not mixed_categories else "non_classe"

        raw_parts = []
        for segment in selected_segments:
            raw_parts.extend(copy.deepcopy(segment.get("parts", [])))

        merged_parts = self.normalize_merged_gpx_parts(raw_parts, tolerance_m=2.0)

        bicolor_signatures = {
            tuple(seg.get("bicolor_categories", []))
            for seg in selected_segments
            if seg.get("display_mode") == "bicolor"
        }
        all_bicolor_same = (
            len(bicolor_signatures) == 1
            and all(seg.get("display_mode") == "bicolor" for seg in selected_segments)
        )

        merged_segment = {
            "id": uuid.uuid4().hex,
            "category_id": category_id,
            "visible": True,
            "display_mode": "bicolor" if all_bicolor_same else "single",
            "bicolor_categories": list(next(iter(bicolor_signatures))) if all_bicolor_same else [],
            "parts": merged_parts,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "merged_from": [seg.get("id") for seg in selected_segments]
        }

        self.snapshot_gpx_segments("Fusion intelligente de segments")
        remaining = [
            segment for segment in workshop.get("segments", [])
            if segment.get("id") not in ids
        ]
        remaining.append(merged_segment)
        workshop["segments"] = sorted(remaining, key=self.gpx_segment_sort_key)

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        if self.gpx_segment_tree:
            self.gpx_segment_tree.selection_set(merged_segment.get("id"))
            self.gpx_segment_tree.focus(merged_segment.get("id"))
        self.draw_gpx_workshop_map()

        if len(merged_parts) == 1:
            continuity = "Les parties jointives ont été recollées en une trace continue."
        else:
            continuity = f"Le segment fusionné reste discontinu en {len(merged_parts)} parties."

        if mixed_categories:
            status = f"Segments fusionnés ; catégories différentes → résultat remis en « Non classé ». {continuity}"
        else:
            status = f"Segments fusionnés. {continuity}"

        self.gpx_workshop_status_var.set(status)

    def split_selected_discontinuous_segment(self):
        ids = self.get_selected_gpx_segment_ids()
        if len(ids) != 1:
            messagebox.showwarning(
                "Dissocier",
                "Sélectionnez exactement un segment discontinu à dissocier."
            )
            return

        segment = self.find_gpx_segment(ids[0])
        if not segment:
            return

        parts = segment.get("parts", [])
        if len(parts) <= 1:
            self.gpx_workshop_status_var.set("Ce segment ne contient qu’une seule partie.")
            return

        new_segments = [
            {
                "id": uuid.uuid4().hex,
                "category_id": segment.get("category_id", "non_classe"),
                "visible": segment.get("visible", True),
                "display_mode": segment.get("display_mode", "single"),
                "bicolor_categories": copy.deepcopy(segment.get("bicolor_categories", [])),
                "parts": [copy.deepcopy(part)],
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "split_from": segment.get("id")
            }
            for part in parts
        ]

        self.snapshot_gpx_segments("Dissociation d’un segment discontinu")
        workshop = self.get_gpx_workshop_state()
        rebuilt = []
        for existing in workshop.get("segments", []):
            if existing.get("id") == segment.get("id"):
                rebuilt.extend(new_segments)
            else:
                rebuilt.append(existing)
        workshop["segments"] = sorted(rebuilt, key=self.gpx_segment_sort_key)

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            f"Segment discontinu dissocié en {len(new_segments)} segment(s)."
        )

    def delete_selected_gpx_segments(self):
        ids = self.get_selected_gpx_segment_ids()
        if not ids:
            return

        if not messagebox.askyesno(
            "Retirer les segments ?",
            f"Retirer {len(ids)} segment(s) de l’atelier ?\n\nCette action est annulable."
        ):
            return

        self.snapshot_gpx_segments("Suppression de segments de l’atelier")
        workshop = self.get_gpx_workshop_state()
        workshop["segments"] = [
            segment for segment in workshop.get("segments", [])
            if segment.get("id") not in ids
        ]
        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(f"{len(ids)} segment(s) retiré(s) de l’atelier.")

    def apply_category_to_selected_segments(self):
        ids = self.get_selected_gpx_segment_ids()
        if not ids:
            messagebox.showwarning("Aucun segment", "Sélectionnez au moins un segment.")
            return

        label = self.gpx_category_combo.get() if self.gpx_category_combo else ""
        category_id = self.category_label_to_id(label)
        category = self.get_gpx_category_by_id(category_id)

        self.snapshot_gpx_segments("Changement de catégorie")
        workshop = self.get_gpx_workshop_state()
        changed = 0
        for segment in workshop.get("segments", []):
            if segment.get("id") in ids:
                segment["category_id"] = category_id
                changed += 1

        self.save_gpx_workshop_state()
        self.refresh_gpx_segment_tree()
        if self.gpx_segment_tree:
            self.gpx_segment_tree.selection_set(ids)
        self.draw_gpx_workshop_map()
        self.gpx_workshop_status_var.set(
            f"{changed} segment(s) classé(s) : {category.get('label', '—')}."
        )

    def add_gpx_category(self):
        """
        Depuis l'atelier GPX, l'ajout ouvre le même éditeur que Paramètres.
        """
        self.open_global_segment_category_editor()

    def export_gpx_workshop_segments(self):
        workshop = self.get_gpx_workshop_state()
        segments = workshop.get("segments", [])

        if not segments:
            messagebox.showwarning(
                "Aucun segment",
                "Préparez et classez d'abord les segments dans l’atelier."
            )
            return

        classified = [
            segment for segment in segments
            if segment.get("category_id", "non_classe") != "non_classe"
        ]
        unclassified_count = len(segments) - len(classified)

        if not classified:
            messagebox.showwarning(
                "Aucun segment classé",
                "Aucun segment n’est encore classé. L’export est donc impossible."
            )
            return

        if unclassified_count:
            if not messagebox.askyesno(
                "Segments non classés",
                (
                    f"{unclassified_count} segment(s) sont encore « Non classé » et ne seront pas exportés.\n\n"
                    "Continuer ?"
                )
            ):
                return

        os.makedirs(self.gpx_folder, exist_ok=True)

        project_name = self.sanitize_filename_part(
            self.catalog_container.get("project", {}).get("bisse_name", "") or os.path.basename(self.base_folder)
        )

        grouped = {}
        for segment in classified:
            grouped.setdefault(segment.get("category_id"), []).append(segment)

        planned_exports = []
        for category_id, category_segments in grouped.items():
            category = self.get_gpx_category_by_id(category_id)
            file_code = self.sanitize_filename_part(category.get("file_code", category_id))
            ordered_segments = sorted(category_segments, key=self.gpx_segment_sort_key)

            for number, segment in enumerate(ordered_segments, start=1):
                filename = f"{project_name}_{file_code}_{number:02d}.gpx"
                target_path = os.path.join(self.gpx_folder, filename)
                planned_exports.append((segment, category, target_path))

        existing_targets = [path for _seg, _cat, path in planned_exports if os.path.exists(path)]
        if existing_targets:
            preview = "\n".join(f"• {os.path.basename(path)}" for path in existing_targets[:8])
            if not messagebox.askyesno(
                "Écraser les exports existants ?",
                (
                    "Certains fichiers d’export existent déjà :\n\n"
                    f"{preview}\n\n"
                    "Les écraser ?"
                )
            ):
                return

        exported_trace_records = []
        for segment, category, target_path in planned_exports:
            self.write_workshop_segment_to_gpx(target_path, segment, category)
            parts_as_segments = [
                copy.deepcopy(part.get("points", []))
                for part in segment.get("parts", [])
                if len(part.get("points", [])) >= 2
            ]
            exported_trace_records.append({
                "id": uuid.uuid4().hex,
                "category": category.get("id"),
                "label": category.get("label"),
                "color": category.get("color"),
                "source_filename": os.path.basename(target_path),
                "source_relative_path": self.relative_to_base(target_path),
                "exported_at": datetime.now().isoformat(timespec="seconds"),
                "point_count": sum(len(part) for part in parts_as_segments),
                "length_m": round(self.calculate_trace_length_m(parts_as_segments), 2),
                "segments": parts_as_segments,
                "workshop_segment_id": segment.get("id")
            })
            self.log(f"📤 GPX exporté : {os.path.basename(target_path)}")

        traces = self.get_trace_sections()
        traces["manual_segments"] = exported_trace_records
        workshop["last_export_at"] = datetime.now().isoformat(timespec="seconds")
        self.save_gpx_workshop_state()

        self.draw_gpx_workshop_map()
        messagebox.showinfo(
            "Export terminé",
            (
                f"{len(exported_trace_records)} fichier(s) GPX exporté(s) dans « Fichiers GPX ».\n\n"
                "Ils sont aussi enregistrés dans catalogue.json et affichables dans l’atelier Photos."
            )
        )
        self.gpx_workshop_status_var.set(
            f"Export terminé : {len(exported_trace_records)} tronçon(s) GPX."
        )

    def write_workshop_segment_to_gpx(self, target_path, segment, category):
        gpx = gpxpy.gpx.GPX()
        track = gpxpy.gpx.GPXTrack()
        track.name = f"{category.get('label', 'Tronçon')} — {os.path.basename(target_path)}"
        gpx.tracks.append(track)

        for part in segment.get("parts", []):
            points = part.get("points", [])
            if len(points) < 2:
                continue

            track_segment = gpxpy.gpx.GPXTrackSegment()
            track.segments.append(track_segment)

            for point in points:
                lat = float(point[0])
                lon = float(point[1])
                ele = point[2] if len(point) > 2 else None
                track_segment.points.append(
                    gpxpy.gpx.GPXTrackPoint(latitude=lat, longitude=lon, elevation=ele)
                )

        with open(target_path, "w", encoding="utf-8") as f:
            f.write(gpx.to_xml())

    # ============================================================
    # MODULE GPS EXIF : PHOTOS DÉJÀ GÉOLOCALISÉES
    # ============================================================

    def rational_to_float(self, value):
        try:
            if isinstance(value, tuple) and len(value) == 2:
                num, den = value
                if den == 0:
                    return 0.0
                return float(num) / float(den)

            if hasattr(value, "numerator") and hasattr(value, "denominator"):
                if value.denominator == 0:
                    return 0.0
                return float(value.numerator) / float(value.denominator)

            return float(value)

        except Exception:
            return 0.0

    def gps_dms_to_decimal(self, dms, ref):
        if not dms or len(dms) != 3:
            return None

        deg = self.rational_to_float(dms[0])
        minutes = self.rational_to_float(dms[1])
        seconds = self.rational_to_float(dms[2])

        decimal = deg + (minutes / 60.0) + (seconds / 3600.0)

        if isinstance(ref, bytes):
            ref = ref.decode("ascii", errors="ignore")

        ref = str(ref).upper().strip()

        if ref in ("S", "W"):
            decimal = -decimal

        return decimal

    def read_gps_metadata_from_jpg(self, image_path):
        result = {
            "ok": False,
            "lat": None,
            "lon": None,
            "ele": None,
            "error": ""
        }

        if not image_path or not os.path.exists(image_path):
            result["error"] = "Image introuvable"
            return result

        if not image_path.lower().endswith((".jpg", ".jpeg")):
            result["error"] = "Le fichier n'est pas un JPG"
            return result

        try:
            exif_dict = piexif.load(image_path)
            gps = exif_dict.get("GPS", {})

            if not gps:
                result["error"] = "Aucun bloc GPS EXIF"
                return result

            lat_tuple = gps.get(piexif.GPSIFD.GPSLatitude)
            lat_ref = gps.get(piexif.GPSIFD.GPSLatitudeRef)

            lon_tuple = gps.get(piexif.GPSIFD.GPSLongitude)
            lon_ref = gps.get(piexif.GPSIFD.GPSLongitudeRef)

            if not lat_tuple or not lat_ref or not lon_tuple or not lon_ref:
                result["error"] = "Coordonnées GPS incomplètes"
                return result

            lat = self.gps_dms_to_decimal(lat_tuple, lat_ref)
            lon = self.gps_dms_to_decimal(lon_tuple, lon_ref)

            if lat is None or lon is None:
                result["error"] = "Coordonnées GPS illisibles"
                return result

            ele = None
            alt_value = gps.get(piexif.GPSIFD.GPSAltitude)
            alt_ref = gps.get(piexif.GPSIFD.GPSAltitudeRef, 0)

            if alt_value is not None:
                ele = self.rational_to_float(alt_value)

                if isinstance(alt_ref, bytes):
                    alt_ref_value = alt_ref[0] if alt_ref else 0
                else:
                    alt_ref_value = int(alt_ref)

                if alt_ref_value == 1:
                    ele = -ele

            result["ok"] = True
            result["lat"] = lat
            result["lon"] = lon
            result["ele"] = ele
            return result

        except Exception as e:
            result["error"] = str(e)
            return result

    def import_gps_from_existing_metadata(self):
        if not os.path.exists(self.catalog_path):
            messagebox.showerror(
                "Erreur",
                "Aucun catalogue.json trouvé. Créez d'abord le catalogue."
            )
            return

        if not messagebox.askyesno(
            "Confirmation",
            (
                "Lire les coordonnées GPS déjà présentes dans les fichiers JPG ?\n\n"
                "Le logiciel va parcourir les images du catalogue et remplir "
                "les coordonnées lorsqu'elles existent déjà dans les métadonnées."
            )
        ):
            return

        try:
            self.catalog_data = self.read_catalog()

            total = len(self.catalog_data)
            found_count = 0
            missing_count = 0
            error_count = 0

            self.progress["value"] = 0
            self.log("📷 Lecture des coordonnées GPS déjà présentes dans les JPG...")

            for i, entry in enumerate(self.catalog_data):
                if entry.get("status") != "OK":
                    continue

                image_path = self.get_entry_image_path(entry)
                filename = os.path.basename(image_path)

                gps_meta = self.read_gps_metadata_from_jpg(image_path)

                if gps_meta.get("ok"):
                    entry["gps_coordinates"] = {
                        "lat": gps_meta["lat"],
                        "lon": gps_meta["lon"],
                        "ele": gps_meta["ele"]
                    }

                    entry["gps_sync"] = "OK_METADATA"
                    entry["gps_source"] = "JPG_EXIF"

                    try:
                        dt = self.get_capture_datetime_for_sort(image_path, entry)
                        if dt and dt != datetime.max:
                            entry["date_taken"] = dt.isoformat()
                    except Exception:
                        pass

                    found_count += 1

                    if gps_meta["ele"] is not None:
                        self.log(
                            f"✅ {filename} -> {gps_meta['lat']:.6f}, "
                            f"{gps_meta['lon']:.6f}, {gps_meta['ele']:.1f} m"
                        )
                    else:
                        self.log(
                            f"✅ {filename} -> {gps_meta['lat']:.6f}, {gps_meta['lon']:.6f}"
                        )

                else:
                    if gps_meta.get("error"):
                        missing_count += 1
                        self.log(f"⚪ {filename} : pas de GPS EXIF ({gps_meta.get('error')})")
                    else:
                        error_count += 1
                        self.log(f"⚠️ {filename} : GPS illisible")

                self.progress["value"] = ((i + 1) / total) * 100
                self.root.update_idletasks()

            self.save_catalog()

            self.log("-" * 40)
            self.log("✅ Lecture GPS EXIF terminée")
            self.log(f"📍 Photos avec GPS trouvé : {found_count}")
            self.log(f"⚪ Photos sans GPS EXIF : {missing_count}")
            self.log(f"⚠️ Erreurs : {error_count}")

            messagebox.showinfo(
                "Lecture GPS terminée",
                (
                    f"Photos avec coordonnées GPS trouvées : {found_count}\n"
                    f"Photos sans GPS EXIF : {missing_count}\n"
                    f"Erreurs : {error_count}"
                )
            )

            self.load_folder(self.base_folder)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur lecture GPS EXIF : {e}")

    # ============================================================
    # MODULE 2 : GPS + FUSEAU HORAIRE + TRACE GPX
    # ============================================================


    # ============================================================
    # MODULE GPS EXIF : ÉCRITURE DES COORDONNÉES DANS LES JPG
    # ============================================================

    def decimal_to_gps_dms_rational(self, value):
        """
        Convertit un degré décimal en DMS rationnel EXIF.
        """
        value = abs(float(value))
        degrees = int(value)
        minutes_float = (value - degrees) * 60
        minutes = int(minutes_float)
        seconds = (minutes_float - minutes) * 60

        # précision au 1/10000e de seconde
        return (
            (degrees, 1),
            (minutes, 1),
            (int(round(seconds * 10000)), 10000)
        )

    def float_to_rational(self, value, precision=100):
        value = float(value)
        return (int(round(value * precision)), precision)

    def parse_iso_datetime_to_utc(self, value):
        if not value:
            return None

        try:
            text_value = str(value).strip()
            if text_value.endswith("Z"):
                text_value = text_value[:-1] + "+00:00"

            dt = datetime.fromisoformat(text_value)

            if dt.tzinfo is None:
                # Dans le catalogue, les dates GPX synchronisées sont normalement UTC.
                dt = dt.replace(tzinfo=timezone.utc)

            return dt.astimezone(timezone.utc)
        except Exception:
            return None

    def write_gps_metadata_to_jpg(self, image_path, lat, lon, ele=None, gps_datetime=None):
        """
        Écrit les coordonnées GPS dans les métadonnées EXIF d'un JPG/JPEG.

        Ne s'applique volontairement pas aux HEIC/HEIF : le logiciel écrit dans
        les JPG utilisés pour le travail, soit les JPG originaux, soit les JPG
        convertis dans Export_JPG.
        """
        if not image_path or not os.path.exists(image_path):
            return {"ok": False, "error": "Image introuvable"}

        if not image_path.lower().endswith((".jpg", ".jpeg")):
            return {"ok": False, "error": "Le fichier n'est pas un JPG/JPEG"}

        if lat is None or lon is None:
            return {"ok": False, "error": "Coordonnées manquantes"}

        try:
            lat = float(lat)
            lon = float(lon)

            if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                return {"ok": False, "error": "Coordonnées hors limites"}

            try:
                exif_dict = piexif.load(image_path)
            except Exception:
                exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}

            gps_ifd = exif_dict.setdefault("GPS", {})

            gps_ifd[piexif.GPSIFD.GPSVersionID] = (2, 3, 0, 0)

            gps_ifd[piexif.GPSIFD.GPSLatitudeRef] = b"N" if lat >= 0 else b"S"
            gps_ifd[piexif.GPSIFD.GPSLatitude] = self.decimal_to_gps_dms_rational(lat)

            gps_ifd[piexif.GPSIFD.GPSLongitudeRef] = b"E" if lon >= 0 else b"W"
            gps_ifd[piexif.GPSIFD.GPSLongitude] = self.decimal_to_gps_dms_rational(lon)

            if ele is not None:
                try:
                    ele_value = float(ele)
                    gps_ifd[piexif.GPSIFD.GPSAltitudeRef] = 1 if ele_value < 0 else 0
                    gps_ifd[piexif.GPSIFD.GPSAltitude] = self.float_to_rational(abs(ele_value), precision=100)
                except Exception:
                    pass

            gps_ifd[piexif.GPSIFD.GPSMapDatum] = b"WGS-84"

            dt_utc = self.parse_iso_datetime_to_utc(gps_datetime)

            if dt_utc:
                gps_ifd[piexif.GPSIFD.GPSDateStamp] = dt_utc.strftime("%Y:%m:%d").encode("ascii")
                gps_ifd[piexif.GPSIFD.GPSTimeStamp] = (
                    (dt_utc.hour, 1),
                    (dt_utc.minute, 1),
                    (dt_utc.second, 1)
                )

            exif_bytes = piexif.dump(exif_dict)
            piexif.insert(exif_bytes, image_path)

            return {"ok": True, "error": ""}

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def write_catalog_gps_to_jpg_metadata(self, ask_confirmation=True, only_ok=True):
        """
        Écrit les coordonnées GPS du catalogue dans les JPG utilisés par le logiciel.
        """
        if not os.path.exists(self.catalog_path):
            messagebox.showerror("Erreur", "Aucun catalogue.json trouvé.")
            return

        try:
            self.catalog_data = self.read_catalog()
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de lire le catalogue :\n{e}")
            return

        candidates = []

        for entry in self.catalog_data:
            if entry.get("status") != "OK":
                continue

            coords = entry.get("gps_coordinates") or {}
            lat = coords.get("lat")
            lon = coords.get("lon")

            if lat is None or lon is None:
                continue

            if only_ok and not str(entry.get("gps_sync", "")).startswith("OK"):
                continue

            image_path = self.get_entry_image_path(entry)

            candidates.append((entry, image_path, coords))

        if not candidates:
            messagebox.showwarning(
                "Aucune coordonnée",
                "Aucune photo du catalogue ne contient des coordonnées GPS à écrire."
            )
            return

        if ask_confirmation:
            if not messagebox.askyesno(
                "Écrire GPS dans les JPG ?",
                (
                    "Écrire les coordonnées GPS du catalogue dans les métadonnées EXIF "
                    "des JPG utilisés par le logiciel ?\n\n"
                    f"Photos concernées : {len(candidates)}\n\n"
                    "Important :\n"
                    "- les JPG originaux seront modifiés directement ;\n"
                    "- les HEIC/HEIF ne seront pas modifiés ;\n"
                    "- pour les HEIC/HEIF, les coordonnées seront écrites dans les JPG convertis "
                    "du dossier Export_JPG.\n\n"
                    "Le catalogue JSON ne sera pas modifié, seules les métadonnées des JPG le seront."
                )
            ):
                return

        self.progress["value"] = 0
        self.log("✍️ Écriture des coordonnées GPS dans les métadonnées EXIF des JPG...")

        written_count = 0
        skipped_count = 0
        error_count = 0
        total = len(candidates)

        for i, (entry, image_path, coords) in enumerate(candidates):
            filename = os.path.basename(image_path)
            lat = coords.get("lat")
            lon = coords.get("lon")
            ele = coords.get("ele")
            gps_datetime = entry.get("date_taken")

            result = self.write_gps_metadata_to_jpg(
                image_path,
                lat,
                lon,
                ele=ele,
                gps_datetime=gps_datetime
            )

            if result.get("ok"):
                written_count += 1
                entry["gps_written_to_jpg_exif"] = True
                entry["gps_written_to_jpg_exif_at"] = datetime.now().isoformat(timespec="seconds")
                self.log(f"✅ GPS écrit dans EXIF : {filename}")
            else:
                error_count += 1
                entry["gps_written_to_jpg_exif"] = False
                entry["gps_written_to_jpg_exif_error"] = result.get("error", "")
                self.log(f"⚠️ GPS non écrit : {filename} · {result.get('error', '')}")

            self.progress["value"] = ((i + 1) / total) * 100
            self.root.update_idletasks()

        # On sauvegarde les petits indicateurs gps_written_to_jpg_exif dans le catalogue.
        try:
            if not isinstance(self.catalog_container, dict):
                self.catalog_container = self.read_catalog_container()
            self.save_catalog()
        except Exception as e:
            self.log(f"⚠️ Les JPG ont été traités, mais le catalogue n'a pas pu être mis à jour : {e}")

        self.log("-" * 40)
        self.log("✅ Écriture GPS EXIF terminée")
        self.log(f"📍 GPS écrits : {written_count}")
        self.log(f"⚠️ Erreurs : {error_count}")
        self.log(f"⚪ Ignorées : {skipped_count}")

        messagebox.showinfo(
            "Écriture GPS terminée",
            (
                "Écriture des coordonnées GPS dans les JPG terminée.\n\n"
                f"GPS écrits : {written_count}\n"
                f"Erreurs : {error_count}"
            )
        )

    def offer_write_gps_to_jpg_after_sync(self, synced_count):
        """
        Après synchronisation GPX, écrit automatiquement dans les JPG si cela
        a été approuvé dans la fenêtre de synchronisation.
        """
        if synced_count <= 0:
            return

        if getattr(self, "write_exif_after_gpx_sync", False):
            self.log("✍️ Écriture EXIF JPG automatique après synchronisation GPX...")
            self.write_catalog_gps_to_jpg_metadata(ask_confirmation=False)
        else:
            self.log("📝 Coordonnées GPS gardées dans le catalogue uniquement. Écriture EXIF JPG non exécutée.")

    def select_gpx(self):
        initialdir = self.gpx_folder if self.gpx_folder and os.path.isdir(self.gpx_folder) else self.base_folder

        file_paths = filedialog.askopenfilenames(
            title="Choisir une ou plusieurs traces GPS horodatées",
            filetypes=[("Fichiers GPX", "*.gpx"), ("Tous les fichiers", "*.*")],
            initialdir=initialdir if initialdir and os.path.isdir(initialdir) else None
        )

        if not file_paths:
            self.gpx_file = None
            self.gpx_files = []
            if hasattr(self, "lbl_gpx"):
                self.lbl_gpx.config(text="Aucune trace", fg="gray")
            if hasattr(self, "btn_sync"):
                self.btn_sync.config(state="disabled")
            return

        self.gpx_files = list(file_paths)
        self.gpx_file = self.gpx_files[0]

        if hasattr(self, "lbl_gpx"):
            if len(self.gpx_files) == 1:
                label = f"Trace : {os.path.basename(self.gpx_files[0])}"
            else:
                first = os.path.basename(self.gpx_files[0])
                label = f"{len(self.gpx_files)} traces sélectionnées · première : {first}"
            self.lbl_gpx.config(text=label, fg="green")

        if hasattr(self, "btn_sync"):
            self.btn_sync.config(state="normal")

        if len(self.gpx_files) == 1:
            self.log(f"🛰️ Trace GPX chargée : {os.path.basename(self.gpx_files[0])}")
        else:
            self.log(f"🛰️ {len(self.gpx_files)} traces GPX chargées :")
            for path in self.gpx_files:
                self.log(f"   • {os.path.basename(path)}")

    def ask_photo_timezone(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Fuseau horaire des photos")
        dialog.geometry("460x270")
        dialog.transient(self.root)
        dialog.grab_set()

        result = {"timezone": None}

        tk.Label(
            dialog,
            text="Fuseau horaire des heures inscrites dans les photos",
            font=("Arial", 12, "bold")
        ).pack(pady=(15, 5))

        tk.Label(
            dialog,
            text=(
                "Les traces GPX sont normalement en UTC.\n"
                "Les photos gardent souvent l'heure locale de l'appareil.\n\n"
                "Pour les bisses en Valais, gardez Europe/Zurich."
            ),
            justify="left",
            wraplength=420
        ).pack(pady=5)

        choices = [
            "Europe/Zurich — Valais / Suisse / choix recommandé",
            "Europe/Paris — France",
            "UTC — si l'appareil photo était réglé en UTC",
            "Autre..."
        ]

        selected = tk.StringVar(value=choices[0])

        combo = ttk.Combobox(
            dialog,
            textvariable=selected,
            values=choices,
            state="readonly",
            width=52
        )
        combo.pack(pady=10)

        custom_var = tk.StringVar(value="Europe/Zurich")

        custom_entry = tk.Entry(dialog, textvariable=custom_var, width=36)
        custom_entry.pack(pady=3)
        custom_entry.pack_forget()

        def on_combo_change(event=None):
            if selected.get() == "Autre...":
                custom_entry.pack(pady=3)
                custom_entry.focus_set()
            else:
                custom_entry.pack_forget()

        combo.bind("<<ComboboxSelected>>", on_combo_change)

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=12)

        def validate():
            choice = selected.get()

            if choice.startswith("Europe/Zurich"):
                tz_name = "Europe/Zurich"
            elif choice.startswith("Europe/Paris"):
                tz_name = "Europe/Paris"
            elif choice.startswith("UTC"):
                tz_name = "UTC"
            else:
                tz_name = custom_var.get().strip()

            try:
                ZoneInfo(tz_name)
            except Exception:
                messagebox.showerror(
                    "Fuseau invalide",
                    (
                        f"Fuseau horaire invalide : {tz_name}\n\n"
                        "Exemples valides :\n"
                        "Europe/Zurich\n"
                        "Europe/Paris\n"
                        "UTC"
                    )
                )
                return

            result["timezone"] = tz_name
            dialog.destroy()

        def cancel():
            result["timezone"] = None
            dialog.destroy()

        tk.Button(btn_frame, text="OK", command=validate, width=12).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Annuler", command=cancel, width=12).pack(side="left", padx=5)

        self.root.wait_window(dialog)
        return result["timezone"]

    def ask_gpx_sync_gap_tolerance(self):
        """
        Demande la tolérance hors plage en minutes.

        0 = aucune position hors plage. Une photo doit tomber dans l'intervalle
        temporel d'un segment GPX pour être synchronisée.
        """
        value = simpledialog.askinteger(
            "Tolérance hors plage GPX",
            (
                "Si une photo est juste avant ou juste après une trace GPX,\n"
                "peut-on utiliser le point GPX le plus proche ?\n\n"
                "Indiquez la tolérance maximale en minutes.\n"
                "0 = ne jamais utiliser de point hors plage.\n\n"
                "Valeur recommandée : 30"
            ),
            parent=self.root,
            initialvalue=getattr(self, "gpx_sync_max_gap_minutes", 30),
            minvalue=0,
            maxvalue=1440
        )

        if value is None:
            return None

        self.gpx_sync_max_gap_minutes = int(value)
        return int(value)


    def ask_gpx_sync_exif_write_approval(self, photo_timezone, max_gap_minutes, gpx_files):
        """
        Fenêtre d'approbation au moment de la synchronisation GPX.

        Retourne :
        - True  : synchroniser + écrire dans les JPG
        - False : synchroniser seulement le catalogue
        - None  : annuler
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("Synchronisation GPX")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        result = {"value": None}

        frame = tk.Frame(dialog, padx=18, pady=16)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text="Synchronisation GPS des photos",
            font=("Arial", 13, "bold")
        ).pack(anchor="w", pady=(0, 8))

        if len(gpx_files) == 1:
            traces_text = os.path.basename(gpx_files[0])
        else:
            traces_text = "\n".join(f"• {os.path.basename(path)}" for path in gpx_files)

        info = (
            f"Nombre de traces GPX : {len(gpx_files)}\n"
            f"Fuseau horaire des photos : {photo_timezone}\n"
            f"Tolérance hors plage : {max_gap_minutes} minute(s)\n\n"
            f"{traces_text}\n\n"
            "Le logiciel va d'abord écrire les coordonnées dans le catalogue JSON.\n"
            "Tu peux aussi approuver l'écriture immédiate dans les métadonnées EXIF "
            "des JPG utilisés par le logiciel."
        )

        tk.Label(
            frame,
            text=info,
            justify="left",
            wraplength=620
        ).pack(anchor="w", pady=(0, 12))

        warning = (
            "Important :\n"
            "• les JPG originaux seront modifiés directement ;\n"
            "• les HEIC/HEIF originaux ne seront pas modifiés ;\n"
            "• pour les HEIC/HEIF, l'écriture se fait dans les JPG convertis d'Export_JPG ;\n"
            "• le catalogue reste la source de contrôle."
        )

        tk.Label(
            frame,
            text=warning,
            justify="left",
            fg="#8a5a00",
            wraplength=620
        ).pack(anchor="w", pady=(0, 14))

        buttons = tk.Frame(frame)
        buttons.pack(fill="x", pady=(6, 0))

        def choose(value):
            result["value"] = value
            dialog.destroy()

        tk.Button(
            buttons,
            text="✅ Synchroniser + écrire dans les JPG",
            command=lambda: choose(True),
            bg="#27ae60",
            fg="white",
            width=34
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            buttons,
            text="📝 Synchroniser le catalogue seulement",
            command=lambda: choose(False),
            width=34
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            buttons,
            text="Annuler",
            command=lambda: choose(None),
            width=12
        ).pack(side="right")

        dialog.protocol("WM_DELETE_WINDOW", lambda: choose(None))

        self.root.wait_window(dialog)
        return result["value"]

    def get_selected_gpx_files(self):
        files = []
        for path in getattr(self, "gpx_files", []) or []:
            if path and os.path.exists(path):
                files.append(path)

        if not files and self.gpx_file and os.path.exists(self.gpx_file):
            files.append(self.gpx_file)

        return files

    def load_gpx_time_segments(self, gpx_paths):
        """
        Charge plusieurs GPX en segments temporels indépendants.

        Important : on ne fusionne pas naïvement tous les points en une seule
        ligne de temps, car cela interpolerait à travers les trous entre deux
        traces séparées.
        """
        segments = []
        total_points = 0

        for gpx_path in gpx_paths:
            source_name = os.path.basename(gpx_path)

            with open(gpx_path, "r", encoding="utf-8") as f:
                gpx = gpxpy.parse(f)

            segment_counter = 0

            for track_index, track in enumerate(gpx.tracks, start=1):
                track_name = track.name or f"Track {track_index}"

                for segment_index, segment in enumerate(track.segments, start=1):
                    points = []

                    for point in segment.points:
                        if not point.time:
                            continue

                        point_time = point.time
                        if point_time.tzinfo is None:
                            point_time = point_time.replace(tzinfo=timezone.utc)
                        else:
                            point_time = point_time.astimezone(timezone.utc)

                        points.append((
                            point_time,
                            point.latitude,
                            point.longitude,
                            point.elevation
                        ))

                    points.sort(key=lambda x: x[0])

                    if not points:
                        continue

                    segment_counter += 1
                    total_points += len(points)
                    segments.append({
                        "source_path": gpx_path,
                        "source_name": source_name,
                        "track_name": track_name,
                        "segment_index": segment_index,
                        "segment_label": f"{source_name} · {track_name} · segment {segment_index}",
                        "points": points,
                        "start": points[0][0],
                        "end": points[-1][0],
                        "points_count": len(points)
                    })

            # Certains GPX peuvent contenir des routes horodatées au lieu de tracks.
            for route_index, route in enumerate(getattr(gpx, "routes", []), start=1):
                points = []
                route_name = route.name or f"Route {route_index}"

                for point in route.points:
                    if not getattr(point, "time", None):
                        continue

                    point_time = point.time
                    if point_time.tzinfo is None:
                        point_time = point_time.replace(tzinfo=timezone.utc)
                    else:
                        point_time = point_time.astimezone(timezone.utc)

                    points.append((
                        point_time,
                        point.latitude,
                        point.longitude,
                        point.elevation
                    ))

                points.sort(key=lambda x: x[0])

                if points:
                    total_points += len(points)
                    segments.append({
                        "source_path": gpx_path,
                        "source_name": source_name,
                        "track_name": route_name,
                        "segment_index": route_index,
                        "segment_label": f"{source_name} · {route_name}",
                        "points": points,
                        "start": points[0][0],
                        "end": points[-1][0],
                        "points_count": len(points)
                    })

            if segment_counter == 0:
                self.log(f"⚠️ Aucun segment track horodaté trouvé dans : {source_name}")

        segments.sort(key=lambda seg: (seg["start"], seg["end"], seg["source_name"]))
        return segments, total_points

    def interpolate_within_points(self, points, photo_time):
        """
        Interpolation stricte à l'intérieur d'un segment.
        Ne retourne rien si photo_time est hors de l'intervalle du segment.
        """
        if not points:
            return None

        if len(points) == 1:
            t, lat, lon, ele = points[0]
            if photo_time == t:
                return lat, lon, ele
            return None

        if photo_time < points[0][0] or photo_time > points[-1][0]:
            return None

        for i in range(len(points) - 1):
            t1, lat1, lon1, ele1 = points[i]
            t2, lat2, lon2, ele2 = points[i + 1]

            if t1 <= photo_time <= t2:
                total_diff = (t2 - t1).total_seconds()

                if total_diff == 0:
                    return lat1, lon1, ele1

                ratio = (photo_time - t1).total_seconds() / total_diff

                lat = lat1 + (lat2 - lat1) * ratio
                lon = lon1 + (lon2 - lon1) * ratio

                if ele1 is not None and ele2 is not None:
                    ele = ele1 + (ele2 - ele1) * ratio
                else:
                    ele = ele1 if ele1 is not None else ele2

                return lat, lon, ele

        return None

    def nearest_endpoint_for_segments(self, segments, photo_time):
        best = None

        for segment in segments:
            for point_index in (0, -1):
                t, lat, lon, ele = segment["points"][point_index]
                gap_seconds = abs((photo_time - t).total_seconds())

                if best is None or gap_seconds < best["gap_seconds"]:
                    best = {
                        "segment": segment,
                        "lat": lat,
                        "lon": lon,
                        "ele": ele,
                        "gap_seconds": gap_seconds,
                        "point_time": t
                    }

        return best

    def get_gps_from_time_segments(self, segments, photo_time, max_gap_seconds):
        """
        Choisit la position GPS pertinente pour une photo.

        Priorité :
        1. segment dont l'intervalle temporel contient la photo ;
        2. si autorisé, endpoint le plus proche dans la tolérance ;
        3. sinon hors plage.
        """
        candidates = []

        for segment in segments:
            if segment["start"] <= photo_time <= segment["end"]:
                gps = self.interpolate_within_points(segment["points"], photo_time)

                if gps:
                    candidates.append((segment, gps))

        if candidates:
            candidates.sort(key=lambda item: (
                (item[0]["end"] - item[0]["start"]).total_seconds(),
                item[0]["source_name"]
            ))
            segment, gps = candidates[0]
            match_type = "INTERPOLATED"
            if len(candidates) > 1:
                match_type = "AMBIGUOUS_OVERLAP"

            lat, lon, ele = gps
            return {
                "ok": True,
                "lat": lat,
                "lon": lon,
                "ele": ele,
                "segment": segment,
                "match_type": match_type,
                "gap_seconds": 0,
                "ambiguous_count": len(candidates)
            }

        nearest = self.nearest_endpoint_for_segments(segments, photo_time)

        if nearest and max_gap_seconds is not None and max_gap_seconds > 0:
            if nearest["gap_seconds"] <= max_gap_seconds:
                return {
                    "ok": True,
                    "lat": nearest["lat"],
                    "lon": nearest["lon"],
                    "ele": nearest["ele"],
                    "segment": nearest["segment"],
                    "match_type": "NEAREST_ENDPOINT",
                    "gap_seconds": nearest["gap_seconds"],
                    "ambiguous_count": 0
                }

        return {
            "ok": False,
            "nearest_gap_seconds": nearest["gap_seconds"] if nearest else None,
            "nearest_segment": nearest["segment"] if nearest else None
        }

    def format_gpx_time_range(self, segment):
        return (
            f"{segment['start'].isoformat()} → {segment['end'].isoformat()} "
            f"({segment['points_count']} points)"
        )

    def run_sync(self):
        gpx_files = self.get_selected_gpx_files()

        if not gpx_files:
            messagebox.showerror("Erreur", "Aucune trace GPX sélectionnée.")
            return

        if not os.path.exists(self.catalog_path):
            messagebox.showerror(
                "Erreur",
                "Aucun catalogue.json trouvé. Lancez d'abord la création du catalogue."
            )
            return

        tz_name = self.ask_photo_timezone()
        if not tz_name:
            return

        max_gap_minutes = self.ask_gpx_sync_gap_tolerance()
        if max_gap_minutes is None:
            return

        approval = self.ask_gpx_sync_exif_write_approval(
            tz_name,
            max_gap_minutes,
            gpx_files
        )

        if approval is None:
            return

        self.photo_timezone_name = tz_name
        self.write_exif_after_gpx_sync = bool(approval)

        self.progress["value"] = 0
        self.log(f"🕒 Fuseau horaire des photos : {self.photo_timezone_name}")
        self.log(f"🛰️ Traces GPX sélectionnées : {len(gpx_files)}")
        self.log(f"⏱️ Tolérance hors plage : {max_gap_minutes} minute(s)")

        if self.write_exif_after_gpx_sync:
            self.log("✍️ Écriture EXIF JPG approuvée : elle sera faite automatiquement après la synchronisation.")
        else:
            self.log("📝 Écriture EXIF JPG non demandée : seul le catalogue sera synchronisé.")

        self.log("📍 [MODULE 2] Démarrage de la synchronisation GPS multi-traces...")
        self.sync_gps()

    def get_photo_time(self, image_path):
        try:
            img = Image.open(image_path)
            info = img._getexif()

            if info:
                date_str = info.get(36867)

                if date_str:
                    dt_local_naive = datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S")
                    photo_tz = ZoneInfo(self.photo_timezone_name)
                    dt_local = dt_local_naive.replace(tzinfo=photo_tz)
                    return dt_local.astimezone(timezone.utc)

            mtime = os.path.getmtime(image_path)
            dt_local = datetime.fromtimestamp(mtime, tz=ZoneInfo(self.photo_timezone_name))
            return dt_local.astimezone(timezone.utc)

        except Exception as e:
            self.log(f"⚠️ Impossible de lire l'heure de {os.path.basename(image_path)} : {e}")
            return None

    def load_gpx_points(self, gpx_path):
        points = []

        with open(gpx_path, "r", encoding="utf-8") as f:
            gpx = gpxpy.parse(f)

        for track in gpx.tracks:
            for segment in track.segments:
                for point in segment.points:
                    if point.time:
                        point_time = point.time

                        if point_time.tzinfo is None:
                            point_time = point_time.replace(tzinfo=timezone.utc)
                        else:
                            point_time = point_time.astimezone(timezone.utc)

                        points.append((
                            point_time,
                            point.latitude,
                            point.longitude,
                            point.elevation
                        ))

        points.sort(key=lambda x: x[0])
        return points

    def get_gps_from_points(self, points, photo_time):
        if not points:
            return None

        if photo_time < points[0][0]:
            return points[0][1], points[0][2], points[0][3]

        if photo_time > points[-1][0]:
            return points[-1][1], points[-1][2], points[-1][3]

        for i in range(len(points) - 1):
            t1, lat1, lon1, ele1 = points[i]
            t2, lat2, lon2, ele2 = points[i + 1]

            if t1 <= photo_time <= t2:
                total_diff = (t2 - t1).total_seconds()

                if total_diff == 0:
                    return lat1, lon1, ele1

                ratio = (photo_time - t1).total_seconds() / total_diff

                lat = lat1 + (lat2 - lat1) * ratio
                lon = lon1 + (lon2 - lon1) * ratio

                if ele1 is not None and ele2 is not None:
                    ele = ele1 + (ele2 - ele1) * ratio
                else:
                    ele = ele1 if ele1 is not None else ele2

                return lat, lon, ele

        return None

    def sync_gps(self):
        try:
            self.catalog_data = self.read_catalog()
            gpx_files = self.get_selected_gpx_files()

            segments, total_points = self.load_gpx_time_segments(gpx_files)

            if not segments:
                messagebox.showerror(
                    "Erreur",
                    "Les traces GPX sélectionnées ne contiennent aucun segment horodaté."
                )
                return

            max_gap_seconds = getattr(self, "gpx_sync_max_gap_minutes", 30) * 60

            self.log("-" * 40)
            self.log("🛰️ Segments GPX horodatés chargés :")
            for segment in segments:
                self.log(f"   • {segment['segment_label']} · {self.format_gpx_time_range(segment)}")
            self.log(f"📍 Total : {len(segments)} segment(s), {total_points} point(s) horodaté(s)")

            total = len(self.catalog_data)
            synced_count = 0
            interpolated_count = 0
            nearest_count = 0
            ambiguous_count = 0
            out_of_range_count = 0
            error_time_count = 0
            missing_file_count = 0
            per_source_count = {}

            for i, entry in enumerate(self.catalog_data):
                if entry.get("status") != "OK":
                    continue

                img_path = self.get_entry_image_path(entry)
                filename = os.path.basename(img_path)

                if not os.path.exists(img_path):
                    entry["gps_sync"] = "FICHIER_INTROUVABLE"
                    missing_file_count += 1
                    self.log(f"⚠️ Fichier introuvable : {filename}")
                    continue

                photo_time = self.get_photo_time(img_path)

                if not photo_time:
                    entry["gps_sync"] = "ERREUR_HEURE"
                    error_time_count += 1
                    continue

                result = self.get_gps_from_time_segments(segments, photo_time, max_gap_seconds)

                if result.get("ok"):
                    lat = result["lat"]
                    lon = result["lon"]
                    ele = result["ele"]
                    segment = result["segment"]
                    match_type = result["match_type"]
                    gap_seconds = result.get("gap_seconds", 0)

                    entry["gps_coordinates"] = {
                        "lat": lat,
                        "lon": lon,
                        "ele": ele
                    }
                    entry["date_taken"] = photo_time.isoformat()
                    entry["gps_sync"] = "OK_GPX_MULTI"
                    entry["gps_source"] = "GPX_MULTI"
                    entry["gps_source_file"] = segment["source_name"]
                    entry["gps_source_segment"] = segment["segment_label"]
                    entry["gps_sync_detail"] = match_type
                    entry["gps_match_seconds"] = round(gap_seconds, 1)

                    synced_count += 1
                    per_source_count[segment["source_name"]] = per_source_count.get(segment["source_name"], 0) + 1

                    if match_type == "INTERPOLATED":
                        interpolated_count += 1
                    elif match_type == "NEAREST_ENDPOINT":
                        nearest_count += 1
                    elif match_type == "AMBIGUOUS_OVERLAP":
                        ambiguous_count += 1

                    if ele is not None:
                        self.log(
                            f"✅ {filename} -> {lat:.5f}, {lon:.5f}, {ele:.1f} m "
                            f"· {segment['source_name']} · {match_type}"
                        )
                    else:
                        self.log(
                            f"✅ {filename} -> {lat:.5f}, {lon:.5f} "
                            f"· {segment['source_name']} · {match_type}"
                        )

                    if match_type == "NEAREST_ENDPOINT":
                        self.log(f"   ⏱️ Point le plus proche utilisé à {gap_seconds/60:.1f} min")

                    if match_type == "AMBIGUOUS_OVERLAP":
                        self.log(
                            f"   ⚠️ Chevauchement temporel : "
                            f"{result.get('ambiguous_count', 0)} segments possibles, premier retenu."
                        )

                else:
                    nearest_gap = result.get("nearest_gap_seconds")
                    entry["gps_sync"] = "HORS_PLAGE_GPX_MULTI"
                    entry["gps_source"] = "GPX_MULTI"
                    entry["gps_sync_detail"] = "OUT_OF_RANGE"
                    entry["gps_match_seconds"] = round(nearest_gap, 1) if nearest_gap is not None else None
                    out_of_range_count += 1

                    if nearest_gap is not None:
                        self.log(f"⚠️ Hors plage GPX : {filename} · plus proche à {nearest_gap/60:.1f} min")
                    else:
                        self.log(f"⚠️ Hors plage GPX : {filename}")

                self.progress["value"] = ((i + 1) / total) * 100
                self.root.update_idletasks()

            if not isinstance(self.catalog_container, dict):
                self.catalog_container = self.read_catalog_container()

            self.catalog_container["last_photo_gpx_sync"] = {
                "synced_at": datetime.now().isoformat(timespec="seconds"),
                "mode": "multi_gpx",
                "photo_timezone": self.photo_timezone_name,
                "max_gap_minutes": getattr(self, "gpx_sync_max_gap_minutes", 30),
                "gpx_files": [os.path.basename(path) for path in gpx_files],
                "segments_count": len(segments),
                "points_count": total_points,
                "synced_count": synced_count,
                "interpolated_count": interpolated_count,
                "nearest_count": nearest_count,
                "ambiguous_count": ambiguous_count,
                "out_of_range_count": out_of_range_count,
                "error_time_count": error_time_count,
                "missing_file_count": missing_file_count,
                "per_source_count": per_source_count
            }

            self.save_catalog()

            self.log("-" * 40)
            self.log("✅ MODULE 2 TERMINÉ — synchronisation multi-GPX")
            self.log(f"📍 Photos géolocalisées : {synced_count}")
            self.log(f"   • Interpolées dans une trace : {interpolated_count}")
            self.log(f"   • Point le plus proche hors plage : {nearest_count}")
            self.log(f"   • Ambiguës / chevauchement : {ambiguous_count}")
            self.log(f"⚪ Hors plage : {out_of_range_count}")
            self.log(f"⚠️ Erreurs d'heure : {error_time_count}")
            self.log(f"⚠️ Fichiers introuvables : {missing_file_count}")

            if per_source_count:
                self.log("📊 Répartition par trace :")
                for source_name, count in sorted(per_source_count.items()):
                    self.log(f"   • {source_name} : {count} photo(s)")

            messagebox.showinfo(
                "Synchronisation terminée",
                (
                    "Synchronisation GPS multi-traces terminée !\n\n"
                    f"Photos géolocalisées : {synced_count}\n"
                    f"Interpolées dans une trace : {interpolated_count}\n"
                    f"Point le plus proche hors plage : {nearest_count}\n"
                    f"Hors plage : {out_of_range_count}\n"
                    f"Erreurs d'heure : {error_time_count}\n"
                    f"Fichiers introuvables : {missing_file_count}"
                )
            )

            self.offer_write_gps_to_jpg_after_sync(synced_count)

            self.load_folder(self.base_folder)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur GPS multi-traces : {e}")

    # ============================================================
    # MODULE 3 : CARTE + VISIONNEUSE + ÉDITION
    # ============================================================

    def load_geolocated_photos(self):
        if not os.path.exists(self.catalog_path):
            return []

        try:
            catalog = self.read_catalog()
        except Exception as e:
            self.log(f"❌ Impossible de lire le catalogue : {e}")
            return []

        self.catalog_data = catalog
        photos = []

        for idx, entry in enumerate(catalog):
            if entry.get("status") != "OK":
                continue

            coords = entry.get("gps_coordinates")
            if not coords:
                continue

            lat = coords.get("lat")
            lon = coords.get("lon")
            ele = coords.get("ele")

            if lat is None or lon is None:
                continue

            image_path = self.get_entry_image_path(entry)
            filename = os.path.basename(image_path)

            jpg_meta = self.read_text_metadata_from_jpg(image_path)

            catalog_title = entry.get("title", "") or ""
            catalog_description = entry.get("description", "") or ""

            jpg_title = jpg_meta.get("title", "") or ""
            jpg_description = jpg_meta.get("description", "") or ""

            display_title = jpg_title if jpg_title else catalog_title
            display_description = jpg_description if jpg_description else catalog_description

            metadata_source = "JPG" if (jpg_title or jpg_description) else "catalogue.json"

            sort_datetime = self.get_capture_datetime_for_sort(image_path, entry)

            photos.append({
                "catalog_index": idx,
                "filename": filename,
                "original_filename": entry.get("original_filename", ""),
                "title": display_title,
                "description": display_description,
                "metadata_source": metadata_source,
                "jpg_title": jpg_title,
                "jpg_description": jpg_description,
                "catalog_title": catalog_title,
                "catalog_description": catalog_description,
                "date_taken": entry.get("date_taken", ""),
                "gps_source": entry.get("gps_source", ""),
                "lat": float(lat),
                "lon": float(lon),
                "ele": ele,
                "image_path": image_path,
                "sort_datetime": sort_datetime,
                "platform_selected": bool(entry.get("platform_selected", False)),
                "platform_order": int(entry.get("platform_order") or 0)
            })

        # Ordre de navigation, de numérotation des marqueurs et de tri dans la carte :
        # date de prise de vue EXIF, avec fallback défini par get_capture_datetime_for_sort.
        photos.sort(
            key=lambda photo: (
                photo.get("sort_datetime", datetime.max),
                photo.get("filename", "").lower()
            )
        )

        return photos

    def show_empty_photo_workshop(self):
        self.clear_main_frame()
        self.status_header.config(text="Atelier Photos · état vide / à compléter", fg="#2980b9")

        panel = tk.Frame(self.main_frame, padx=24, pady=24)
        panel.pack(fill="both", expand=True)

        tk.Button(
            panel,
            text="↩️ Tableau de bord",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(anchor="w", pady=(0, 20))

        tk.Label(
            panel,
            text="Atelier Photos",
            font=("Arial", 18, "bold")
        ).pack(anchor="w", pady=(0, 8))

        has_catalog = os.path.exists(self.catalog_path)
        has_images = self.folder_has_images(self.photos_folder)

        if not has_catalog:
            message = (
                "Aucun catalogue photo n’existe encore pour ce dossier.\n"
                "Vous pouvez déjà travailler dans l’atelier GPX, ou créer le catalogue photo quand les images sont prêtes."
            )
        elif not has_images:
            message = (
                "Le catalogue existe, mais aucun dossier photo actif contenant des images n’a été trouvé.\n"
                "Choisissez un dossier photos ou ajoutez des images au dossier du bisse."
            )
        else:
            message = (
                "Des photos existent, mais aucune photo géolocalisée n’est actuellement disponible pour la carte.\n"
                "Vous pouvez lire les coordonnées GPS déjà inscrites dans les JPG ou synchroniser les photos avec une ou plusieurs traces GPX."
            )

        tk.Label(
            panel,
            text=message,
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=900
        ).pack(anchor="w", fill="x", pady=(0, 18))

        actions = tk.Frame(panel)
        actions.pack(anchor="w", fill="x")

        tk.Button(
            actions,
            text="📂 Choisir manuellement le dossier photos",
            command=self.choose_manual_photos_folder
        ).pack(fill="x", pady=4)

        if has_images:
            tk.Button(
                actions,
                text="🚀 Créer / recréer le catalogue photo",
                command=self.run_conversion,
                bg="#2c3e50",
                fg="white"
            ).pack(fill="x", pady=4)

        if has_catalog:
            tk.Button(
                actions,
                text="📷 Lire GPS des JPG",
                command=self.import_gps_from_existing_metadata,
                bg="#16a085",
                fg="white"
            ).pack(fill="x", pady=4)

            tk.Button(
                actions,
                text="🛰️ Géolocaliser avec une ou plusieurs traces GPX",
                command=self.show_gpx_sync_selector,
                bg="#27ae60",
                fg="white"
            ).pack(fill="x", pady=4)

            tk.Button(
                actions,
                text="✍️ Écrire GPS du catalogue dans les JPG",
                command=self.write_catalog_gps_to_jpg_metadata,
                bg="#2980b9",
                fg="white"
            ).pack(fill="x", pady=4)

            tk.Button(
                actions,
                text="🔤 Renommer les photos après tri",
                command=self.show_rename_interface,
                bg="#8e44ad",
                fg="white"
            ).pack(fill="x", pady=4)

        tk.Button(
            actions,
            text="🧭 Ouvrir l’atelier GPX",
            command=self.show_gpx_workshop,
            bg="#d35400",
            fg="white"
        ).pack(fill="x", pady=(14, 4))


    def show_photo_preparation_dialog(self):
        """
        Regroupe les fonctions de préparation photos hors du tableau de bord.

        Cette fenêtre accueille les commandes techniques liées aux photos :
        choix du dossier, catalogue, GPS, géolocalisation et renommage.
        La visionneuse reste ainsi centrée sur l'édition des métadonnées.
        """
        if not self.base_folder:
            messagebox.showwarning("Aucun dossier", "Ouvrez d'abord un dossier bisse.")
            return

        has_images = self.folder_has_images(self.photos_folder)
        has_catalog = os.path.exists(self.catalog_path)
        is_geolocated = False

        if has_catalog:
            try:
                catalog = self.read_catalog()
                is_geolocated = any(str(entry.get("gps_sync", "")).startswith("OK") for entry in catalog)
            except Exception:
                is_geolocated = False

        window = tk.Toplevel(self.root)
        window.title("Préparation photos")
        window.geometry("520x430")
        window.minsize(460, 360)
        window.transient(self.root)

        frame = tk.Frame(window, padx=14, pady=12)
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(0, weight=1)

        tk.Label(
            frame,
            text="Préparation photos",
            font=("Arial", 15, "bold"),
            anchor="w"
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

        status_lines = [
            f"Dossier photos actif : {self.photos_folder or '—'}",
            f"Images détectées : {'oui' if has_images else 'non'}",
            f"Catalogue : {'oui' if has_catalog else 'non'}",
            f"Géolocalisation catalogue : {'oui' if is_geolocated else 'non'}"
        ]

        tk.Label(
            frame,
            text="\n".join(status_lines),
            justify="left",
            anchor="w",
            fg="#555555",
            wraplength=470
        ).grid(row=1, column=0, sticky="ew", pady=(0, 12))

        actions = tk.LabelFrame(frame, text="Actions", padx=10, pady=8)
        actions.grid(row=2, column=0, sticky="ew")
        actions.grid_columnconfigure(0, weight=1)

        def close_then(callback):
            def wrapped():
                try:
                    window.destroy()
                except Exception:
                    pass
                callback()
            return wrapped

        tk.Button(
            actions,
            text="📂 Choisir le dossier photos",
            command=close_then(self.choose_manual_photos_folder)
        ).grid(row=0, column=0, sticky="ew", pady=3)

        tk.Button(
            actions,
            text="🚀 Créer / recréer le catalogue photo",
            command=close_then(self.run_conversion),
            bg="#2c3e50",
            fg="white"
        ).grid(row=1, column=0, sticky="ew", pady=3)

        gps_state = "normal" if has_catalog else "disabled"
        tk.Button(
            actions,
            text="📷 Lire GPS des JPG",
            command=close_then(self.import_gps_from_existing_metadata),
            bg="#16a085",
            fg="white",
            state=gps_state
        ).grid(row=2, column=0, sticky="ew", pady=3)

        tk.Button(
            actions,
            text="🛰️ Géolocaliser / resynchroniser avec une ou plusieurs traces GPX",
            command=close_then(self.show_resync_interface if is_geolocated else self.show_gpx_sync_selector),
            bg="#27ae60",
            fg="white",
            state=gps_state
        ).grid(row=3, column=0, sticky="ew", pady=3)

        tk.Button(
            actions,
            text="✍️ Écrire GPS du catalogue dans les JPG",
            command=close_then(self.write_catalog_gps_to_jpg_metadata),
            bg="#2980b9",
            fg="white",
            state=gps_state
        ).grid(row=4, column=0, sticky="ew", pady=3)

        tk.Button(
            actions,
            text="🔤 Renommer les photos après tri",
            command=close_then(self.show_rename_interface),
            bg="#8e44ad",
            fg="white",
            state=gps_state
        ).grid(row=5, column=0, sticky="ew", pady=3)

        tk.Label(
            frame,
            text=(
                "Cette fenêtre remplace les anciennes commandes photos placées provisoirement "
                "dans les options avancées du tableau de bord."
            ),
            justify="left",
            anchor="w",
            fg="#666666",
            wraplength=470
        ).grid(row=3, column=0, sticky="ew", pady=(12, 0))

        tk.Button(
            frame,
            text="Fermer",
            command=window.destroy
        ).grid(row=4, column=0, sticky="e", pady=(12, 0))

    def show_map_interface(self):
        photos = self.load_geolocated_photos()

        if not photos:
            self.show_empty_photo_workshop()
            return

        self.geolocated_photos = photos
        self.current_photo = None
        self.gpx_workshop_active = False

        self.clear_main_frame()
        self.status_header.config(text="Atelier Photos · carte · visionneuse · métadonnées", fg="#2980b9")

        toolbar = tk.Frame(self.main_frame)
        toolbar.pack(fill="x", pady=(0, 8))

        row_a = tk.Frame(toolbar)
        row_a.pack(fill="x", pady=(0, 4))

        tk.Button(
            row_a,
            text="↩️ Tableau de bord",
            command=lambda: self.load_folder(self.base_folder)
        ).pack(side="left", padx=(0, 4))

        tk.Button(
            row_a,
            text="📂 Dossier",
            command=self.select_base_folder
        ).pack(side="left", padx=4)

        tk.Button(
            row_a,
            text="🧭 Atelier GPX",
            command=self.show_gpx_workshop,
            bg="#d35400",
            fg="white"
        ).pack(side="left", padx=4)

        tk.Button(
            row_a,
            text="⚙️ Préparation photos",
            command=self.show_photo_preparation_dialog
        ).pack(side="left", padx=4)

        row_b = tk.Frame(toolbar)
        row_b.pack(fill="x")

        map_group = tk.LabelFrame(row_b, text="Fond de carte", padx=6, pady=3)
        map_group.pack(side="left", padx=(0, 8))

        tk.Button(
            map_group,
            text="Auto",
            command=lambda: self.set_swisstopo_layer("color_auto")
        ).pack(side="left", padx=2)

        tk.Button(
            map_group,
            text="10k",
            command=lambda: self.set_swisstopo_layer("color_10k")
        ).pack(side="left", padx=2)

        tk.Button(
            map_group,
            text="25k",
            command=lambda: self.set_swisstopo_layer("color_detail")
        ).pack(side="left", padx=2)

        tk.Button(
            map_group,
            text="Standard",
            command=lambda: self.set_swisstopo_layer("color")
        ).pack(side="left", padx=2)

        tk.Button(
            map_group,
            text="Grise",
            command=lambda: self.set_swisstopo_layer("grey")
        ).pack(side="left", padx=2)

        tk.Button(
            map_group,
            text="Aérienne",
            command=lambda: self.set_swisstopo_layer("image")
        ).pack(side="left", padx=2)

        nav_group = tk.LabelFrame(row_b, text="Navigation", padx=6, pady=3)
        nav_group.pack(side="left", padx=(0, 8))

        tk.Button(
            nav_group,
            text="◀",
            command=self.select_previous_photo,
            width=4
        ).pack(side="left", padx=2)

        tk.Label(
            nav_group,
            textvariable=self.photo_index_var,
            width=18,
            anchor="center"
        ).pack(side="left", padx=4)

        tk.Button(
            nav_group,
            text="▶",
            command=self.select_next_photo,
            width=4
        ).pack(side="left", padx=2)

        tk.Button(
            row_b,
            text="📍 Recentrer photo",
            command=self.center_on_current_photo
        ).pack(side="left", padx=4)

        tk.Button(
            row_b,
            text="🧭 Cadrer photos + tracés",
            command=self.fit_map_to_content
        ).pack(side="left", padx=4)

        tk.Button(
            row_b,
            text="ℹ️ Détails photo",
            command=self.show_current_photo_technical_details
        ).pack(side="left", padx=4)

        paned = tk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL, sashwidth=7)
        paned.pack(fill="both", expand=True)

        map_panel = tk.Frame(paned, bg="#eeeeee")
        viewer_panel = tk.Frame(paned, bg="#222222")
        meta_panel = tk.Frame(paned, bg="#f4f4f4", padx=12, pady=10)

        paned.add(map_panel, minsize=380, width=470)
        paned.add(viewer_panel, minsize=620, width=760)
        paned.add(meta_panel, minsize=310, width=340)

        self.build_map_panel(map_panel)
        self.build_viewer_panel(viewer_panel)
        self.build_metadata_panel(meta_panel)

        self.set_swisstopo_layer("color")
        self.draw_bisse_traces_on_map()
        # select_photo_on_map() rafraîchit déjà les repères photo.
        # Ne pas appeler refresh_map_markers() juste avant évite un double rendu
        # inutile à l'ouverture de la carte.
        # À l'ouverture, on remplit le panneau photo sans déclencher
        # un premier chargement de tuiles centré sur la photo 1.
        # Le cadrage global juste après suffit.
        self.select_photo_on_map(photos[0], center_map=False)
        self.fit_map_to_content()
        self.start_photo_layer_watch("photo")

        self.log(f"🗺️ Atelier Photos ouvert avec {len(photos)} photo(s).")

    def build_map_panel(self, parent):
        header = tk.Frame(parent, bg="#eeeeee")
        header.pack(fill="x", padx=8, pady=(6, 3))

        tk.Label(
            header,
            text="Carte",
            font=("Arial", 11, "bold"),
            bg="#eeeeee"
        ).pack(side="left")

        tk.Label(
            parent,
            text=(
                "Pastilles bleues = photos. Les photos proches sont regroupées ; "
                "un clic zoome, puis déploie virtuellement les positions encore superposées."
            ),
            font=("Arial", 9),
            fg="#555555",
            bg="#eeeeee",
            wraplength=470,
            justify="left"
        ).pack(anchor="w", padx=8, pady=(0, 4))

        trace_controls = tk.LabelFrame(parent, text="Affichage", bg="#eeeeee", padx=6, pady=5)
        trace_controls.pack(fill="x", padx=8, pady=(0, 6))

        tk.Checkbutton(
            trace_controls,
            text="Photos",
            variable=self.show_photos_on_map_var,
            command=self.refresh_map_markers,
            bg="#eeeeee"
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))

        tk.Checkbutton(
            trace_controls,
            text="Ciel ouvert",
            variable=self.show_trace_ciel_var,
            command=self.refresh_trace_display,
            bg="#eeeeee"
        ).grid(row=0, column=1, sticky="w", padx=(0, 8))

        tk.Checkbutton(
            trace_controls,
            text="Canalisé",
            variable=self.show_trace_canalise_var,
            command=self.refresh_trace_display,
            bg="#eeeeee"
        ).grid(row=0, column=2, sticky="w", padx=(0, 8))

        tk.Checkbutton(
            trace_controls,
            text="Abandonné",
            variable=self.show_trace_abandonne_var,
            command=self.refresh_trace_display,
            bg="#eeeeee"
        ).grid(row=1, column=0, sticky="w", padx=(0, 8))

        tk.Checkbutton(
            trace_controls,
            text="Trace topo",
            variable=self.show_trace_topo_var,
            command=self.refresh_trace_display,
            bg="#eeeeee"
        ).grid(row=1, column=1, sticky="w", padx=(0, 8))

        tk.Checkbutton(
            trace_controls,
            text="Inconnus / autres",
            variable=self.show_trace_inconnu_var,
            command=self.refresh_trace_display,
            bg="#eeeeee"
        ).grid(row=1, column=2, sticky="w", padx=(0, 8))

        self.map_widget = tkintermapview.TkinterMapView(
            parent,
            width=500,
            height=640,
            corner_radius=0
        )
        self.map_widget.pack(fill="both", expand=True, padx=6, pady=6)
        self.map_widget.add_left_click_map_command(
            lambda coords: self.handle_photo_map_background_click("photo", coords)
        )

    def build_viewer_panel(self, parent):
        top = tk.Frame(parent, bg="#222222")
        top.pack(fill="x", pady=4)

        tk.Button(top, text="➖", command=lambda: self.viewer_change_zoom(0.8), width=4).pack(side="left", padx=3)
        tk.Button(top, text="➕", command=lambda: self.viewer_change_zoom(1.25), width=4).pack(side="left", padx=3)
        tk.Button(top, text="Adapter", command=self.viewer_fit_to_panel).pack(side="left", padx=3)
        tk.Button(top, text="100 %", command=self.viewer_actual_size).pack(side="left", padx=3)

        tk.Label(
            top,
            textvariable=self.viewer_zoom_var,
            bg="#222222",
            fg="white"
        ).pack(side="left", padx=12)

        tk.Label(
            parent,
            textvariable=self.viewer_info_var,
            bg="#222222",
            fg="white",
            font=("Arial", 10, "bold"),
            wraplength=560,
            justify="left"
        ).pack(fill="x", padx=8, pady=(0, 4))

        canvas_frame = tk.Frame(parent, bg="#222222")
        canvas_frame.pack(fill="both", expand=True)

        self.viewer_canvas = tk.Canvas(canvas_frame, bg="#222222", highlightthickness=0)
        self.viewer_canvas.pack(side="left", fill="both", expand=True)

        v_scroll = tk.Scrollbar(canvas_frame, orient="vertical", command=self.viewer_canvas.yview)
        v_scroll.pack(side="right", fill="y")

        h_scroll = tk.Scrollbar(parent, orient="horizontal", command=self.viewer_canvas.xview)
        h_scroll.pack(fill="x")

        self.viewer_canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.viewer_canvas.bind("<MouseWheel>", self.viewer_mousewheel)
        self.viewer_canvas.bind("<ButtonPress-1>", self.viewer_start_pan)
        self.viewer_canvas.bind("<B1-Motion>", self.viewer_do_pan)
        self.viewer_canvas.bind("<Button-4>", lambda event: self.viewer_change_zoom(1.10))
        self.viewer_canvas.bind("<Button-5>", lambda event: self.viewer_change_zoom(0.90))

    def build_metadata_panel(self, parent):
        """
        Panneau d'édition photo compact, sans scroll interne.

        Les détails techniques sont consultables via la barre supérieure
        ("ℹ️ Détails photo"). Le panneau droit reste centré sur l'édition :
        titre, description, sélection plateforme, sauvegarde et écartement.
        """
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        tk.Label(
            parent,
            text="Métadonnées",
            font=("Arial", 14, "bold"),
            bg="#f4f4f4"
        ).grid(row=0, column=0, sticky="w", pady=(0, 4))

        self.photo_technical_summary_var = tk.StringVar(value="")
        tk.Label(
            parent,
            textvariable=self.photo_technical_summary_var,
            bg="#f4f4f4",
            fg="#555555",
            wraplength=290,
            justify="left"
        ).grid(row=1, column=0, sticky="ew", pady=(0, 8))

        edit = tk.Frame(parent, bg="#f4f4f4")
        edit.grid(row=2, column=0, sticky="nsew")
        edit.grid_columnconfigure(0, weight=1)
        edit.grid_rowconfigure(3, weight=1)

        tk.Label(
            edit,
            text="Titre",
            font=("Arial", 10, "bold"),
            bg="#f4f4f4"
        ).grid(row=0, column=0, sticky="w", pady=(0, 2))

        self.photo_title_entry = tk.Entry(edit, font=("Arial", 10))
        self.photo_title_entry.grid(row=1, column=0, sticky="ew", pady=(0, 8))

        tk.Label(
            edit,
            text="Description",
            font=("Arial", 10, "bold"),
            bg="#f4f4f4"
        ).grid(row=2, column=0, sticky="w", pady=(0, 2))

        self.photo_desc_text = tk.Text(edit, height=8, wrap="word", font=("Arial", 10))
        self.photo_desc_text.grid(row=3, column=0, sticky="nsew", pady=(0, 8))

        platform_frame = tk.LabelFrame(edit, text="Plateforme", bg="#f4f4f4", padx=6, pady=6)
        platform_frame.grid(row=4, column=0, sticky="ew", pady=(0, 8))
        platform_frame.grid_columnconfigure(1, weight=1)

        tk.Checkbutton(
            platform_frame,
            text="⭐ Choisie",
            variable=self.platform_selected_var,
            bg="#f4f4f4"
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        tk.Label(platform_frame, text="Ordre :", bg="#f4f4f4").grid(row=1, column=0, sticky="w", pady=(4, 0))
        tk.Spinbox(
            platform_frame,
            from_=0,
            to=999,
            textvariable=self.platform_order_var,
            width=6
        ).grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(4, 0))

        tk.Label(
            platform_frame,
            text="0 = non défini.",
            bg="#f4f4f4",
            fg="#555555",
            justify="left"
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(3, 0))

        buttons = tk.Frame(parent, bg="#f4f4f4")
        buttons.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        buttons.grid_columnconfigure(0, weight=1)

        tk.Button(
            buttons,
            text="💾 Sauvegarder titre / description / plateforme",
            command=self.save_current_photo_metadata,
            bg="#27ae60",
            fg="white",
            height=2
        ).grid(row=0, column=0, sticky="ew", pady=(0, 5))

        tk.Button(
            buttons,
            text="🗑️ Écarter cette photo du corpus",
            command=self.discard_current_photo,
            bg="#c0392b",
            fg="white"
        ).grid(row=1, column=0, sticky="ew", pady=3)

        tk.Label(
            buttons,
            textvariable=self.photo_status_var,
            bg="#f4f4f4",
            fg="#2c3e50",
            wraplength=300,
            justify="left"
        ).grid(row=2, column=0, sticky="ew", pady=(6, 0))

    def set_swisstopo_layer(self, layer_name):
        if not self.map_widget:
            return

        if layer_name == "color_auto":
            self.set_swisstopo_auto_mode("photo", True)
            return

        self.set_swisstopo_auto_mode("photo", False)
        self.set_tile_server_for_widget(self.map_widget, "photo", layer_name, force=True)

    def get_photo_marker_icon(self, number, selected=False, mode="normal"):
        """
        Pastille photo bleue. La photo active est simplement rouge, comme dans
        l'ancien comportement, sans halo cyan complexe.
        """
        cache_key = (str(number), bool(selected), str(mode), "v48")
        if cache_key in self.photo_marker_icon_cache:
            return self.photo_marker_icon_cache[cache_key]

        if mode == "discrete":
            size = 17 if selected else 14
            image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
            draw = ImageDraw.Draw(image)
            margin = 2
            fill = (220, 53, 69, 220) if selected else (45, 156, 219, 115)
            edge = (255, 255, 255, 240) if selected else (31, 97, 141, 155)
            draw.ellipse(
                (margin, margin, size - 1 - margin, size - 1 - margin),
                fill=fill,
                outline=edge,
                width=2 if selected else 1
            )
            icon = ImageTk.PhotoImage(image)
            self.photo_marker_icon_cache[cache_key] = icon
            return icon

        size = 42 if selected else 38
        image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)

        outer = 2
        inner = 4
        fill = "#e53935" if selected else "#2d9cdb"
        edge = "#8e1b17" if selected else "#1f618d"

        draw.ellipse(
            (outer, outer, size - 1 - outer, size - 1 - outer),
            fill="#ffffff",
            outline="#ffffff"
        )
        draw.ellipse(
            (inner, inner, size - 1 - inner, size - 1 - inner),
            fill=fill,
            outline=edge,
            width=2
        )

        label = str(number)
        try:
            font_size = 15 if len(label) <= 2 else 12
            font = ImageFont.truetype("arialbd.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", 15 if len(label) <= 2 else 12)
            except Exception:
                font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), label, font=font)
        text_x = (size - (bbox[2] - bbox[0])) / 2
        text_y = (size - (bbox[3] - bbox[1])) / 2 - 1
        draw.text((text_x, text_y), label, fill="#ffffff", font=font)

        icon = ImageTk.PhotoImage(image)
        self.photo_marker_icon_cache[cache_key] = icon
        return icon

    def get_photo_cluster_icon(self, count, selected=False):
        """
        Agrégat bleu rond avec trois points et compteur.
        L'agrégat contenant la photo active devient rouge.
        """
        cache_key = (int(count), bool(selected), "v48")
        if cache_key in self.photo_cluster_icon_cache:
            return self.photo_cluster_icon_cache[cache_key]

        size = 52 if selected else (48 if count < 100 else 52)
        image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)

        outer = 2
        inner = 4
        fill = "#d32f2f" if selected else "#176fa8"
        edge = "#7f1717" if selected else "#073f68"

        draw.ellipse(
            (outer, outer, size - 1 - outer, size - 1 - outer),
            fill="#ffffff",
            outline="#ffffff"
        )
        draw.ellipse(
            (inner, inner, size - 1 - inner, size - 1 - inner),
            fill=fill,
            outline=edge,
            width=2
        )

        dot_y = inner + 10
        dot_radius = 2
        center_x = size / 2
        for dx in (-8, 0, 8):
            draw.ellipse(
                (
                    center_x + dx - dot_radius,
                    dot_y - dot_radius,
                    center_x + dx + dot_radius,
                    dot_y + dot_radius,
                ),
                fill="#ffffff"
            )

        label = str(count)
        try:
            font = ImageFont.truetype("arialbd.ttf", 14 if len(label) <= 2 else 11)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", 14 if len(label) <= 2 else 11)
            except Exception:
                font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), label, font=font)
        text_x = (size - (bbox[2] - bbox[0])) / 2
        text_y = inner + 18
        draw.text((text_x, text_y), label, fill="#ffffff", font=font)

        icon = ImageTk.PhotoImage(image)
        self.photo_cluster_icon_cache[cache_key] = icon
        return icon

    def get_photo_anchor_icon(self):
        if "anchor" in self.photo_anchor_icon_cache:
            return self.photo_anchor_icon_cache["anchor"]

        size = 12
        image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)
        draw.ellipse((1, 1, size - 2, size - 2), fill="#0b4f8a", outline="#ffffff", width=2)
        icon = ImageTk.PhotoImage(image)
        self.photo_anchor_icon_cache["anchor"] = icon
        return icon

    def get_photo_map_widget(self, context):
        return self.map_widget if context == "photo" else self.gpx_editor_map

    def get_photo_layer_mode(self, context):
        if context == "photo":
            return "visible" if self.show_photos_on_map_var.get() else "hidden"
        return self.gpx_photo_display_mode_var.get()

    def get_photo_layer_photos(self, context):
        if context == "photo":
            return self.geolocated_photos
        return self.gpx_workshop_photos

    def get_selected_photo_catalog_index(self, context):
        if context == "photo" and self.current_photo:
            return self.current_photo.get("catalog_index")
        if context == "gpx" and self.gpx_photo_viewer_current_photo:
            return self.gpx_photo_viewer_current_photo.get("catalog_index")
        return None

    def latlon_to_world_pixel(self, lat, lon, zoom):
        latitude = max(-85.05112878, min(85.05112878, float(lat)))
        longitude = float(lon)
        world = 256.0 * (2.0 ** float(zoom))
        x = (longitude + 180.0) / 360.0 * world
        sin_lat = math.sin(math.radians(latitude))
        y = (0.5 - math.log((1 + sin_lat) / (1 - sin_lat)) / (4 * math.pi)) * world
        return x, y

    def world_pixel_to_latlon(self, x, y, zoom):
        world = 256.0 * (2.0 ** float(zoom))
        lon = float(x) / world * 360.0 - 180.0
        mercator = math.pi - 2.0 * math.pi * float(y) / world
        lat = math.degrees(math.atan(math.sinh(mercator)))
        return lat, lon

    def build_photo_clusters(self, photos, zoom, threshold_px=None):
        threshold = float(threshold_px or self.photo_cluster_threshold_px)
        clusters = []

        for number, photo in enumerate(photos, start=1):
            x, y = self.latlon_to_world_pixel(photo["lat"], photo["lon"], zoom)
            item = {
                "photo": photo,
                "number": number,
                "x": x,
                "y": y
            }

            best = None
            best_distance = None
            for cluster in clusters:
                distance = math.hypot(x - cluster["center_x"], y - cluster["center_y"])
                if distance <= threshold and (best_distance is None or distance < best_distance):
                    best = cluster
                    best_distance = distance

            if best is None:
                clusters.append({
                    "items": [item],
                    "sum_x": x,
                    "sum_y": y,
                    "center_x": x,
                    "center_y": y,
                    "min_x": x,
                    "max_x": x,
                    "min_y": y,
                    "max_y": y,
                })
            else:
                best["items"].append(item)
                best["sum_x"] += x
                best["sum_y"] += y
                count = len(best["items"])
                best["center_x"] = best["sum_x"] / count
                best["center_y"] = best["sum_y"] / count
                best["min_x"] = min(best["min_x"], x)
                best["max_x"] = max(best["max_x"], x)
                best["min_y"] = min(best["min_y"], y)
                best["max_y"] = max(best["max_y"], y)

        for cluster in clusters:
            lat, lon = self.world_pixel_to_latlon(
                cluster["center_x"],
                cluster["center_y"],
                zoom
            )
            cluster["lat"] = lat
            cluster["lon"] = lon
            cluster["span_px"] = max(
                cluster["max_x"] - cluster["min_x"],
                cluster["max_y"] - cluster["min_y"]
            )
            cluster["photo_ids"] = tuple(
                item["photo"].get("catalog_index")
                for item in cluster["items"]
            )
            cluster["min_number"] = min(item["number"] for item in cluster["items"])

        return sorted(clusters, key=lambda c: c["min_number"])

    def clear_photo_layer_objects(self, context):
        marker_list = self.map_markers if context == "photo" else self.gpx_editor_photo_markers
        for marker in marker_list:
            try:
                marker.delete()
            except Exception:
                pass
        marker_list.clear()

        for path in self.photo_spider_paths.get(context, []):
            try:
                path.delete()
            except Exception:
                pass
        self.photo_spider_paths[context] = []

    def set_context_photo_marker(self, context, lat, lon, icon, command=None):
        widget = self.get_photo_map_widget(context)
        if not widget:
            return None

        kwargs = {
            "icon": icon,
            "icon_anchor": "center"
        }
        if command is not None:
            kwargs["command"] = command

        try:
            marker = widget.set_marker(lat, lon, text="", **kwargs)
        except TypeError:
            marker = widget.set_marker(lat, lon, text="", command=command)

        if context == "photo":
            self.map_markers.append(marker)
        else:
            self.gpx_editor_photo_markers.append(marker)
        return marker

    def mark_photo_marker_interaction(self, context):
        """
        Empêche le clic de fond de carte, émis juste après un clic marqueur,
        de refermer le déploiement pendant le relâchement de la souris.
        """
        self.photo_map_click_guard_until[context] = time.monotonic() + 0.65

    def consume_photo_map_click_guard(self, context):
        deadline = float(self.photo_map_click_guard_until.get(context, 0.0) or 0.0)
        self.photo_map_click_guard_until[context] = 0.0
        return time.monotonic() <= deadline

    def handle_photo_map_background_click(self, context, _coords=None):
        if self.consume_photo_map_click_guard(context):
            return
        if self.photo_spider_state.get(context) is not None:
            self.close_photo_spider(context, redraw=True)

    def close_photo_spider_from_anchor(self, context):
        self.mark_photo_marker_interaction(context)
        self.close_photo_spider(context, redraw=True)

    def open_photo_from_context(self, context, photo):
        self.mark_photo_marker_interaction(context)
        if context == "photo":
            self.select_photo_on_map(photo)
        else:
            self.show_gpx_photo_quick_view(photo)

    def render_normal_photo_clusters(self, context, clusters):
        selected_id = self.get_selected_photo_catalog_index(context)
        ordered = sorted(
            clusters,
            key=lambda c: selected_id in c["photo_ids"]
        )

        for cluster in ordered:
            items = cluster["items"]
            selected = selected_id in cluster["photo_ids"]

            if len(items) == 1:
                item = items[0]
                icon = self.get_photo_marker_icon(item["number"], selected=selected)
                self.set_context_photo_marker(
                    context,
                    item["photo"]["lat"],
                    item["photo"]["lon"],
                    icon,
                    command=lambda _marker, p=item["photo"], ctx=context: self.open_photo_from_context(ctx, p)
                )
            else:
                icon = self.get_photo_cluster_icon(len(items), selected=selected)
                self.set_context_photo_marker(
                    context,
                    cluster["lat"],
                    cluster["lon"],
                    icon,
                    command=lambda _marker, c=cluster, ctx=context: self.on_photo_cluster_click(ctx, c)
                )

    def render_discrete_photos(self, context, photos):
        selected_id = self.get_selected_photo_catalog_index(context)
        for number, photo in enumerate(photos, start=1):
            selected = photo.get("catalog_index") == selected_id
            icon = self.get_photo_marker_icon(number, selected=selected, mode="discrete")
            # Aucun callback : les petits repères restent informatifs et les
            # clics sur les segments demeurent prioritaires.
            self.set_context_photo_marker(
                context,
                photo["lat"],
                photo["lon"],
                icon,
                command=None
            )

    def refresh_photo_layer(self, context, force=False):
        widget = self.get_photo_map_widget(context)
        if not widget:
            return

        mode = self.get_photo_layer_mode(context)
        photos = self.get_photo_layer_photos(context) or []

        self.clear_photo_layer_objects(context)

        if mode == "hidden" or not photos:
            self.photo_layer_clusters[context] = []
            self.photo_spider_state[context] = None
            return

        zoom = self.get_map_zoom_value(widget)
        if zoom is None:
            zoom = 17

        if mode == "discrete":
            self.photo_spider_state[context] = None
            self.photo_layer_clusters[context] = []
            self.render_discrete_photos(context, photos)
            return

        clusters = self.build_photo_clusters(photos, zoom)
        self.photo_layer_clusters[context] = clusters

        if self.photo_spider_state.get(context):
            if self.render_photo_spider(context, clusters):
                return
            self.photo_spider_state[context] = None

        self.render_normal_photo_clusters(context, clusters)

    def cluster_max_distance_m(self, cluster):
        items = cluster.get("items", [])
        if len(items) <= 1:
            return 0.0

        center_lat = cluster["lat"]
        center_lon = cluster["lon"]
        return max(
            self.haversine_distance_m(
                center_lat,
                center_lon,
                item["photo"]["lat"],
                item["photo"]["lon"]
            )
            for item in items
        )

    def find_cluster_for_photo_ids(self, clusters, photo_ids):
        wanted = set(photo_ids)
        for cluster in clusters:
            if wanted.issubset(set(cluster.get("photo_ids", ()))):
                return cluster
        return None

    def on_photo_cluster_click(self, context, cluster):
        self.mark_photo_marker_interaction(context)

        widget = self.get_photo_map_widget(context)
        if not widget:
            return

        wanted_ids = tuple(cluster.get("photo_ids", ()))
        current_state = self.photo_spider_state.get(context) or {}
        if tuple(current_state.get("photo_ids", ())) == wanted_ids:
            self.close_photo_spider(context, redraw=True)
            return

        current_zoom = self.get_map_zoom_value(widget) or 17
        max_zoom = int(getattr(widget, "max_zoom", 26) or 26)
        max_distance = self.cluster_max_distance_m(cluster)

        if max_distance <= 1.5 or current_zoom >= max_zoom:
            self.open_photo_spider(context, cluster)
            return

        span = max(1.0, float(cluster.get("span_px", 1.0)))
        desired_span = self.photo_cluster_threshold_px * 1.8
        zoom_gain = max(1, int(math.ceil(math.log(desired_span / span, 2)))) if span < desired_span else 1
        target_zoom = min(max_zoom, current_zoom + zoom_gain)

        self.animate_photo_cluster_zoom(
            context,
            cluster,
            target_zoom,
            open_spider_if_still_grouped=(target_zoom >= max_zoom)
        )

    def animate_photo_cluster_zoom(self, context, cluster, target_zoom, open_spider_if_still_grouped=False):
        widget = self.get_photo_map_widget(context)
        if not widget:
            return

        self.photo_zoom_animation_token += 1
        token = self.photo_zoom_animation_token
        current_zoom = self.get_map_zoom_value(widget) or target_zoom
        center = (cluster["lat"], cluster["lon"])
        wanted_ids = tuple(cluster["photo_ids"])

        try:
            widget.set_position(center[0], center[1])
        except Exception:
            pass

        steps = list(range(current_zoom + 1, int(target_zoom) + 1))

        def advance(index=0):
            if token != self.photo_zoom_animation_token:
                return

            if index < len(steps):
                try:
                    widget.set_position(center[0], center[1])
                    widget.set_zoom(steps[index])
                except Exception:
                    return
                self.root.after(135, lambda: advance(index + 1))
                return

            self.photo_layer_last_signature[context] = None
            self.refresh_photo_layer(context, force=True)

            if open_spider_if_still_grouped:
                clusters = self.photo_layer_clusters.get(context, [])
                remaining = self.find_cluster_for_photo_ids(clusters, wanted_ids)
                if remaining and len(remaining.get("items", [])) > 1:
                    self.open_photo_spider(context, remaining)

        advance()

    def spider_offsets(self, count):
        offsets = []
        remaining = int(count)
        ring = 1
        marker_spacing = 42.0

        while remaining > 0:
            radius = 42.0 + (ring - 1) * 40.0
            capacity = max(6, int((2 * math.pi * radius) // marker_spacing))
            use = min(remaining, capacity)
            angle_offset = -math.pi / 2 + (ring % 2) * 0.16

            for i in range(use):
                angle = angle_offset + (2 * math.pi * i / use)
                offsets.append((math.cos(angle) * radius, math.sin(angle) * radius))

            remaining -= use
            ring += 1

        return offsets

    def open_photo_spider(self, context, cluster):
        self.photo_spider_state[context] = {
            "photo_ids": tuple(cluster.get("photo_ids", ())),
            "zoom": self.get_map_zoom_value(self.get_photo_map_widget(context)) or 17
        }
        self.refresh_photo_layer(context, force=True)

    def render_photo_spider(self, context, clusters):
        state = self.photo_spider_state.get(context)
        widget = self.get_photo_map_widget(context)
        if not state or not widget:
            return False

        cluster = self.find_cluster_for_photo_ids(clusters, state.get("photo_ids", ()))
        if not cluster or len(cluster.get("items", [])) <= 1:
            return False

        target_ids = set(cluster["photo_ids"])
        other_clusters = [
            candidate for candidate in clusters
            if set(candidate.get("photo_ids", ())) != target_ids
        ]
        self.render_normal_photo_clusters(context, other_clusters)

        zoom = self.get_map_zoom_value(widget) or state.get("zoom", 17)
        center_x, center_y = self.latlon_to_world_pixel(cluster["lat"], cluster["lon"], zoom)
        offsets = self.spider_offsets(len(cluster["items"]))
        selected_id = self.get_selected_photo_catalog_index(context)

        # Lignes d'attache d'abord, puis les marqueurs par-dessus.
        deployed = []
        for item, (offset_x, offset_y) in zip(cluster["items"], offsets):
            lat, lon = self.world_pixel_to_latlon(center_x + offset_x, center_y + offset_y, zoom)
            deployed.append((item, lat, lon))
            try:
                path = widget.set_path(
                    [(cluster["lat"], cluster["lon"]), (lat, lon)],
                    color="#86b7d8",
                    width=2
                )
                self.photo_spider_paths[context].append(path)
            except Exception:
                pass

        self.set_context_photo_marker(
            context,
            cluster["lat"],
            cluster["lon"],
            self.get_photo_anchor_icon(),
            command=lambda _marker, ctx=context: self.close_photo_spider_from_anchor(ctx)
        )

        for item, lat, lon in deployed:
            selected = item["photo"].get("catalog_index") == selected_id
            icon = self.get_photo_marker_icon(item["number"], selected=selected)
            self.set_context_photo_marker(
                context,
                lat,
                lon,
                icon,
                command=lambda _marker, p=item["photo"], ctx=context: self.open_photo_from_context(ctx, p)
            )

        return True

    def close_photo_spider(self, context, redraw=True):
        if self.photo_spider_state.get(context) is None:
            return

        self.photo_spider_state[context] = None
        if redraw:
            self.refresh_photo_layer(context, force=True)

    def photo_layer_signature(self, context):
        widget = self.get_photo_map_widget(context)
        if not widget:
            return None

        try:
            width = int(widget.winfo_width())
            height = int(widget.winfo_height())
        except Exception:
            width = height = 0

        return (
            self.get_map_zoom_value(widget),
            width,
            height,
            self.get_photo_layer_mode(context)
        )

    def start_photo_layer_watch(self, context):
        self.stop_photo_layer_watch(context)

        def tick():
            widget = self.get_photo_map_widget(context)
            if not widget:
                self.photo_layer_after_id[context] = None
                return

            signature = self.photo_layer_signature(context)
            if signature != self.photo_layer_last_signature.get(context):
                old = self.photo_layer_last_signature.get(context)
                self.photo_layer_last_signature[context] = signature

                if old and signature and old[0] != signature[0]:
                    self.photo_spider_state[context] = None

                self.refresh_photo_layer(context, force=True)

            try:
                self.photo_layer_after_id[context] = self.root.after(240, tick)
            except Exception:
                self.photo_layer_after_id[context] = None

        try:
            self.photo_layer_after_id[context] = self.root.after(120, tick)
        except Exception:
            self.photo_layer_after_id[context] = None

    def stop_photo_layer_watch(self, context):
        after_id = self.photo_layer_after_id.get(context)
        if after_id:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass
        self.photo_layer_after_id[context] = None
        self.photo_layer_last_signature[context] = None

    def is_coordinate_visible_on_map(self, widget, lat, lon, margin_px=55):
        if not widget or lat is None or lon is None:
            return False

        try:
            zoom = self.get_map_zoom_value(widget) or 17
            center_lat, center_lon = widget.get_position()
            center_x, center_y = self.latlon_to_world_pixel(center_lat, center_lon, zoom)
            point_x, point_y = self.latlon_to_world_pixel(lat, lon, zoom)
            half_w = max(1, widget.winfo_width() / 2 - margin_px)
            half_h = max(1, widget.winfo_height() / 2 - margin_px)
            return abs(point_x - center_x) <= half_w and abs(point_y - center_y) <= half_h
        except Exception:
            return False

    def center_map_on_photo_if_needed(self, widget, lat, lon):
        if not widget or lat is None or lon is None:
            return

        if not self.is_coordinate_visible_on_map(widget, lat, lon):
            try:
                widget.set_position(lat, lon)
            except Exception:
                pass

    def refresh_map_markers(self):
        self.refresh_photo_layer("photo", force=True)

    def fit_map_to_photos(self, photos):
        if not photos or not self.map_widget:
            return

        lats = [p["lat"] for p in photos]
        lons = [p["lon"] for p in photos]

        min_lat = min(lats)
        max_lat = max(lats)
        min_lon = min(lons)
        max_lon = max(lons)

        center_lat = (min_lat + max_lat) / 2
        center_lon = (min_lon + max_lon) / 2

        self.map_widget.set_position(center_lat, center_lon)

        span = max(max_lat - min_lat, max_lon - min_lon)

        if span < 0.002:
            zoom = 18
        elif span < 0.005:
            zoom = 17
        elif span < 0.01:
            zoom = 16
        elif span < 0.03:
            zoom = 15
        elif span < 0.08:
            zoom = 14
        elif span < 0.15:
            zoom = 13
        elif span < 0.3:
            zoom = 12
        else:
            zoom = 11

        self.map_widget.set_zoom(zoom)

    def select_photo_on_map(self, photo, center_map=True):
        if self.current_photo is not None and self.photo_title_entry is not None:
            self.save_current_photo_metadata(silent=True)

        self.current_photo = photo
        self.refresh_map_markers()

        current_index = self.geolocated_photos.index(photo) + 1
        total = len(self.geolocated_photos)

        filename = photo.get("filename", "")
        original_filename = photo.get("original_filename", "")
        title = photo.get("title", "")
        description = photo.get("description", "")
        date_taken = photo.get("date_taken", "")
        metadata_source = photo.get("metadata_source", "")
        gps_source = photo.get("gps_source", "")
        lat = photo.get("lat")
        lon = photo.get("lon")
        ele = photo.get("ele")
        image_path = photo.get("image_path", "")

        self.photo_index_var.set(f"Photo {current_index} / {total}")

        filename_lines = []
        if filename:
            filename_lines.append(f"Fichier : {filename}")
        if original_filename and original_filename != filename:
            filename_lines.append(f"Original : {original_filename}")
        if metadata_source:
            filename_lines.append(f"Métadonnées texte lues depuis : {metadata_source}")

        self.photo_filename_var.set("\n".join(filename_lines))

        meta_lines = []

        if date_taken:
            meta_lines.append(f"Date : {date_taken}")

        if gps_source:
            meta_lines.append(f"Source GPS : {gps_source}")

        if ele is not None:
            try:
                meta_lines.append(f"Altitude : {float(ele):.1f} m")
            except Exception:
                meta_lines.append(f"Altitude : {ele}")

        self.photo_meta_var.set("\n".join(meta_lines))

        if lat is not None and lon is not None:
            self.photo_coords_var.set(f"Coordonnées :\n{lat:.6f}, {lon:.6f}")
        else:
            self.photo_coords_var.set("")

        if hasattr(self, "photo_technical_summary_var"):
            summary_parts = [f"Photo {current_index} / {total}"]
            if date_taken:
                try:
                    # Affichage court si date ISO complète.
                    summary_parts.append(str(date_taken).split("T")[0])
                except Exception:
                    summary_parts.append(str(date_taken))
            summary_parts.append("GPS OK" if lat is not None and lon is not None else "GPS —")
            if ele is not None:
                try:
                    summary_parts.append(f"{float(ele):.0f} m")
                except Exception:
                    pass
            self.photo_technical_summary_var.set(" · ".join(summary_parts))

        self.photo_title_entry.delete(0, tk.END)
        self.photo_title_entry.insert(0, title)

        self.photo_desc_text.delete("1.0", tk.END)
        self.photo_desc_text.insert("1.0", description)

        self.platform_selected_var.set(bool(photo.get("platform_selected", False)))
        try:
            self.platform_order_var.set(int(photo.get("platform_order") or 0))
        except Exception:
            self.platform_order_var.set(0)

        self.photo_status_var.set("")

        self.load_photo_in_viewer(image_path)

        if center_map:
            self.center_map_on_photo_if_needed(self.map_widget, lat, lon)

    def load_photo_in_viewer(self, image_path):
        self.viewer_original_image = None
        self.viewer_display_image = None

        if not os.path.exists(image_path):
            self.viewer_info_var.set("Image introuvable")
            if self.viewer_canvas:
                self.viewer_canvas.delete("all")
            return

        try:
            img = Image.open(image_path)
            img = ImageOps.exif_transpose(img)
            self.viewer_original_image = img.copy()

            filename = os.path.basename(image_path)
            w, h = self.viewer_original_image.size
            self.viewer_info_var.set(f"{filename} · {w} × {h}px")

            self.viewer_canvas.after(100, self.viewer_fit_to_panel)

        except Exception as e:
            self.viewer_info_var.set(f"Aperçu impossible : {e}")
            if self.viewer_canvas:
                self.viewer_canvas.delete("all")

    def viewer_render_image(self):
        if self.viewer_original_image is None or self.viewer_canvas is None:
            return

        original_w, original_h = self.viewer_original_image.size

        new_w = max(1, int(original_w * self.viewer_zoom))
        new_h = max(1, int(original_h * self.viewer_zoom))

        img = self.viewer_original_image.resize(
            (new_w, new_h),
            Image.Resampling.LANCZOS
        )

        self.viewer_display_image = ImageTk.PhotoImage(img)

        self.viewer_canvas.delete("all")
        self.viewer_canvas.create_image(0, 0, anchor="nw", image=self.viewer_display_image)
        self.viewer_canvas.config(scrollregion=(0, 0, new_w, new_h))
        self.viewer_zoom_var.set(f"Zoom : {self.viewer_zoom * 100:.0f} %")

    def viewer_change_zoom(self, factor):
        if self.viewer_original_image is None:
            return

        self.viewer_zoom *= factor
        self.viewer_zoom = max(0.05, min(8.0, self.viewer_zoom))
        self.viewer_render_image()

    def viewer_fit_to_panel(self):
        if self.viewer_original_image is None or self.viewer_canvas is None:
            return

        canvas_w = self.viewer_canvas.winfo_width()
        canvas_h = self.viewer_canvas.winfo_height()

        if canvas_w <= 1 or canvas_h <= 1:
            return

        original_w, original_h = self.viewer_original_image.size

        self.viewer_zoom = min(canvas_w / original_w, canvas_h / original_h)

        if self.viewer_zoom <= 0:
            self.viewer_zoom = 1.0

        self.viewer_render_image()

    def viewer_actual_size(self):
        if self.viewer_original_image is None:
            return

        self.viewer_zoom = 1.0
        self.viewer_render_image()

    def viewer_mousewheel(self, event):
        if self.viewer_canvas is None:
            return

        if event.state & 0x0004:
            if event.delta > 0:
                self.viewer_change_zoom(1.15)
            else:
                self.viewer_change_zoom(0.87)
        else:
            self.viewer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def viewer_start_pan(self, event):
        if self.viewer_canvas:
            self.viewer_canvas.scan_mark(event.x, event.y)

    def viewer_do_pan(self, event):
        if self.viewer_canvas:
            self.viewer_canvas.scan_dragto(event.x, event.y, gain=1)

    # ============================================================
    # MÉTADONNÉES TEXTE JPG
    # ============================================================

    def _value_to_bytes(self, value):
        if value is None:
            return b""
        if isinstance(value, bytes):
            return value
        if isinstance(value, bytearray):
            return bytes(value)
        if isinstance(value, (tuple, list)):
            try:
                return bytes(value)
            except Exception:
                return b""
        if isinstance(value, str):
            return value.encode("utf-8", errors="replace")
        return b""

    def decode_windows_xp_field(self, value):
        raw = self._value_to_bytes(value)

        if not raw:
            return ""

        try:
            text = raw.decode("utf-16le", errors="ignore")
            return text.replace("\x00", "").strip()
        except Exception:
            try:
                return raw.decode("utf-8", errors="ignore").replace("\x00", "").strip()
            except Exception:
                return ""

    def encode_windows_xp_field(self, text):
        if not text:
            return b""
        return text.encode("utf-16le") + b"\x00\x00"

    def decode_exif_text_field(self, value):
        if value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        raw = self._value_to_bytes(value)

        if not raw:
            return ""

        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                return raw.decode(enc, errors="ignore").replace("\x00", "").strip()
            except Exception:
                pass

        return ""

    def decode_user_comment(self, value):
        raw = self._value_to_bytes(value)

        if not raw:
            return ""

        try:
            if raw.startswith(b"UNICODE\x00"):
                text = raw[8:].decode("utf-16be", errors="ignore")
                return text.replace("\x00", "").strip()

            if raw.startswith(b"ASCII\x00\x00\x00"):
                text = raw[8:].decode("ascii", errors="ignore")
                return text.replace("\x00", "").strip()

            for enc in ("utf-8", "latin-1", "cp1252"):
                try:
                    text = raw.decode(enc, errors="ignore")
                    return text.replace("\x00", "").strip()
                except Exception:
                    pass

            return ""

        except Exception:
            return ""

    def read_text_metadata_from_jpg(self, image_path):
        result = {
            "title": "",
            "description": "",
            "details": {},
            "ok": False,
            "error": ""
        }

        if not image_path or not os.path.exists(image_path):
            result["error"] = "Image introuvable"
            return result

        if not image_path.lower().endswith((".jpg", ".jpeg")):
            result["error"] = "Le fichier n'est pas un JPG"
            return result

        try:
            exif_dict = piexif.load(image_path)

            zeroth = exif_dict.get("0th", {})
            exif = exif_dict.get("Exif", {})

            xp_title = self.decode_windows_xp_field(zeroth.get(piexif.ImageIFD.XPTitle))
            xp_comment = self.decode_windows_xp_field(zeroth.get(piexif.ImageIFD.XPComment))
            image_description = self.decode_exif_text_field(zeroth.get(piexif.ImageIFD.ImageDescription))

            xp_subject = ""
            if hasattr(piexif.ImageIFD, "XPSubject"):
                xp_subject = self.decode_windows_xp_field(zeroth.get(piexif.ImageIFD.XPSubject))

            user_comment = self.decode_user_comment(exif.get(piexif.ExifIFD.UserComment))

            details = {
                "XPTitle": xp_title,
                "XPComment": xp_comment,
                "ImageDescription": image_description,
                "XPSubject": xp_subject,
                "UserComment": user_comment
            }

            title = xp_title or image_description or xp_subject
            description = xp_comment or user_comment

            result["title"] = title
            result["description"] = description
            result["details"] = details
            result["ok"] = True
            return result

        except Exception as e:
            result["error"] = str(e)
            return result

    def write_text_metadata_to_jpg(self, image_path, title, description):
        if not image_path or not os.path.exists(image_path):
            return False, "Image introuvable"

        if not image_path.lower().endswith((".jpg", ".jpeg")):
            return False, "Le fichier n'est pas un JPG"

        try:
            try:
                exif_dict = piexif.load(image_path)
            except Exception:
                exif_dict = {
                    "0th": {},
                    "Exif": {},
                    "GPS": {},
                    "1st": {},
                    "thumbnail": None
                }

            title = title or ""
            description = description or ""

            if title:
                exif_dict["0th"][piexif.ImageIFD.XPTitle] = self.encode_windows_xp_field(title)
            else:
                exif_dict["0th"].pop(piexif.ImageIFD.XPTitle, None)

            if description:
                exif_dict["0th"][piexif.ImageIFD.XPComment] = self.encode_windows_xp_field(description)
            else:
                exif_dict["0th"].pop(piexif.ImageIFD.XPComment, None)

            if description:
                user_comment_prefix = b"UNICODE\x00"
                user_comment_text = description.encode("utf-16be", errors="replace")
                exif_dict["Exif"][piexif.ExifIFD.UserComment] = user_comment_prefix + user_comment_text
            else:
                exif_dict["Exif"].pop(piexif.ExifIFD.UserComment, None)

            exif_bytes = piexif.dump(exif_dict)
            piexif.insert(exif_bytes, image_path)

            return True, "Métadonnées JPG écrites"

        except Exception as e:
            return False, str(e)

    def diagnose_current_photo_metadata(self):
        if self.current_photo is None:
            messagebox.showwarning("Aucune photo", "Aucune photo sélectionnée.")
            return

        image_path = self.current_photo.get("image_path", "")

        jpg_meta = self.read_text_metadata_from_jpg(image_path)
        details = jpg_meta.get("details", {})
        gps_meta = self.read_gps_metadata_from_jpg(image_path)

        self.log("-" * 60)
        self.log(f"🔎 Diagnostic métadonnées : {self.current_photo.get('filename')}")
        self.log(f"📁 Fichier : {image_path}")

        if not jpg_meta.get("ok"):
            self.log(f"❌ Lecture texte JPG impossible : {jpg_meta.get('error')}")
        else:
            self.log("📷 Champs texte lus dans le JPG avec piexif :")
            self.log(f"   XPTitle          : {details.get('XPTitle', '')}")
            self.log(f"   XPComment        : {details.get('XPComment', '')}")
            self.log(f"   ImageDescription : {details.get('ImageDescription', '')}")
            self.log(f"   XPSubject        : {details.get('XPSubject', '')}")
            self.log(f"   UserComment      : {details.get('UserComment', '')}")
            self.log(f"   → Titre retenu       : {jpg_meta.get('title', '')}")
            self.log(f"   → Description retenue: {jpg_meta.get('description', '')}")

        if gps_meta.get("ok"):
            self.log("📍 GPS EXIF lu dans le JPG :")
            self.log(f"   Latitude  : {gps_meta.get('lat')}")
            self.log(f"   Longitude : {gps_meta.get('lon')}")
            self.log(f"   Altitude  : {gps_meta.get('ele')}")
        else:
            self.log(f"📍 GPS EXIF : aucun / illisible ({gps_meta.get('error')})")

        self.log("📒 Valeurs catalogue.json / interface :")
        self.log(f"   catalog_title       : {self.current_photo.get('catalog_title', '')}")
        self.log(f"   catalog_description : {self.current_photo.get('catalog_description', '')}")
        self.log(f"   affiché title       : {self.current_photo.get('title', '')}")
        self.log(f"   affiché description : {self.current_photo.get('description', '')}")
        self.log(f"   source affichage    : {self.current_photo.get('metadata_source', '')}")
        self.log(f"   source GPS          : {self.current_photo.get('gps_source', '')}")
        self.log("-" * 60)

        self.photo_status_var.set(
            "🔎 Diagnostic écrit dans le journal."
        )

    def save_current_photo_metadata(self, silent=False):
        if self.current_photo is None:
            if not silent:
                messagebox.showwarning("Aucune photo", "Aucune photo sélectionnée.")
            return

        if self.photo_title_entry is None or self.photo_desc_text is None:
            return

        catalog_index = self.current_photo.get("catalog_index")

        if catalog_index is None:
            if not silent:
                messagebox.showerror("Erreur", "Index catalogue introuvable.")
            return

        requested_title = self.photo_title_entry.get().strip()
        requested_description = self.photo_desc_text.get("1.0", tk.END).strip()

        try:
            image_path = self.current_photo.get("image_path", "")
            jpg_ok, jpg_message = self.write_text_metadata_to_jpg(
                image_path,
                requested_title,
                requested_description
            )

            jpg_meta = self.read_text_metadata_from_jpg(image_path)

            if jpg_meta.get("ok"):
                actual_title = jpg_meta.get("title", "")
                actual_description = jpg_meta.get("description", "")
            else:
                actual_title = requested_title
                actual_description = requested_description

            self.catalog_data = self.read_catalog()

            if catalog_index >= len(self.catalog_data):
                if not silent:
                    messagebox.showerror("Erreur", "Index catalogue invalide.")
                return

            platform_selected = bool(self.platform_selected_var.get())
            try:
                platform_order = int(self.platform_order_var.get())
            except Exception:
                platform_order = 0

            self.catalog_data[catalog_index]["title"] = actual_title
            self.catalog_data[catalog_index]["description"] = actual_description
            self.catalog_data[catalog_index]["platform_selected"] = platform_selected
            self.catalog_data[catalog_index]["platform_order"] = platform_order

            self.save_catalog()

            self.current_photo["title"] = actual_title
            self.current_photo["description"] = actual_description
            self.current_photo["jpg_title"] = jpg_meta.get("title", "")
            self.current_photo["jpg_description"] = jpg_meta.get("description", "")
            self.current_photo["catalog_title"] = actual_title
            self.current_photo["catalog_description"] = actual_description
            self.current_photo["metadata_source"] = "JPG" if jpg_ok else "catalogue.json"
            self.current_photo["platform_selected"] = platform_selected
            self.current_photo["platform_order"] = platform_order

            for photo in self.geolocated_photos:
                if photo.get("catalog_index") == catalog_index:
                    photo["title"] = actual_title
                    photo["description"] = actual_description
                    photo["jpg_title"] = jpg_meta.get("title", "")
                    photo["jpg_description"] = jpg_meta.get("description", "")
                    photo["catalog_title"] = actual_title
                    photo["catalog_description"] = actual_description
                    photo["metadata_source"] = "JPG" if jpg_ok else "catalogue.json"
                    photo["platform_selected"] = platform_selected
                    photo["platform_order"] = platform_order
                    break

            self.photo_title_entry.delete(0, tk.END)
            self.photo_title_entry.insert(0, actual_title)

            self.photo_desc_text.delete("1.0", tk.END)
            self.photo_desc_text.insert("1.0", actual_description)

            if not silent:
                if jpg_ok and jpg_meta.get("ok"):
                    self.photo_status_var.set(
                        "✅ Sauvegardé dans le JPG, relu depuis le JPG, puis synchronisé avec le catalogue."
                    )
                    self.log(f"💾 Métadonnées JPG + catalogue sauvegardées : {self.current_photo.get('filename')}")
                elif jpg_ok:
                    self.photo_status_var.set(
                        "⚠️ JPG écrit, mais relecture JPG incomplète. Catalogue mis à jour avec les valeurs demandées."
                    )
                    self.log(f"⚠️ JPG écrit mais relecture incomplète : {self.current_photo.get('filename')}")
                else:
                    self.photo_status_var.set(
                        f"⚠️ Catalogue sauvegardé, mais JPG non modifié : {jpg_message}"
                    )
                    self.log(
                        f"⚠️ Métadonnées catalogue OK, JPG non modifié pour {self.current_photo.get('filename')} : {jpg_message}"
                    )

        except Exception as e:
            if not silent:
                messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur sauvegarde métadonnées : {e}")

    def discard_current_photo(self):
        """
        Écarte la photo actuellement affichée.

        Sécurité :
        - ne supprime pas définitivement ;
        - déplace le fichier JPG dans Photos_supprimees/ ;
        - marque l'entrée du catalogue comme SUPPRIMEE ;
        - retire la photo de la carte en cours.
        """
        if self.current_photo is None:
            messagebox.showwarning("Aucune photo", "Aucune photo sélectionnée.")
            return

        catalog_index = self.current_photo.get("catalog_index")

        if catalog_index is None:
            messagebox.showerror("Erreur", "Index catalogue introuvable.")
            return

        image_path = self.current_photo.get("image_path", "")

        if not image_path or not os.path.exists(image_path):
            messagebox.showerror(
                "Erreur",
                f"Fichier introuvable :\n{image_path}"
            )
            return

        filename = os.path.basename(image_path)

        if not messagebox.askyesno(
            "Écarter cette photo ?",
            (
                f"Écarter cette photo du corpus ?\n\n"
                f"{filename}\n\n"
                "Le fichier ne sera pas supprimé définitivement.\n"
                "Il sera déplacé dans un dossier Photos_supprimees."
            )
        ):
            return

        try:
            # Sauvegarde silencieuse des métadonnées en cours avant déplacement.
            self.save_current_photo_metadata(silent=True)

            self.catalog_data = self.read_catalog()

            if catalog_index >= len(self.catalog_data):
                messagebox.showerror("Erreur", "Index catalogue invalide.")
                return

            trash_folder = os.path.join(self.base_folder, "Photos_supprimees")
            os.makedirs(trash_folder, exist_ok=True)

            target_path = os.path.join(trash_folder, filename)

            # Évite d'écraser une photo déjà écartée portant le même nom.
            if os.path.exists(target_path):
                base, ext = os.path.splitext(filename)
                counter = 2

                while True:
                    candidate = os.path.join(
                        trash_folder,
                        f"{base}_supprimee_{counter}{ext}"
                    )
                    if not os.path.exists(candidate):
                        target_path = candidate
                        break
                    counter += 1

            os.rename(image_path, target_path)

            entry = self.catalog_data[catalog_index]

            entry["status"] = "SUPPRIMEE"
            entry["discarded"] = True
            entry["discard_date"] = datetime.now().isoformat(timespec="seconds")
            entry["discarded_from_relative_path"] = self.relative_to_base(image_path)
            entry["discarded_to_relative_path"] = self.relative_to_base(target_path)

            # On garde les métadonnées, mais la photo n'apparaîtra plus dans la carte,
            # car load_geolocated_photos ignore les entrées dont status != OK.
            self.save_catalog()

            self.log(f"🗑️ Photo écartée : {filename} -> {self.relative_to_base(target_path)}")

            old_photo = self.current_photo
            old_index = self.geolocated_photos.index(old_photo)

            self.geolocated_photos = [
                p for p in self.geolocated_photos
                if p.get("catalog_index") != catalog_index
            ]

            self.current_photo = None

            if not self.geolocated_photos:
                messagebox.showinfo(
                    "Tri terminé",
                    "Il n'y a plus de photos géolocalisées à afficher."
                )
                self.load_folder(self.base_folder)
                return

            new_index = min(old_index, len(self.geolocated_photos) - 1)

            self.refresh_map_markers()
            self.select_photo_on_map(self.geolocated_photos[new_index])

            self.photo_status_var.set(
                f"🗑️ Photo écartée dans Photos_supprimees : {filename}"
            )

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log(f"❌ Erreur lors de l'écartement de la photo : {e}")

    def center_on_current_photo(self):
        if self.current_photo is None or not self.map_widget:
            return

        lat = self.current_photo.get("lat")
        lon = self.current_photo.get("lon")

        if lat is not None and lon is not None:
            self.map_widget.set_position(lat, lon)
            self.map_widget.set_zoom(17)


    def show_current_photo_technical_details(self):
        """
        Ouvre une petite fenêtre avec les détails techniques de la photo active.
        Ces informations sont utiles ponctuellement, mais ne doivent pas encombrer
        le panneau d'édition.
        """
        if not self.current_photo:
            messagebox.showwarning("Aucune photo", "Aucune photo n'est sélectionnée.")
            return

        filename = self.photo_filename_var.get().strip()
        meta = self.photo_meta_var.get().strip()
        coords = self.photo_coords_var.get().strip()

        lines = []
        if filename:
            lines.append(filename)
        if meta:
            if lines:
                lines.append("")
            lines.append(meta)
        if coords:
            if lines:
                lines.append("")
            lines.append(coords)

        if not lines:
            lines = ["Aucun détail technique disponible pour cette photo."]

        window = tk.Toplevel(self.root)
        window.title("Détails techniques de la photo")
        window.geometry("520x360")
        window.minsize(420, 260)
        window.transient(self.root)

        frame = tk.Frame(window, padx=14, pady=12)
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        tk.Label(
            frame,
            text="Détails techniques de la photo",
            font=("Arial", 13, "bold"),
            anchor="w"
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

        text_box = tk.Text(frame, wrap="word", height=10)
        text_box.grid(row=1, column=0, sticky="nsew")
        text_box.insert("1.0", "\n".join(lines))
        text_box.config(state="disabled")

        tk.Button(
            frame,
            text="Fermer",
            command=window.destroy
        ).grid(row=2, column=0, sticky="e", pady=(10, 0))

    def select_previous_photo(self):
        if not self.geolocated_photos or self.current_photo is None:
            return

        current_index = self.geolocated_photos.index(self.current_photo)
        new_index = current_index - 1

        if new_index < 0:
            new_index = len(self.geolocated_photos) - 1

        self.select_photo_on_map(self.geolocated_photos[new_index])

    def select_next_photo(self):
        if not self.geolocated_photos or self.current_photo is None:
            return

        current_index = self.geolocated_photos.index(self.current_photo)
        new_index = current_index + 1

        if new_index >= len(self.geolocated_photos):
            new_index = 0

        self.select_photo_on_map(self.geolocated_photos[new_index])


if __name__ == "__main__":
    root = tk.Tk()
    app = BisseManagerApp(root)
    root.mainloop()
